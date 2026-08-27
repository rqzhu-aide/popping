import os
import csv
import copy
import gzip
import hashlib
import hmac
import io
import json
import math
import mimetypes
import re
import sqlite3
import threading
import time
import uuid
from datetime import datetime, timedelta, timezone
from functools import wraps
from fractions import Fraction
from pathlib import Path
from urllib.parse import urlsplit

from flask import (
    Flask, render_template, request, redirect,
    url_for, session, jsonify, flash, g, make_response, send_file
)
from werkzeug.exceptions import HTTPException, RequestEntityTooLarge
from werkzeug.middleware.proxy_fix import ProxyFix
import yaml

import config
from database import (
    ensure_schema,
    execute_db,
    forget_schema,
    get_db,
    get_max_members_per_team,
    get_max_teams,
    init_app,
    init_db,
    validate_current_schema,
    validate_legacy_adoption_candidate,
    inspect_schema_version,
    is_sqlite_busy_error,
    query_db,
    SQLITE_BUSY_RETRY_AFTER_SECONDS,
)
from pin_policy import is_valid_instructor_pin
from demo_instance import (
    DemoLifecycleBusy,
    DemoResetCooldown,
    canonical_class_slug,
    course_class_dir,
    create_bounded_demo_instance,
    is_demo_instance_slug,
    reset_demo_instance,
    touch_demo_instance,
)
from question_catalog import (
    QuestionParseError,
    parse_question_blocks,
    parse_week_questions,
    read_week_questions,
    validate_question_catalog,
)
from versioning import (
    APP_VERSION,
    BASELINE_DATA_VERSION,
    BASELINE_SCHEMA_VERSION,
    EXPORT_FORMAT_VERSION,
    SCHEMA_VERSION,
    compatibility_label,
    compatible_series,
    parse_version,
    public_version,
)

mimetypes.add_type('image/webp', '.webp')

app = Flask(__name__)
app.config.from_object(config)
if os.environ.get('RENDER'):
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
init_app(app)
app.jinja_env.globals['app_version'] = public_version(APP_VERSION)


class CourseTemporarilyUnavailable(Exception):
    """A short-lived storage failure that must not invalidate a login."""


@app.errorhandler(RequestEntityTooLarge)
def request_too_large(_error):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Request is too large'}), 413
    return 'Request is too large', 413


@app.errorhandler(CourseTemporarilyUnavailable)
def course_temporarily_unavailable(_error):
    message = 'Course data is temporarily unavailable. Please try again.'
    if request.path.startswith('/api/'):
        response = jsonify({'error': message})
    else:
        response = make_response(message)
    response.status_code = 503
    response.headers['Retry-After'] = '5'
    return response


@app.errorhandler(sqlite3.OperationalError)
def handle_sqlite_operational_error(error):
    """Return a recoverable response when SQLite remains write-locked."""
    if not is_sqlite_busy_error(error):
        return handle_unexpected_error(error)
    message = 'The class database is briefly busy. Please try again.'
    app.logger.warning(
        'Temporary database contention on %s %s',
        request.method, request.path,
    )
    if request.path.startswith('/api/'):
        response = jsonify({
            'error': message,
            'retry_after': SQLITE_BUSY_RETRY_AFTER_SECONDS,
        })
    else:
        response = make_response(message)
    response.status_code = 503
    response.headers['Retry-After'] = str(SQLITE_BUSY_RETRY_AFTER_SECONDS)
    return response


@app.errorhandler(HTTPException)
def handle_http_exception(error):
    """HTTP errors (404, 403, 405 …): JSON for API routes, default page otherwise."""
    if request.path.startswith('/api/'):
        return jsonify({'error': error.name or 'Request failed'}), error.code
    return error.get_response()


@app.errorhandler(Exception)
def handle_unexpected_error(error):
    """Genuine unhandled exception: log server-side, return a safe message."""
    app.logger.exception('Unhandled error on %s %s', request.method, request.path)
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Something went wrong. Please try again.'}), 500
    return 'Internal Server Error', 500

PHASES = ['setup', 'discussion', 'competition', 'ended']

PHASE_LABELS = {
    'setup': 'Setup',
    'discussion': 'Group Discussion',
    'competition': 'Present and Challenge',
    'ended': 'End Session'
}

# Duration (seconds) of the presentation rating window.
POLL_DURATION = 40
# This fixed interval drains requests that normally reached the server before
# a close cutoff. A severe database stall still returns a retryable error
# instead of allowing a write to appear after an instructor transition.
POLL_SUBMISSION_GRACE_SECONDS = 3
POLL_SETTLING_MESSAGE = (
    'Final ratings are still being saved. Try again in a few seconds.'
)
STUDENT_ONLINE_WINDOW = timedelta(minutes=1)

# mtime-keyed cache of per-course poll_duration, so the ~1s poll loop doesn't
# re-parse course.yaml on every request.
_poll_duration_cache = {}


def get_poll_duration(slug):
    """Rating-window length (seconds) for this course.

    Reads the optional ``poll_duration`` field from course.yaml (clamped to
    5-300s; anything missing or invalid falls back to POLL_DURATION). Cached by
    the file's mtime so frequent polls don't re-parse YAML.
    """
    yaml_path = os.path.join(_course_class_dir(slug), 'course.yaml')
    try:
        mtime = os.path.getmtime(yaml_path)
    except OSError:
        return POLL_DURATION
    cached = _poll_duration_cache.get(slug)
    if cached and cached[0] == mtime:
        return cached[1]
    value = POLL_DURATION
    try:
        with open(yaml_path, encoding='utf-8') as config_file:
            cfg = yaml.safe_load(config_file) or {}
        raw = cfg.get('poll_duration')
        if raw is not None:
            value = int(raw)
    except Exception:
        value = POLL_DURATION
    if not (5 <= value <= 300):
        value = POLL_DURATION
    if len(_poll_duration_cache) >= POLL_DURATION_CACHE_LIMIT:
        _poll_duration_cache.clear()
    _poll_duration_cache[slug] = (mtime, value)
    return value

# Compress JSON responses large enough to benefit. Poll responses are the
# dominant classroom traffic; gzip reduces repeated keys and question HTML
# substantially without adding a third-party dependency.
JSON_COMPRESSION_MIN_BYTES = 500
LOGIN_FAILURE_LIMIT = 3
LOGIN_CLIENT_FAILURE_LIMIT = 300
LOGIN_WINDOW_SECONDS = 60
MAX_ROSTER_BYTES = 1024 * 1024
MAX_ROSTER_ROWS = 500
MAX_EXPORT_ROWS = 100000
MAX_EXPORT_BYTES = 50 * 1024 * 1024
COURSE_AVAILABILITY_TTL = 30
COURSE_UNAVAILABLE_TTL = 5
# Students and instructors in a live demo poll /api/poll continuously, so the
# .last-used TTL marker is refreshed at most once per minute per instance
# instead of on every request.
DEMO_TOUCH_INTERVAL = 60
DEMO_TOUCH_FAILURE_RETRY_INTERVAL = 5
DEMO_TOUCH_CACHE_LIMIT = 128
# Unauthenticated routes key these caches by attacker-controlled slugs, so
# cap them like the demo touch cache: clear once the limit is reached.
POLL_DURATION_CACHE_LIMIT = 256
COURSE_AVAILABILITY_CACHE_LIMIT = 256
_SPREADSHEET_FORMULA_PREFIXES = ('=', '+', '-', '@')
REQUIRED_COURSE_SCHEMA = {
    'instructors': {'id', 'username', 'name', 'pin'},
    'courses': {
        'id', 'name', 'code', 'semester', 'slug', 'instructor_id', 'is_active'
    },
    'teams': {'id', 'course_id', 'name', 'color'},
    'students': {
        'id', 'course_id', 'student_id', 'name', 'pin', 'team_id'
    },
    'questions': {'id', 'course_id', 'question_num', 'question_text'},
    'course_state': {
        'id', 'course_id', 'phase', 'active_team_id', 'active_question_id',
        'current_question', 'presentation_started_at'
    },
}
REQUIRED_COURSE_TABLES = frozenset(REQUIRED_COURSE_SCHEMA)

_course_availability_cache = {}
_course_availability_lock = threading.RLock()
_demo_instance_touch_lock = threading.Lock()
_demo_instance_touch_last = {}
_demo_instance_touch_failed = {}
_demo_instance_touch_inflight = set()
_instructor_catalog_sync_lock = threading.Lock()
_instructor_catalog_sync_attempted = set()


def _spreadsheet_safe_value(value):
    """Neutralize text that spreadsheet programs could execute as a formula."""
    if not isinstance(value, str):
        return value
    candidate = value.lstrip(' \t\r\n')
    if candidate.startswith(_SPREADSHEET_FORMULA_PREFIXES):
        return "'" + value
    return value


def _spreadsheet_safe_row(values):
    return [_spreadsheet_safe_value(value) for value in values]


def _export_phase_error(db, course_id):
    state = db.execute(
        'SELECT phase FROM course_state WHERE course_id = ?',
        [course_id],
    ).fetchone()
    if not state or state['phase'] not in ('setup', 'ended'):
        return 'Data can only be exported during Setup or after End Session.'
    return None


def is_valid_slug(slug):
    return isinstance(slug, str) and re.fullmatch(r'[A-Za-z0-9_-]+', slug)


def _parse_db_datetime(dt_str):
    """Parse a SQLite datetime string (space or T separator) to a UTC datetime."""
    if not dt_str:
        return None
    return datetime.fromisoformat(str(dt_str).replace(' ', 'T'))


def _utcnow():
    """Current UTC time as a naive datetime.

    Stored timestamps (SQLite CURRENT_TIMESTAMP, session ISO strings) are
    naive UTC, so keeping ``now`` naive preserves comparison behavior while
    replacing the deprecated ``datetime.utcnow()``.
    """
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _student_is_online(last_active_at, now=None):
    """Return whether a saved activity time is inside the presence window."""
    if not last_active_at:
        return False
    checked_at = now if now is not None else _utcnow()
    try:
        last_seen = _parse_db_datetime(last_active_at)
        age = checked_at - last_seen
        return timedelta(0) <= age < STUDENT_ONLINE_WINDOW
    except (TypeError, ValueError):
        return False


def course_db_path(slug):
    if not is_valid_slug(slug):
        return None
    return os.path.join(config.DATA_DIR, slug, 'popping.db')


def _course_class_dir(slug):
    return course_class_dir(config.CLASSES_DIR, slug)


def active_presentation_key(state):
    """Stable key for the current presentation, even if its timer is paused."""
    if not state:
        return None
    if 'poll_question_key' in state.keys() and state['poll_question_key']:
        return state['poll_question_key']
    if 'presentation_started_at' in state.keys() and state['presentation_started_at']:
        return f"pres-{state['presentation_started_at']}"
    return None


def _history_presentation_key(item):
    """Return the stable presentation key stored by one history item."""
    if not isinstance(item, dict):
        return ''
    return (
        item.get('presentation_key')
        or f"pres-{item.get('started_at', '')}"
    )


def _history_data_version(item):
    """Return stored history provenance, adopting only a missing v1 baseline."""
    if not isinstance(item, dict):
        return None
    value = item.get('data_version')
    return BASELINE_DATA_VERSION if value is None else value


def _data_version_is_compatible(value):
    """Fail closed unless a stored data version matches this schema series."""
    try:
        return (
            compatible_series(value) == compatible_series(SCHEMA_VERSION)
        )
    except (TypeError, ValueError):
        return False


def _history_item_is_compatible(item):
    """Return whether a well-formed history item is live for this schema."""
    version = _history_data_version(item)
    return version is not None and _data_version_is_compatible(version)


def _positive_lecture_week(value):
    """Return a lecture week only for an exact positive integer value."""
    return value if type(value) is int and value > 0 else None


def _rating_integer(value):
    """Parse one rating without truncating floats or accepting booleans."""
    if type(value) is int:
        return value
    if isinstance(value, str) and re.fullmatch(r'[0-9]+', value.strip()):
        return int(value.strip())
    raise ValueError('Rating must be an integer')


def _resolve_history_week(item, question_weeks=None, rating_weeks=None):
    """Resolve a history item's lecture week using current-export inference."""
    if not isinstance(item, dict):
        return None
    week_num = item.get('week_num')
    if week_num is None and question_weeks:
        week_num = question_weeks.get(item.get('question_id'))
    if week_num is None and rating_weeks:
        week_num = rating_weeks.get(_history_presentation_key(item))
    return _positive_lecture_week(week_num)


def _compatible_rating_weeks(db, course_id):
    """Map unambiguous presentation keys to compatible lecture weeks."""
    rows = db.execute(
        '''SELECT question_key, MIN(week_num) AS week_num,
                  COUNT(DISTINCT week_num) AS week_count
           FROM presentation_ratings
           WHERE course_id = ? AND week_num IS NOT NULL
             AND popping_version_compatible(data_version, ?) = 1
           GROUP BY question_key''',
        [course_id, SCHEMA_VERSION],
    ).fetchall()
    return {
        row['question_key']: row['week_num']
        for row in rows if row['week_count'] == 1
    }


def _presentation_guard(data, state):
    """Validate that an instructor action still targets the displayed presentation."""
    expected_key = str(data.get('presentation_key') or '').strip()
    if not expected_key:
        return 'Presentation key is required', 400
    current_key = active_presentation_key(state)
    if (not state or state['phase'] != 'competition' or
            not state['active_team_id'] or not state['active_question_id'] or
            not current_key):
        return 'No active presentation', 409
    if expected_key != current_key:
        return 'This instructor page is stale; reload before controlling the presentation', 409
    return None


def _expected_state_guard(data, state):
    """Reject a state-changing request sent from a stale instructor page."""
    expected_phase = data.get('expected_phase')
    expected_session_key = data.get('expected_session_key')
    if expected_phase is None or expected_session_key is None:
        return 'Expected phase and session key are required', 400
    try:
        expected_session_key = int(expected_session_key)
    except (TypeError, ValueError):
        return 'Invalid session key', 400
    if (not state or expected_phase != state['phase'] or
            expected_session_key != (state['session_key'] or 0)):
        return 'This instructor page is stale; reload before changing the course state', 409
    return None


def _expected_roster_state_guard(data, state):
    """Reject roster changes based on an out-of-date instructor roster."""
    guard = _expected_state_guard(data, state)
    if guard:
        return guard
    expected_version = data.get('expected_roster_version')
    if expected_version is None:
        return 'Expected roster version is required', 400
    try:
        expected_version = int(expected_version)
    except (TypeError, ValueError):
        return 'Invalid roster version', 400
    current_version = state['roster_version'] or 0
    if expected_version != current_version:
        return 'The roster changed in another page; reload before continuing', 409
    return None


def _prune_reset_backups(slug, keep=3):
    """Best-effort pruning of validated pre-reset backups."""
    backup_dir = os.path.join(config.DATA_DIR, slug, 'reset-backups')
    try:
        backups = sorted(
            os.path.join(backup_dir, name)
            for name in os.listdir(backup_dir)
            if (name.startswith('popping-before-reset-') and
                name.endswith('.db'))
        )
    except OSError:
        app.logger.warning(
            'Could not list reset backups for pruning in %s', backup_dir
        )
        return
    for old_path in backups[:-keep]:
        try:
            os.remove(old_path)
        except OSError:
            app.logger.warning('Could not remove old reset backup %s', old_path)


def _create_reset_backup(slug, *, prune=True):
    """Create a consistent SQLite backup immediately before destructive reset."""
    source_path = course_db_path(slug)
    backup_dir = os.path.join(config.DATA_DIR, slug, 'reset-backups')
    os.makedirs(backup_dir, exist_ok=True)
    stamp = _utcnow().strftime('%Y%m%dT%H%M%S%f')
    backup_path = os.path.join(backup_dir, f'popping-before-reset-{stamp}.db')
    source = sqlite3.connect(source_path, timeout=30)
    target = sqlite3.connect(backup_path)
    try:
        source.backup(target)
        integrity = [row[0] for row in target.execute(
            'PRAGMA integrity_check'
        ).fetchall()]
        foreign_keys = target.execute('PRAGMA foreign_key_check').fetchall()
        if integrity != ['ok'] or foreign_keys:
            raise RuntimeError('Reset backup validation failed')
    except Exception:
        target.close()
        source.close()
        try:
            os.remove(backup_path)
        except OSError:
            pass
        raise
    else:
        target.close()
        source.close()
    if prune:
        _prune_reset_backups(slug)
    return os.path.basename(backup_path)


def _discard_reset_backup(slug, backup_name):
    """Remove one backup created by this process after a stale reset abort."""
    if not backup_name or os.path.basename(backup_name) != backup_name:
        return
    backup_path = os.path.join(
        config.DATA_DIR, slug, 'reset-backups', backup_name
    )
    try:
        os.remove(backup_path)
    except FileNotFoundError:
        pass
    except OSError:
        app.logger.warning('Could not remove stale reset backup %s', backup_path)


def _poll_cutoff(state, poll_duration=None):
    """Return the fixed cutoff for the current presentation poll."""
    if not state or not state['poll_started_at']:
        return None
    try:
        started = _parse_db_datetime(state['poll_started_at'])
        natural_cutoff = started + timedelta(
            seconds=poll_duration or POLL_DURATION
        )
        manual_cutoff = _parse_db_datetime(state['poll_closed_at'])
        return min(natural_cutoff, manual_cutoff) if manual_cutoff \
            else natural_cutoff
    except (TypeError, ValueError):
        return None


def _challenge_ratings_cutoff(state, poll_duration=None):
    """Return the earliest current challenge or main-poll cutoff."""
    if not state:
        return None
    try:
        explicit_cutoff = _parse_db_datetime(
            state['challenge_ratings_closed_at']
        )
    except (TypeError, ValueError):
        explicit_cutoff = None
    main_cutoff = _poll_cutoff(state, poll_duration=poll_duration)
    if explicit_cutoff and main_cutoff:
        return min(explicit_cutoff, main_cutoff)
    return explicit_cutoff or main_cutoff


def _challenge_ratings_are_open(state, now=None, poll_duration=None):
    """Return whether active challenge controls should remain available."""
    cutoff = _challenge_ratings_cutoff(
        state, poll_duration=poll_duration
    )
    return cutoff is None or (now or _utcnow()) < cutoff


def _has_active_challenges(state):
    if not state:
        return False
    try:
        challenges = json.loads(state['active_challenges_json'] or '[]')
    except (TypeError, ValueError):
        return False
    return bool(challenges)


def _poll_is_open(state, now=None, poll_duration=None):
    """Return whether the persisted visible rating window is open."""
    if not state or not state['poll_active']:
        return False
    try:
        started = _parse_db_datetime(state['poll_started_at'])
        cutoff = _poll_cutoff(state, poll_duration=poll_duration)
        checked_at = now or _utcnow()
        return bool(started and cutoff and started <= checked_at < cutoff)
    except (TypeError, ValueError):
        return False


def _poll_accepts_rating(state, now=None, poll_duration=None):
    """Accept arrivals within the fixed post-cutoff drain envelope."""
    try:
        started = _parse_db_datetime(state['poll_started_at']) if state else None
        cutoff = _poll_cutoff(state, poll_duration=poll_duration)
        arrived_at = now or _utcnow()
        deadline = cutoff + timedelta(
            seconds=POLL_SUBMISSION_GRACE_SECONDS
        ) if cutoff else None
        return bool(started and deadline and started <= arrived_at < deadline)
    except (TypeError, ValueError):
        return False


def _challenge_ratings_accept(
        state, now=None, poll_duration=None):
    """Accept challenge arrivals within the same fixed drain envelope."""
    if not state:
        return False
    try:
        cutoff = _challenge_ratings_cutoff(
            state, poll_duration=poll_duration
        )
        deadline = cutoff + timedelta(
            seconds=POLL_SUBMISSION_GRACE_SECONDS
        ) if cutoff else None
        return cutoff is None or (now or _utcnow()) < deadline
    except (TypeError, ValueError):
        return False


def _ratings_settling_state(state, now=None, poll_duration=None):
    """Describe the fixed drain window after any current rating cutoff."""
    checked_at = now or _utcnow()
    cutoffs = [_poll_cutoff(state, poll_duration=poll_duration)]
    if _has_active_challenges(state):
        cutoffs.append(_challenge_ratings_cutoff(
            state, poll_duration=poll_duration
        ))

    remaining = 0
    for cutoff in cutoffs:
        if cutoff is None:
            continue
        deadline = cutoff + timedelta(seconds=POLL_SUBMISSION_GRACE_SECONDS)
        if cutoff <= checked_at < deadline:
            remaining = max(
                remaining,
                max(1, math.ceil((deadline - checked_at).total_seconds())),
            )
    return {
        'ratings_settling': remaining > 0,
        'ratings_settling_remaining': remaining,
    }


def _poll_transition_error(
        state, active_message, now=None, poll_duration=None):
    """Return why an instructor action must preserve this rating window."""
    duration = poll_duration or POLL_DURATION
    if _poll_is_open(state, now=now, poll_duration=duration):
        return active_message
    settling = _ratings_settling_state(
        state, now=now, poll_duration=duration
    )
    if settling['ratings_settling']:
        return POLL_SETTLING_MESSAGE
    return None


def _prepare_rating_transition(
        db, state, active_message, now=None, poll_duration=None):
    """Close open challenge ratings and report whether a transition must wait."""
    checked_at = now or _utcnow()
    duration = poll_duration or POLL_DURATION
    if _poll_is_open(state, now=checked_at, poll_duration=duration):
        return ({
            'error': active_message,
            'ratings_settling': False,
            'ratings_settling_remaining': 0,
        }, False)

    changed = False
    if (_has_active_challenges(state) and
            _challenge_ratings_cutoff(
                state, poll_duration=duration
            ) is None):
        db.execute(
            """UPDATE course_state
               SET challenge_ratings_closed_at = ?
               WHERE course_id = ?""",
            [
                checked_at.strftime('%Y-%m-%d %H:%M:%S.%f'),
                state['course_id'],
            ],
        )
        state = db.execute(
            'SELECT * FROM course_state WHERE course_id = ?',
            [state['course_id']],
        ).fetchone()
        changed = True

    settling = _ratings_settling_state(
        state, now=checked_at, poll_duration=duration
    )
    if settling['ratings_settling']:
        return ({'error': POLL_SETTLING_MESSAGE, **settling}, changed)
    return (None, changed)


def _derive_timing_state(state, now=None, poll_duration=None):
    """Return server-authoritative timer values for one shared UTC instant."""
    if not state:
        return {
            'presentation_remaining': None,
            'poll_remaining': 0,
            'session_elapsed': None,
            'ratings_settling': False,
            'ratings_settling_remaining': 0,
        }
    now = now or _utcnow()
    presentation_remaining = state['presentation_remaining']
    if state['presentation_started_at'] and state['presentation_time_cap']:
        try:
            started = _parse_db_datetime(state['presentation_started_at'])
            remaining = (state['presentation_time_cap'] or 300) - (
                now - started
            ).total_seconds()
            presentation_remaining = max(0, math.ceil(remaining))
        except (TypeError, ValueError):
            pass

    poll_remaining = 0
    if state['poll_active'] and state['poll_started_at']:
        try:
            started = _parse_db_datetime(state['poll_started_at'])
            remaining = (poll_duration or POLL_DURATION) - (
                now - started).total_seconds()
            poll_remaining = max(0, math.ceil(remaining))
        except (TypeError, ValueError):
            pass

    session_elapsed = None
    if state['session_started_at']:
        try:
            started = _parse_db_datetime(state['session_started_at'])
            session_elapsed = max(0, math.floor((now - started).total_seconds()))
        except (TypeError, ValueError):
            pass
    settling = _ratings_settling_state(
        state, now=now, poll_duration=poll_duration
    )
    return {
        'presentation_remaining': presentation_remaining,
        'poll_remaining': poll_remaining,
        'session_elapsed': session_elapsed,
        **settling,
    }


# Students poll about once a second (cheap state-version path). Refresh
# presence at most this often so last_active_at stays truthful without a write
# on every poll. The is_online cutoff used elsewhere is 1 minute, so a 30s
# cadence keeps that readout accurate.
STUDENT_ACTIVITY_SYNC_INTERVAL = timedelta(seconds=30)
STUDENT_ACTIVITY_FAILURE_RETRY_INTERVAL = timedelta(seconds=5)
STUDENT_ACTIVITY_WRITE_TIMEOUT = 0.05


def _sync_student_activity(slug, session_key=None):
    """Best-effort refresh of the logged-in student's last_active_at.

    Writes when the course session changed or when the previous write is older
    than STUDENT_ACTIVITY_SYNC_INTERVAL. Bounding writes this way means 1s
    student polling cannot cause a write storm, yet the instructor's "online
    now" counts stay correct. A dedicated short-timeout connection makes
    presence tracking non-blocking and keeps its failure separate from the
    classroom state response.
    """
    student_id = session.get('student_id')
    if not student_id:
        return False

    now = _utcnow()
    session_changed = (
        session_key is not None
        and session.get('activity_session_key') != session_key
    )
    stale = True
    last_synced = session.get('last_active_synced_at')
    if last_synced:
        try:
            stale = now - _parse_db_datetime(last_synced) \
                >= STUDENT_ACTIVITY_SYNC_INTERVAL
        except (TypeError, ValueError):
            stale = True
    if not session_changed and not stale:
        return True

    last_failed = session.get('last_active_sync_failed_at')
    if last_failed:
        try:
            if (now - _parse_db_datetime(last_failed)
                    < STUDENT_ACTIVITY_FAILURE_RETRY_INTERVAL):
                return False
        except (TypeError, ValueError):
            pass

    db = None
    try:
        db_path = course_db_path(slug)
        if not db_path or not os.path.isfile(db_path):
            return False
        db = sqlite3.connect(
            db_path,
            timeout=STUDENT_ACTIVITY_WRITE_TIMEOUT,
        )
        if session_key is None:
            state = db.execute(
                'SELECT session_key FROM course_state LIMIT 1'
            ).fetchone()
            session_key = (state[0] or 0) if state else 0
        db.execute(
            '''UPDATE students SET last_active_at = CURRENT_TIMESTAMP
               WHERE student_id = ? AND is_active = 1''',
            [student_id],
        )
        db.commit()
    except Exception:
        if db is not None:
            try:
                db.rollback()
            except sqlite3.Error:
                pass
        session['last_active_sync_failed_at'] = now.isoformat()
        return False
    finally:
        if db is not None:
            db.close()

    session['activity_session_key'] = session_key
    session['last_active_synced_at'] = now.isoformat()
    session.pop('last_active_sync_failed_at', None)
    return True


def _serialize_question_blocks(entries):
    if not entries:
        return ''
    return ''.join(
        f"---\n{frontmatter.strip()}\n---\n\n{body.strip()}\n\n"
        for frontmatter, body in entries
    ).rstrip() + '\n'


def _write_text_atomic(path, content):
    """Replace a text file atomically on the same filesystem."""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    temporary = f'{path}.tmp-{uuid.uuid4().hex}'
    try:
        with open(temporary, 'w', encoding='utf-8', newline='\n') as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        if os.path.exists(temporary):
            os.remove(temporary)


def _write_bytes_atomic(path, content):
    """Replace a binary file atomically on the same filesystem."""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    temporary = f'{path}.tmp-{uuid.uuid4().hex}'
    try:
        with open(temporary, 'wb') as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        if os.path.exists(temporary):
            os.remove(temporary)


MAX_QUESTION_UPLOAD_BYTES = 1024 * 1024
MAX_WEEK_QUESTIONS = 100


def _persistent_questions_dir(slug):
    return os.path.join(config.DATA_DIR, slug, 'questions')


def _persistent_week_question_path(slug, week_num):
    return os.path.join(
        _persistent_questions_dir(slug),
        f'week-{week_num}-questions.md',
    )


def _bundled_week_question_path(slug, week_num):
    return os.path.join(
        _course_class_dir(slug), f'week-{week_num}-questions.md'
    )


def _resolve_week_question_path(slug, week_num):
    """Resolve a durable upload first, then the bundled course fallback."""
    persistent = _persistent_week_question_path(slug, week_num)
    return persistent if os.path.exists(persistent) else \
        _bundled_week_question_path(slug, week_num)


def _validate_course_question_catalog(slug, weeks=None):
    return validate_question_catalog(
        _persistent_questions_dir(slug),
        weeks=weeks,
        fallback_dir=_course_class_dir(slug),
    )


def read_presentation_question_index(slug, week_num):
    """Compatibility name for the canonical weekly Markdown loader."""
    return read_week_questions(
        _resolve_week_question_path(slug, week_num), week_num=week_num
    )


def sync_presentation_questions(
        slug, course_id, week_num, db=None, commit=True,
        bump_discussion_version=False):
    """Sync the canonical weekly file into current presentation rows."""
    owns_transaction = db is None
    if owns_transaction:
        ensure_schema(slug)
        db = get_db(slug)
        db.execute('BEGIN IMMEDIATE')
    try:
        questions = read_presentation_question_index(slug, week_num)
        canonical_prefix = f'week-{week_num}-q-'
        existing = db.execute(
            '''SELECT * FROM questions
               WHERE course_id = ? AND COALESCE(week_num, 1) = ?
                 AND (source_key IS NULL
                      OR source_key LIKE 'presentation:%'
                      OR source_key LIKE ?)''',
            [course_id, week_num, f'{canonical_prefix}%']
        ).fetchall()
        by_source = {
            row['source_key']: row for row in existing if row['source_key']
        }
        legacy = [
            row for row in existing
            if not row['source_key']
            or str(row['source_key']).startswith('presentation:')
        ]

        retained_ids = set()
        changed = False
        for question in questions:
            source_key = question['source_key']
            row = by_source.get(source_key)
            if row is None:
                normalized_title = ' '.join(
                    question['title'].split()
                ).casefold()
                candidates = [
                    candidate for candidate in legacy
                    if candidate['id'] not in retained_ids
                    and ' '.join(str(
                        candidate['title']
                        or candidate['question_text']
                        or ''
                    ).split()).casefold() == normalized_title
                ]
                candidates.sort(key=lambda candidate: (
                    candidate['question_num'] != question['num'],
                    candidate['id'],
                ))
                row = candidates[0] if candidates else None

            question_text = question['title'][:200]
            if row is None:
                cursor = db.execute(
                    '''INSERT INTO questions
                       (course_id, question_num, question_text, title, content,
                        week_num, source_key)
                       VALUES (?, ?, ?, ?, ?, ?, ?)''',
                    [
                        course_id, question['num'], question_text,
                        question['title'], question['content'], week_num,
                        source_key,
                    ]
                )
                retained_ids.add(cursor.lastrowid)
                changed = True
            else:
                retained_ids.add(row['id'])
                if (row['question_num'] != question['num'] or
                        row['question_text'] != question_text or
                        row['title'] != question['title'] or
                        row['content'] != question['content'] or
                        row['week_num'] != week_num or
                        row['source_key'] != source_key):
                    db.execute(
                        '''UPDATE questions
                           SET question_num = ?, question_text = ?, title = ?,
                               content = ?, week_num = ?, source_key = ?
                           WHERE id = ?''',
                        [
                            question['num'], question_text, question['title'],
                            question['content'], week_num, source_key, row['id'],
                        ]
                    )
                    changed = True

        for row in existing:
            if (row['id'] not in retained_ids and row['source_key'] and
                    str(row['source_key']).startswith(canonical_prefix)):
                db.execute('DELETE FROM questions WHERE id = ?', [row['id']])
                changed = True

        if changed and bump_discussion_version:
            db.execute(
                '''UPDATE course_state
                   SET discussion_questions_version =
                       COALESCE(discussion_questions_version, 0) + 1
                   WHERE course_id = ?''',
                [course_id],
            )
        if commit:
            db.commit()
        return len(questions)
    except Exception:
        if owns_transaction:
            db.rollback()
        raise


def _read_appendix_question_rows(slug, week_num):
    """Read valid appendix blocks and assign stable presentation source keys."""
    appendix_path = _appendix_path(slug, week_num)
    if not os.path.exists(appendix_path):
        return []

    with open(appendix_path, 'r', encoding='utf-8-sig') as f:
        entries = parse_question_blocks(f.read())

    rows = []
    seen_numbers = set()
    for position, (fm_block, body_block) in enumerate(entries, 1):
        try:
            metadata = yaml.safe_load(fm_block) or {}
        except yaml.YAMLError as exc:
            raise QuestionParseError(
                f'Invalid appendix frontmatter in question {position}'
            ) from exc
        title = str(metadata.get('title') or '').strip()
        if not title:
            raise QuestionParseError(
                f'Appendix question {position} has no title'
            )

        body = body_block.strip()
        label_match = re.match(r'^A(\d+)\s*:', title, re.IGNORECASE)
        if not label_match:
            raise QuestionParseError(
                f'Appendix question {position} must start with an A-number label'
            )
        question_num = int(label_match.group(1))
        if question_num in seen_numbers:
            raise QuestionParseError(
                f'Duplicate appendix label A{question_num}'
            )
        seen_numbers.add(question_num)
        rows.append({
            'source_key': f'appendix:{week_num}:A{question_num}',
            'question_num': question_num,
            'question_text': title[:200],
            'title': title,
            'content': body,
        })
    return rows


def sync_appendix_questions(
        slug, course_id, week_num, db=None, commit=True,
        bump_discussion_version=False):
    """Make a week's appendix questions selectable during presentations."""
    owns_transaction = db is None
    if owns_transaction:
        ensure_schema(slug)
        db = get_db(slug)
        db.execute('BEGIN IMMEDIATE')
    try:
        desired = _read_appendix_question_rows(slug, week_num)
        source_prefix = f'appendix:{week_num}:%'
        existing = db.execute(
            '''SELECT * FROM questions
               WHERE course_id = ? AND source_key LIKE ?''',
            [course_id, source_prefix]
        ).fetchall()
        by_source = {row['source_key']: row for row in existing}
        retained_ids = set()
        changed = False

        for question in desired:
            row = by_source.get(question['source_key'])
            if row is None:
                # Preserve IDs from the short-lived content-hash key scheme.
                row = next((candidate for candidate in existing
                            if candidate['id'] not in retained_ids and
                            not candidate['source_key'].startswith(
                                f'appendix:{week_num}:A') and
                            candidate['title'] == question['title'] and
                            candidate['content'] == question['content']), None)
            if row is None:
                cursor = db.execute(
                    '''INSERT INTO questions
                       (course_id, question_num, question_text, title, content,
                        week_num, source_key)
                       VALUES (?, ?, ?, ?, ?, ?, ?)''',
                    [course_id, question['question_num'],
                     question['question_text'], question['title'],
                     question['content'], week_num, question['source_key']]
                )
                retained_ids.add(cursor.lastrowid)
                changed = True
            else:
                retained_ids.add(row['id'])
                if (row['question_num'] != question['question_num'] or
                        row['question_text'] != question['question_text'] or
                        row['title'] != question['title'] or
                        row['content'] != question['content'] or
                        row['week_num'] != week_num or
                        row['source_key'] != question['source_key']):
                    db.execute(
                        '''UPDATE questions
                           SET question_num = ?, question_text = ?, title = ?,
                               content = ?, week_num = ?, source_key = ?
                           WHERE id = ?''',
                        [question['question_num'], question['question_text'],
                         question['title'], question['content'], week_num,
                         question['source_key'], row['id']]
                    )
                    changed = True

        for row in existing:
            if row['id'] not in retained_ids:
                db.execute('DELETE FROM questions WHERE id = ?', [row['id']])
                changed = True

        if changed and bump_discussion_version:
            db.execute(
                '''UPDATE course_state
                   SET discussion_questions_version =
                       COALESCE(discussion_questions_version, 0) + 1
                   WHERE course_id = ?''',
                [course_id],
            )
        if commit:
            db.commit()
        return len(desired)
    except Exception:
        if owns_transaction:
            db.rollback()
        raise


def _sync_instructor_catalog_once(slug):
    '''Sync bundled questions once per worker without touching a live turn.'''
    course = query_db(slug, 'SELECT id FROM courses LIMIT 1', one=True)
    if not course:
        return
    state = query_db(
        slug,
        '''SELECT discussion_week, active_question_id FROM course_state
           WHERE course_id = ?''',
        [course['id']], one=True,
    )
    if not state or state['active_question_id']:
        return
    week = state['discussion_week'] or 1
    key = (
        os.path.abspath(config.CLASSES_DIR), os.path.abspath(config.DATA_DIR),
        slug, course['id'], week,
    )
    with _instructor_catalog_sync_lock:
        if key in _instructor_catalog_sync_attempted:
            return
        if len(_instructor_catalog_sync_attempted) >= COURSE_AVAILABILITY_CACHE_LIMIT:
            _instructor_catalog_sync_attempted.clear()
        _instructor_catalog_sync_attempted.add(key)

    try:
        catalog_week = _validate_course_question_catalog(
            slug, weeks=[week]
        ).get_week(week)
        if not catalog_week or not catalog_week.presentation.ready:
            return
        sync_presentation_questions(
            slug, course['id'], week,
            bump_discussion_version=True,
        )
    except (OSError, ValueError, QuestionParseError) as exc:
        app.logger.warning(
            'Skipped one-time question sync for %s week %s: %s',
            slug, week, exc,
        )
    except sqlite3.Error:
        with _instructor_catalog_sync_lock:
            _instructor_catalog_sync_attempted.discard(key)
        app.logger.warning(
            'Deferred one-time question sync for %s week %s',
            slug, week, exc_info=True,
        )


@app.before_request
def mark_request_arrival():
    g.request_arrived_at = _utcnow()


@app.after_request
def compress_json_response(response):
    """Gzip sizeable JSON responses when the client advertises support."""
    response.headers.setdefault('X-Content-Type-Options', 'nosniff')
    response.headers.setdefault('Referrer-Policy', 'same-origin')
    response.headers.setdefault('X-Frame-Options', 'DENY')
    response.headers.setdefault(
        'X-Popping-Version', public_version(APP_VERSION)
    )
    authenticated = (
        getattr(g, 'authenticated_role', None) is not None
        or _exclusive_session_role() is not None
    )
    if (authenticated and request.endpoint != 'static'
            and request.path != '/healthz'):
        response.headers['Cache-Control'] = 'no-store, private'
    if response.direct_passthrough or response.status_code < 200 or response.status_code >= 300:
        return response
    if response.mimetype != 'application/json' or response.headers.get('Content-Encoding'):
        return response
    response.vary.add('Accept-Encoding')
    if 'gzip' not in request.headers.get('Accept-Encoding', '').lower():
        return response

    data = response.get_data()
    if len(data) < JSON_COMPRESSION_MIN_BYTES:
        return response
    compressed = gzip.compress(data, compresslevel=5)
    if len(compressed) >= len(data):
        return response

    response.set_data(compressed)
    response.headers['Content-Encoding'] = 'gzip'
    response.headers['Content-Length'] = len(compressed)
    return response


@app.template_filter('phase_label')
def phase_label_filter(phase):
    return PHASE_LABELS.get(phase, phase.upper())


def _student_identity_value(student, field):
    """Return one student identity field from a row, mapping, or object."""
    if student is None:
        return None
    if hasattr(student, 'keys'):
        keys = student.keys()
        return student[field] if field in keys else None
    if isinstance(student, dict):
        return student.get(field)
    return getattr(student, field, None)


def _normalized_identity_text(value):
    return str(value).strip() if value is not None else ''


def _student_display_identity(student):
    """Resolve one public label and record which identity field supplied it."""
    student_id = _normalized_identity_text(
        _student_identity_value(student, 'student_id')
    )
    source_hint = _normalized_identity_text(
        _student_identity_value(student, 'identity_source')
    )
    if source_hint == 'student_id' and student_id:
        return student_id, 'student_id'

    display_name = _normalized_identity_text(
        _student_identity_value(student, 'display_name')
    )
    if display_name:
        return display_name, 'display_name'

    roster_name = _normalized_identity_text(
        _student_identity_value(student, 'roster_name')
    ) or _normalized_identity_text(_student_identity_value(student, 'name'))
    if roster_name:
        return roster_name, 'roster_name'

    snapshot_name = _normalized_identity_text(
        _student_identity_value(student, 'student_name')
    ) or _normalized_identity_text(
        _student_identity_value(student, 'challenger_name')
    )
    if snapshot_name:
        source = source_hint or (
            'student_id' if snapshot_name == student_id else 'snapshot_name'
        )
        return snapshot_name, source
    if student_id:
        return student_id, 'student_id'
    return 'Unknown', 'unknown'


def _student_display_name(student):
    """Resolve student display name, roster name, then public student ID."""
    return _student_display_identity(student)[0]


def _enrich_live_challenge_identities(slug, course_id, challenges):
    """Overlay current names in responses without rewriting event snapshots."""
    student_db_ids = set()
    for challenge in challenges:
        if not isinstance(challenge, dict):
            continue
        try:
            student_db_id = int(challenge.get('challenger_id'))
        except (TypeError, ValueError):
            continue
        if student_db_id > 0:
            student_db_ids.add(student_db_id)
    if not student_db_ids:
        return challenges

    student_db_ids = sorted(student_db_ids)
    placeholders = ','.join('?' * len(student_db_ids))
    identities = query_db(
        slug,
        f'''SELECT id, student_id, name AS roster_name, display_name
            FROM students
            WHERE course_id = ? AND id IN ({placeholders})''',
        [course_id, *student_db_ids],
    )
    identities_by_id = {
        str(identity['id']): identity for identity in identities
    }

    enriched = []
    for challenge in challenges:
        if not isinstance(challenge, dict):
            enriched.append(challenge)
            continue
        identity = identities_by_id.get(str(challenge.get('challenger_id')))
        if identity is None:
            enriched.append(challenge)
            continue
        current = dict(challenge)
        challenger_name, identity_source = _student_display_identity(identity)
        current['challenger_name'] = challenger_name
        current['challenger_student_id'] = identity['student_id']
        current['identity_source'] = identity_source
        enriched.append(current)
    return enriched


@app.template_filter('display_name')
def display_name_filter(student):
    """Return the single student-facing display name."""
    return _student_display_name(student)


@app.template_filter('instructor_display_name')
def instructor_display_name_filter(student):
    """Return display_name plus ID for instructor-facing identity labels."""
    display_name, identity_source = _student_display_identity(student)
    student_id = _normalized_identity_text(
        _student_identity_value(student, 'student_id')
    )
    if student_id and identity_source != 'student_id':
        return f"{display_name} ({student_id})"
    return student_id or display_name


def _auth_failure():
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Not logged in'}), 401
    return redirect(url_for('index'))


def _instructor_session_token(slug, instructor_id, pin):
    """Return a credential-bound marker for one instructor session."""
    secret = app.config['SECRET_KEY']
    if isinstance(secret, str):
        secret = secret.encode('utf-8')
    message = (
        f'instructor-session\0{slug}\0{instructor_id}\0{pin}'
    ).encode('utf-8')
    return hmac.new(secret, message, hashlib.sha256).hexdigest()


def _authenticated_instructor(slug):
    """Return the active instructor only if its login credential is current."""
    instructor = query_db(
        slug,
        '''SELECT i.id, i.pin FROM instructors i
           JOIN courses c ON c.instructor_id = i.id
           WHERE i.id = ? AND c.slug = ? AND c.is_active = 1''',
        [session.get('instructor_id'), slug], one=True
    )
    supplied_token = session.get('instructor_auth_token')
    if (not instructor or not isinstance(supplied_token, str) or
            not hmac.compare_digest(
                supplied_token,
                _instructor_session_token(
                    slug, instructor['id'], instructor['pin']
                ),
            )):
        session.clear()
        return None
    return instructor


def _student_session_token(slug, row_id, student_id, pin):
    '''Return a credential-bound marker for one student session.'''
    secret = app.config['SECRET_KEY']
    if isinstance(secret, str):
        secret = secret.encode('utf-8')
    message = (
        f'student-session\0{slug}\0{row_id}\0{student_id}\0{pin}'
    ).encode('utf-8')
    return hmac.new(secret, message, hashlib.sha256).hexdigest()


def _authenticated_student(slug):
    '''Return the active student only if its login credential is current.'''
    student = query_db(
        slug,
        '''SELECT s.* FROM students s JOIN courses c ON s.course_id = c.id
           WHERE s.student_id = ? AND c.slug = ? AND s.is_active = 1
             AND c.is_active = 1''',
        [session.get('student_id'), slug], one=True
    )
    supplied_token = session.get('student_auth_token')
    if (not student or not isinstance(supplied_token, str) or
            not hmac.compare_digest(
                supplied_token,
                _student_session_token(
                    slug, student['id'], student['student_id'], student['pin']
                ),
            )):
        session.clear()
        return None
    display_name = _student_display_name(student)
    if session.get('display_name') != display_name:
        session['display_name'] = display_name
    if session.get('name') != display_name:
        session['name'] = display_name
    return student


def _login_origin_is_allowed():
    """Reject login posts with an explicit cross-origin browser source."""
    source = request.headers.get('Origin') or request.headers.get('Referer')
    if not source:
        # Preserve non-browser clients. SameSite=Lax remains the baseline when
        # a user agent omits both source headers.
        return True
    try:
        supplied = urlsplit(source)
    except ValueError:
        return False
    return (
        supplied.scheme.casefold() == request.scheme.casefold()
        and supplied.netloc.casefold() == request.host.casefold()
        and not supplied.username
        and not supplied.password
    )


def _login_client_hash():
    address = request.remote_addr or 'unknown'
    return hashlib.sha256(address.encode('utf-8')).hexdigest()


def _login_retry_after_locked(
        db, course_id, login_type, principal, client_hash, now):
    """Return the current retry delay while holding the login write lock."""
    row = db.execute(
        '''SELECT failed_count, window_started_at, blocked_until
           FROM login_attempts
           WHERE course_id = ? AND login_type = ? AND principal = ?
             AND client_hash = ?''',
        [course_id, login_type, principal, client_hash],
    ).fetchone()
    if not row:
        return 0

    window_start = now - timedelta(seconds=LOGIN_WINDOW_SECONDS)
    started_at = _parse_db_datetime(row['window_started_at'])
    blocked_until = _parse_db_datetime(row['blocked_until'])
    if blocked_until and blocked_until > now:
        return max(1, int((blocked_until - now).total_seconds()) + 1)
    if (started_at and started_at > window_start
            and row['failed_count'] >= LOGIN_FAILURE_LIMIT):
        return max(
            1,
            int((started_at + timedelta(seconds=LOGIN_WINDOW_SECONDS) - now)
                .total_seconds()) + 1,
        )
    return 0


def _record_login_failure_locked(
        db, course_id, login_type, principal, client_hash, now):
    """Record one failure inside the caller's serialized transaction."""
    db.execute(
        '''DELETE FROM login_attempts
           WHERE window_started_at <= datetime('now', ?)
             AND (blocked_until IS NULL OR blocked_until <= CURRENT_TIMESTAMP)''',
        [f'-{LOGIN_WINDOW_SECONDS} seconds'],
    )
    row = db.execute(
        '''SELECT * FROM login_attempts
           WHERE course_id = ? AND login_type = ? AND principal = ?
             AND client_hash = ?''',
        [course_id, login_type, principal, client_hash],
    ).fetchone()
    window_start = now - timedelta(seconds=LOGIN_WINDOW_SECONDS)
    if not row or _parse_db_datetime(row['window_started_at']) <= window_start:
        db.execute(
            '''INSERT INTO login_attempts
               (course_id, login_type, principal, client_hash, failed_count,
                window_started_at, blocked_until)
               VALUES (?, ?, ?, ?, 1, CURRENT_TIMESTAMP, NULL)
               ON CONFLICT(course_id, login_type, principal, client_hash)
               DO UPDATE SET failed_count = 1,
                   window_started_at = CURRENT_TIMESTAMP,
                   blocked_until = NULL''',
            [course_id, login_type, principal, client_hash],
        )
        return

    failed_count = row['failed_count'] + 1
    blocked_until = None
    if failed_count >= LOGIN_FAILURE_LIMIT:
        blocked_until = (
            now + timedelta(seconds=LOGIN_WINDOW_SECONDS)
        ).strftime('%Y-%m-%d %H:%M:%S')
    db.execute(
        '''UPDATE login_attempts
           SET failed_count = ?, blocked_until = ?
           WHERE id = ?''',
        [failed_count, blocked_until, row['id']],
    )


def _login_client_retry_after_locked(
        db, course_id, login_type, client_hash, now):
    '''Bound rapid principal rotation from one client without new schema.'''
    window_start = now - timedelta(seconds=LOGIN_WINDOW_SECONDS)
    rows = db.execute(
        '''SELECT failed_count, window_started_at FROM login_attempts
           WHERE course_id = ? AND login_type = ? AND client_hash = ?''',
        [course_id, login_type, client_hash],
    ).fetchall()
    active = []
    for row in rows:
        started_at = _parse_db_datetime(row['window_started_at'])
        if started_at and started_at > window_start:
            active.append((row['failed_count'], started_at))
    if sum(count for count, _started_at in active) < LOGIN_CLIENT_FAILURE_LIMIT:
        return 0
    earliest_expiry = min(
        started_at + timedelta(seconds=LOGIN_WINDOW_SECONDS)
        for _count, started_at in active
    )
    return max(1, int((earliest_expiry - now).total_seconds()) + 1)


def _authenticate_with_throttle(
        slug, course_id, login_type, principal, client_hash,
        lookup_sql, lookup_params, touch_student=False,
        student_display_name=None, clear_student_display_name=False):
    """Serialize throttle check, credential lookup, and counter update."""
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        now = _utcnow()
        retry_after = max(
            _login_retry_after_locked(
                db, course_id, login_type, principal, client_hash, now
            ),
            _login_client_retry_after_locked(
                db, course_id, login_type, client_hash, now
            ),
        )
        if retry_after:
            db.commit()
            return None, retry_after

        identity = db.execute(lookup_sql, lookup_params).fetchone()
        if identity:
            identity = dict(identity)
            db.execute(
                '''DELETE FROM login_attempts
                   WHERE course_id = ? AND login_type = ? AND principal = ?
                     AND client_hash = ?''',
                [course_id, login_type, principal, client_hash],
            )
            if touch_student:
                db.execute(
                    '''UPDATE students SET last_login_at = CURRENT_TIMESTAMP
                       WHERE id = ?''',
                    [identity['id']],
                )
                current_display_name = _normalized_identity_text(
                    identity.get('display_name')
                ) or None
                desired_display_name = current_display_name
                if clear_student_display_name:
                    desired_display_name = None
                elif student_display_name is not None:
                    desired_display_name = student_display_name
                if desired_display_name != current_display_name:
                    db.execute(
                        'UPDATE students SET display_name = ? WHERE id = ?',
                        [desired_display_name, identity['id']],
                    )
                    _bump_roster_version(slug, course_id, db=db)
                    identity['display_name'] = desired_display_name
        else:
            _record_login_failure_locked(
                db, course_id, login_type, principal, client_hash, now
            )
        db.commit()
        return identity, 0
    except Exception:
        db.rollback()
        raise


def _rate_limited_login_response(template, slug, course, retry_after):
    if retry_after >= 60:
        unit = 'minute' if retry_after < 120 else 'minutes'
        amount = max(1, math.ceil(retry_after / 60))
        wait = f'{amount} {unit}'
    else:
        wait = f'{retry_after} second{"s" if retry_after != 1 else ""}'
    flash(
        f'Too many failed login attempts. '
        f'Please try again in {wait}.',
        'error'
    )
    response = make_response(
        render_template(template, course=course, slug=slug), 429
    )
    response.headers['Retry-After'] = str(retry_after)
    return response


def _exclusive_session_role():
    """Return the role only when exactly one authenticated role is present."""
    role = session.get('role')
    has_student = bool(session.get('student_id'))
    has_instructor = bool(session.get('instructor_id'))
    if not session.get('slug') or has_student == has_instructor:
        return None
    if role == 'student' and has_student:
        return 'student'
    if role == 'instructor' and has_instructor:
        return 'instructor'
    return None


def _session_course_is_ready(slug):
    """Validate a signed-in course without discarding transient sessions."""
    availability = _course_availability(slug)
    status = availability['status']
    if (status == 'unavailable' or
            (status == 'missing' and not is_demo_instance_slug(slug))):
        raise CourseTemporarilyUnavailable()
    if availability['status'] != 'ready':
        session.clear()
        return False
    return True


def student_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        slug = session.get('slug')
        if _exclusive_session_role() != 'student':
            return _auth_failure()
        if not _session_course_is_ready(slug):
            return _auth_failure()
        ensure_schema(slug)
        student = _authenticated_student(slug)
        if not student:
            return _auth_failure()
        g.current_student = student
        g.authenticated_role = 'student'
        _sync_student_activity(slug)
        return f(*args, **kwargs)
    return decorated


def instructor_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        slug = session.get('slug')
        route_slug = kwargs.get('slug')
        if (_exclusive_session_role() != 'instructor' or
                (route_slug is not None and route_slug != slug)):
            return _auth_failure()
        if not _session_course_is_ready(slug):
            return _auth_failure()
        instructor = _authenticated_instructor(slug)
        if not instructor:
            return _auth_failure()
        ensure_schema(slug)
        g.authenticated_role = 'instructor'
        return f(*args, **kwargs)
    return decorated


def _course_with_config_metadata(course, course_config):
    """Overlay current display metadata without rewriting durable DB history."""
    if not course:
        return course
    effective = dict(course)
    if isinstance(course_config, dict):
        for field in ('name', 'code', 'semester'):
            value = course_config.get(field)
            if isinstance(value, str) and value.strip():
                effective[field] = value.strip()
    return effective


def _inspect_course_availability(slug):
    """Verify one configured course without modifying its database."""
    result = {
        'slug': slug,
        'status': 'invalid',
        'message': 'Course setup could not be verified.',
        'configured_active': False,
        'config': None,
        'course': None,
        'has_db': False,
        'actual_schema_version': None,
    }
    if not is_valid_slug(slug):
        return result

    class_slug = canonical_class_slug(slug)
    class_dir = _course_class_dir(slug)
    yaml_path = os.path.join(class_dir, 'course.yaml')
    if not os.path.isdir(class_dir) or not os.path.isfile(yaml_path):
        return result

    try:
        with open(yaml_path, encoding='utf-8') as config_file:
            course_config = yaml.safe_load(config_file)
    except OSError:
        result.update({
            'status': 'unavailable',
            'message': 'Course data is temporarily unavailable.',
        })
        return result
    except (UnicodeError, yaml.YAMLError):
        return result
    if not isinstance(course_config, dict):
        return result

    result['config'] = course_config
    result['configured_active'] = course_config.get('active') is True
    if course_config.get('slug') != class_slug:
        return result
    if not result['configured_active']:
        result.update({
            'status': 'inactive',
            'message': 'Course is not currently active.',
        })
        return result

    db_path = course_db_path(slug)
    result['has_db'] = bool(db_path and os.path.isfile(db_path))
    if not result['has_db']:
        result.update({
            'status': 'missing',
            'message': 'Course database has not been initialized.',
        })
        return result

    connection = None
    try:
        db_uri = f'{Path(db_path).resolve().as_uri()}?mode=ro'
        connection = sqlite3.connect(db_uri, uri=True, timeout=2)
        connection.row_factory = sqlite3.Row
        connection.execute('PRAGMA query_only = ON')
        quick_check = connection.execute('PRAGMA quick_check').fetchall()
        if [row[0] for row in quick_check] != ['ok']:
            return result
        tables = {
            row[0] for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            ).fetchall()
        }
        if not REQUIRED_COURSE_TABLES.issubset(tables):
            return result
        for table, required_columns in REQUIRED_COURSE_SCHEMA.items():
            columns = {
                row[1] for row in connection.execute(
                    f'PRAGMA table_info({table})'
                ).fetchall()
            }
            if not required_columns.issubset(columns):
                return result
        recorded_version = inspect_schema_version(
            connection, allow_unversioned=True
        )
        if recorded_version is None:
            validate_legacy_adoption_candidate(connection)
        actual_schema_version = recorded_version or BASELINE_SCHEMA_VERSION
        result['actual_schema_version'] = actual_schema_version
        rows = connection.execute(
            '''SELECT c.*, i.name AS instructor_name
               FROM courses c
               JOIN instructors i ON i.id = c.instructor_id
               LIMIT 2''',
        ).fetchall()
        if (len(rows) != 1 or rows[0]['slug'] != slug or
                rows[0]['is_active'] != 1):
            return result
        state_rows = connection.execute(
            'SELECT course_id FROM course_state LIMIT 2'
        ).fetchall()
        if (len(state_rows) != 1 or
                state_rows[0]['course_id'] != rows[0]['id']):
            return result
        if actual_schema_version != SCHEMA_VERSION:
            result.update({
                'status': 'migration_required',
                'message': (
                    'Course database maintenance is required before login.'
                ),
            })
            return result
        validate_current_schema(connection)
        result.update({
            'status': 'ready',
            'message': '',
            'course': _course_with_config_metadata(rows[0], course_config),
        })
    except (OSError, sqlite3.OperationalError):
        result.update({
            'status': 'unavailable',
            'message': 'Course data is temporarily unavailable.',
        })
        return result
    except (sqlite3.Error, RuntimeError):
        return result
    finally:
        if connection is not None:
            connection.close()
    return result


def _clear_course_availability_cache(slug=None):
    """Clear cached validation results after an in-process setup change."""
    with _course_availability_lock:
        if slug is None:
            _course_availability_cache.clear()
            return
        prefix = (os.path.abspath(config.CLASSES_DIR),
                  os.path.abspath(config.DATA_DIR), slug)
        _course_availability_cache.pop(prefix, None)


def _course_availability(slug):
    """Return a short-lived, process-wide validation result for one course."""
    cache_key = (
        os.path.abspath(config.CLASSES_DIR),
        os.path.abspath(config.DATA_DIR),
        slug,
    )
    now = time.monotonic()
    with _course_availability_lock:
        cached = _course_availability_cache.get(cache_key)
        if cached:
            cache_ttl = (
                COURSE_UNAVAILABLE_TTL
                if cached['result']['status'] in (
                    'unavailable', 'missing', 'migration_required')
                else COURSE_AVAILABILITY_TTL
            )
            if now - cached['checked_at'] < cache_ttl:
                return copy.deepcopy(cached['result'])
        result = _inspect_course_availability(slug)
        if len(_course_availability_cache) >= COURSE_AVAILABILITY_CACHE_LIMIT:
            _course_availability_cache.clear()
        _course_availability_cache[cache_key] = {
            'checked_at': time.monotonic(),
            'result': copy.deepcopy(result),
        }
        return result


def _scan_courses():
    """Return active configured courses with their verified availability."""
    courses = []
    if not os.path.isdir(config.CLASSES_DIR):
        return courses
    for slug in sorted(os.listdir(config.CLASSES_DIR)):
        # The demo course is reachable only via /demo — never list it publicly.
        if slug == 'demo':
            continue
        availability = _course_availability(slug)
        cfg = availability.get('config')
        if not availability['configured_active'] or not isinstance(cfg, dict):
            continue
        verified_course = availability.get('course') or {}
        courses.append({
            'id': slug,
            'slug': slug,
            'name': cfg.get('name') or slug,
            'code': cfg.get('code'),
            'semester': cfg.get('semester'),
            'url': cfg.get('url'),
            'has_db': availability['has_db'],
            'instructor_name': verified_course.get('instructor_name', ''),
            'availability_status': availability['status'],
            'availability_message': availability['message'],
        })
    return courses


@app.route('/healthz')
def healthz():
    """Readiness check using the same identity invariants as course login."""
    if (not os.path.isdir(config.CLASSES_DIR)
            or not os.path.isdir(config.DATA_DIR)):
        return jsonify({'status': 'unavailable'}), 503

    checked = 0
    course_schema_versions = {}
    try:
        for slug in sorted(os.listdir(config.CLASSES_DIR)):
            if slug == 'demo' or not is_valid_slug(slug):
                continue
            yaml_path = os.path.join(_course_class_dir(slug), 'course.yaml')
            if not os.path.isfile(yaml_path):
                continue

            availability = _inspect_course_availability(slug)
            course_config = availability.get('config')
            if (isinstance(course_config, dict)
                    and course_config.get('active') is not True):
                continue
            if availability['status'] not in ('ready', 'migration_required'):
                return jsonify({'status': 'unavailable'}), 503
            actual_schema_version = availability.get('actual_schema_version')
            if not actual_schema_version:
                return jsonify({'status': 'unavailable'}), 503
            course_schema_versions[slug] = actual_schema_version
            checked += 1
    except (OSError, UnicodeError, yaml.YAMLError, sqlite3.Error, RuntimeError):
        return jsonify({'status': 'unavailable'}), 503

    actual_versions = set(course_schema_versions.values())
    if len(actual_versions) == 1:
        reported_schema_version = public_version(next(iter(actual_versions)))
    elif actual_versions:
        reported_schema_version = 'mixed'
    else:
        reported_schema_version = public_version(SCHEMA_VERSION)
    payload = {
        'status': 'ok',
        'courses_checked': checked,
        'website_version': public_version(APP_VERSION),
        'database_schema_version': reported_schema_version,
    }
    if any(
        version != SCHEMA_VERSION
        for version in course_schema_versions.values()
    ):
        payload['schema_migration_pending'] = True
        payload['status'] = 'unavailable'
        payload['course_database_schema_versions'] = {
            slug: public_version(version)
            for slug, version in sorted(course_schema_versions.items())
        }
        return jsonify(payload), 503
    return jsonify(payload)


@app.route('/')
def index():
    role = _exclusive_session_role()
    if role == 'student':
        return redirect(url_for('dashboard'))
    if role == 'instructor':
        return redirect(url_for('instructor_course', slug=session['slug']))
    if session.get('student_id') or session.get('instructor_id') or session.get('role'):
        session.clear()
    if request.args.get('session') == 'ended':
        flash(
            'Your session ended — please log in again. '
            'If you were in a class, your instructor may have updated the roster.',
            'success',
        )
    courses = _scan_courses()
    return render_template('index.html', courses=courses)


@app.route('/login/<slug>', methods=['GET', 'POST'])
def login(slug):
    availability = _course_availability(slug)
    if availability['status'] != 'ready':
        flash(availability['message'], 'error')
        return redirect(url_for('index'))
    ensure_schema(slug)
    course = _course_with_config_metadata(
        query_db(slug, 'SELECT * FROM courses WHERE slug = ?', [slug], one=True),
        availability.get('config'),
    )

    if request.method == 'POST':
        if not _login_origin_is_allowed():
            return 'Forbidden', 403
        student_id = request.form.get('student_id', '').strip()
        pin = request.form.get('pin', '').strip()
        raw_display_name = request.form.get('display_name')
        if raw_display_name is None:
            raw_display_name = request.form.get('name', '')
        display_name = raw_display_name.strip()
        clear_display_name = (
            request.form.get('clear_display_name', '').casefold()
            in ('1', 'true', 'on', 'yes')
        )
        if not student_id or not pin:
            flash('Please enter both ID and PIN.', 'error')
            return redirect(url_for('login', slug=slug))
        if display_name and clear_display_name:
            flash('Enter a display name or clear it, not both.', 'error')
            return redirect(url_for('login', slug=slug))
        if len(display_name) > 200:
            flash('Display name must be 200 characters or fewer.', 'error')
            return redirect(url_for('login', slug=slug))
        principal = student_id.casefold()[:200]
        client_hash = _login_client_hash()
        student, retry_after = _authenticate_with_throttle(
            slug,
            course['id'],
            'student',
            principal,
            client_hash,
            '''SELECT s.* FROM students s JOIN courses c ON c.id = s.course_id
               WHERE s.student_id = ? COLLATE NOCASE
                 AND s.pin = ? AND c.slug = ? AND s.is_active = 1''',
            [student_id, pin, slug],
            touch_student=True,
            student_display_name=display_name or None,
            clear_student_display_name=clear_display_name,
        )
        if retry_after:
            return _rate_limited_login_response(
                'login.html', slug, course, retry_after
            )
        if student:
            session.clear()
            session['role'] = 'student'
            session['student_id'] = student['student_id']
            display_name = _student_display_name(student)
            session['display_name'] = display_name
            session['name'] = display_name
            session['slug'] = slug
            session['student_auth_token'] = _student_session_token(
                slug, student['id'], student['student_id'], student['pin']
            )
            return redirect(url_for('dashboard'))
        flash('Invalid login for this course.', 'error')
        return redirect(url_for('login', slug=slug))

    return render_template('login.html', course=course, slug=slug)


@app.route('/instructor_login/<slug>', methods=['GET', 'POST'])
def instructor_login(slug):
    availability = _course_availability(slug)
    if availability['status'] != 'ready':
        flash(availability['message'], 'error')
        return redirect(url_for('index'))
    ensure_schema(slug)
    course = _course_with_config_metadata(
        query_db(slug, 'SELECT * FROM courses WHERE slug = ?', [slug], one=True),
        availability.get('config'),
    )

    if request.method == 'POST':
        if not _login_origin_is_allowed():
            return 'Forbidden', 403
        username = request.form.get('username', '').strip()
        pin = request.form.get('pin', '')
        if not username or not pin:
            flash('Please enter both username and PIN.', 'error')
            return redirect(url_for('instructor_login', slug=slug))
        lookup_pin = pin if is_valid_instructor_pin(pin) else None
        principal = username.casefold()[:200]
        client_hash = _login_client_hash()
        instructor, retry_after = _authenticate_with_throttle(
            slug,
            course['id'],
            'instructor',
            principal,
            client_hash,
            '''SELECT i.* FROM instructors i
               JOIN courses c ON c.instructor_id = i.id
               WHERE i.username = ? AND i.pin = ? AND c.slug = ?''',
            [username, lookup_pin, slug],
        )
        if retry_after:
            return _rate_limited_login_response(
                'instructor_login.html', slug, course, retry_after
            )
        if instructor:
            session.clear()
            session['role'] = 'instructor'
            session['instructor_id'] = instructor['id']
            session['instructor_name'] = instructor['name']
            session['slug'] = slug
            session['instructor_auth_token'] = _instructor_session_token(
                slug, instructor['id'], instructor['pin']
            )
            return redirect(url_for('instructor_course', slug=slug))
        flash('Invalid login for this course.', 'error')
        return redirect(url_for('instructor_login', slug=slug))

    return render_template('instructor_login.html', course=course, slug=slug)


@app.route('/demo')
def demo():
    return render_template('demo.html', instance_slug=None)


def _remembered_demo_instance():
    for candidate in (
            session.get('demo_instance_slug'), session.get('slug')):
        if is_demo_instance_slug(candidate):
            return candidate
    return None


def _clear_session_preserving_demo_instance(instance_slug=None):
    remembered = instance_slug or _remembered_demo_instance()
    session.clear()
    if is_demo_instance_slug(remembered):
        session['demo_instance_slug'] = remembered


@app.route('/demo/start', methods=['POST'])
def demo_start():
    """Create or reuse one bounded private demo without spawning a process."""
    try:
        instance_slug, removed = create_bounded_demo_instance(
            config.DATA_DIR,
            config.CLASSES_DIR,
            config.DATABASE_SCHEMA,
            reuse_slug=_remembered_demo_instance(),
        )
    except DemoLifecycleBusy:
        flash('A demo is being prepared. Please try again in a moment.', 'error')
        return redirect(url_for('demo'))
    except Exception:
        app.logger.exception('Could not create a private demo instance')
        flash('The demo could not be prepared. Please try again.', 'error')
        return redirect(url_for('demo'))

    for removed_slug in removed:
        _clear_course_availability_cache(removed_slug)
        forget_schema(removed_slug)
    if instance_slug is None:
        flash('All demo spaces are in use. Please try again later.', 'error')
        return redirect(url_for('demo'))
    _clear_course_availability_cache(instance_slug)

    _clear_session_preserving_demo_instance(instance_slug)
    return redirect(url_for(
        'demo_instance_home', instance_slug=instance_slug
    ))


def _demo_instance_ready(instance_slug):
    if not is_demo_instance_slug(instance_slug):
        return False
    return _course_availability(instance_slug)['status'] == 'ready'


def _touch_demo_instance_throttled(slug):
    """Refresh a live demo's .last-used marker at most once a minute."""
    now = time.monotonic()
    with _demo_instance_touch_lock:
        if len(_demo_instance_touch_last) >= DEMO_TOUCH_CACHE_LIMIT:
            _demo_instance_touch_last.clear()
        if len(_demo_instance_touch_failed) >= DEMO_TOUCH_CACHE_LIMIT:
            _demo_instance_touch_failed.clear()
        last_success = _demo_instance_touch_last.get(slug)
        if (last_success is not None
                and now - last_success < DEMO_TOUCH_INTERVAL):
            return
        last_failure = _demo_instance_touch_failed.get(slug)
        if (last_failure is not None
                and now - last_failure < DEMO_TOUCH_FAILURE_RETRY_INTERVAL):
            return
        if slug in _demo_instance_touch_inflight:
            return
        _demo_instance_touch_inflight.add(slug)
    try:
        touch_demo_instance(config.DATA_DIR, slug)
    except OSError:
        # The instance may have been reaped concurrently; the poll itself
        # still succeeded, so a missing marker is not an error here.
        with _demo_instance_touch_lock:
            _demo_instance_touch_failed[slug] = time.monotonic()
    else:
        with _demo_instance_touch_lock:
            _demo_instance_touch_last[slug] = time.monotonic()
            _demo_instance_touch_failed.pop(slug, None)
    finally:
        with _demo_instance_touch_lock:
            _demo_instance_touch_inflight.discard(slug)


@app.route('/demo/<instance_slug>')
def demo_instance_home(instance_slug):
    if not _demo_instance_ready(instance_slug):
        flash('This private demo is no longer available.', 'error')
        return redirect(url_for('demo'))
    try:
        touch_demo_instance(config.DATA_DIR, instance_slug)
    except OSError:
        flash('This private demo is no longer available.', 'error')
        return redirect(url_for('demo'))
    session['demo_instance_slug'] = instance_slug
    return render_template('demo.html', instance_slug=instance_slug)


def _start_demo_session(instance_slug, role, principal):
    if not _demo_instance_ready(instance_slug):
        flash('This private demo is no longer available.', 'error')
        return redirect(url_for('demo'))
    if not principal:
        flash('This private demo is no longer available.', 'error')
        return redirect(url_for('demo'))

    session.clear()
    session['slug'] = instance_slug
    session['demo_instance_slug'] = instance_slug
    session['role'] = role
    session['is_demo'] = True
    if role == 'instructor':
        session['instructor_id'] = principal['id']
        session['instructor_name'] = principal['name']
        session['instructor_auth_token'] = _instructor_session_token(
            instance_slug, principal['id'], principal['pin']
        )
        return redirect(url_for('instructor_course', slug=instance_slug))

    session['student_id'] = principal['student_id']
    display_name = _student_display_name(principal)
    session['display_name'] = display_name
    session['name'] = display_name
    session['student_auth_token'] = _student_session_token(
        instance_slug, principal['id'], principal['student_id'], principal['pin']
    )
    execute_db(
        instance_slug,
        'UPDATE students SET last_login_at = CURRENT_TIMESTAMP WHERE id = ?',
        [principal['id']],
    )
    return redirect(url_for('dashboard'))


@app.route('/demo/<instance_slug>/instructor', methods=['POST'])
def demo_instructor(instance_slug):
    if not _demo_instance_ready(instance_slug):
        flash('This private demo is no longer available.', 'error')
        return redirect(url_for('demo'))
    instructor = query_db(
        instance_slug, 'SELECT * FROM instructors ORDER BY id LIMIT 1', one=True
    )
    return _start_demo_session(instance_slug, 'instructor', instructor)


@app.route('/demo/<instance_slug>/student/<int:student_number>',
           methods=['POST'])
def demo_student(instance_slug, student_number):
    if student_number not in (1, 2) or not _demo_instance_ready(instance_slug):
        flash('This private demo student is not available.', 'error')
        return redirect(url_for('demo'))
    ensure_schema(instance_slug)
    student_id = f'demo00{student_number}'
    student = query_db(
        instance_slug,
        '''SELECT * FROM students
           WHERE student_id = ? AND is_active = 1''',
        [student_id],
        one=True,
    )
    if not student:
        flash('This private demo student is not available.', 'error')
        return redirect(url_for('demo_instance_home', instance_slug=instance_slug))
    return _start_demo_session(instance_slug, 'student', student)


@app.route('/demo/instructor')
@app.route('/demo/student')
def legacy_demo_role():
    flash('Start a private demo first.', 'error')
    return redirect(url_for('demo'))


@app.route('/demo/exit', methods=['POST'])
def demo_exit():
    """Exit demo mode while retaining this browser's private demo slot."""
    _clear_session_preserving_demo_instance()
    return redirect(url_for('demo'))


@app.route('/demo/<instance_slug>/reset', methods=['POST'])
def demo_reset(instance_slug):
    """Reset only the authenticated visitor's tiny private demo."""
    if (not is_demo_instance_slug(instance_slug) or
            not session.get('is_demo') or
            session.get('slug') != instance_slug or
            _exclusive_session_role() not in ('student', 'instructor')):
        return 'Forbidden', 403

    try:
        reset_demo_instance(
            config.DATA_DIR, config.CLASSES_DIR, instance_slug
        )
        _clear_course_availability_cache(instance_slug)
        flash('Demo has been reset to its initial state.', 'success')
    except DemoResetCooldown as exc:
        response = make_response('Please wait before resetting again.', 429)
        response.headers['Retry-After'] = str(max(1, math.ceil(exc.retry_after)))
        return response
    except (OSError, sqlite3.Error):
        app.logger.exception('Could not reset demo instance %s', instance_slug)
        flash('The demo is busy. Please try again.', 'error')
        return redirect(url_for(
            'demo_instance_home', instance_slug=instance_slug
        ))

    _clear_session_preserving_demo_instance(instance_slug)
    return redirect(url_for(
        'demo_instance_home', instance_slug=instance_slug
    ))


@app.route('/demo/reset', methods=['POST'])
def legacy_demo_reset():
    """The former public reset is deliberately inert."""
    return 'Forbidden', 403


@app.route('/logout', methods=['POST'])
def logout():
    if session.get('is_demo') or _remembered_demo_instance():
        _clear_session_preserving_demo_instance()
    else:
        session.clear()
    return redirect(url_for('index'))


@app.route('/dashboard')
@student_login_required
def dashboard():
    slug = session['slug']
    from database import ensure_schema
    ensure_schema(slug)
    student = query_db(slug,
        'SELECT * FROM students WHERE student_id = ? AND is_active = 1',
        [session['student_id']], one=True
    )
    team = None
    if student and student['team_id']:
        team = query_db(slug,
            'SELECT * FROM teams WHERE id = ?', [student['team_id']], one=True
        )
    course = _course_with_config_metadata(
        query_db(slug, 'SELECT * FROM courses WHERE slug = ?', [slug], one=True),
        _course_availability(slug).get('config'),
    )
    state = query_db(slug,
        'SELECT * FROM course_state WHERE course_id = ?', [course['id']], one=True
    )
    max_teams = get_max_teams(slug, course['id'])
    teams = query_db(
        slug,
        '''SELECT team.*, COUNT(student.id) AS member_count
           FROM teams team
           LEFT JOIN students student
             ON student.team_id = team.id AND student.is_active = 1
           WHERE team.course_id = ?
           GROUP BY team.id ORDER BY team.id LIMIT ?''',
        [course['id'], max_teams]
    )
    teams_locked = state['teams_locked'] if state and 'teams_locked' in state.keys() else 0
    teammates = []
    if team:
        teammates = query_db(slug,
            '''SELECT student_id, name, display_name FROM students
               WHERE team_id = ? AND id != ? AND is_active = 1
               ORDER BY COALESCE(NULLIF(TRIM(display_name), ''),
                                 NULLIF(TRIM(name), ''), student_id),
                        student_id''',
            [team['id'], student['id']]
        )
    return render_template(
        'dashboard.html',
        student=student, team=team, teams=teams,
        state=state, course=course, phases=PHASES,
        teams_locked=teams_locked, teammates=teammates,
        max_teams=max_teams,
        max_members=get_max_members_per_team(slug, course['id'])
    )


@app.route('/instructor/<slug>')
@instructor_login_required
def instructor_course(slug):
    if session.get('slug') != slug:
        flash('Unauthorized.', 'error')
        return redirect(url_for('index'))

    course = _course_with_config_metadata(
        query_db(slug, 'SELECT * FROM courses WHERE slug = ?', [slug], one=True),
        _course_availability(slug).get('config'),
    )
    ensure_schema(slug)
    teams = query_db(slug,
        'SELECT * FROM teams WHERE course_id = ? ORDER BY id', [course['id']]
    )
    state = query_db(slug,
        'SELECT * FROM course_state WHERE course_id = ?', [course['id']], one=True
    )
    poll_duration = get_poll_duration(slug)
    if state:
        state = dict(state)
        now = _utcnow()
        state['poll_active'] = _poll_is_open(
            state, now=now, poll_duration=poll_duration)
        state.update(_derive_timing_state(
            state, now=now, poll_duration=poll_duration))
        try:
            full_history = json.loads(state.get('presentation_history') or '[]')
        except (TypeError, ValueError):
            full_history = []
        state['presentation_history'] = json.dumps([
            item for item in full_history
            if (_history_item_is_compatible(item)
                and item.get('session_key', 0) ==
                (state.get('session_key') or 0))
        ])
    selected_week = state['discussion_week'] if state and state['discussion_week'] else 1
    try:
        catalog_week = _validate_course_question_catalog(
            slug,
            weeks=[selected_week],
        ).get_week(selected_week)
    except (OSError, ValueError):
        catalog_week = None
    presentation_catalog_ready = bool(
        catalog_week and catalog_week.presentation.ready
    )
    if state is not None:
        state['presentation_catalog_ready'] = presentation_catalog_ready
    if (not state or state['phase'] != 'competition' or
            not state['active_question_id']):
        if presentation_catalog_ready:
            sync_presentation_questions(
                slug, course['id'], selected_week,
                bump_discussion_version=True,
            )
        try:
            sync_appendix_questions(
                slug, course['id'], selected_week,
                bump_discussion_version=True,
            )
        except QuestionParseError as exc:
            flash(f'Appendix question file needs attention: {exc}', 'error')
    max_teams = get_max_teams(slug, course['id'])
    teams_locked = state['teams_locked'] if state and 'teams_locked' in state.keys() else 0
    students = query_db(slug,
        '''SELECT s.*, t.name as team_name, t.color as team_color
           FROM students s LEFT JOIN teams t ON s.team_id = t.id
           WHERE s.course_id = ? AND s.is_active = 1
           ORDER BY COALESCE(NULLIF(TRIM(s.display_name), ''),
                             NULLIF(TRIM(s.name), ''), s.student_id)''',
        [course['id']]
    )
    questions = query_db(slug,
        '''SELECT * FROM questions
           WHERE course_id = ? AND COALESCE(week_num, 1) = ?
             AND (source_key LIKE ? OR source_key LIKE 'appendix:%' OR id = ?)
           ORDER BY CASE WHEN source_key LIKE 'appendix:%' THEN 1 ELSE 0 END,
                    question_num, id''',
        [course['id'], selected_week, f'week-{selected_week}-q-%',
         state['active_question_id'] if state else None]
    )
    if not presentation_catalog_ready:
        active_question_id = state['active_question_id'] if state else None
        questions = [
            question for question in questions
            if str(question['source_key'] or '').startswith('appendix:')
            or question['id'] == active_question_id
        ]
    presence_checked_at = _utcnow()
    participation_counts = _participation_counts_by_student(
        get_db(slug), course['id'], [student['id'] for student in students]
    )
    students_enhanced = []
    for s in students:
        d = dict(s)
        d['roster_name'] = s['name']
        d['is_online'] = _student_is_online(
            s['last_active_at'],
            now=presence_checked_at,
        )
        counts = participation_counts.get(s['id'], {})
        d['presentation_count'] = counts.get('presentation_count', 0)
        d['challenger_count'] = counts.get('challenger_count', 0)
        students_enhanced.append(d)

    # End session stats
    end_stats = None
    if state and state['phase'] == 'ended':
        # Participants are students who submitted a response in this session.
        participants = query_db(slug,
            '''SELECT COUNT(DISTINCT student_id) AS c FROM (
                   SELECT grader_id AS student_id FROM teammate_thumbs
                   WHERE course_id = ? AND session_key = ?
                     AND popping_version_compatible(data_version, ?) = 1
                   UNION
                   SELECT student_id FROM presentation_ratings
                   WHERE course_id = ? AND session_key = ?
                     AND popping_version_compatible(data_version, ?) = 1
                   UNION
                   SELECT rater_id AS student_id FROM challenge_ratings
                   WHERE course_id = ? AND session_key = ?
                     AND popping_version_compatible(data_version, ?) = 1
               )''',
            [course['id'], state['session_key'] or 0, SCHEMA_VERSION,
             course['id'], state['session_key'] or 0, SCHEMA_VERSION,
             course['id'], state['session_key'] or 0, SCHEMA_VERSION],
            one=True)
        # Top teams by avg presentation rating (extracted to shared helper)
        hist = json.loads(state['presentation_history']) if state and state['presentation_history'] else []
        hist = [
            item for item in hist
            if item.get('session_key', 0) == (state['session_key'] or 0)
        ]
        team_ratings = _compute_top_teams(
            slug, course['id'], hist, state['session_key'] or 0
        )
        challenger_ratings = _compute_top_challengers(
            slug, course['id'], state['session_key'] or 0
        )
        end_stats = {
            'participants': participants['c'] if participants else 0,
            'top_teams': [team for team in team_ratings if team['rank'] <= 3],
            'top_challengers': [
                challenger for challenger in challenger_ratings
                if challenger['rank'] == 1
            ],
        }

    # Track which questions have already been presented
    presented_question_ids = set()
    if state and state['presentation_history']:
        try:
            for h in json.loads(state['presentation_history']):
                if (h.get('session_key', 0) == (state['session_key'] or 0)
                        and 'question_id' in h):
                    presented_question_ids.add(h['question_id'])
        except Exception:
            pass

    return render_template(
        'instructor.html',
        course=course, teams=teams, students=students_enhanced,
        state=state, phases=PHASES, questions=questions,
        max_teams=max_teams,
        max_members=get_max_members_per_team(slug, course['id']),
        teams_locked=teams_locked,
        session_started_at=state['session_started_at'] if state and 'session_started_at' in state.keys() else None,
        end_stats=end_stats,
        presented_question_ids=list(presented_question_ids),
        POLL_DURATION=poll_duration
    )


# ---------------------------------------------------------------------------
# Student API
# ---------------------------------------------------------------------------

def _authenticated_role(slug):
    """Return the role only when the session principal belongs to this course."""
    if not slug or session.get('slug') != slug:
        return None
    role = _exclusive_session_role()
    if role and not _session_course_is_ready(slug):
        return None
    if role:
        ensure_schema(slug)
    if role == 'student' and session.get('student_id'):
        row = _authenticated_student(slug)
        if row:
            g.current_student = row
            g.authenticated_role = 'student'
            _sync_student_activity(slug)
            return 'student'
        return None
    if role == 'instructor' and session.get('instructor_id'):
        if _authenticated_instructor(slug):
            g.authenticated_role = 'instructor'
            return 'instructor'
        return None
    return None


def _bump_roster_version(slug, course_id, db=None):
    """Mark roster/team membership data as changed for polling clients."""
    ensure_schema(slug)
    sql = ('UPDATE course_state '
           'SET roster_version = COALESCE(roster_version, 0) + 1 '
           'WHERE course_id = ?')
    if db is None:
        execute_db(slug, sql, [course_id])
        row = query_db(
            slug,
            'SELECT roster_version FROM course_state WHERE course_id = ?',
            [course_id], one=True,
        )
    else:
        db.execute(sql, [course_id])
        row = db.execute(
            'SELECT roster_version FROM course_state WHERE course_id = ?',
            [course_id],
        ).fetchone()
    return int(row['roster_version'] or 0) if row else 0


def _current_roster_version(db, course_id):
    row = db.execute(
        'SELECT roster_version FROM course_state WHERE course_id = ?',
        [course_id],
    ).fetchone()
    return int(row['roster_version'] or 0) if row else 0


_SESSION_ACTIVITY_TABLES = (
    'teammate_thumbs',
    'presentation_ratings',
    'challenge_rounds',
    'challenge_ratings',
    'presentation_participants',
)
_ROSTER_FROZEN_MESSAGE = (
    'This session already has saved activity. End Session, then start a new '
    'session before changing teams, roster, or capacity.'
)


def _current_session_has_durable_activity(db, course_id, session_key):
    """Return whether saved feedback already depends on this session roster."""
    session_key = session_key or 0
    history_row = db.execute(
        'SELECT presentation_history FROM course_state WHERE course_id = ?',
        [course_id],
    ).fetchone()
    raw_history = history_row['presentation_history'] if history_row else None
    if raw_history and str(raw_history).strip() not in ('', '[]'):
        try:
            history = json.loads(raw_history)
        except (TypeError, ValueError):
            return True
        if not isinstance(history, list):
            return True
        for item in history:
            if not isinstance(item, dict):
                return True
            item_session = item.get('session_key')
            if item_session is None:
                if session_key == 0:
                    return True
                continue
            try:
                item_session = int(item_session)
            except (TypeError, ValueError):
                return True
            if item_session == session_key:
                return True

    for table in _SESSION_ACTIVITY_TABLES:
        if db.execute(
            f'''SELECT 1 FROM {table}
                WHERE course_id = ? AND session_key = ? LIMIT 1''',
            [course_id, session_key],
        ).fetchone():
            return True
    return False


def _session_roster_mutation_guard(db, course_id, state):
    """Protect the team/roster attribution of current-session feedback."""
    if not state:
        return 'Course state is unavailable; reload before changing the roster', 409
    if _current_session_has_durable_activity(
            db, course_id, state['session_key'] or 0):
        return _ROSTER_FROZEN_MESSAGE, 409
    return None


def _participation_counts_by_student(db, course_id, student_ids=None):
    """Return trustworthy course-wide participation counts by roster row."""
    target_ids = []
    if student_ids is not None:
        for student_id in student_ids:
            if student_id is None:
                continue
            try:
                target_ids.append(int(student_id))
            except (TypeError, ValueError):
                continue
        target_ids = list(dict.fromkeys(target_ids))
        if not target_ids:
            return {}

    target_filter = ''
    params = [course_id]
    if target_ids:
        placeholders = ','.join('?' * len(target_ids))
        target_filter = f' AND id IN ({placeholders})'
        params.extend(target_ids)
    params.extend([
        course_id, SCHEMA_VERSION,
        course_id, SCHEMA_VERSION,
    ])
    rows = db.execute(
        f'''WITH target_students AS (
                SELECT id FROM students
                WHERE course_id = ? {target_filter}
            ),
            presentation_counts AS (
                SELECT participant.student_id, COUNT(*) AS presentation_count
                FROM presentation_participants participant
                JOIN target_students target
                  ON target.id = participant.student_id
                WHERE participant.course_id = ?
                  AND typeof(participant.week_num) = 'integer'
                  AND participant.week_num > 0
                  AND popping_version_compatible(
                          participant.data_version, ?) = 1
                GROUP BY participant.student_id
            ),
            challenger_counts AS (
                SELECT challenge.challenger_id AS student_id,
                       COUNT(*) AS challenger_count
                FROM challenge_rounds challenge
                JOIN target_students target
                  ON target.id = challenge.challenger_id
                WHERE challenge.course_id = ?
                  AND typeof(challenge.week_num) = 'integer'
                  AND challenge.week_num > 0
                  AND popping_version_compatible(
                          challenge.data_version, ?) = 1
                GROUP BY challenge.challenger_id
            )
            SELECT target.id AS student_id,
                   COALESCE(presentation.presentation_count, 0)
                       AS presentation_count,
                   COALESCE(challenger.challenger_count, 0)
                       AS challenger_count
            FROM target_students target
            LEFT JOIN presentation_counts presentation
              ON presentation.student_id = target.id
            LEFT JOIN challenger_counts challenger
              ON challenger.student_id = target.id''',
        params,
    ).fetchall()
    return {
        row['student_id']: {
            'presentation_count': int(row['presentation_count'] or 0),
            'challenger_count': int(row['challenger_count'] or 0),
        }
        for row in rows
    }


# Thumbs-up during discussion recognizes a teammate for the whole discussion
# phase (one per teammate per session), not per question.
_DISCUSSION_THUMB_KEY = 'discussion'


def _bump_discussion_questions_version(db, course_id):
    """Signal that the visible discussion-question set changed (hide/show,
    add, delete, or week change). Also trips the state_version trigger, so
    polling students refetch the list within ~1s."""
    db.execute(
        'UPDATE course_state '
        'SET discussion_questions_version = '
        '    COALESCE(discussion_questions_version, 0) + 1 '
        'WHERE course_id = ?',
        [course_id]
    )


def _archive_students(slug, db_ids, bump_roster=True, db=None, commit=True):
    """Remove students from the live roster while preserving historical responses."""
    if not db_ids:
        return
    db = db or get_db(slug)
    ph = ','.join('?' * len(db_ids))
    course_ids = [row['course_id'] for row in db.execute(
        f'SELECT DISTINCT course_id FROM students WHERE id IN ({ph})', db_ids
    ).fetchall()]
    db.execute(
        f'''UPDATE students
            SET is_active = 0,
                last_team_id = COALESCE(last_team_id, team_id),
                team_id = NULL
            WHERE id IN ({ph})''',
        db_ids
    )
    if bump_roster:
        for course_id in course_ids:
            _bump_roster_version(slug, course_id, db=db)
    if commit:
        db.commit()


# ---------------------------------------------------------------------------
# Shared helpers for state/teams computation (used by /api/state, /api/teams,
# and the consolidated /api/poll endpoint).
# ---------------------------------------------------------------------------

def _compute_top_teams(slug, course_id, history, session_key):
    """Compute team rankings by average presentation rating.

    ``history`` is the already-parsed presentation_history list (each item
    is a dict with 'started_at' and 'team' keys).  Pass the parsed list
    directly to avoid a redundant course_state re-query on the poll path.
    Returns ranked team averages with the number of submitted rating forms.
    """
    history_json = history or []
    key_to_team = {}
    for h in history_json:
        if not _history_item_is_compatible(h):
            continue
        qkey = _history_presentation_key(h)
        key_to_team[qkey] = {
            'id': h.get('team_id'),
            'name': h.get('team', 'Unknown'),
        }

    all_ratings = query_db(slug,
        '''SELECT question_key, presenting_team_id, presenting_team_name,
                  q1_developed, q2_easy
           FROM presentation_ratings WHERE course_id = ? AND session_key = ?
             AND popping_version_compatible(data_version, ?) = 1''',
        [course_id, session_key, SCHEMA_VERSION])

    team_scores = {}
    for r in all_ratings:
        historical = key_to_team.get(r['question_key']) or {}
        team_id = r['presenting_team_id'] or historical.get('id')
        team_name = r['presenting_team_name'] or historical.get('name')
        if (not team_name or r['q1_developed'] is None or
                r['q2_easy'] is None):
            continue
        identity = ('id', team_id) if team_id is not None else (
            'name', team_name.casefold()
        )
        score = team_scores.setdefault(identity, {
            'id': team_id, 'name': team_name, 'score_sum': 0,
            'score_count': 0, 'rating_count': 0,
        })
        score['score_sum'] += r['q1_developed'] + r['q2_easy']
        score['score_count'] += 2
        score['rating_count'] += 1

    team_ratings = list(team_scores.values())
    for score in team_ratings:
        score['_fraction'] = Fraction(score['score_sum'], score['score_count'])
    team_ratings.sort(key=lambda score: (
        -score['_fraction'], score['name'].casefold(), score['id'] or 0
    ))
    prior_fraction = None
    for position, score in enumerate(team_ratings, 1):
        if prior_fraction is None or score['_fraction'] != prior_fraction:
            rank = position
        score['rank'] = rank
        score['avg_score'] = round(float(score['_fraction']), 2)
        prior_fraction = score['_fraction']
        score.pop('_fraction')
    return team_ratings


def _compute_top_challengers(slug, course_id, session_key):
    """Compute challenger rankings by average challenge rating.

    Follows the _compute_top_teams pattern: exact Fraction-based averages,
    ties share a rank. Students see name + rank only (no scores) — same
    grading-privacy rule as the team leaderboard.
    """
    all_ratings = query_db(slug,
        '''SELECT cr.challenger_id,
                  student.student_id AS student_id,
                  student.display_name AS display_name,
                  student.name AS roster_name,
                  cr.challenger_name AS snapshot_name,
                  cr.score
           FROM challenge_ratings cr
           LEFT JOIN students student ON student.id = cr.challenger_id
           WHERE cr.course_id = ? AND cr.session_key = ?
             AND popping_version_compatible(cr.data_version, ?) = 1''',
        [course_id, session_key, SCHEMA_VERSION])

    challenger_scores = {}
    for r in all_ratings:
        if r['student_id']:
            name, identity_source = _student_display_identity(r)
        else:
            name = _normalized_identity_text(r['snapshot_name'])
            identity_source = 'snapshot_name'
        if not name or r['score'] is None:
            continue
        identity = ('id', r['challenger_id']) if r['challenger_id'] is not None else (
            'name', name.casefold()
        )
        score = challenger_scores.setdefault(identity, {
            'id': r['challenger_id'], 'student_id': r['student_id'],
            'display_name': r['display_name'],
            'roster_name': r['roster_name'],
            'identity_source': identity_source,
            'name': name, 'score_sum': 0, 'rating_count': 0,
        })
        score['score_sum'] += r['score']
        score['rating_count'] += 1

    ranked = list(challenger_scores.values())
    for score in ranked:
        score['_fraction'] = Fraction(score['score_sum'], score['rating_count'])
    ranked.sort(key=lambda score: (
        -score['_fraction'], score['name'].casefold(), score['id'] or 0
    ))
    prior_fraction = None
    for position, score in enumerate(ranked, 1):
        if prior_fraction is None or score['_fraction'] != prior_fraction:
            rank = position
        score['rank'] = rank
        score['avg_score'] = round(float(score['_fraction']), 2)
        prior_fraction = score['_fraction']
        score.pop('_fraction')
    return ranked


def _compute_state(slug, include_poll_count=True, known_question_id=None,
                   known_question_revision=None):
    """Compute the course-state dict — shared by /api/state and /api/poll.

    This is the single source of truth for presentation timer, poll status,
    active team/question, and the student's own team. Question content is
    omitted only when the client confirms both its ID and content revision.
    """
    state = query_db(slug, 'SELECT * FROM course_state LIMIT 1', one=True)
    if state:
        state = dict(state)  # mutable copy — avoids re-query on poll auto-close

    active_team = None
    if state and state['active_team_id']:
        active_team = query_db(slug,
            'SELECT * FROM teams WHERE id = ?', [state['active_team_id']], one=True)

    my_team = None
    me = None
    if session.get('role') == 'student' and 'student_id' in session:
        me = getattr(g, 'current_student', None)
        if me is None:
            me = query_db(slug,
                'SELECT * FROM students WHERE student_id = ? AND is_active = 1',
                [session['student_id']], one=True)
        if me and me['team_id']:
            my_team = query_db(slug,
                'SELECT * FROM teams WHERE id = ?', [me['team_id']], one=True)

    active_question = None
    if state and state['active_question_id']:
        aq = query_db(slug,
            'SELECT * FROM questions WHERE id = ?', [state['active_question_id']], one=True)
        if aq:
            full_question = dict(aq)
            revision_source = '\0'.join(str(value or '') for value in (
                full_question.get('title'), full_question.get('question_text'),
                full_question.get('content'),
            ))
            revision = hashlib.sha256(
                revision_source.encode('utf-8')
            ).hexdigest()
            if (known_question_id == full_question['id'] and
                    known_question_revision == revision):
                active_question = {
                    'id': full_question['id'],
                    'revision': revision,
                    'content_unchanged': True,
                }
            else:
                full_question['revision'] = revision
                active_question = full_question

    now = _utcnow()
    poll_duration = get_poll_duration(slug)
    timing = _derive_timing_state(state, now=now, poll_duration=poll_duration)
    presentation_remaining = timing['presentation_remaining']

    # Expiry is derived, not written by every polling client.  The instructor's
    # next start/stop action persists the next transition without a write storm.
    poll_active_bool = _poll_is_open(state, now=now, poll_duration=poll_duration)

    # Poll count — use course_id from the already-fetched state row
    # (no separate SELECT needed)
    poll_count = 0
    pres_key = active_presentation_key(state)
    if (include_poll_count and state and pres_key
            and state.get('active_team_id')):
        cid = state['course_id']
        cnt = query_db(slug,
            '''SELECT COUNT(DISTINCT student_id) AS c
               FROM presentation_ratings
               WHERE course_id = ? AND session_key = ? AND question_key = ?
                 AND popping_version_compatible(data_version, ?) = 1''',
            [cid, state.get('session_key', 0), pres_key, SCHEMA_VERSION],
            one=True)
        poll_count = cnt['c'] if cnt else 0

    needs_history = bool(include_poll_count or (state and state['phase'] == 'ended'))
    try:
        all_history = json.loads(state['presentation_history']) \
            if needs_history and state and state['presentation_history'] else []
    except (TypeError, ValueError):
        all_history = []
    current_session_key = state.get('session_key', 0) if state else 0
    history = [
        item for item in all_history
        if (_history_item_is_compatible(item)
            and item.get('session_key', 0) == current_session_key)
    ]

    active_challenges = (
        json.loads(state['active_challenges_json'] or '[]')
        if state else []
    )
    if state and active_challenges:
        active_challenges = _enrich_live_challenge_identities(
            slug, state['course_id'], active_challenges
        )
    result = {
        'phase': state['phase'] if state else 'setup',
        'active_team': dict(active_team) if active_team else None,
        'active_question': active_question,
        'my_team': dict(my_team) if my_team else None,
        'current_question': state['current_question'] if state else None,
        'discussion_questions_version': (
            state.get('discussion_questions_version') or 0) if state else 0,
        'presentation_started_at': state['presentation_started_at'] if state else None,
        'session_started_at': state['session_started_at'] if state else None,
        'presentation_time_cap': state['presentation_time_cap'] if state else 300,
        'presentation_remaining': presentation_remaining,
        'teams_locked': bool(state['teams_locked']) if state else False,
        'max_teams': (state.get('max_teams') or 6) if state else 6,
        'max_members': (
            state.get('max_members_per_team') or 10
        ) if state else 10,
        'discussion_week': state.get('discussion_week', 1) if state else 1,
        'poll_active': poll_active_bool,
        'poll_started_at': state['poll_started_at'] if state else None,
        'poll_duration': poll_duration,
        'poll_remaining': timing['poll_remaining'],
        'ratings_settling': timing['ratings_settling'],
        'ratings_settling_remaining': timing['ratings_settling_remaining'],
        'poll_question_key': state['poll_question_key'] if state else None,
        'roster_version': state.get('roster_version', 0) if state else 0,
        'session_key': state.get('session_key', 0) if state else 0,
        'state_version': state.get('state_version', 0) if state else 0,
        'session_elapsed': timing['session_elapsed'],
        'presentation_history': history,
        # Challenge feature: active challengers for this presentation.
        'active_challenges': active_challenges,
        'challenge_ratings_open': bool(
            state and _has_active_challenges(state)
            and _challenge_ratings_are_open(
                state, now=now, poll_duration=poll_duration
            )
        ),
        # Internal metadata used by api_poll for ended-phase ranking.
        '_course_id': state['course_id'] if state else None,
    }
    if me:
        result['current_student_display_name'] = _student_display_name(me)
        my_student_id = str(me['id'])
        result['is_active_challenger'] = any(
            isinstance(challenge, dict) and
            str(challenge.get('challenger_id')) == my_student_id
            for challenge in active_challenges
        )
    if include_poll_count:
        result['poll_count'] = poll_count or 0
        cid = state['course_id'] if state else None
        if cid:
            unassigned = query_db(
                slug,
                '''SELECT COUNT(*) AS c FROM students
                   WHERE course_id = ? AND team_id IS NULL AND is_active = 1''',
                [cid], one=True
            )
            result['unassigned_count'] = unassigned['c'] if unassigned else 0

            poll_eligible = 0
            if state.get('active_team_id'):
                eligible = query_db(
                    slug,
                    '''SELECT COUNT(*) AS c FROM students
                       WHERE course_id = ? AND team_id IS NOT NULL AND team_id != ?
                         AND is_active = 1''',
                    [cid, state['active_team_id']], one=True
                )
                poll_eligible = eligible['c'] if eligible else 0
            result['poll_eligible_count'] = poll_eligible

            thumb_participants = 0
            thumb_eligible = 0
            thumb_team_progress = []
            if state.get('phase') == 'discussion':
                team_rows = query_db(
                    slug,
                    '''WITH visible_teams AS (
                           SELECT id AS team_id, name AS team_name
                           FROM teams
                           WHERE course_id = ?
                           ORDER BY id
                           LIMIT ?
                       ),
                       team_members AS (
                           SELECT student.team_id,
                                  COUNT(*) AS member_count
                           FROM students student
                           JOIN visible_teams team
                             ON team.team_id = student.team_id
                           WHERE student.course_id = ? AND student.is_active = 1
                           GROUP BY student.team_id
                       ),
                       team_thumbs AS (
                           SELECT grader.team_id,
                                  COUNT(*) AS thumb_count,
                                  COUNT(DISTINCT thumb.grader_id)
                                      AS participant_count
                           FROM teammate_thumbs thumb
                           JOIN students grader
                             ON grader.id = thumb.grader_id
                            AND grader.course_id = thumb.course_id
                            AND grader.is_active = 1
                            AND grader.team_id IS NOT NULL
                           JOIN students recipient
                             ON recipient.id = thumb.recipient_id
                            AND recipient.course_id = thumb.course_id
                            AND recipient.is_active = 1
                            AND recipient.team_id = grader.team_id
                           JOIN visible_teams team
                             ON team.team_id = grader.team_id
                           WHERE thumb.course_id = ? AND thumb.session_key = ?
                             AND thumb.question_key = ?
                             AND popping_version_compatible(
                                     thumb.data_version, ?) = 1
                           GROUP BY grader.team_id
                       )
                       SELECT team.team_id, team.team_name,
                              COALESCE(member.member_count, 0) AS member_count,
                              CASE WHEN COALESCE(member.member_count, 0) > 1
                                   THEN member.member_count ELSE 0 END
                                  AS eligible_count,
                              COALESCE(thumb.participant_count, 0)
                                  AS participant_count,
                              COALESCE(thumb.thumb_count, 0) AS thumb_count
                       FROM visible_teams team
                       LEFT JOIN team_members member
                         ON member.team_id = team.team_id
                       LEFT JOIN team_thumbs thumb
                         ON thumb.team_id = team.team_id
                       ORDER BY team.team_id''',
                    [cid, state.get('max_teams') or 6, cid,
                     cid, state.get('session_key', 0),
                     _DISCUSSION_THUMB_KEY, SCHEMA_VERSION]
                )
                thumb_team_progress = [
                    {
                        'team_id': row['team_id'],
                        'team_name': row['team_name'],
                        'member_count': row['member_count'],
                        'eligible_count': row['eligible_count'],
                        'participant_count': row['participant_count'],
                        'thumb_count': row['thumb_count'],
                    }
                    for row in team_rows
                ]
                thumb_participants = sum(
                    row['participant_count'] for row in team_rows
                )
                thumb_eligible = sum(row['eligible_count'] for row in team_rows)
            result['thumb_participant_count'] = thumb_participants
            result['thumb_eligible_count'] = thumb_eligible
            result['thumb_team_progress'] = thumb_team_progress

            # Challenge feature: instructor-only raised-hands list + per-challenge
            # rating counts, shown alongside the presentation poll controls.
            if pres_key and state.get('active_team_id'):
                hands = query_db(
                    slug,
                    '''SELECT hand.student_id,
                              student.student_id AS student_identifier,
                              student.display_name AS display_name,
                              student.name AS roster_name,
                              COALESCE(NULLIF(TRIM(student.display_name), ''),
                                       NULLIF(TRIM(student.name), ''),
                                       student.student_id,
                                       NULLIF(TRIM(hand.student_name), ''))
                                  AS student_name,
                              hand.student_team_id, hand.student_team_name
                       FROM challenge_hands hand
                       LEFT JOIN students student ON student.id = hand.student_id
                       WHERE hand.course_id = ? AND hand.presentation_key = ?
                       ORDER BY hand.raised_at''',
                    [cid, pres_key])
                active_challenges = result.get('active_challenges') or []
                participation_ids = [hand['student_id'] for hand in hands]
                participation_ids.extend(
                    challenge.get('challenger_id')
                    for challenge in active_challenges
                    if isinstance(challenge, dict)
                )
                participation_counts = _participation_counts_by_student(
                    get_db(slug), cid, participation_ids
                )
                result['challenge_hands'] = [
                    {'student_id': h['student_id'],
                     'student_identifier': h['student_identifier'],
                     'display_name': h['display_name'],
                     'roster_name': h['roster_name'],
                     'student_name': h['student_name'],
                     'student_team_id': h['student_team_id'],
                     'student_team_name': h['student_team_name'],
                     **participation_counts.get(h['student_id'], {
                         'presentation_count': 0,
                         'challenger_count': 0,
                     })}
                    for h in hands
                ]
                result['active_challenges'] = [
                    {
                        **challenge,
                        **participation_counts.get(
                            challenge.get('challenger_id'), {
                                'presentation_count': 0,
                                'challenger_count': 0,
                            }
                        ),
                    }
                    for challenge in active_challenges
                    if isinstance(challenge, dict)
                ]
                challenge_keys = list(dict.fromkeys(
                    str(ch.get('challenge_key'))
                    for ch in result['active_challenges']
                    if ch.get('challenge_key')
                ))
                challenge_summaries = {
                    key: {
                        'submitted_count': 0,
                        'eligible_count': 0,
                    }
                    for key in challenge_keys
                }
                if challenge_keys:
                    placeholders = ','.join('?' * len(challenge_keys))
                    summary_rows = query_db(
                        slug,
                        f'''WITH selected_challenges AS (
                                SELECT course_id, challenge_key, challenger_id,
                                       challenger_team_id, presenting_team_id
                                FROM challenge_rounds
                                WHERE course_id = ?
                                  AND popping_version_compatible(
                                          data_version, ?) = 1
                                  AND challenge_key IN ({placeholders})
                            ),
                            submitted AS (
                                SELECT rating.challenge_key,
                                       COUNT(*) AS submitted_count
                                FROM challenge_ratings rating
                                JOIN selected_challenges challenge
                                  ON challenge.challenge_key = rating.challenge_key
                                 AND challenge.course_id = rating.course_id
                                WHERE popping_version_compatible(
                                          rating.data_version, ?) = 1
                                GROUP BY rating.challenge_key
                            ),
                            eligible AS (
                                SELECT challenge.challenge_key,
                                       COUNT(student.id) AS eligible_count
                                FROM selected_challenges challenge
                                LEFT JOIN students student
                                  ON student.course_id = challenge.course_id
                                 AND student.is_active = 1
                                 AND student.team_id IS NOT NULL
                                 AND (challenge.challenger_id IS NULL
                                      OR student.id != challenge.challenger_id)
                                 AND (challenge.challenger_team_id IS NULL
                                      OR student.team_id !=
                                         challenge.challenger_team_id)
                                 AND (challenge.presenting_team_id IS NULL
                                      OR student.team_id !=
                                         challenge.presenting_team_id)
                                GROUP BY challenge.challenge_key
                            )
                            SELECT challenge.challenge_key,
                                   COALESCE(submitted.submitted_count, 0)
                                       AS submitted_count,
                                   COALESCE(eligible.eligible_count, 0)
                                       AS eligible_count
                            FROM selected_challenges challenge
                            LEFT JOIN submitted
                              ON submitted.challenge_key = challenge.challenge_key
                            LEFT JOIN eligible
                              ON eligible.challenge_key = challenge.challenge_key''',
                        [cid, SCHEMA_VERSION] + challenge_keys + [SCHEMA_VERSION]
                    )
                    for row in summary_rows:
                        challenge_summaries[row['challenge_key']] = {
                            'submitted_count': row['submitted_count'],
                            'eligible_count': row['eligible_count'],
                        }
                result['challenge_rating_summaries'] = challenge_summaries
            else:
                result['challenge_hands'] = []
                result['challenge_rating_summaries'] = {}
        else:
            result.update({
                'unassigned_count': 0, 'poll_eligible_count': 0,
                'thumb_participant_count': 0, 'thumb_eligible_count': 0,
                'thumb_team_progress': [],
                'challenge_hands': [],
                'challenge_rating_summaries': {},
            })
        result['completed_presentation_count'] = len(history)
        result['presentation_number'] = len(history) + 1 \
            if active_team and active_question else None
    return result


def _compute_teams(slug, course_id, max_teams=None, member_team_id=None,
                   include_all_members=True,
                   include_participation_counts=False):
    """Compute the teams + members list for the versioned roster endpoint.

    Uses two roster queries instead of N+1, plus one aggregate participation
    query when instructor-only counts are requested.
    Pass ``max_teams`` when the caller already has the visible-team limit.
    """
    if max_teams is None:
        max_teams = get_max_teams(slug, course_id)
    teams = query_db(slug,
        '''SELECT team.*, COUNT(student.id) AS member_count
           FROM teams team
           LEFT JOIN students student
             ON student.team_id = team.id AND student.is_active = 1
           WHERE team.course_id = ?
           GROUP BY team.id ORDER BY team.id LIMIT ?''',
        [course_id, max_teams])
    if not teams:
        return []
    team_ids = [team['id'] for team in teams]
    visible_member_team_ids = team_ids if include_all_members else (
        [member_team_id] if member_team_id in team_ids else []
    )
    all_members = []
    if visible_member_team_ids:
        placeholders = ','.join('?' * len(visible_member_team_ids))
        all_members = query_db(slug,
            f'''SELECT id, student_id, name, display_name, team_id
                FROM students
                WHERE team_id IN ({placeholders}) AND is_active = 1
                ORDER BY team_id, COALESCE(NULLIF(TRIM(display_name), ''),
                                           NULLIF(TRIM(name), ''), student_id)''',
            visible_member_team_ids)
    participation_counts = {}
    if include_participation_counts:
        participation_counts = _participation_counts_by_student(
            get_db(slug), course_id, [member['id'] for member in all_members]
        )
    members_by_team = {}
    for m in all_members:
        resolved_name = _student_display_name(m)
        member = {
            'student_id': m['student_id'],
            'display_name': m['display_name'],
            'name': resolved_name,
        }
        if include_participation_counts:
            member['roster_name'] = m['name']
        if include_participation_counts:
            member['id'] = m['id']
            member.update(participation_counts.get(m['id'], {
                'presentation_count': 0,
                'challenger_count': 0,
            }))
        members_by_team.setdefault(m['team_id'], []).append(member)
    return [
        {'id': team['id'], 'name': team['name'], 'color': team['color'],
         'member_count': team['member_count'],
         'members_visible': include_all_members or team['id'] == member_team_id,
         'members': members_by_team.get(team['id'], [])}
        for team in teams
    ]


@app.route('/api/teams', methods=['GET'])
def api_teams():
    slug = session.get('slug')
    role = _authenticated_role(slug)
    if not role:
        return jsonify({'error': 'Not logged in'}), 401
    course = query_db(slug, 'SELECT * FROM courses LIMIT 1', one=True)
    if role == 'instructor':
        return jsonify(_compute_teams(
            slug, course['id'],
            include_participation_counts=True,
        ))
    student = query_db(
        slug,
        '''SELECT team_id FROM students
           WHERE course_id = ? AND student_id = ? AND is_active = 1''',
        [course['id'], session['student_id']], one=True
    )
    return jsonify(_compute_teams(
        slug, course['id'],
        member_team_id=student['team_id'] if student else None,
        include_all_members=True,
    ))


@app.route('/api/join_team', methods=['POST'])
@student_login_required
def join_team():
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    team_id = data.get('team_id')
    if team_id is None:
        return jsonify({'error': 'Team ID required'}), 400
    if team_id:
        try:
            team_id = int(team_id)
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid team ID'}), 400

    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        student = db.execute(
            '''SELECT s.* FROM students s JOIN courses c ON c.id = s.course_id
               WHERE s.student_id = ? AND c.slug = ? AND s.is_active = 1''',
            [session['student_id'], slug]
        ).fetchone()
        state = db.execute(
            'SELECT * FROM course_state WHERE course_id = ?',
            [student['course_id'] if student else -1]
        ).fetchone()
        if not student:
            db.rollback()
            return jsonify({'error': 'Student not found'}), 404
        if not state or state['phase'] != 'setup':
            db.rollback()
            return jsonify({'error': 'Team selection is closed'}), 403
        if state['teams_locked']:
            db.rollback()
            return jsonify({'error': 'Teams are currently locked by the instructor'}), 403

        # Leaving team (team_id = 0 means unassign).
        if not team_id:
            if student['team_id'] is None:
                roster_version = _current_roster_version(db, state['course_id'])
                db.commit()
                return jsonify({'success': True, 'roster_version': roster_version})
            freeze_guard = _session_roster_mutation_guard(
                db, state['course_id'], state
            )
            if freeze_guard:
                db.rollback()
                return jsonify({'error': freeze_guard[0]}), freeze_guard[1]
            db.execute(
                '''UPDATE students
                   SET last_team_id = COALESCE(last_team_id, team_id),
                       team_id = NULL
                   WHERE id = ?''',
                [student['id']]
            )
            roster_version = _bump_roster_version(
                slug, state['course_id'], db=db
            )
            db.commit()
            return jsonify({'success': True, 'roster_version': roster_version})
        if student['team_id'] == team_id:
            roster_version = _current_roster_version(db, state['course_id'])
            db.commit()
            return jsonify({'success': True, 'roster_version': roster_version})

        visible = db.execute(
            'SELECT id FROM teams WHERE course_id = ? ORDER BY id LIMIT ?',
            [state['course_id'], state['max_teams'] or 6]
        ).fetchall()
        if team_id not in {row['id'] for row in visible}:
            db.rollback()
            return jsonify({'error': 'Team is not available'}), 400

        member_count = db.execute(
            'SELECT COUNT(*) AS c FROM students WHERE team_id = ? AND is_active = 1', [team_id]
        ).fetchone()['c']
        if member_count >= (state['max_members_per_team'] or 10):
            db.rollback()
            return jsonify({'error': 'That team is full'}), 409

        freeze_guard = _session_roster_mutation_guard(
            db, state['course_id'], state
        )
        if freeze_guard:
            db.rollback()
            return jsonify({'error': freeze_guard[0]}), freeze_guard[1]
        db.execute(
            '''UPDATE students
               SET team_id = ?, last_team_id = ?, last_team_joined_at = CURRENT_TIMESTAMP
               WHERE id = ?''',
            [team_id, team_id, student['id']]
        )
        roster_version = _bump_roster_version(slug, state['course_id'], db=db)
        db.commit()
        return jsonify({'success': True, 'roster_version': roster_version})
    except Exception:
        db.rollback()
        raise


@app.route('/api/state', methods=['GET'])
def api_state():
    """Course state — accessible to both students and instructors."""
    slug = session.get('slug')
    if not slug:
        return jsonify({'error': 'Not logged in'}), 401
    role = _authenticated_role(slug)
    if not role:
        return jsonify({'error': 'Not logged in'}), 401
    is_instructor = role == 'instructor'
    state_data = _compute_state(
        slug,
        include_poll_count=is_instructor,
        known_question_id=request.args.get('known_question_id', type=int),
        known_question_revision=request.args.get('known_question_revision'),
    )
    # Strip internal + grading metadata from student responses
    state_data.pop('_course_id', None)
    if role == 'student':
        state_data.pop('presentation_history', None)
    return jsonify(state_data)


@app.route('/api/poll', methods=['GET'])
def api_poll():
    """Lightweight polling endpoint returning frequently changing state.

    Roster data is fetched separately only when ``roster_version`` changes.
    Clients may send a known question ID and revision so unchanged question
    bodies are omitted. ``poll_interval`` lets clients adapt to the phase.
    """
    slug = session.get('slug')
    if not slug:
        return jsonify({'error': 'Not logged in'}), 401
    role = _authenticated_role(slug)
    if not role:
        return jsonify({'error': 'Not logged in'}), 401

    # Active use keeps a demo instance alive even past the 2h TTL marker.
    if is_demo_instance_slug(slug):
        _touch_demo_instance_throttled(slug)

    ensure_schema(slug)

    is_instructor = role == 'instructor'
    if is_instructor:
        _sync_instructor_catalog_once(slug)

    # --- Cheap "nothing changed" path (students only) ---
    # A student that already has a full snapshot may send ``since=<version>``.
    # When course_state hasn't been written since, we skip the expensive
    # ``_compute_state`` work and return a tiny ``changed: false`` response.
    # This makes ~1s student polling affordable, so instructor actions (phase
    # changes, posted questions, started polls, team selections) reach students
    # within about a second instead of the old 2-5s per-phase interval.
    #
    # Instructors always use the full path: their participation counts change
    # as students act, without any course_state write.
    known_version = request.args.get('since', type=int)
    if known_version is not None and not is_instructor:
        version_row = query_db(
            slug,
            '''SELECT state_version, phase, session_key,
                      poll_active, poll_started_at, poll_closed_at
               FROM course_state LIMIT 1''',
            one=True,
        )
        current_version = (
            version_row['state_version'] or 0
        ) if version_row else 0
        current_phase = (
            version_row['phase'] or 'setup'
        ) if version_row else 'setup'
        # An open rating window's expiry is *derived* from time, so it can flip
        # without a course_state write. Students already run the countdown
        # locally, so an unchanged version can still use the compact path. The
        # poll_closed flag below supplies the authoritative close signal when
        # the server-side cutoff is reached.
        poll_window_persisted = bool(
            version_row
            and version_row['poll_active']
            and version_row['poll_started_at']
        )
        poll_window_live = _poll_is_open(
            version_row,
            now=_utcnow(),
            poll_duration=get_poll_duration(slug),
        )
        if (version_row is not None
                and current_version == known_version):
            # Keep presence fresh. This is throttled to ~30s and only touches
            # the students table (no course_state write), so the state-version
            # short-circuit above stays valid.
            _sync_student_activity(slug, version_row['session_key'] or 0)
            cheap_interval = 5000 if current_phase == 'ended' else 1000
            if is_demo_instance_slug(slug):
                cheap_interval = max(cheap_interval, 5000)
            return jsonify({
                'changed': False,
                'state_version': current_version,
                'poll_interval': cheap_interval,
                'poll_closed': poll_window_persisted and not poll_window_live,
            })

    # --- State ---
    state_data = _compute_state(
        slug,
        include_poll_count=is_instructor,
        known_question_id=request.args.get('known_question_id', type=int),
        known_question_revision=request.args.get('known_question_revision'),
    )
    if role == 'student':
        _sync_student_activity(slug, state_data.get('session_key'))
    course_id = state_data.pop('_course_id')

    # Strip grading metadata from student responses — students never see
    # poll counts or presentation history (rating counts per team).
    # Save history first — needed for top-teams computation below.
    pres_history = state_data.get('presentation_history', [])
    if role == 'student':
        state_data.pop('presentation_history', None)

    # --- Top 3 teams (only when session has ended, students only) ---
    top_teams = None
    if state_data['phase'] == 'ended' and role == 'student':
        all_ranked = _compute_top_teams(
            slug, course_id, pres_history, state_data.get('session_key', 0)
        )
        # Students see names and tied medal ranks, but no scores.
        top_teams = [
            {'name': team['name'], 'rank': team['rank']}
            for team in all_ranked if team['rank'] <= 3
        ]

    # --- Best Challenger (only when session has ended, students only) ---
    # Same privacy rule as top teams: name + tied medal rank, no scores.
    top_challengers = None
    if state_data['phase'] == 'ended' and role == 'student':
        all_challengers = _compute_top_challengers(
            slug, course_id, state_data.get('session_key', 0)
        )
        top_challengers = [
            {
                'name': c['name'], 'student_id': c.get('student_id'),
                'rank': c['rank'],
            }
            for c in all_challengers if c['rank'] == 1
        ]

    # --- Adaptive interval hint ---
    # 1s during all active phases so instructor actions (phase changes, posting
    # questions, starting polls, selecting teams) appear within 1 second.
    # Keep a low-frequency recovery poll after End Session so another session
    # can start without requiring every student to reload manually. ~5s is
    # slow enough to be cheap but fast enough that students sitting on the
    # results screen see a new session start promptly.
    stop_polling = False
    if state_data['phase'] == 'ended':
        poll_interval = 5000
    elif role == 'instructor':
        poll_interval = 1000
    else:
        # The cheap path above makes ~1s student polling affordable, so phase
        # transitions and posted questions reach students within ~1 second.
        poll_interval = 1000
    if is_demo_instance_slug(slug):
        poll_interval = max(poll_interval, 5000)

    return jsonify({
        'changed': True,
        'state': state_data,
        'state_version': state_data.get('state_version', 0),
        'top_teams': top_teams,
        'top_challengers': top_challengers,
        'poll_interval': poll_interval,
        'stop_polling': stop_polling,
    })


@app.route('/api/grade_peer', methods=['POST'])
@student_login_required
def grade_peer():
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    recipient_sid = data.get('recipient_id')
    selected = data.get('selected')
    if recipient_sid is None:
        return jsonify({'error': 'Recipient is required'}), 400
    if not isinstance(selected, bool):
        return jsonify({'error': 'Selected must be true or false'}), 400

    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        grader = db.execute(
            'SELECT * FROM students WHERE student_id = ? AND is_active = 1',
            [session['student_id']]
        ).fetchone()
        recipient = db.execute(
            '''SELECT * FROM students
               WHERE student_id = ? AND course_id = ? AND is_active = 1''',
            [str(recipient_sid), grader['course_id'] if grader else -1]
        ).fetchone()
        state = db.execute(
            'SELECT * FROM course_state WHERE course_id = ?',
            [grader['course_id'] if grader else -1]
        ).fetchone()
        if not grader:
            db.rollback()
            return jsonify({'error': 'Grader not found'}), 404
        if not recipient:
            db.rollback()
            return jsonify({'error': 'Recipient not found'}), 404
        if grader['id'] == recipient['id']:
            db.rollback()
            return jsonify({'error': 'Cannot grade yourself'}), 400
        if not state or state['phase'] != 'discussion':
            db.rollback()
            return jsonify({'error': 'Teammate thumbs are only open during discussion'}), 403
        if not grader['team_id'] or recipient['team_id'] != grader['team_id']:
            db.rollback()
            return jsonify({'error': 'You can only grade teammates'}), 403

        identity_params = [
            grader['course_id'], state['session_key'] or 0,
            _DISCUSSION_THUMB_KEY, grader['id'], recipient['id'],
        ]
        if selected:
            team_rows = db.execute(
                '''SELECT id, name FROM teams
                   WHERE course_id = ? AND id IN (?, ?)''',
                [grader['course_id'], grader['team_id'], recipient['team_id']]
            ).fetchall()
            team_names = {row['id']: row['name'] for row in team_rows}
            db.execute(
                '''INSERT INTO teammate_thumbs
                   (course_id, session_key, week_num, question_key,
                    source_question_key, question_title, grader_id, recipient_id,
                    grader_team_id, grader_team_name,
                    recipient_team_id, recipient_team_name, data_version)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                   ON CONFLICT(course_id, session_key, question_key, grader_id, recipient_id)
                   DO UPDATE SET
                       week_num = COALESCE(
                           teammate_thumbs.week_num,
                           excluded.week_num
                       ),
                       source_question_key = COALESCE(
                           teammate_thumbs.source_question_key,
                           excluded.source_question_key
                       ),
                       question_title = COALESCE(
                           teammate_thumbs.question_title,
                           excluded.question_title
                       ),
                       grader_team_id = COALESCE(
                           teammate_thumbs.grader_team_id,
                           excluded.grader_team_id
                       ),
                       grader_team_name = COALESCE(
                           teammate_thumbs.grader_team_name,
                           excluded.grader_team_name
                       ),
                       recipient_team_id = COALESCE(
                           teammate_thumbs.recipient_team_id,
                           excluded.recipient_team_id
                       ),
                       recipient_team_name = COALESCE(
                           teammate_thumbs.recipient_team_name,
                           excluded.recipient_team_name
                       ),
                       updated_at = CURRENT_TIMESTAMP''',
                [
                    grader['course_id'], state['session_key'] or 0,
                    state['discussion_week'] or 1,
                    _DISCUSSION_THUMB_KEY,
                    _DISCUSSION_THUMB_KEY,
                    '',
                    grader['id'], recipient['id'],
                    grader['team_id'], team_names.get(grader['team_id'], ''),
                    recipient['team_id'], team_names.get(recipient['team_id'], ''),
                    APP_VERSION,
                ]
            )
        else:
            db.execute(
                '''DELETE FROM teammate_thumbs
                   WHERE course_id = ? AND session_key = ? AND question_key = ?
                     AND grader_id = ? AND recipient_id = ?''',
                identity_params
            )
        db.commit()
        return jsonify({'success': True, 'selected': selected})
    except Exception:
        db.rollback()
        raise


@app.route('/api/my_responses', methods=['GET'])
@student_login_required
def my_responses():
    """Return only this student's saved controls for the current live context."""
    slug = session['slug']
    ensure_schema(slug)
    student = query_db(
        slug, 'SELECT * FROM students WHERE student_id = ? AND is_active = 1',
        [session['student_id']], one=True
    )
    state = query_db(
        slug, 'SELECT * FROM course_state WHERE course_id = ?',
        [student['course_id']], one=True
    )
    thumb_recipient_ids = []
    if state and state['phase'] == 'discussion':
        rows = query_db(
            slug,
            '''SELECT recipient.student_id FROM teammate_thumbs thumb
               JOIN students recipient ON recipient.id = thumb.recipient_id
               WHERE thumb.course_id = ? AND thumb.session_key = ?
                 AND thumb.question_key = ? AND thumb.grader_id = ?
                 AND popping_version_compatible(
                         thumb.data_version, ?) = 1
               ORDER BY recipient.student_id''',
            [student['course_id'], state['session_key'] or 0,
             _DISCUSSION_THUMB_KEY, student['id'], SCHEMA_VERSION]
        )
        thumb_recipient_ids = [row['student_id'] for row in rows]

    presentation_key = active_presentation_key(state)
    rating = None
    if presentation_key:
        saved = query_db(
            slug,
            '''SELECT q1_developed, q2_easy FROM presentation_ratings
               WHERE course_id = ? AND student_id = ? AND question_key = ?
                 AND popping_version_compatible(data_version, ?) = 1''',
            [student['course_id'], student['id'], presentation_key,
             SCHEMA_VERSION], one=True
        )
        if saved:
            rating = {
                'q1_developed': saved['q1_developed'],
                'q2_easy': saved['q2_easy'],
            }
    challenge_ratings = {}
    challenge_hand_raised = False
    if presentation_key:
        saved_challenges = query_db(
            slug,
            '''SELECT challenge_key, score FROM challenge_ratings
               WHERE course_id = ? AND rater_id = ?
                 AND presentation_key = ?
                 AND popping_version_compatible(data_version, ?) = 1''',
            [student['course_id'], student['id'], presentation_key,
             SCHEMA_VERSION],
        )
        challenge_ratings = {
            row['challenge_key']: row['score'] for row in saved_challenges
        }
        raised = query_db(
            slug,
            '''SELECT 1 FROM challenge_hands
               WHERE course_id = ? AND student_id = ?
                 AND presentation_key = ?''',
            [student['course_id'], student['id'], presentation_key],
            one=True,
        )
        challenge_hand_raised = bool(raised)

    return jsonify({
        'phase': state['phase'] if state else 'setup',
        'session_key': (state['session_key'] or 0) if state else 0,
        'thumb_recipient_ids': thumb_recipient_ids,
        'presentation_key': presentation_key,
        'rating': rating,
        'challenge_ratings': challenge_ratings,
        'challenge_hand_raised': challenge_hand_raised,
    })



# ---------------------------------------------------------------------------
# Instructor API
# ---------------------------------------------------------------------------

def _presentation_history_match(state, presentation_key):
    """Classify a history key as belonging to this or another session."""
    try:
        history = json.loads(state['presentation_history'] or '[]')
    except (TypeError, ValueError):
        return None
    current_session_key = state['session_key'] or 0
    matched_other_session = False
    for item in history:
        if (not isinstance(item, dict) or
                _history_presentation_key(item) != presentation_key):
            continue
        try:
            item_session_key = int(item.get('session_key') or 0)
        except (TypeError, ValueError):
            item_session_key = 0
        if item_session_key == current_session_key:
            return 'current_session'
        matched_other_session = True
    return 'other_session' if matched_other_session else None


def _finalize_active_presentation(slug, course_id, db=None):
    """Append the active presentation to history and clear it in one transaction."""
    owns_transaction = db is None
    if db is None:
        db = get_db(slug)
        db.execute('BEGIN IMMEDIATE')
    try:
        state = db.execute(
            'SELECT * FROM course_state WHERE course_id = ?', [course_id]
        ).fetchone()
        if not state:
            if owns_transaction:
                db.commit()
            return False

        try:
            history = json.loads(state['presentation_history'] or '[]')
        except (TypeError, ValueError):
            history = []

        if state['active_team_id'] and state['active_question_id']:
            team = db.execute(
                'SELECT name FROM teams WHERE id = ? AND course_id = ?',
                [state['active_team_id'], course_id]
            ).fetchone()
            question = db.execute(
                'SELECT question_text, title FROM questions WHERE id = ? AND course_id = ?',
                [state['active_question_id'], course_id]
            ).fetchone()
            presentation_key = active_presentation_key(state)
            team_name = team['name'] if team else 'Unknown'
            # Deliberate dedup: finalizing the same presentation_key twice
            # (e.g. a double-finalize race) must not grant a second turn.
            # Genuine re-presentations always get a fresh pres-<uuid> key
            # from start_presentation, so they still count separately.
            db.execute(
                '''INSERT INTO presentation_participants
                   (course_id, session_key, week_num, presentation_key,
                    student_id, student_identifier, student_name,
                    team_id, team_name, data_version)
                   SELECT student.course_id, ?, ?, ?, student.id,
                          student.student_id,
                          COALESCE(NULLIF(TRIM(student.display_name), ''),
                                   NULLIF(TRIM(student.name), ''),
                                   student.student_id),
                          ?, ?, ?
                   FROM students student
                   WHERE student.course_id = ? AND student.team_id = ?
                     AND student.is_active = 1
                   ON CONFLICT(course_id, presentation_key, student_id)
                   DO NOTHING''',
                [state['session_key'] or 0, state['discussion_week'] or 1,
                 presentation_key, state['active_team_id'], team_name,
                 APP_VERSION, course_id, state['active_team_id']]
            )
            count = db.execute(
                '''SELECT COUNT(DISTINCT student_id) AS c FROM presentation_ratings
                   WHERE course_id = ? AND question_key = ?
                     AND popping_version_compatible(data_version, ?) = 1''',
                [course_id, presentation_key, SCHEMA_VERSION]
            ).fetchone()
            title = (question['title'] or question['question_text']) \
                if question else (state['current_question'] or '')
            started_at = state['presentation_created_at'] or state['presentation_started_at'] or ''
            history.append({
                'data_version': APP_VERSION,
                'presentation_key': presentation_key,
                'session_key': state['session_key'] or 0,
                'week_num': state['discussion_week'] or 1,
                'title': title,
                'team_id': state['active_team_id'],
                'team': team_name,
                'responses': count['c'] if count else 0,
                'started_at': started_at,
                'question_id': state['active_question_id'],
                'ended_at': _utcnow().strftime('%Y-%m-%d %H:%M:%S'),
                'challenges': [
                    {
                        'challenge_key': ch['challenge_key'],
                        'challenge_num': ch['challenge_num'],
                        'challenger_id': ch['challenger_id'],
                        'challenger_name': ch['challenger_name'],
                        'challenger_team_id': ch['challenger_team_id'],
                        'challenger_team_name': ch['challenger_team_name'],
                    }
                    for ch in (
                        db.execute(
                            '''SELECT challenge_key, challenge_num,
                                      challenger_id, challenger_name,
                                      challenger_team_id, challenger_team_name
                               FROM challenge_rounds
                               WHERE course_id = ? AND presentation_key = ?
                                 AND popping_version_compatible(
                                         data_version, ?) = 1
                               ORDER BY challenge_num''',
                            [course_id, presentation_key, SCHEMA_VERSION]
                        ).fetchall()
                    )
                ],
            })

        # Clean up stale raised-hand records for this presentation
        # (do this BEFORE _clear_active_presentation nulls the state fields).
        if state['active_team_id'] and state['active_question_id']:
            db.execute(
                '''DELETE FROM challenge_hands
                   WHERE course_id = ? AND presentation_key = ?''',
                [course_id, presentation_key]
            )
        _clear_active_presentation(db, course_id, json.dumps(history))
        if owns_transaction:
            db.commit()
        return bool(state['active_team_id'] and state['active_question_id'])
    except Exception:
        if owns_transaction:
            db.rollback()
        raise


def _clear_active_presentation(db, course_id, history_json=None):
    """Clear live presentation state, optionally replacing stored history."""
    assignments = [
        'active_team_id = NULL', 'active_question_id = NULL',
        'current_question = NULL', 'presentation_started_at = NULL',
        'presentation_created_at = NULL', 'presentation_time_cap = 300',
        'presentation_remaining = NULL', 'poll_active = 0',
        'poll_question_key = NULL', 'poll_started_at = NULL',
        'poll_closed_at = NULL',
        'challenge_ratings_closed_at = NULL',
        'active_challenges_json = ?',
    ]
    params = ["[]"]
    if history_json is not None:
        assignments.append('presentation_history = ?')
        params.append(history_json)
    params.append(course_id)
    db.execute(
        f"UPDATE course_state SET {', '.join(assignments)} WHERE course_id = ?",
        params
    )

@app.route('/api/set_phase', methods=['POST'])
@instructor_login_required
def set_phase():
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    phase = data.get('phase')
    if phase not in PHASES:
        return jsonify({'error': 'Invalid phase'}), 400
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        course = db.execute('SELECT id FROM courses LIMIT 1').fetchone()
        state = db.execute(
            'SELECT * FROM course_state WHERE course_id = ?', [course['id']]
        ).fetchone()
        guard = _expected_state_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        displayed_presentation = str(data.get('presentation_key') or '').strip()
        current_presentation = active_presentation_key(state) or ''
        if displayed_presentation != current_presentation:
            db.rollback()
            return jsonify({
                'error': 'This instructor page is stale; reload before changing the course state'
            }), 409
        old_phase = state['phase'] if state else 'setup'
        entering_interactive_phase = (
            old_phase in ('setup', 'ended')
            and phase in ('discussion', 'competition')
        )
        if entering_interactive_phase:
            roster_guard = _expected_roster_state_guard(data, state)
            if roster_guard:
                db.rollback()
                return jsonify({'error': roster_guard[0]}), roster_guard[1]
            unassigned_count = db.execute(
                '''SELECT COUNT(*) AS c FROM students
                   WHERE course_id = ? AND is_active = 1 AND team_id IS NULL''',
                [course['id']],
            ).fetchone()['c']
            if (unassigned_count > 0 and
                    data.get('confirm_unassigned_students') is not True):
                db.rollback()
                return jsonify({
                    'error': (
                        f'{unassigned_count} enrolled student'
                        f'{"s are" if unassigned_count != 1 else " is"} '
                        'not assigned to a team. Confirm before starting '
                        f'{PHASE_LABELS[phase]}.'
                    ),
                    'requires_confirmation': True,
                    'confirmation_type': 'unassigned_students',
                    'unassigned_count': unassigned_count,
                }), 409
        if (phase == 'ended' and old_phase != 'ended' and
                data.get('confirm_end_session') is not True):
            db.rollback()
            return jsonify({
                'error': 'Confirm before ending the session',
                'requires_confirmation': True,
            }), 409
        if old_phase == 'competition' and phase != 'competition':
            if state['active_team_id'] or state['active_question_id']:
                guard = _presentation_guard(data, state)
                if guard:
                    db.rollback()
                    return jsonify({'error': guard[0]}), guard[1]
                transition, ratings_changed = _prepare_rating_transition(
                    db,
                    state,
                    'Stop the active rating poll before leaving this phase',
                    now=g.request_arrived_at,
                    poll_duration=get_poll_duration(slug),
                )
                if transition:
                    if ratings_changed:
                        db.commit()
                    else:
                        db.rollback()
                    return jsonify(transition), 409
            _finalize_active_presentation(slug, course['id'], db=db)

        session_increment = 1 if old_phase == 'ended' and phase != 'ended' else 0
        if old_phase != phase:
            db.execute(
                '''UPDATE course_state
                   SET phase = ?,
                       session_key = COALESCE(session_key, 0) + ?,
                       current_question = NULL,
                       current_discussion_key = NULL,
                       current_discussion_source_key = NULL,
                       current_discussion_title = NULL,
                       current_discussion_content = NULL,
                       poll_active = CASE WHEN ? = 'competition' THEN poll_active ELSE 0 END,
                       poll_started_at = CASE WHEN ? = 'competition' THEN poll_started_at ELSE NULL END,
                       session_started_at = CASE
                           WHEN ? = 'ended' OR ? > 0 THEN NULL
                           ELSE session_started_at END
                   WHERE course_id = ?''',
                [phase, session_increment, phase, phase, phase,
                 session_increment, course['id']]
            )
        fresh = db.execute(
            'SELECT session_key FROM course_state WHERE course_id = ?', [course['id']]
        ).fetchone()
        db.commit()
    except Exception:
        db.rollback()
        raise
    return jsonify({
        'success': True,
        'phase': phase,
        'session_key': fresh['session_key'] if fresh else 0,
    })


@app.route('/api/set_max_teams', methods=['POST'])
@instructor_login_required
def set_max_teams():
    slug = session['slug']
    ensure_schema(slug)
    data = request.get_json(silent=True) or {}
    new_max = data.get('max_teams')
    if not isinstance(new_max, int) or new_max < 1 or new_max > 20:
        return jsonify({'error': 'Team count must be between 1 and 20'}), 400
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        course = db.execute('SELECT id FROM courses LIMIT 1').fetchone()
        state = db.execute(
            '''SELECT phase, session_key, roster_version, max_teams
               FROM course_state WHERE course_id = ?''',
            [course['id']]
        ).fetchone()
        guard = _expected_roster_state_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        if not state or state['phase'] != 'setup':
            db.rollback()
            return jsonify({'error': 'Team settings can only change during setup'}), 409
        available = db.execute(
            'SELECT COUNT(*) AS c FROM teams WHERE course_id = ?', [course['id']]
        ).fetchone()['c']
        if new_max > available:
            db.rollback()
            return jsonify({'error': f'Only {available} teams are available'}), 400
        current_max = state['max_teams'] or available
        if new_max == current_max:
            roster_version = _current_roster_version(db, course['id'])
            db.commit()
            return jsonify({
                'success': True,
                'max_teams': new_max,
                'roster_version': roster_version,
            })
        freeze_guard = _session_roster_mutation_guard(db, course['id'], state)
        if freeze_guard:
            db.rollback()
            return jsonify({'error': freeze_guard[0]}), freeze_guard[1]
        if new_max < current_max:
            hidden = db.execute(
                'SELECT id FROM teams WHERE course_id = ? ORDER BY id LIMIT -1 OFFSET ?',
                [course['id'], new_max]
            ).fetchall()
            if hidden:
                ids = [team['id'] for team in hidden]
                placeholders = ','.join('?' * len(ids))
                db.execute(
                    f'''UPDATE students
                        SET last_team_id = COALESCE(last_team_id, team_id),
                            team_id = NULL
                        WHERE course_id = ? AND team_id IN ({placeholders})
                          AND is_active = 1''',
                    [course['id']] + ids
                )
        db.execute(
            '''UPDATE course_state
               SET max_teams = ?, roster_version = COALESCE(roster_version, 0) + 1
               WHERE course_id = ?''',
            [new_max, course['id']]
        )
        roster_version = _current_roster_version(db, course['id'])
        db.commit()
        return jsonify({
            'success': True,
            'max_teams': new_max,
            'roster_version': roster_version,
        })
    except Exception:
        db.rollback()
        raise


@app.route('/api/random_assign', methods=['POST'])
@instructor_login_required
def random_assign():
    slug = session['slug']
    import random as rnd
    data = request.get_json(silent=True) or {}
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        course = db.execute('SELECT id FROM courses LIMIT 1').fetchone()
        state = db.execute(
            '''SELECT phase, session_key, roster_version, max_teams,
                      max_members_per_team
               FROM course_state WHERE course_id = ?''',
            [course['id']]
        ).fetchone()
        guard = _expected_roster_state_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        if not state or state['phase'] != 'setup':
            db.rollback()
            return jsonify({'error': 'Random assignment is only available during setup'}), 409
        teams = db.execute(
            'SELECT id FROM teams WHERE course_id = ? ORDER BY id LIMIT ?',
            [course['id'], state['max_teams'] or 6]
        ).fetchall()
        if not teams:
            db.rollback()
            return jsonify({'error': 'No teams available'}), 400

        team_ids = [team['id'] for team in teams]
        unassigned = db.execute(
            '''SELECT id, last_active_at FROM students
               WHERE course_id = ? AND team_id IS NULL AND is_active = 1''',
            [course['id']]
        ).fetchall()
        presence_checked_at = _utcnow()
        student_ids = [
            student['id'] for student in unassigned
            if _student_is_online(
                student['last_active_at'], now=presence_checked_at
            )
        ]
        skipped_offline = len(unassigned) - len(student_ids)
        if not student_ids:
            roster_version = _current_roster_version(db, course['id'])
            db.commit()
            return jsonify({
                'success': True,
                'assigned': 0,
                'remaining': 0,
                'skipped_offline': skipped_offline,
                'roster_version': roster_version,
            })

        rnd.shuffle(student_ids)
        placeholders = ','.join('?' * len(team_ids))
        count_rows = db.execute(
            f'''SELECT team_id, COUNT(*) AS c FROM students
                WHERE team_id IN ({placeholders}) AND is_active = 1
                GROUP BY team_id''',
            team_ids
        ).fetchall()
        counts = {row['team_id']: row['c'] for row in count_rows}
        for team_id in team_ids:
            counts.setdefault(team_id, 0)

        max_members = state['max_members_per_team'] or 10
        assignments = {}
        for student_id in student_ids:
            candidates = [
                team_id for team_id in team_ids if counts[team_id] < max_members
            ]
            if not candidates:
                break
            minimum = min(counts[team_id] for team_id in candidates)
            team_id = rnd.choice([
                candidate for candidate in candidates if counts[candidate] == minimum
            ])
            assignments[student_id] = team_id
            counts[team_id] += 1

        if not assignments:
            roster_version = _current_roster_version(db, course['id'])
            db.commit()
            return jsonify({
                'success': True,
                'assigned': 0,
                'remaining': len(student_ids),
                'skipped_offline': skipped_offline,
                'roster_version': roster_version,
            })
        freeze_guard = _session_roster_mutation_guard(db, course['id'], state)
        if freeze_guard:
            db.rollback()
            return jsonify({'error': freeze_guard[0]}), freeze_guard[1]

        by_team = {}
        for student_id, team_id in assignments.items():
            by_team.setdefault(team_id, []).append(student_id)
        for team_id, assigned_ids in by_team.items():
            placeholders = ','.join('?' * len(assigned_ids))
            db.execute(
                f'''UPDATE students SET team_id = ?, last_team_id = ?,
                    last_team_joined_at = CURRENT_TIMESTAMP
                    WHERE id IN ({placeholders})''',
                [team_id, team_id] + assigned_ids
            )
        roster_version = _bump_roster_version(slug, course['id'], db=db)
        db.commit()
        return jsonify({
            'success': True,
            'assigned': len(assignments),
            'remaining': len(student_ids) - len(assignments),
            'skipped_offline': skipped_offline,
            'roster_version': roster_version,
        })
    except Exception:
        db.rollback()
        raise


@app.route('/api/start_session_timer', methods=['POST'])
@instructor_login_required
def start_session_timer():
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        state = db.execute('SELECT * FROM course_state LIMIT 1').fetchone()
        guard = _expected_state_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        if state['phase'] not in ('setup', 'discussion'):
            db.rollback()
            return jsonify({'error': f"The session timer is only available before the {PHASE_LABELS['competition']} phase"}), 409
        db.execute(
            '''UPDATE course_state
               SET session_started_at = COALESCE(session_started_at, CURRENT_TIMESTAMP)
               WHERE course_id = ?''',
            [state['course_id']]
        )
        fresh = db.execute(
            'SELECT session_started_at FROM course_state WHERE course_id = ?',
            [state['course_id']]
        ).fetchone()
        db.commit()
        return jsonify({
            'success': True,
            'session_started_at': fresh['session_started_at'] if fresh else None,
        })
    except Exception:
        db.rollback()
        raise


@app.route('/api/stop_session_timer', methods=['POST'])
@instructor_login_required
def stop_session_timer():
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        state = db.execute('SELECT * FROM course_state LIMIT 1').fetchone()
        guard = _expected_state_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        db.execute(
            'UPDATE course_state SET session_started_at = NULL WHERE course_id = ?',
            [state['course_id']]
        )
        db.commit()
        return jsonify({'success': True, 'session_started_at': None})
    except Exception:
        db.rollback()
        raise


def _session_has_week_scoped_activity(
        db, course_id, session_key, presentation_history):
    """Return whether this session already owns durable week-scoped data."""
    recorded = db.execute(
        '''SELECT (
               EXISTS(SELECT 1 FROM teammate_thumbs
                      WHERE course_id = ? AND session_key = ?)
               OR EXISTS(SELECT 1 FROM presentation_ratings
                         WHERE course_id = ? AND session_key = ?)
               OR EXISTS(SELECT 1 FROM challenge_rounds
                         WHERE course_id = ? AND session_key = ?)
               OR EXISTS(SELECT 1 FROM challenge_ratings
                         WHERE course_id = ? AND session_key = ?)
               OR EXISTS(SELECT 1 FROM presentation_participants
                         WHERE course_id = ? AND session_key = ?)
           ) AS found''',
        [course_id, session_key, course_id, session_key,
         course_id, session_key, course_id, session_key,
         course_id, session_key],
    ).fetchone()
    if recorded and recorded['found']:
        return True

    try:
        history = json.loads(presentation_history or '[]')
    except (TypeError, ValueError):
        # Preserve an unparseable nonempty history rather than detach it from
        # the lecture week where the presentation occurred.
        return str(presentation_history or '').strip() not in ('', '[]')
    if not isinstance(history, list):
        return bool(history)
    return any(
        isinstance(item, dict)
        and item.get('session_key', 0) == session_key
        for item in history
    )


@app.route('/api/set_discussion_week', methods=['POST'])
@instructor_login_required
def set_discussion_week():
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    week = data.get('week')
    if not isinstance(week, int) or week < 1:
        return jsonify({'error': 'Invalid week'}), 400
    try:
        catalog_week = _validate_course_question_catalog(
            slug, weeks=[week]
        ).get_week(week)
    except (OSError, ValueError) as exc:
        return jsonify({'error': f'Could not validate week {week}: {exc}'}), 422
    if not catalog_week or not catalog_week.discussion.ready:
        issues = [] if not catalog_week else [
            issue.message for issue in catalog_week.discussion.issues
        ]
        detail = f": {issues[0]}" if issues else ''
        return jsonify({
            'error': (
                f'Week {week} questions are not ready{detail}'
            )
        }), 422
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        course = db.execute('SELECT id FROM courses LIMIT 1').fetchone()
        state = db.execute(
            '''SELECT phase, session_key, discussion_week,
                      presentation_history
               FROM course_state WHERE course_id = ?''',
            [course['id']]
        ).fetchone()
        guard = _expected_state_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        if not state or state['phase'] != 'setup':
            db.rollback()
            return jsonify({'error': 'The discussion week can only change during setup'}), 409
        current_week = state['discussion_week'] or 1
        current_session = state['session_key'] or 0
        if (week != current_week and _session_has_week_scoped_activity(
                db, course['id'], current_session,
                state['presentation_history'])):
            db.rollback()
            return jsonify({
                'error': (
                    'The current session already contains recorded activity. '
                    'End Session and return to Setup before changing the '
                    'lecture week.'
                )
            }), 409
        try:
            _read_appendix_question_rows(slug, week)
        except QuestionParseError as exc:
            db.rollback()
            return jsonify({'error': str(exc)}), 422
        sync_presentation_questions(
            slug, course['id'], week, db=db, commit=False
        )
        sync_appendix_questions(
            slug, course['id'], week, db=db, commit=False
        )
        db.execute(
            '''UPDATE course_state
               SET discussion_week = ?, current_question = NULL,
                   current_discussion_key = NULL,
                   current_discussion_source_key = NULL,
                   current_discussion_title = NULL,
                   current_discussion_content = NULL
               WHERE course_id = ?''',
            [week, course['id']]
        )
        _bump_discussion_questions_version(db, course['id'])
        total = db.execute(
            '''SELECT COUNT(*) AS c FROM questions
               WHERE course_id = ? AND COALESCE(week_num, 1) = ?
                 AND (source_key LIKE ? OR source_key LIKE 'appendix:%')''',
            [course['id'], week, f'week-{week}-q-%']
        ).fetchone()['c']
        db.commit()
    except Exception:
        db.rollback()
        raise
    return jsonify({
        'success': True,
        'question_count': total,
        'question_sync': 'synced',
        'presentation_ready': True,
    })


def _question_destination_baseline(path):
    """Return a stable digest for the persistent file being replaced."""
    if not os.path.isfile(path):
        return 'missing'
    digest = hashlib.sha256()
    with open(path, 'rb') as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b''):
            digest.update(chunk)
    return f'file:{digest.hexdigest()}'


def _question_upload_preview_token(
        slug, week, content_digest, destination_baseline):
    secret = app.config['SECRET_KEY']
    if isinstance(secret, str):
        secret = secret.encode('utf-8')
    message = (
        f'question-upload\0{slug}\0{week}\0{content_digest}'
        f'\0{destination_baseline}'
    ).encode('utf-8')
    return hmac.new(secret, message, hashlib.sha256).hexdigest()


@app.route('/api/upload_questions', methods=['POST'])
@instructor_login_required
def upload_questions():
    """Preview, then persist and sync one canonical weekly Markdown file."""
    slug = session['slug']
    if session.get('is_demo') or is_demo_instance_slug(slug):
        return jsonify({
            'error': 'Question uploads are not available in the demo'
        }), 403

    try:
        week = int(request.form.get('week'))
    except (TypeError, ValueError):
        return jsonify({'error': 'A positive week number is required'}), 400
    if week < 1:
        return jsonify({'error': 'A positive week number is required'}), 400

    uploaded = request.files.get('file')
    if uploaded is None or not uploaded.filename:
        return jsonify({'error': 'No question file uploaded'}), 400
    if not uploaded.filename.lower().endswith('.md'):
        return jsonify({'error': 'Please upload a Markdown (.md) file'}), 400

    raw = uploaded.stream.read(MAX_QUESTION_UPLOAD_BYTES + 1)
    if len(raw) > MAX_QUESTION_UPLOAD_BYTES:
        return jsonify({
            'error': 'Question file must be 1 MB or smaller'
        }), 413
    if not raw:
        return jsonify({'error': 'Question file is empty'}), 400
    try:
        text = raw.decode('utf-8-sig')
    except UnicodeDecodeError:
        return jsonify({'error': 'Question file must be valid UTF-8'}), 422
    try:
        questions = parse_week_questions(
            text,
            source_path=f'week-{week}-questions.md',
            week_num=week,
            max_questions=MAX_WEEK_QUESTIONS,
        )
    except QuestionParseError as exc:
        return jsonify({'error': str(exc)}), 422

    question_summary = [
        {'id': question['id'], 'title': question['title']}
        for question in questions
    ]
    digest = hashlib.sha256(raw).hexdigest()
    destination = _persistent_week_question_path(slug, week)
    destination_baseline = _question_destination_baseline(destination)
    expected_token = _question_upload_preview_token(
        slug, week, digest, destination_baseline
    )
    ensure_schema(slug)
    course = query_db(slug, 'SELECT id FROM courses LIMIT 1', one=True)
    state = query_db(
        slug,
        'SELECT phase, session_key FROM course_state WHERE course_id = ?',
        [course['id']],
        one=True,
    )
    guard = _expected_state_guard(request.form, state)
    if guard:
        return jsonify({'error': guard[0]}), guard[1]
    if not state or state['phase'] != 'setup':
        return jsonify({
            'error': 'Questions can only be uploaded during setup'
        }), 409

    confirmed = str(request.form.get('confirm') or '').strip().lower() == 'true'
    if not confirmed:
        return jsonify({
            'requires_confirmation': True,
            'preview_token': expected_token,
            'week': week,
            'question_count': len(questions),
            'questions': question_summary,
        })

    supplied_token = str(request.form.get('preview_token') or '')
    if not supplied_token or not hmac.compare_digest(
            supplied_token, expected_token):
        return jsonify({
            'error': 'The question file changed after preview; preview it again'
        }), 409

    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    previous_exists = False
    previous_content = b''
    file_written = False
    try:
        state = db.execute(
            'SELECT phase, session_key FROM course_state WHERE course_id = ?',
            [course['id']],
        ).fetchone()
        guard = _expected_state_guard(request.form, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        if not state or state['phase'] != 'setup':
            db.rollback()
            return jsonify({
                'error': 'Questions can only be uploaded during setup'
            }), 409

        locked_baseline = _question_destination_baseline(destination)
        locked_token = _question_upload_preview_token(
            slug, week, digest, locked_baseline
        )
        if not hmac.compare_digest(supplied_token, locked_token):
            db.rollback()
            return jsonify({
                'error': 'Questions changed after preview; preview them again'
            }), 409

        previous_exists = os.path.isfile(destination)
        if previous_exists:
            with open(destination, 'rb') as handle:
                previous_content = handle.read()
        _write_bytes_atomic(destination, raw)
        file_written = True
        sync_presentation_questions(
            slug, course['id'], week, db=db, commit=False
        )
        _bump_discussion_questions_version(db, course['id'])
        db.commit()
    except Exception:
        try:
            if file_written:
                if previous_exists:
                    _write_bytes_atomic(destination, previous_content)
                elif os.path.exists(destination):
                    os.remove(destination)
        finally:
            db.rollback()
        raise

    return jsonify({
        'success': True,
        'week': week,
        'question_count': len(questions),
        'questions': question_summary,
    })


@app.route('/api/toggle_lock_teams', methods=['POST'])
@instructor_login_required
def toggle_lock_teams():
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    locked = 1 if data.get('locked') else 0
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        course = db.execute('SELECT id FROM courses LIMIT 1').fetchone()
        state = db.execute(
            'SELECT phase, session_key FROM course_state WHERE course_id = ?',
            [course['id']]
        ).fetchone()
        guard = _expected_state_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        if not state or state['phase'] != 'setup':
            db.rollback()
            return jsonify({'error': 'Team locking is only available during setup'}), 409
        db.execute(
            'UPDATE course_state SET teams_locked = ? WHERE course_id = ?',
            [locked, course['id']]
        )
        db.commit()
        return jsonify({'success': True, 'locked': bool(locked)})
    except Exception:
        db.rollback()
        raise


@app.route('/api/set_max_members', methods=['POST'])
@instructor_login_required
def set_max_members():
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    new_max = data.get('max_members')
    if not isinstance(new_max, int) or new_max < 1 or new_max > 99:
        return jsonify({'error': 'Max members must be between 1 and 99'}), 400
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        course = db.execute('SELECT id FROM courses LIMIT 1').fetchone()
        state = db.execute(
            '''SELECT phase, session_key, roster_version, max_members_per_team
               FROM course_state
               WHERE course_id = ?''',
            [course['id']]
        ).fetchone()
        guard = _expected_roster_state_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        if not state or state['phase'] != 'setup':
            db.rollback()
            return jsonify({'error': 'Team settings can only change during setup'}), 409

        current_max = state['max_members_per_team'] or 10
        if new_max == current_max:
            roster_version = _current_roster_version(db, course['id'])
            db.commit()
            return jsonify({
                'success': True,
                'max_members': new_max,
                'roster_version': roster_version,
            })
        freeze_guard = _session_roster_mutation_guard(db, course['id'], state)
        if freeze_guard:
            db.rollback()
            return jsonify({'error': freeze_guard[0]}), freeze_guard[1]

        excess_ids = []
        if new_max < current_max:
            full_teams = db.execute(
                '''SELECT team_id, COUNT(*) AS cnt FROM students
                   WHERE course_id = ? AND team_id IS NOT NULL AND is_active = 1
                   GROUP BY team_id HAVING COUNT(*) > ?''',
                [course['id'], new_max]
            ).fetchall()
            for team in full_teams:
                excess = db.execute(
                    '''SELECT id FROM students
                       WHERE course_id = ? AND team_id = ? AND is_active = 1
                       ORDER BY last_team_joined_at DESC NULLS LAST, id DESC LIMIT ?''',
                    [course['id'], team['team_id'], team['cnt'] - new_max]
                ).fetchall()
                excess_ids.extend(student['id'] for student in excess)
            if excess_ids:
                placeholders = ','.join('?' * len(excess_ids))
                db.execute(
                    f'''UPDATE students
                        SET last_team_id = COALESCE(last_team_id, team_id),
                            team_id = NULL
                        WHERE id IN ({placeholders})''',
                    excess_ids
                )

        db.execute(
            '''UPDATE course_state
               SET max_members_per_team = ?,
                   roster_version = COALESCE(roster_version, 0) + 1
               WHERE course_id = ?''',
            [new_max, course['id']]
        )
        roster_version = _current_roster_version(db, course['id'])
        db.commit()
        return jsonify({
            'success': True,
            'max_members': new_max,
            'roster_version': roster_version,
        })
    except Exception:
        db.rollback()
        raise


# ---------------------------------------------------------------------------
# Question Bank API
# ---------------------------------------------------------------------------

_LEGACY_QUESTION_REVISION_RE = re.compile(r'^[0-9a-f]{8}$')


def _is_legacy_question_alias(question_key, legacy_prefix):
    if not question_key or not legacy_prefix:
        return False
    suffix = question_key[len(legacy_prefix):] \
        if question_key.startswith(legacy_prefix) else ''
    return bool(_LEGACY_QUESTION_REVISION_RE.fullmatch(suffix))


def _legacy_question_alias_glob(legacy_prefix):
    return legacy_prefix + ('[0-9a-f]' * 8)


def _reconcile_hidden_question_aliases_in_transaction(
        db, course_id, week_num, stable_key, legacy_key, legacy_prefix=None):
    """Move matching hidden rows to one stable question key."""
    params = [course_id, week_num, legacy_key]
    alias_clause = 'question_key = ?'
    if legacy_prefix:
        alias_clause += ' OR question_key GLOB ?'
        params.append(_legacy_question_alias_glob(legacy_prefix))
    aliases = db.execute(
        f'''SELECT question_key FROM hidden_discussion_questions
            WHERE course_id = ? AND week_num = ?
              AND ({alias_clause})''',
        params,
    ).fetchall()
    if not aliases:
        return False
    db.execute(
        '''INSERT OR IGNORE INTO hidden_discussion_questions
           (course_id, week_num, question_key) VALUES (?, ?, ?)''',
        [course_id, week_num, stable_key],
    )
    db.executemany(
        '''DELETE FROM hidden_discussion_questions
           WHERE course_id = ? AND week_num = ? AND question_key = ?''',
        [
            (course_id, week_num, row['question_key'])
            for row in aliases
            if row['question_key'] != stable_key
        ],
    )
    return True


def _migrate_hidden_question_aliases(
        slug, course_id, week_num, alias_pairs):
    """Best-effort replacement of matching legacy visibility keys."""
    alias_pairs = {
        (legacy_key, stable_key)
        for legacy_key, stable_key in alias_pairs
        if legacy_key and stable_key and legacy_key != stable_key
    }
    db_path = course_db_path(slug)
    if not alias_pairs or not db_path or not os.path.isfile(db_path):
        return

    db = None
    try:
        db = sqlite3.connect(db_path, timeout=0.1)
        db.execute('BEGIN IMMEDIATE')
        for legacy_key, stable_key in alias_pairs:
            exists = db.execute(
                '''SELECT 1 FROM hidden_discussion_questions
                   WHERE course_id = ? AND week_num = ?
                     AND question_key = ?''',
                [course_id, week_num, legacy_key],
            ).fetchone()
            if not exists:
                continue
            db.execute(
                '''INSERT OR IGNORE INTO hidden_discussion_questions
                   (course_id, week_num, question_key)
                   VALUES (?, ?, ?)''',
                [course_id, week_num, stable_key],
            )
            db.execute(
                '''DELETE FROM hidden_discussion_questions
                   WHERE course_id = ? AND week_num = ?
                     AND question_key = ?''',
                [course_id, week_num, legacy_key],
            )
        db.commit()
    except (OSError, sqlite3.Error):
        if db is not None:
            try:
                db.rollback()
            except sqlite3.Error:
                pass
    finally:
        if db is not None:
            db.close()


@app.route('/api/discussion_questions', methods=['GET'])
def discussion_questions():
    """Weekly discussion questions for the current week.

    Instructors see every question (with a ``hidden`` flag) plus the week list
    for the selector. Students see only the visible questions. Both receive
    ``version`` so clients refetch only when the visible set changes.
    """
    slug = session.get('slug')
    role = _authenticated_role(slug)
    if not slug or role is None:
        return jsonify({'error': 'Not logged in'}), 401
    week_param = request.args.get('week')
    course = query_db(slug, 'SELECT id FROM courses LIMIT 1', one=True)
    ensure_schema(slug)
    state = query_db(
        slug,
        '''SELECT discussion_week, discussion_questions_version
           FROM course_state WHERE course_id = ?''',
        [course['id']], one=True,
    )
    saved_week = state['discussion_week'] if state and state['discussion_week'] else 1
    version = (state['discussion_questions_version'] or 0) if state else 0

    selected_week = saved_week
    if role == 'instructor' and week_param is not None:
        normalized_week = week_param.strip()
        if not re.fullmatch(r'[1-9]\d*', normalized_week):
            return jsonify({'error': 'Week must be a positive integer'}), 400
        selected_week = int(normalized_week)

    catalog_weeks = None if role == 'instructor' else [saved_week]
    try:
        catalog = _validate_course_question_catalog(
            slug, weeks=catalog_weeks
        )
    except (OSError, ValueError) as exc:
        return jsonify({'error': f'Could not validate question catalog: {exc}'}), 422

    target = catalog.get_week(selected_week)
    if target is None:
        try:
            target = _validate_course_question_catalog(
                slug, weeks=[selected_week]
            ).get_week(selected_week)
        except (OSError, ValueError) as exc:
            return jsonify({
                'error': f'Could not validate question catalog: {exc}'
            }), 422

    catalog_statuses = list(catalog.weeks)
    if (target is not None
            and all(week.week != target.week for week in catalog_statuses)):
        catalog_statuses.append(target)
        catalog_statuses.sort(key=lambda week: week.week)

    questions_list = []
    weeks = [
        {
            'num': week.week,
            'file': os.path.basename(week.discussion.path),
            'ready': week.discussion.ready,
            'discussion_ready': week.discussion.ready,
            'presentation_ready': week.presentation.ready,
            'issues': [issue.message for issue in week.discussion.issues],
        }
        for week in catalog_statuses
    ]

    def _load_md(filepath, week, is_appendix=False):
        """Load questions with stable keys plus current legacy aliases."""
        out = []
        if not os.path.exists(filepath):
            return out
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            content = f.read()
        seen_identities = set()
        source = 'a' if is_appendix else 'q'
        for position, (fm_block, body_block) in enumerate(
                parse_question_blocks(content), 1):
            try:
                fm = yaml.safe_load(fm_block) or {}
            except yaml.YAMLError as exc:
                raise QuestionParseError(
                    f'Invalid question frontmatter in question {position}'
                ) from exc
            title = str(fm.get('title') or '').strip()
            if not title:
                raise QuestionParseError(
                    f'Question {position} has no title'
                )
            revision = hashlib.sha256(
                (title + '\0' + body_block).encode('utf-8')
            ).hexdigest()[:16]
            explicit_id = str(fm.get('id') or '').strip()
            has_explicit_id = bool(re.fullmatch(
                r'[A-Za-z0-9_-]+', explicit_id
            ))
            label = re.match(r'^A(\d+)\s*:', title, re.IGNORECASE) \
                if is_appendix else None
            if is_appendix and not label:
                raise QuestionParseError(
                    f'Appendix question {position} must start with an A-number label'
                )
            if is_appendix:
                identity = f'A{label.group(1)}'
                legacy_identity = explicit_id if has_explicit_id else identity
            else:
                identity = (
                    explicit_id if has_explicit_id
                    else f'position-{position}'
                )
                legacy_identity = explicit_id if has_explicit_id else revision
            if identity in seen_identities:
                label_name = (
                    'appendix label' if is_appendix else 'question ID'
                )
                raise QuestionParseError(
                    f'Duplicate {label_name} {identity}'
                )
            seen_identities.add(identity)
            stable_key = f'week-{week}-{source}-{identity}'
            legacy_prefix = (
                f'week-{week}-{source}-{legacy_identity}-'
                if is_appendix or has_explicit_id else None
            )
            legacy_keys = {
                f'week-{week}-{source}-{legacy_identity}-{revision[:8]}',
                f'week-{week}-{source}-{revision}-{revision[:8]}',
            }
            item = {
                'key': stable_key,
                'title': title,
                'content': body_block,
                'revision': revision,
                '_legacy_keys': legacy_keys,
                '_legacy_prefix': legacy_prefix,
            }
            if is_appendix:
                item['appendix_id'] = identity
            out.append(item)
        return out

    try:
        if target and target.discussion.ready:
            q_path = _resolve_week_question_path(slug, target.week)
            questions_list = _load_md(q_path, target.week)

        # Load appendix from the persistent data disk (survives deploys)
        appendix_week = selected_week
        appendix_path = _appendix_path(slug, appendix_week)
        appendix = _load_md(
            appendix_path, appendix_week, is_appendix=True
        )
        questions_list.extend(appendix)
    except QuestionParseError as exc:
        return jsonify({'error': str(exc)}), 422

    for display_number, question in enumerate(questions_list, 1):
        question['display_number'] = display_number
        question['hidden'] = False
        question['source'] = (
            'appendix' if 'appendix_id' in question else 'bank'
        )
        question.pop('_legacy_keys', None)
        question.pop('_legacy_prefix', None)

    if role == 'instructor' and questions_list:
        presentation_rows = query_db(
            slug,
            '''SELECT id, source_key FROM questions
               WHERE course_id = ? AND COALESCE(week_num, 1) = ?
                 AND (source_key LIKE ? OR source_key LIKE ?)''',
            [
                course['id'], selected_week,
                f'week-{selected_week}-q-%',
                f'appendix:{selected_week}:%',
            ],
        )
        question_ids = {
            row['source_key']: row['id'] for row in presentation_rows
        }
        for question in questions_list:
            presentation_source_key = (
                f"appendix:{selected_week}:{question['appendix_id']}"
                if question.get('appendix_id') else question['key']
            )
            question['question_id'] = question_ids.get(
                presentation_source_key
            )

    if role == 'student':
        weeks = []

    return jsonify({
        'weeks': weeks,
        'current_week': appendix_week,
        'ready': bool(target and target.discussion.ready),
        'issues': (
            [issue.message for issue in target.discussion.issues]
            if target else [f'Week {selected_week} questions are not ready']
        ),
        'version': version,
        'questions': questions_list,
    })


@app.route('/api/toggle_discussion_question', methods=['POST'])
@instructor_login_required
def toggle_discussion_question():
    '''Keep the retired visibility endpoint harmless for older clients.'''
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    question_key = str(data.get('question_key') or '').strip()
    if not question_key:
        return jsonify({'error': 'question_key is required'}), 400
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        course = db.execute('SELECT id FROM courses LIMIT 1').fetchone()
        state = db.execute(
            'SELECT phase, session_key, discussion_week FROM course_state '
            'WHERE course_id = ?', [course['id']]
        ).fetchone()
        guard = _expected_state_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        week = state['discussion_week'] if state and state['discussion_week'] else 1
        cursor = db.execute(
            '''DELETE FROM hidden_discussion_questions
               WHERE course_id = ? AND week_num = ?
                 AND (question_key = ? OR question_key GLOB ?)''',
            [course['id'], week, question_key,
             _legacy_question_alias_glob(f'{question_key}-')],
        )
        if cursor.rowcount:
            _bump_discussion_questions_version(db, course['id'])
        db.commit()
    except Exception:
        db.rollback()
        raise
    return jsonify({
        'success': True,
        'visible': True,
        'hide_supported': False,
    })


def _appendix_dir(slug):
    """Directory for appendix question files on the persistent data disk."""
    d = os.path.join(config.DATA_DIR, slug, 'appendix')
    os.makedirs(d, exist_ok=True)
    # Publish legacy appendix seeds atomically without changing source files.
    # Linking a fully written temporary file creates the destination only if
    # another worker has not already published it.
    for week in range(1, 20):
        old = os.path.join(_course_class_dir(slug), f'week-{week}-appendix.md')
        if os.path.exists(old):
            new = os.path.join(d, f'week-{week}-appendix.md')
            if not os.path.exists(new):
                temporary = os.path.join(
                    d, f'.week-{week}-appendix.{uuid.uuid4().hex}.tmp'
                )
                try:
                    with open(old, 'rb') as source, open(temporary, 'xb') as target:
                        for chunk in iter(lambda: source.read(1024 * 1024), b''):
                            target.write(chunk)
                        target.flush()
                        os.fsync(target.fileno())
                    try:
                        os.link(temporary, new)
                    except FileExistsError:
                        pass
                finally:
                    try:
                        os.remove(temporary)
                    except FileNotFoundError:
                        pass
    return d


def _appendix_path(slug, week):
    """File path for a given week's appendix questions."""
    return os.path.join(_appendix_dir(slug), f'week-{week}-appendix.md')


@app.route('/api/questions', methods=['POST'])
@instructor_login_required
def add_question():
    """Add an appendix question to the persistent question file."""
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    title = str(data.get('title') or '').strip()
    content = str(data.get('content') or '').strip()
    raw_client_request_id = data.get(
        'client_request_id', data.get('request_key')
    )
    client_request_id = None
    if raw_client_request_id is not None:
        client_request_id = str(raw_client_request_id).strip()
        if not re.fullmatch(r'[A-Za-z0-9][A-Za-z0-9._:-]{0,127}',
                            client_request_id):
            return jsonify({'error': 'Invalid client_request_id'}), 400
    try:
        requested_week = int(data.get('week'))
    except (TypeError, ValueError):
        return jsonify({'error': 'Valid week required'}), 400
    if not title or not content:
        return jsonify({'error': 'Title and content required'}), 400
    if len(title) > 500 or len(content) > 50000:
        return jsonify({'error': 'Title or content is too long'}), 400

    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    original_content = ''
    file_written = False
    appendix_path = None
    try:
        course = db.execute('SELECT id FROM courses LIMIT 1').fetchone()
        state = db.execute(
            '''SELECT phase, session_key, discussion_week
               FROM course_state WHERE course_id = ?''',
            [course['id']]
        ).fetchone()
        guard = _expected_state_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        week = state['discussion_week'] if state and state['discussion_week'] else 1
        if requested_week != week:
            db.rollback()
            return jsonify({'error': 'The displayed appendix week is stale'}), 409
        appendix_path = _appendix_path(slug, week)
        if os.path.exists(appendix_path):
            with open(appendix_path, 'r', encoding='utf-8-sig') as handle:
                original_content = handle.read()
        existing_entries = parse_question_blocks(original_content) \
            if original_content.strip() else []

        highest_label = 0
        replayed_label = None
        for position, (frontmatter, existing_body) in enumerate(
                existing_entries, 1):
            metadata = yaml.safe_load(frontmatter) or {}
            match = re.match(
                r'^A(\d+)\s*:', str(metadata.get('title') or ''), re.IGNORECASE
            )
            if not match:
                raise QuestionParseError(
                    f'Appendix question {position} must start with an A-number label'
                )
            label_number = int(match.group(1))
            highest_label = max(highest_label, label_number)
            stored_request_id = str(
                metadata.get('client_request_id')
                or metadata.get('request_key')
                or ''
            ).strip()
            if client_request_id and stored_request_id == client_request_id:
                if replayed_label is not None:
                    raise QuestionParseError(
                        'Duplicate appendix client_request_id in the file'
                    )
                replayed_label = f'A{label_number}'
                expected_title = f'{replayed_label}: {title}'
                if (str(metadata.get('title') or '').strip() != expected_title
                        or existing_body.strip() != content):
                    db.rollback()
                    return jsonify({
                        'error': (
                            'client_request_id was already used for a '
                            'different appendix question'
                        )
                    }), 409

        if replayed_label is not None:
            sync_appendix_questions(
                slug, course['id'], week, db=db, commit=False,
                bump_discussion_version=True,
            )
            question_row = db.execute(
                '''SELECT id, title FROM questions
                   WHERE course_id = ? AND source_key = ?''',
                [course['id'], f'appendix:{week}:{replayed_label}']
            ).fetchone()
            db.commit()
            return jsonify({
                'success': True,
                'replayed': True,
                'label': replayed_label,
                'appendix_id': replayed_label,
                'question_id': question_row['id'] if question_row else None,
                'title': (
                    question_row['title'] if question_row
                    else f'{replayed_label}: {title}'
                ),
            })

        label = f'A{highest_label + 1}'
        metadata = {'title': f'{label}: {title}'}
        if client_request_id:
            metadata['client_request_id'] = client_request_id
        frontmatter = yaml.safe_dump(
            metadata, allow_unicode=True, sort_keys=False
        ).strip()
        new_entries = existing_entries + [(frontmatter, content)]
        _write_text_atomic(
            appendix_path, _serialize_question_blocks(new_entries)
        )
        file_written = True
        sync_appendix_questions(
            slug, course['id'], week, db=db, commit=False
        )
        _bump_discussion_questions_version(db, course['id'])
        # The competition question select is keyed by the numeric question
        # id, so return it for an in-place option rebuild (no reload).
        question_row = db.execute(
            '''SELECT id, title FROM questions
               WHERE course_id = ? AND source_key = ?''',
            [course['id'], f'appendix:{week}:{label}']
        ).fetchone()
        db.commit()
        return jsonify({
            'success': True,
            'replayed': False,
            'label': label,
            'appendix_id': label,
            'question_id': question_row['id'] if question_row else None,
            'title': (
                question_row['title'] if question_row else f'{label}: {title}'
            ),
        })
    except QuestionParseError as exc:
        try:
            if file_written:
                _write_text_atomic(appendix_path, original_content)
        finally:
            db.rollback()
        return jsonify({'error': str(exc)}), 422
    except Exception:
        try:
            if file_written:
                _write_text_atomic(appendix_path, original_content)
        finally:
            db.rollback()
        raise


@app.route('/api/delete_appendix_question', methods=['POST'])
@instructor_login_required
def delete_appendix_question():
    """Delete one appendix question by its stable A-number label."""
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    appendix_id = str(data.get('appendix_id') or '').strip().upper()
    if not re.fullmatch(r'A\d+', appendix_id):
        return jsonify({'error': 'Appendix question ID required'}), 400
    try:
        requested_week = int(data.get('week'))
    except (ValueError, TypeError):
        return jsonify({'error': 'Valid week required'}), 400

    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    original_content = ''
    file_written = False
    appendix_path = None
    try:
        course = db.execute('SELECT id FROM courses LIMIT 1').fetchone()
        state = db.execute(
            '''SELECT discussion_week, phase, session_key, current_discussion_key,
                      current_discussion_source_key FROM course_state
               WHERE course_id = ?''',
            [course['id']]
        ).fetchone()
        guard = _expected_state_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        if state and state['phase'] == 'competition':
            db.rollback()
            return jsonify({'error': f"Appendix questions cannot be deleted during the {PHASE_LABELS['competition']} phase"}), 409
        week = state['discussion_week'] if state and state['discussion_week'] else 1
        if requested_week != week:
            db.rollback()
            return jsonify({'error': 'The displayed appendix week is stale'}), 409
        appendix_path = _appendix_path(slug, week)
        if not os.path.exists(appendix_path):
            db.rollback()
            return jsonify({'error': 'Appendix file not found'}), 404

        _read_appendix_question_rows(slug, week)
        with open(appendix_path, 'r', encoding='utf-8-sig') as handle:
            original_content = handle.read()
        entries = parse_question_blocks(original_content)
        selected_index = None
        selected_key = None
        selected_legacy_key = None
        selected_legacy_prefix = None
        for index, (frontmatter, body) in enumerate(entries):
            metadata = yaml.safe_load(frontmatter) or {}
            title = str(metadata.get('title') or '').strip()
            label = re.match(r'^A(\d+)\s*:', title, re.IGNORECASE)
            identity = f'A{label.group(1)}' if label else None
            if identity == appendix_id:
                digest = hashlib.sha256(
                    (title + '\0' + body).encode('utf-8')
                ).hexdigest()[:16]
                explicit_id = str(metadata.get('id') or '').strip()
                legacy_identity = (
                    explicit_id if re.fullmatch(
                        r'[A-Za-z0-9_-]+', explicit_id
                    ) else identity
                )
                selected_index = index
                selected_key = f'week-{week}-a-{identity}'
                selected_legacy_prefix = (
                    f'week-{week}-a-{legacy_identity}-'
                )
                selected_legacy_key = (
                    f'{selected_legacy_prefix}{digest[:8]}'
                )
                break
        if selected_index is None:
            db.rollback()
            return jsonify({'error': 'Appendix question not found'}), 404

        del entries[selected_index]
        _write_text_atomic(appendix_path, _serialize_question_blocks(entries))
        file_written = True
        sync_appendix_questions(
            slug, course['id'], week, db=db, commit=False
        )
        db.execute(
            '''DELETE FROM hidden_discussion_questions
               WHERE course_id = ? AND week_num = ?
                 AND (question_key = ? OR question_key = ?
                      OR question_key GLOB ?)''',
            [
                course['id'], week, selected_key, selected_legacy_key,
                _legacy_question_alias_glob(selected_legacy_prefix),
            ],
        )
        # Legacy posted-question columns: clear them if the deleted question
        # was the one last posted by the removed single-question flow.
        posted_key = (
            state['current_discussion_source_key']
            or state['current_discussion_key']
        ) if state else None
        unposted = bool(state) and (
            posted_key in (selected_key, selected_legacy_key)
            or (
                posted_key
                and _is_legacy_question_alias(
                    posted_key, selected_legacy_prefix
                )
            )
        )
        if unposted:
            db.execute(
                '''UPDATE course_state
                   SET current_question = NULL, current_discussion_key = NULL,
                       current_discussion_source_key = NULL,
                       current_discussion_title = NULL,
                       current_discussion_content = NULL
                   WHERE course_id = ?''',
                [course['id']]
            )
        _bump_discussion_questions_version(db, course['id'])
        db.commit()
        return jsonify({'success': True, 'unposted': unposted})
    except QuestionParseError as exc:
        try:
            if file_written:
                _write_text_atomic(appendix_path, original_content)
        finally:
            db.rollback()
        return jsonify({'error': str(exc)}), 422
    except Exception:
        try:
            if file_written:
                _write_text_atomic(appendix_path, original_content)
        finally:
            db.rollback()
        raise


@app.route('/api/edit_appendix_question', methods=['POST'])
@instructor_login_required
def edit_appendix_question():
    """Edit one appendix question in place, keeping its A-number label."""
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    appendix_id = str(data.get('appendix_id') or '').strip().upper()
    title = str(data.get('title') or '').strip()
    content = str(data.get('content') or '').strip()
    if not re.fullmatch(r'A\d+', appendix_id):
        return jsonify({'error': 'Appendix question ID required'}), 400
    try:
        requested_week = int(data.get('week'))
    except (ValueError, TypeError):
        return jsonify({'error': 'Valid week required'}), 400
    if not title or not content:
        return jsonify({'error': 'Title and content required'}), 400
    if len(title) > 500 or len(content) > 50000:
        return jsonify({'error': 'Title or content is too long'}), 400

    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    original_content = ''
    file_written = False
    appendix_path = None
    try:
        course = db.execute('SELECT id FROM courses LIMIT 1').fetchone()
        state = db.execute(
            '''SELECT discussion_week, phase, session_key FROM course_state
               WHERE course_id = ?''',
            [course['id']]
        ).fetchone()
        guard = _expected_state_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        if state and state['phase'] == 'competition':
            db.rollback()
            return jsonify({'error': f"Appendix questions cannot be edited during the {PHASE_LABELS['competition']} phase"}), 409
        week = state['discussion_week'] if state and state['discussion_week'] else 1
        if requested_week != week:
            db.rollback()
            return jsonify({'error': 'The displayed appendix week is stale'}), 409
        appendix_path = _appendix_path(slug, week)
        if not os.path.exists(appendix_path):
            db.rollback()
            return jsonify({'error': 'Appendix file not found'}), 404

        _read_appendix_question_rows(slug, week)
        with open(appendix_path, 'r', encoding='utf-8-sig') as handle:
            original_content = handle.read()
        entries = parse_question_blocks(original_content)
        selected_index = None
        selected_label = None
        selected_key = None
        selected_legacy_key = None
        selected_legacy_prefix = None
        for index, (frontmatter, body) in enumerate(entries):
            metadata = yaml.safe_load(frontmatter) or {}
            entry_title = str(metadata.get('title') or '').strip()
            label = re.match(r'^A(\d+)\s*:', entry_title, re.IGNORECASE)
            identity = f'A{label.group(1)}' if label else None
            if identity == appendix_id:
                digest = hashlib.sha256(
                    (entry_title + '\0' + body).encode('utf-8')
                ).hexdigest()[:16]
                explicit_id = str(metadata.get('id') or '').strip()
                legacy_identity = (
                    explicit_id if re.fullmatch(
                        r'[A-Za-z0-9_-]+', explicit_id
                    ) else identity
                )
                selected_index = index
                selected_label = identity
                selected_metadata = dict(metadata)
                selected_key = f'week-{week}-a-{identity}'
                selected_legacy_prefix = (
                    f'week-{week}-a-{legacy_identity}-'
                )
                selected_legacy_key = (
                    f'{selected_legacy_prefix}{digest[:8]}'
                )
                break
        if selected_index is None:
            db.rollback()
            return jsonify({'error': 'Appendix question not found'}), 404

        _reconcile_hidden_question_aliases_in_transaction(
            db,
            course['id'],
            week,
            selected_key,
            selected_legacy_key,
            selected_legacy_prefix,
        )
        selected_metadata['title'] = f'{selected_label}: {title}'
        frontmatter = yaml.safe_dump(
            selected_metadata, allow_unicode=True, sort_keys=False
        ).strip()
        entries[selected_index] = (frontmatter, content)
        _write_text_atomic(appendix_path, _serialize_question_blocks(entries))
        file_written = True
        sync_appendix_questions(
            slug, course['id'], week, db=db, commit=False
        )
        _bump_discussion_questions_version(db, course['id'])
        # The competition question select is keyed by the numeric question
        # id, so return it for an in-place option rebuild (no reload).
        question_row = db.execute(
            '''SELECT id, title FROM questions
               WHERE course_id = ? AND source_key = ?''',
            [course['id'], f'appendix:{week}:{selected_label}']
        ).fetchone()
        db.commit()
        return jsonify({
            'success': True,
            'label': selected_label,
            'appendix_id': selected_label,
            'question_id': question_row['id'] if question_row else None,
            'title': (
                question_row['title'] if question_row
                else f'{selected_label}: {title}'
            ),
        })
    except QuestionParseError as exc:
        try:
            if file_written:
                _write_text_atomic(appendix_path, original_content)
        finally:
            db.rollback()
        return jsonify({'error': str(exc)}), 422
    except Exception:
        try:
            if file_written:
                _write_text_atomic(appendix_path, original_content)
        finally:
            db.rollback()
        raise


@app.route('/api/unassign_all', methods=['POST'])
@instructor_login_required
def unassign_all():
    """Clear all active team assignments and lock student joining."""
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        course = db.execute('SELECT id FROM courses LIMIT 1').fetchone()
        state = db.execute(
            '''SELECT phase, session_key, roster_version
               FROM course_state WHERE course_id = ?''',
            [course['id']]
        ).fetchone()
        guard = _expected_roster_state_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        if not state or state['phase'] != 'setup':
            db.rollback()
            return jsonify({'error': 'Teams can only be changed during setup'}), 409
        count = db.execute(
            '''SELECT COUNT(*) AS c FROM students
               WHERE course_id = ? AND is_active = 1 AND team_id IS NOT NULL''',
            [course['id']],
        ).fetchone()['c']
        if count:
            freeze_guard = _session_roster_mutation_guard(
                db, course['id'], state
            )
            if freeze_guard:
                db.rollback()
                return jsonify({'error': freeze_guard[0]}), freeze_guard[1]
            db.execute(
                '''UPDATE students
                   SET last_team_id = COALESCE(last_team_id, team_id),
                       team_id = NULL
                   WHERE course_id = ? AND is_active = 1''', [course['id']]
            )
            roster_version = _bump_roster_version(
                slug, course['id'], db=db
            )
        else:
            roster_version = _current_roster_version(db, course['id'])
        db.execute(
            'UPDATE course_state SET teams_locked = 1 WHERE course_id = ?',
            [course['id']],
        )
        db.commit()
        return jsonify({
            'success': True,
            'count': count,
            'locked': True,
            'roster_version': roster_version,
        })
    except Exception:
        db.rollback()
        raise


# ---------------------------------------------------------------------------
# Competition / Presentation Control API
# ---------------------------------------------------------------------------

@app.route('/roster_template.csv')
@instructor_login_required
def roster_template():
    """Download a CSV template with example rows for roster upload."""
    csv_content = (
        'student_id,name,pin\n'
        '1001,Alice Chen,4271\n'
        '1002,Bob Garcia,8839\n'
        '1003,Cara Singh,1056\n'
    )
    return (
        csv_content,
        200,
        {
            'Content-Type': 'text/csv',
            'Content-Disposition': 'attachment; filename=roster_template.csv'
        }
    )


_ROSTER_CASE_COLLISION_MESSAGE = (
    'The existing roster contains student IDs that differ only by letter '
    'case. Resolve those duplicate IDs before uploading a roster.'
)
_ROSTER_MERGE_TOKEN_DOMAIN = b'popping-roster-merge-v1\0'


def _roster_rows_by_casefolded_id(rows):
    """Index roster rows without silently choosing between ambiguous IDs."""
    indexed = {}
    for row in rows:
        student_id_key = row['student_id'].casefold()
        if student_id_key in indexed:
            return None
        indexed[student_id_key] = row
    return indexed


def _roster_merge_preview_token(raw_content, slug, session_key, roster_version):
    """Bind confirmation to this merge protocol, file, and roster state."""
    state_context = (
        f"{slug}:{session_key or 0}:{roster_version or 0}"
    ).encode('utf-8')
    return hashlib.sha256(
        _ROSTER_MERGE_TOKEN_DOMAIN + raw_content + b'\0' + state_context
    ).hexdigest()


@app.route('/api/upload_roster', methods=['POST'])
@instructor_login_required
def upload_roster():
    """Merge a validated CSV into the roster without changing omitted IDs."""
    slug = session['slug']
    if is_demo_instance_slug(slug):
        return jsonify({'error': 'The demo roster is fixed at two students'}), 403
    ensure_schema(slug)
    course = query_db(slug, 'SELECT id FROM courses LIMIT 1', one=True)
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if file.filename == '' or not file.filename.lower().endswith('.csv'):
        return jsonify({'error': 'Please upload a CSV file'}), 400
    try:
        raw_content = file.read(MAX_ROSTER_BYTES + 1)
        if len(raw_content) > MAX_ROSTER_BYTES:
            return jsonify({'error': 'Roster file must be 1 MB or smaller'}), 413
        content = raw_content.decode('utf-8-sig')
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
    except Exception:
        return jsonify({'error': 'Failed to read CSV file'}), 400

    if not rows:
        return jsonify({'error': 'Empty CSV'}), 400
    if len(rows) - 1 > MAX_ROSTER_ROWS:
        return jsonify({'error': f'Roster cannot contain more than {MAX_ROSTER_ROWS} students'}), 400

    # Detect header format
    header = [h.strip().lower() for h in rows[0]]
    if header[:3] == ['student_id', 'name', 'pin']:
        sid_col, name_col, pin_col = 0, 1, 2
    elif header[:3] == ['id', 'name', 'pin']:
        sid_col, name_col, pin_col = 0, 1, 2
    else:
        return jsonify({'error': 'CSV header must be: student_id, name, pin (or ID, Name, PIN)'}), 400

    # Parse and validate
    parsed = []
    errors = []
    seen_ids = set()
    for i, row in enumerate(rows[1:], start=2):
        if not row or not any(cell.strip() for cell in row):
            continue
        if len(row) <= sid_col or not row[sid_col].strip():
            errors.append(f'Line {i}: student ID is required')
            continue
        sid = row[sid_col].strip()
        name = row[name_col].strip() if len(row) > name_col else ''
        pin = row[pin_col].strip() if len(row) > pin_col else ''
        if len(sid) > 100:
            errors.append(f'Line {i}: student ID must be 100 characters or fewer')
            continue
        if len(name) > 200:
            errors.append(f'Line {i}: name must be 200 characters or fewer')
            continue
        sid_key = sid.casefold()
        if sid_key in seen_ids:
            errors.append(f'Line {i}: duplicate student ID "{sid}"')
            continue
        seen_ids.add(sid_key)
        if not re.fullmatch(r'[0-9]{4}', pin):
            errors.append(f'Line {i}: PIN must be exactly 4 digits for "{sid}"')
            continue
        parsed.append({'student_id': sid, 'name': name or None, 'pin': pin})

    if errors:
        return jsonify({
            'error': 'Validation failed',
            'details': errors[:20],
            'error_count': len(errors),
        }), 400
    if not parsed:
        return jsonify({'error': 'Roster must contain at least one student'}), 400

    state = query_db(
        slug,
        '''SELECT phase, session_key, roster_version
           FROM course_state WHERE course_id = ?''',
        [course['id']], one=True
    )
    form_state = {
        'expected_phase': request.form.get('expected_phase'),
        'expected_session_key': request.form.get('expected_session_key'),
        'expected_roster_version': request.form.get('expected_roster_version'),
    }
    guard = _expected_roster_state_guard(form_state, state)
    if guard:
        return jsonify({'error': guard[0]}), guard[1]
    if not state or state['phase'] != 'setup':
        return jsonify({'error': 'The roster can only be updated during setup'}), 409
    preview_token = _roster_merge_preview_token(
        raw_content, slug, state['session_key'], state['roster_version']
    )
    existing_rows = query_db(slug,
        '''SELECT id, student_id, name, pin, is_active FROM students
           WHERE course_id = ?''',
        [course['id']])
    existing_by_sid = _roster_rows_by_casefolded_id(existing_rows)
    if existing_by_sid is None:
        return jsonify({'error': _ROSTER_CASE_COLLISION_MESSAGE}), 409
    csv_sids = {p['student_id'].casefold() for p in parsed}
    preview_new = 0
    preview_restored = 0
    preview_updated = 0
    preview_unchanged = 0
    preview_pin_changed = 0
    for person in parsed:
        existing = existing_by_sid.get(person['student_id'].casefold())
        if not existing:
            preview_new += 1
            continue
        if existing['pin'] != person['pin']:
            preview_pin_changed += 1
        merged_name = (
            existing['name'] if person['name'] is None else person['name']
        )
        if not existing['is_active']:
            preview_restored += 1
        elif ((existing['name'] or None) != (merged_name or None)
                or existing['pin'] != person['pin']):
            preview_updated += 1
        else:
            preview_unchanged += 1
    preview_omitted_unchanged = sum(
        1 for sid, row in existing_by_sid.items()
        if row['is_active'] and sid not in csv_sids
    )

    confirmed = str(request.form.get('confirm', '')).lower() in ('1', 'true', 'yes')
    if not confirmed:
        return jsonify({
            'success': True,
            'requires_confirmation': True,
            'roster_upload_mode': 'merge',
            'preview_token': preview_token,
            'new': preview_new,
            # Preserve the old inclusive key for cached clients.
            'added': preview_new + preview_restored,
            'restored': preview_restored,
            'reactivated': preview_restored,
            'updated': preview_updated,
            'unchanged': preview_unchanged,
            'pin_changed': preview_pin_changed,
            'omitted_unchanged': preview_omitted_unchanged,
            'removed': 0,
            'total': len(parsed),
        })
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        locked_state = db.execute(
            '''SELECT phase, session_key, roster_version
               FROM course_state WHERE course_id = ?''',
            [course['id']]
        ).fetchone()
        guard = _expected_roster_state_guard(form_state, locked_state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        if not locked_state or locked_state['phase'] != 'setup':
            db.rollback()
            return jsonify({'error': 'The roster can only be updated during setup'}), 409
        expected_token = _roster_merge_preview_token(
            raw_content,
            slug,
            locked_state['session_key'],
            locked_state['roster_version'],
        )
        if request.form.get('preview_token') != expected_token:
            db.rollback()
            return jsonify({'error': 'Roster changed after preview; preview it again'}), 409
        existing_rows = db.execute(
            '''SELECT id, student_id, name, pin, is_active FROM students
               WHERE course_id = ?''',
            [course['id']]
        ).fetchall()
        existing_by_sid = _roster_rows_by_casefolded_id(existing_rows)
        if existing_by_sid is None:
            db.rollback()
            return jsonify({'error': _ROSTER_CASE_COLLISION_MESSAGE}), 409

        to_update = []
        to_restore = []
        to_insert = []
        unchanged = 0
        pin_changed = 0
        for person in parsed:
            existing = existing_by_sid.get(person['student_id'].casefold())
            if existing:
                merged_name = (
                    existing['name']
                    if person['name'] is None else person['name']
                )
                if existing['pin'] != person['pin']:
                    pin_changed += 1
                if not existing['is_active']:
                    to_restore.append((
                        merged_name or None, person['pin'], existing['id']
                    ))
                elif ((existing['name'] or None) != (merged_name or None)
                        or existing['pin'] != person['pin']):
                    to_update.append((
                        merged_name or None, person['pin'], existing['id']
                    ))
                else:
                    unchanged += 1
            else:
                to_insert.append((
                    course['id'], person['student_id'], person['name'] or None,
                    person['pin'],
                ))

        if to_insert or to_restore:
            freeze_guard = _session_roster_mutation_guard(
                db, course['id'], locked_state
            )
            if freeze_guard:
                db.rollback()
                return jsonify({'error': freeze_guard[0]}), freeze_guard[1]
        if to_update:
            db.executemany(
                'UPDATE students SET name = ?, pin = ? WHERE id = ?',
                to_update
            )
        if to_restore:
            db.executemany(
                '''UPDATE students
                   SET name = ?,
                       pin = ?,
                       is_active = 1,
                       last_team_id = COALESCE(last_team_id, team_id),
                       team_id = NULL,
                       last_active_at = NULL
                   WHERE id = ?''',
                to_restore
            )
        if to_insert:
            db.executemany(
                'INSERT INTO students (course_id, student_id, name, pin) VALUES (?, ?, ?, ?)',
                to_insert
            )
        if to_update or to_restore or to_insert:
            roster_version = _bump_roster_version(
                slug, course['id'], db=db
            )
        else:
            roster_version = _current_roster_version(db, course['id'])
        db.commit()
    except Exception:
        db.rollback()
        raise

    return jsonify({
        'success': True,
        'requires_confirmation': False,
        'roster_upload_mode': 'merge',
        'new': len(to_insert),
        # Preserve the old inclusive key for cached clients.
        'added': len(to_insert) + len(to_restore),
        'restored': len(to_restore),
        'reactivated': len(to_restore),
        'updated': len(to_update),
        'unchanged': unchanged,
        'pin_changed': pin_changed,
        'omitted_unchanged': sum(
            1 for sid, row in existing_by_sid.items()
            if row['is_active'] and sid not in csv_sids
        ),
        'removed': 0,
        'total': len(parsed),
        'roster_version': roster_version,
    })


@app.route('/api/start_presentation', methods=['POST'])
@instructor_login_required
def start_presentation():
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    team_id = data.get('team_id')
    question_id = data.get('question_id')
    time_cap = data.get('time_cap', 300)
    if not team_id or not question_id:
        return jsonify({'error': 'Team and question required'}), 400
    try:
        team_id = int(team_id)
        question_id = int(question_id)
        time_cap = int(time_cap)
    except (TypeError, ValueError):
        return jsonify({'error': 'Invalid team, question, or time cap'}), 400
    time_cap_was_clamped = time_cap < 10 or time_cap > 3600
    if time_cap_was_clamped:
        time_cap = max(10, min(3600, time_cap))
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        course = db.execute('SELECT id FROM courses LIMIT 1').fetchone()
        state = db.execute(
            'SELECT * FROM course_state WHERE course_id = ?', [course['id']]
        ).fetchone()
        guard = _expected_state_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        if not state or state['phase'] != 'competition':
            db.rollback()
            return jsonify({'error': f"Switch to the {PHASE_LABELS['competition']} phase first"}), 409
        if state and (state['active_team_id'] or state['active_question_id']):
            db.rollback()
            return jsonify({'error': 'Finish the active presentation first'}), 409
        max_teams = state['max_teams'] or 6
        visible_teams = db.execute(
            '''SELECT t.id, COUNT(s.id) AS member_count
               FROM teams t
               LEFT JOIN students s ON s.team_id = t.id AND s.is_active = 1
               WHERE t.course_id = ?
               GROUP BY t.id
               ORDER BY t.id LIMIT ?''',
            [course['id'], max_teams]
        ).fetchall()
        selected_team = next(
            (team for team in visible_teams if team['id'] == team_id), None
        )
        if selected_team is None:
            db.rollback()
            return jsonify({'error': 'Team is not available'}), 400
        if not selected_team['member_count']:
            db.rollback()
            return jsonify({'error': 'Cannot start a presentation for an empty team'}), 409
        question = db.execute(
            '''SELECT question_text, title, week_num, source_key
               FROM questions WHERE id = ? AND course_id = ?''',
            [question_id, course['id']]
        ).fetchone()
        if not question:
            db.rollback()
            return jsonify({'error': 'Question not found'}), 404
        selected_week = state['discussion_week'] or 1
        if (question['week_num'] or 1) != selected_week:
            db.rollback()
            return jsonify({
                'error': 'The selected question belongs to a different week'
            }), 409

        is_appendix = str(question['source_key'] or '').startswith('appendix:')
        is_canonical = str(question['source_key'] or '').startswith(
            f'week-{selected_week}-q-'
        )
        if not is_appendix and not is_canonical:
            db.rollback()
            return jsonify({
                'error': 'The selected question is no longer in this week'
            }), 409
        if is_appendix:
            try:
                sync_appendix_questions(
                    slug, course['id'], selected_week, db=db, commit=False
                )
            except QuestionParseError as exc:
                db.rollback()
                return jsonify({'error': str(exc)}), 422
        else:
            try:
                catalog_week = _validate_course_question_catalog(
                    slug,
                    weeks=[selected_week],
                ).get_week(selected_week)
            except (OSError, ValueError):
                catalog_week = None
            if not catalog_week or not catalog_week.presentation.ready:
                db.rollback()
                return jsonify({
                    'error': f'Week {selected_week} questions are not ready'
                }), 409
            sync_presentation_questions(
                slug, course['id'], selected_week, db=db, commit=False
            )

        question = db.execute(
            '''SELECT question_text, title FROM questions
               WHERE id = ? AND course_id = ?''',
            [question_id, course['id']]
        ).fetchone()
        if not question:
            db.rollback()
            return jsonify({
                'error': 'The question list changed; reload before starting'
            }), 409

        question_text = question['title'] or question['question_text']
        started_at = _utcnow().strftime('%Y-%m-%d %H:%M:%S.%f')
        presentation_key = f'pres-{uuid.uuid4().hex}'
        db.execute(
            '''UPDATE course_state
               SET phase = 'competition', active_team_id = ?, active_question_id = ?,
                   current_question = ?, presentation_started_at = ?,
                   presentation_created_at = ?,
                   presentation_time_cap = ?, presentation_remaining = NULL,
                   poll_active = 0, poll_question_key = ?, poll_started_at = NULL,
                   poll_closed_at = NULL,
                   challenge_ratings_closed_at = NULL,
                   current_discussion_key = NULL,
                   current_discussion_source_key = NULL,
                   current_discussion_title = NULL,
                   current_discussion_content = NULL
               WHERE course_id = ?''',
            [team_id, question_id, question_text, started_at, started_at, time_cap,
             presentation_key, course['id']]
        )
        db.commit()
        response = {
            'success': True,
            'presentation_key': presentation_key,
            'time_cap': time_cap,
        }
        if time_cap_was_clamped:
            response['notice'] = (
                f'Time cap adjusted to {time_cap}s (allowed range 10 to 3600 seconds)'
            )
        return jsonify(response)
    except Exception:
        db.rollback()
        raise


@app.route('/api/stop_presentation', methods=['POST'])
@instructor_login_required
def stop_presentation():
    """Pause the presentation timer and save remaining time."""
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        state = db.execute('SELECT * FROM course_state LIMIT 1').fetchone()
        guard = _presentation_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        if not state['presentation_started_at']:
            db.rollback()
            return jsonify({'error': 'Presentation timer is already paused'}), 409
        try:
            started = _parse_db_datetime(state['presentation_started_at'])
        except (ValueError, TypeError):
            db.rollback()
            return jsonify({'error': 'Presentation start time is corrupted'}), 409
        elapsed = (_utcnow() - started).total_seconds()
        cap = state['presentation_time_cap'] or 300
        remaining = max(0, int(cap - elapsed))
        db.execute(
            '''UPDATE course_state
               SET presentation_started_at = NULL, presentation_remaining = ?
               WHERE course_id = ?''',
            [remaining, state['course_id']]
        )
        db.commit()
        return jsonify({'success': True, 'remaining': remaining})
    except Exception:
        db.rollback()
        raise


@app.route('/api/resume_presentation', methods=['POST'])
@instructor_login_required
def resume_presentation():
    """Resume a paused presentation while keeping the original time cap."""
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        state = db.execute('SELECT * FROM course_state LIMIT 1').fetchone()
        guard = _presentation_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        remaining = state['presentation_remaining']
        if remaining is None or remaining <= 0:
            db.rollback()
            return jsonify({'error': 'No remaining time to resume'}), 409
        cap = state['presentation_time_cap'] or 300
        consumed = cap - remaining
        shifted_start = _utcnow() - timedelta(seconds=consumed)
        db.execute(
            '''UPDATE course_state
               SET presentation_started_at = ?, presentation_remaining = NULL
               WHERE course_id = ?''',
            [shifted_start.strftime('%Y-%m-%d %H:%M:%S'), state['course_id']]
        )
        db.commit()
        return jsonify({'success': True, 'remaining': remaining})
    except Exception:
        db.rollback()
        raise


@app.route('/api/reset_presentation_timer', methods=['POST'])
@instructor_login_required
def reset_presentation_timer():
    """Reset the timer to the original time cap (paused state)."""
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        state = db.execute('SELECT * FROM course_state LIMIT 1').fetchone()
        guard = _presentation_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        cap = state['presentation_time_cap'] or 300
        db.execute(
            '''UPDATE course_state
               SET presentation_started_at = NULL, presentation_remaining = ?
               WHERE course_id = ?''',
            [cap, state['course_id']]
        )
        db.commit()
        return jsonify({'success': True, 'cap': cap})
    except Exception:
        db.rollback()
        raise


@app.route('/api/next_presentation', methods=['POST'])
@instructor_login_required
def next_presentation():
    """Stop current presentation, save to history, clear for next."""
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        state = db.execute('SELECT * FROM course_state LIMIT 1').fetchone()
        displayed_presentation = str(
            data.get('presentation_key') or ''
        ).strip()
        if (state and state['phase'] == 'competition' and
                not state['active_team_id'] and
                not state['active_question_id'] and
                not active_presentation_key(state) and
                displayed_presentation):
            guard = _expected_state_guard(data, state)
            if guard:
                db.rollback()
                return jsonify({'error': guard[0]}), guard[1]
            history_match = _presentation_history_match(
                state, displayed_presentation
            )
            db.rollback()
            if history_match == 'current_session':
                return jsonify({'success': True, 'already_finished': True})
            if history_match == 'other_session':
                return jsonify({
                    'error': (
                        'This presentation belongs to an earlier session. '
                        'Reload before continuing.'
                    ),
                    'outcome': 'stale_session',
                }), 409
            return jsonify({
                'error': (
                    'Another page canceled this presentation. '
                    'No participation was recorded.'
                ),
                'outcome': 'canceled',
            }), 409
        guard = _presentation_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        transition, ratings_changed = _prepare_rating_transition(
            db,
            state,
            'Stop the active rating poll before finishing this presentation',
            now=g.request_arrived_at,
            poll_duration=get_poll_duration(slug),
        )
        if transition:
            if ratings_changed:
                db.commit()
            else:
                db.rollback()
            return jsonify(transition), 409
        _finalize_active_presentation(slug, state['course_id'], db=db)
        db.commit()
        return jsonify({'success': True})
    except Exception:
        db.rollback()
        raise


@app.route('/api/start_poll', methods=['POST'])
@instructor_login_required
def start_poll():
    """Open the configured rating window for the active presentation."""
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        state = db.execute('SELECT * FROM course_state LIMIT 1').fetchone()
        guard = _presentation_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        poll_duration = get_poll_duration(slug)
        if _poll_is_open(
                state, now=g.request_arrived_at,
                poll_duration=poll_duration):
            started_at = state['poll_started_at']
            db.rollback()
            return jsonify({
                'success': True,
                'already_active': True,
                'poll_started_at': started_at,
                'poll_duration': poll_duration,
                'poll_remaining': _derive_timing_state(
                    state, now=g.request_arrived_at,
                    poll_duration=poll_duration)['poll_remaining'],
            })
        settling = _ratings_settling_state(
            state, now=g.request_arrived_at, poll_duration=poll_duration
        )
        if settling['ratings_settling']:
            db.rollback()
            return jsonify({
                'error': POLL_SETTLING_MESSAGE,
                **settling,
            }), 409
        # Derive the rating key from server state — never trust client input.
        question_key = active_presentation_key(state)
        if not question_key:
            db.rollback()
            return jsonify({'error': 'No active presentation'}), 400
        started_at = _utcnow().strftime('%Y-%m-%d %H:%M:%S.%f')
        db.execute(
            '''UPDATE course_state
               SET poll_active = 1, poll_question_key = ?, poll_started_at = ?,
                   poll_closed_at = NULL, challenge_ratings_closed_at = NULL
               WHERE course_id = ?''',
            [question_key, started_at, state['course_id']]
        )
        fresh = db.execute(
            'SELECT * FROM course_state WHERE course_id = ?',
            [state['course_id']]
        ).fetchone()
        # Build the response before committing so that a failure during
        # response construction rolls back the poll-start rather than
        # leaving a committed poll with no response to the instructor.
        timing = _derive_timing_state(fresh, poll_duration=poll_duration)
        response = jsonify({
            'success': True,
            'poll_started_at': fresh['poll_started_at'] if fresh else None,
            'poll_duration': poll_duration,
            'poll_remaining': timing['poll_remaining'],
        })
        db.commit()
        return response
    except Exception:
        db.rollback()
        raise


@app.route('/api/stop_poll', methods=['POST'])
@instructor_login_required
def stop_poll():
    """Close the active rating window."""
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        state = db.execute('SELECT * FROM course_state LIMIT 1').fetchone()
        guard = _presentation_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]

        already_closed = bool(state['poll_closed_at'])
        if state['poll_started_at'] and not already_closed:
            cutoff = g.request_arrived_at.strftime('%Y-%m-%d %H:%M:%S.%f')
            natural_cutoff = _poll_cutoff(
                state, poll_duration=get_poll_duration(slug)
            )
            challenge_cutoff = min(
                g.request_arrived_at, natural_cutoff or g.request_arrived_at
            ).strftime('%Y-%m-%d %H:%M:%S.%f')
            db.execute(
                '''UPDATE course_state
                   SET poll_active = 0, poll_closed_at = ?
                   WHERE course_id = ?''',
                [cutoff, state['course_id']],
            )
            if (_has_active_challenges(state) and
                    not state['challenge_ratings_closed_at']):
                db.execute(
                    '''UPDATE course_state
                       SET challenge_ratings_closed_at = ?
                       WHERE course_id = ?''',
                    [challenge_cutoff, state['course_id']],
                )
        elif state['poll_active']:
            db.execute(
                'UPDATE course_state SET poll_active = 0 WHERE course_id = ?',
                [state['course_id']],
            )

        fresh = db.execute(
            'SELECT * FROM course_state WHERE course_id = ?',
            [state['course_id']],
        ).fetchone()
        settling = _ratings_settling_state(
            fresh,
            now=g.request_arrived_at,
            poll_duration=get_poll_duration(slug),
        )
        db.commit()
        return jsonify({
            'success': True,
            'already_stopped': already_closed,
            'poll_closed_at': fresh['poll_closed_at'],
            **settling,
        })
    except Exception:
        db.rollback()
        raise


@app.route('/api/cancel_presentation', methods=['POST'])
@instructor_login_required
def cancel_presentation():
    """Clear a mistaken presentation without adding it to history."""
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        state = db.execute('SELECT * FROM course_state LIMIT 1').fetchone()
        displayed_presentation = str(
            data.get('presentation_key') or ''
        ).strip()
        if (state and state['phase'] == 'competition' and
                not state['active_team_id'] and
                not state['active_question_id'] and
                not active_presentation_key(state) and
                displayed_presentation):
            guard = _expected_state_guard(data, state)
            if guard:
                db.rollback()
                return jsonify({'error': guard[0]}), guard[1]
            history_match = _presentation_history_match(
                state, displayed_presentation
            )
            db.rollback()
            if history_match == 'current_session':
                return jsonify({
                    'error': (
                        'Another page already finished this presentation. '
                        'Its participation records were kept.'
                    ),
                    'outcome': 'finished',
                }), 409
            if history_match == 'other_session':
                return jsonify({
                    'error': (
                        'This presentation belongs to an earlier session. '
                        'Reload before continuing.'
                    ),
                    'outcome': 'stale_session',
                }), 409
            return jsonify({'success': True, 'already_canceled': True})
        guard = _presentation_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        transition, ratings_changed = _prepare_rating_transition(
            db,
            state,
            'Stop the active rating poll before cancelling this presentation',
            now=g.request_arrived_at,
            poll_duration=get_poll_duration(slug),
        )
        if transition:
            if ratings_changed:
                db.commit()
            else:
                db.rollback()
            return jsonify(transition), 409
        presentation_key = active_presentation_key(state)
        counts = db.execute(
            '''SELECT
                   (SELECT COUNT(*) FROM presentation_ratings
                    WHERE course_id = ? AND question_key = ?)
                       AS presentation_rating_count,
                   (SELECT COUNT(*) FROM challenge_ratings
                    WHERE course_id = ? AND presentation_key = ?)
                       AS challenge_rating_count''',
            [state['course_id'], presentation_key,
             state['course_id'], presentation_key]
        ).fetchone()
        presentation_count = counts['presentation_rating_count']
        challenge_count = counts['challenge_rating_count']
        total_count = presentation_count + challenge_count
        if total_count and not data.get('discard_ratings'):
            db.rollback()
            return jsonify({
                'error': 'This presentation already has saved ratings',
                'rating_count': total_count,
                'presentation_rating_count': presentation_count,
                'challenge_rating_count': challenge_count,
                'requires_discard': True,
            }), 409
        db.execute(
            '''DELETE FROM presentation_ratings
               WHERE course_id = ? AND question_key = ?''',
            [state['course_id'], presentation_key]
        )
        # Clean up any challenge data for this presentation.
        db.execute(
            '''DELETE FROM challenge_ratings
               WHERE course_id = ? AND presentation_key = ?''',
            [state['course_id'], presentation_key]
        )
        db.execute(
            '''DELETE FROM challenge_rounds
               WHERE course_id = ? AND presentation_key = ?''',
            [state['course_id'], presentation_key]
        )
        db.execute(
            '''DELETE FROM challenge_hands
               WHERE course_id = ? AND presentation_key = ?''',
            [state['course_id'], presentation_key]
        )
        _clear_active_presentation(db, state['course_id'])
        db.commit()
        return jsonify({
            'success': True,
            'discarded_ratings': total_count,
            'discarded_presentation_ratings': presentation_count,
            'discarded_challenge_ratings': challenge_count,
        })
    except Exception:
        db.rollback()
        raise


# ---------------------------------------------------------------------------
# Challenge API — raise hand, select challenger, rate challenger
# ---------------------------------------------------------------------------

def _student_has_active_challenge(
        db, course_id, presentation_key, student_id):
    """Return whether this student is already selected in the presentation."""
    return db.execute(
        '''SELECT 1 FROM challenge_rounds
           WHERE course_id = ? AND presentation_key = ?
             AND challenger_id = ?
           LIMIT 1''',
        [course_id, presentation_key, student_id],
    ).fetchone() is not None


@app.route('/api/raise_hand', methods=['POST'])
@student_login_required
def raise_hand():
    """Student raises their hand to challenge the presenting team."""
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    expected_key = str(data.get('presentation_key') or '').strip()
    if not expected_key:
        return jsonify({'error': 'Presentation key is required'}), 400
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        state = db.execute('SELECT * FROM course_state LIMIT 1').fetchone()
        if not state or state['phase'] != 'competition':
            db.rollback()
            return jsonify({'error': f"Not in the {PHASE_LABELS['competition']} phase"}), 403
        pres_key = active_presentation_key(state)
        if not pres_key or pres_key != expected_key:
            db.rollback()
            return jsonify({'error': 'The presentation has changed'}), 409
        student = db.execute(
            'SELECT * FROM students WHERE student_id = ? AND is_active = 1',
            [session['student_id']]
        ).fetchone()
        if not student:
            db.rollback()
            return jsonify({'error': 'Student not found'}), 404
        if not student['team_id']:
            db.rollback()
            return jsonify({'error': 'Join a team first'}), 403
        if student['team_id'] == state['active_team_id']:
            db.rollback()
            return jsonify({'error': 'You cannot challenge your own team'}), 403
        if _student_has_active_challenge(
                db, student['course_id'], pres_key, student['id']):
            db.rollback()
            return jsonify({
                'error': 'You are already the active challenger',
                'already_selected': True,
            }), 409
        # Insert (idempotent via UNIQUE constraint)
        team = db.execute(
            'SELECT name FROM teams WHERE id = ?', [student['team_id']]
        ).fetchone()
        student_name = _student_display_name(student)
        db.execute(
            '''INSERT INTO challenge_hands
               (course_id, session_key, presentation_key, student_id,
                student_name, student_team_id, student_team_name)
               VALUES (?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(course_id, presentation_key, student_id) DO NOTHING''',
            [student['course_id'], state['session_key'] or 0, pres_key,
             student['id'], student_name, student['team_id'],
             team['name'] if team else 'Unknown']
        )
        db.commit()
        return jsonify({'success': True})
    except Exception:
        db.rollback()
        raise


@app.route('/api/lower_hand', methods=['POST'])
@student_login_required
def lower_hand():
    """Student lowers their hand (withdraws challenge request)."""
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    expected_key = str(data.get('presentation_key') or '').strip()
    if not expected_key:
        return jsonify({'error': 'Presentation key is required'}), 400
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        state = db.execute('SELECT * FROM course_state LIMIT 1').fetchone()
        if not state or state['phase'] != 'competition':
            db.rollback()
            return jsonify({'error': f"Not in the {PHASE_LABELS['competition']} phase"}), 403
        pres_key = active_presentation_key(state)
        if not pres_key or pres_key != expected_key:
            db.rollback()
            return jsonify({'error': 'The presentation has changed'}), 409
        student = db.execute(
            'SELECT id FROM students WHERE student_id = ? AND is_active = 1',
            [session['student_id']]
        ).fetchone()
        if not student:
            db.rollback()
            return jsonify({'error': 'Student not found'}), 404
        db.execute(
            '''DELETE FROM challenge_hands
               WHERE course_id = ? AND presentation_key = ? AND student_id = ?''',
            [state['course_id'], pres_key, student['id']]
        )
        db.commit()
        return jsonify({'success': True})
    except Exception:
        db.rollback()
        raise


@app.route('/api/select_challenger', methods=['POST'])
@instructor_login_required
def select_challenger():
    """Instructor picks a student from the raised-hands list as a challenger."""
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    target_student_db_id = data.get('student_id')
    if target_student_db_id is None:
        return jsonify({'error': 'Student ID is required'}), 400
    try:
        target_student_db_id = int(target_student_db_id)
    except (TypeError, ValueError):
        return jsonify({'error': 'Invalid student ID'}), 400
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        state = db.execute('SELECT * FROM course_state LIMIT 1').fetchone()
        guard = _presentation_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        settling = _ratings_settling_state(
            state,
            now=g.request_arrived_at,
            poll_duration=get_poll_duration(slug),
        )
        if settling['ratings_settling']:
            db.rollback()
            return jsonify({
                'error': POLL_SETTLING_MESSAGE,
                **settling,
            }), 409
        pres_key = active_presentation_key(state)
        # Verify the student has raised their hand.
        hand = db.execute(
            '''SELECT hand.*, student.student_id AS roster_student_id,
                      student.name AS current_roster_name,
                      student.display_name AS current_display_name
               FROM challenge_hands hand
               LEFT JOIN students student ON student.id = hand.student_id
               WHERE hand.course_id = ? AND hand.presentation_key = ?
                 AND hand.student_id = ?''',
            [state['course_id'], pres_key, target_student_db_id]
        ).fetchone()
        if not hand:
            db.rollback()
            return jsonify({'error': 'This student has not raised their hand'}), 409
        if _student_has_active_challenge(
                db, state['course_id'], pres_key, hand['student_id']):
            # Normalize a stale hand left by an older client or data set.
            db.execute(
                '''DELETE FROM challenge_hands
                   WHERE course_id = ? AND presentation_key = ?
                     AND student_id = ?''',
                [state['course_id'], pres_key, hand['student_id']],
            )
            db.commit()
            return jsonify({
                'error': 'This student is already the active challenger',
                'already_selected': True,
            }), 409
        challenger_name = _student_display_name({
            'display_name': hand['current_display_name'],
            'roster_name': hand['current_roster_name'],
            'student_id': hand['roster_student_id'],
        })
        if challenger_name == 'Unknown':
            challenger_name = (
                _normalized_identity_text(hand['student_name'])
                or 'Unknown'
            )
        # Compute next challenge number.
        max_num_row = db.execute(
            '''SELECT MAX(challenge_num) AS m FROM challenge_rounds
               WHERE course_id = ? AND presentation_key = ?''',
            [state['course_id'], pres_key]
        ).fetchone()
        challenge_num = (max_num_row['m'] or 0) + 1
        challenge_key = f"{pres_key}-ch{challenge_num}-{uuid.uuid4().hex}"
        presenting_team = db.execute(
            'SELECT name FROM teams WHERE id = ?',
            [state['active_team_id']]
        ).fetchone()
        question = db.execute(
            'SELECT title, question_text FROM questions WHERE id = ?',
            [state['active_question_id']]
        ).fetchone()
        db.execute(
            '''INSERT INTO challenge_rounds
               (course_id, session_key, week_num, presentation_key,
                challenge_key, challenge_num, challenger_id, challenger_name,
                challenger_team_id, challenger_team_name,
                presenting_team_id, presenting_team_name,
                question_id, question_title, data_version)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            [state['course_id'], state['session_key'] or 0,
             state['discussion_week'] or 1, pres_key,
             challenge_key, challenge_num,
             hand['student_id'], challenger_name,
             hand['student_team_id'], hand['student_team_name'],
             state['active_team_id'],
             presenting_team['name'] if presenting_team else 'Unknown',
             state['active_question_id'],
             (question['title'] or question['question_text']) if question else '',
             APP_VERSION]
        )
        # Remove from raised hands.
        db.execute(
            '''DELETE FROM challenge_hands
               WHERE course_id = ? AND presentation_key = ? AND student_id = ?''',
            [state['course_id'], pres_key, hand['student_id']]
        )
        # Append to active_challenges_json (bumps state_version via trigger).
        challenges = json.loads(state['active_challenges_json'] or '[]')
        challenges.append({
            'challenge_key': challenge_key,
            'challenge_num': challenge_num,
            'challenger_id': hand['student_id'],
            'challenger_name': challenger_name,
            'challenger_student_id': hand['roster_student_id'],
            'challenger_team_id': hand['student_team_id'],
            'challenger_team_name': hand['student_team_name'],
        })
        db.execute(
            """UPDATE course_state
               SET active_challenges_json = ?, challenge_ratings_closed_at = NULL
               WHERE course_id = ?""",
            [json.dumps(challenges), state['course_id']]
        )
        fresh = db.execute(
            '''SELECT poll_started_at, poll_closed_at,
                      challenge_ratings_closed_at
               FROM course_state WHERE course_id = ?''',
            [state['course_id']],
        ).fetchone()
        challenge_ratings_open = _challenge_ratings_are_open(
            fresh, now=g.request_arrived_at, poll_duration=get_poll_duration(slug)
        )
        db.commit()
        return jsonify({
            'success': True,
            'challenge_key': challenge_key,
            'challenge_num': challenge_num,
            'challenge_ratings_open': challenge_ratings_open,
        })
    except Exception:
        db.rollback()
        raise


@app.route('/api/clear_challenger', methods=['POST'])
@instructor_login_required
def clear_challenger():
    """Instructor removes a challenger (mistake selection)."""
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    challenge_key = str(data.get('challenge_key') or '').strip()
    if not challenge_key:
        return jsonify({'error': 'Challenge key is required'}), 400
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        state = db.execute('SELECT * FROM course_state LIMIT 1').fetchone()
        guard = _presentation_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        challenge = db.execute(
            '''SELECT presentation_key, challenger_id FROM challenge_rounds
               WHERE course_id = ? AND challenge_key = ?''',
            [state['course_id'], challenge_key]
        ).fetchone()
        if not challenge:
            db.rollback()
            return jsonify({'success': True, 'already_cleared': True})
        if challenge['presentation_key'] != active_presentation_key(state):
            db.rollback()
            return jsonify({
                'error': 'This challenge belongs to a different presentation'
            }), 409
        transition, ratings_changed = _prepare_rating_transition(
            db,
            state,
            'Stop the active rating poll before clearing this challenger',
            now=g.request_arrived_at,
            poll_duration=get_poll_duration(slug),
        )
        if transition:
            if ratings_changed:
                db.commit()
            else:
                db.rollback()
            return jsonify(transition), 409
        challenge_count = db.execute(
            '''SELECT COUNT(*) AS c FROM challenge_ratings
               WHERE course_id = ? AND challenge_key = ?''',
            [state['course_id'], challenge_key]
        ).fetchone()['c']
        if challenge_count and not data.get('discard_ratings'):
            db.rollback()
            return jsonify({
                'error': 'This challenger already has saved ratings',
                'rating_count': challenge_count,
                'challenge_rating_count': challenge_count,
                'requires_discard': True,
            }), 409
        db.execute(
            '''DELETE FROM challenge_ratings
               WHERE course_id = ? AND challenge_key = ?''',
            [state['course_id'], challenge_key]
        )
        db.execute(
            '''DELETE FROM challenge_rounds
               WHERE course_id = ? AND challenge_key = ?''',
            [state['course_id'], challenge_key]
        )
        db.execute(
            '''DELETE FROM challenge_hands
               WHERE course_id = ? AND presentation_key = ?
                 AND student_id = ?''',
            [state['course_id'], challenge['presentation_key'],
             challenge['challenger_id']],
        )
        # Remove from active_challenges_json
        challenges = json.loads(state['active_challenges_json'] or '[]')
        challenges = [c for c in challenges if c.get('challenge_key') != challenge_key]
        if challenges and not state['poll_started_at']:
            db.execute(
                '''UPDATE course_state
                   SET active_challenges_json = ?,
                       challenge_ratings_closed_at = NULL
                   WHERE course_id = ?''',
                [json.dumps(challenges), state['course_id']]
            )
        else:
            db.execute(
                '''UPDATE course_state SET active_challenges_json = ?
                   WHERE course_id = ?''',
                [json.dumps(challenges), state['course_id']]
            )
        db.commit()
        return jsonify({
            'success': True,
            'discarded_challenge_ratings': challenge_count,
        })
    except Exception:
        db.rollback()
        raise


@app.route('/api/submit_challenge_rating', methods=['POST'])
@student_login_required
def submit_challenge_rating():
    """Student submits a 1-5 rating for a challenger's question quality."""
    request_arrived_at = g.request_arrived_at
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    score = data.get('score')
    challenge_key = str(data.get('challenge_key') or '').strip()
    if score is None:
        return jsonify({'error': 'Score is required'}), 400
    try:
        score = _rating_integer(score)
    except (ValueError, TypeError):
        return jsonify({'error': 'Score must be an integer'}), 400
    if not (1 <= score <= 5):
        return jsonify({'error': 'Score must be 1-5'}), 400
    if not challenge_key:
        return jsonify({'error': 'Challenge key is required'}), 400
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        state = db.execute('SELECT * FROM course_state LIMIT 1').fetchone()
        if not state or state['phase'] != 'competition':
            db.rollback()
            return jsonify({'error': f"Not in the {PHASE_LABELS['competition']} phase"}), 403
        pres_key = active_presentation_key(state)
        if not pres_key:
            db.rollback()
            return jsonify({'error': 'No active presentation'}), 403
        student = db.execute(
            'SELECT * FROM students WHERE student_id = ? AND is_active = 1',
            [session['student_id']]
        ).fetchone()
        if not student:
            db.rollback()
            return jsonify({'error': 'Student not found'}), 404
        if not student['team_id']:
            db.rollback()
            return jsonify({'error': 'Join a team before rating'}), 403
        # Verify the challenge exists and is active.
        challenge = db.execute(
            '''SELECT * FROM challenge_rounds
               WHERE course_id = ? AND challenge_key = ?
                 AND popping_version_compatible(data_version, ?) = 1''',
            [student['course_id'], challenge_key, SCHEMA_VERSION]
        ).fetchone()
        if not challenge:
            db.rollback()
            return jsonify({'error': 'Challenge not found'}), 404
        # Validate it belongs to the current presentation.
        if challenge['presentation_key'] != pres_key:
            db.rollback()
            return jsonify({'error': 'This challenge is from a different presentation'}), 409
        if not _challenge_ratings_accept(
                state,
                now=request_arrived_at,
                poll_duration=get_poll_duration(slug),
        ):
            db.rollback()
            return jsonify({'error': 'Challenge rating is closed'}), 403
        # Rating eligibility: rater cannot be the challenger or on
        # the challenger's team, and cannot be on the presenting team.
        if student['id'] == challenge['challenger_id']:
            db.rollback()
            return jsonify({'error': 'You cannot rate your own challenge'}), 403
        if student['team_id'] == challenge['challenger_team_id']:
            db.rollback()
            return jsonify({'error': 'You cannot rate a teammate\'s challenge'}), 403
        if student['team_id'] == challenge['presenting_team_id']:
            db.rollback()
            return jsonify({'error': 'Presenting team cannot rate challenges'}), 403
        rater_team = db.execute(
            'SELECT name FROM teams WHERE id = ?', [student['team_id']]
        ).fetchone()
        db.execute(
            '''INSERT INTO challenge_ratings
               (course_id, session_key, week_num, challenge_key,
                presentation_key, challenger_id, challenger_name,
                challenger_team_id, challenger_team_name,
                rater_id, rater_name, rater_team_id, rater_team_name, score,
                data_version)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(course_id, challenge_key, rater_id)
               DO UPDATE SET score = excluded.score,
                             created_at = CURRENT_TIMESTAMP''',
            [student['course_id'], state['session_key'] or 0,
             state['discussion_week'] or 1, challenge_key,
             pres_key, challenge['challenger_id'],
             challenge['challenger_name'],
             challenge['challenger_team_id'],
             challenge['challenger_team_name'],
             student['id'], _student_display_name(student),
             student['team_id'],
             rater_team['name'] if rater_team else 'Unknown',
             score, APP_VERSION]
        )
        db.commit()
        return jsonify({'success': True})
    except Exception:
        db.rollback()
        raise


@app.route('/api/submit_rating', methods=['POST'])
@student_login_required
def submit_rating():
    """Student submits star ratings (1–5) for the current presentation."""
    request_arrived_at = g.request_arrived_at
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    q1 = data.get('q1_developed')
    q2 = data.get('q2_easy')
    expected_key = data.get('presentation_key')
    if q1 is None or q2 is None:
        return jsonify({'error': 'Both ratings required'}), 400
    try:
        q1 = _rating_integer(q1)
        q2 = _rating_integer(q2)
    except (ValueError, TypeError):
        return jsonify({'error': 'Ratings must be integers'}), 400
    if not (1 <= q1 <= 5 and 1 <= q2 <= 5):
        return jsonify({'error': 'Ratings must be 1–5'}), 400

    if not isinstance(expected_key, str) or not expected_key:
        return jsonify({'error': 'Presentation key is required'}), 400

    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        state = db.execute('SELECT * FROM course_state LIMIT 1').fetchone()
        # Read and write under one lock so a delayed request cannot cross into the
        # next presentation after it passed validation.
        if not state or state['phase'] != 'competition' or \
           not state['active_team_id'] or not state['active_question_id']:
            db.rollback()
            return jsonify({'error': 'No active presentation to rate'}), 403

        student = db.execute(
            'SELECT * FROM students WHERE student_id = ? AND is_active = 1',
            [session['student_id']]
        ).fetchone()
        if not student:
            db.rollback()
            return jsonify({'error': 'Student not found'}), 404

        # Block self-grading — the presenting team cannot rate their own presentation.
        if student['team_id'] and state['active_team_id'] and \
           student['team_id'] == state['active_team_id']:
            db.rollback()
            return jsonify({'error': 'You cannot rate your own presentation'}), 403
        question_key = active_presentation_key(state)
        if not question_key:
            db.rollback()
            return jsonify({'error': 'No active presentation to rate'}), 403
        if expected_key != question_key:
            db.rollback()
            return jsonify({'error': 'The presentation has changed; refresh and try again'}), 409
        if not _poll_accepts_rating(
                state,
                now=request_arrived_at,
                poll_duration=get_poll_duration(slug)):
            db.rollback()
            return jsonify({'error': 'The rating poll is closed'}), 403
        if not student['team_id']:
            db.rollback()
            return jsonify({'error': 'Join a team before rating presentations'}), 403

        presenting_team = db.execute(
            'SELECT name FROM teams WHERE id = ? AND course_id = ?',
            [state['active_team_id'], student['course_id']]
        ).fetchone()
        rater_team = db.execute(
            'SELECT name FROM teams WHERE id = ? AND course_id = ?',
            [student['team_id'], student['course_id']]
        ).fetchone()
        question = db.execute(
            '''SELECT title, question_text FROM questions
               WHERE id = ? AND course_id = ?''',
            [state['active_question_id'], student['course_id']]
        ).fetchone()
        db.execute(
            '''INSERT INTO presentation_ratings
               (course_id, student_id, question_key, session_key, week_num,
                presenting_team_id, presenting_team_name, question_id,
                question_title, rater_team_id, rater_team_name,
                q1_developed, q2_easy, data_version)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(course_id, student_id, question_key)
               DO UPDATE SET q1_developed = excluded.q1_developed,
                             q2_easy = excluded.q2_easy,
                             week_num = excluded.week_num,
                             rater_team_id = excluded.rater_team_id,
                             rater_team_name = excluded.rater_team_name,
                             created_at = CURRENT_TIMESTAMP''',
            [student['course_id'], student['id'], question_key,
             state['session_key'] or 0, state['discussion_week'] or 1,
             state['active_team_id'],
             presenting_team['name'] if presenting_team else 'Unknown',
             state['active_question_id'],
             (question['title'] or question['question_text']) if question else '',
             student['team_id'], rater_team['name'] if rater_team else 'Unknown',
             q1, q2, APP_VERSION]
        )
        db.commit()
        return jsonify({'success': True, 'presentation_key': question_key})
    except Exception:
        db.rollback()
        raise


def _escape_like(term):
    """Escape LIKE wildcards so a search term matches literally."""
    return term.replace('\\', '\\\\').replace('%', '\\%').replace('_', '\\_')


@app.route('/api/students', methods=['GET'])
@instructor_login_required
def api_students():
    slug = session['slug']
    course = query_db(slug, 'SELECT id FROM courses LIMIT 1', one=True)
    page = max(1, request.args.get('page', 1, type=int))
    per_page = min(max(1, request.args.get('per_page', 10, type=int)), 100)
    sort_col = request.args.get('sort', 'student_id')
    order = request.args.get('order', 'asc')
    search = request.args.get('search', '').strip()
    team_filter = request.args.get('team', '')

    allowed_sorts = {
        'student_id': 's.student_id',
        'name': "COALESCE(NULLIF(TRIM(s.display_name), ''), "
                "NULLIF(TRIM(s.name), ''), s.student_id) "
                "COLLATE NOCASE",
        'last_active_at': 's.last_active_at',
        'presentation_count': 'presentation_count',
        'challenger_count': 'challenger_count',
    }
    if sort_col not in allowed_sorts:
        sort_col = 'student_id'
    sort_sql = allowed_sorts[sort_col]
    order_sql = 'DESC' if order == 'desc' else 'ASC'

    presence_checked_at = _utcnow()

    where_clause = ''
    params = [course['id']]
    if search:
        escaped_search = f'%{_escape_like(search)}%'
        where_clause += (
            " AND (s.student_id LIKE ? ESCAPE '\\' "
            "OR COALESCE(s.display_name, '') LIKE ? ESCAPE '\\' "
            "OR COALESCE(s.name, '') LIKE ? ESCAPE '\\')"
        )
        params.extend((escaped_search, escaped_search, escaped_search))
    if team_filter == 'none':
        where_clause += ' AND s.team_id IS NULL'
    elif team_filter and team_filter != 'none':
        try:
            where_clause += ' AND s.team_id = ?'
            params.append(int(team_filter))
        except ValueError:
            pass  # invalid team filter — show all

    count = query_db(slug,
        f'''SELECT COUNT(*) as c FROM students s
            WHERE s.course_id = ? AND s.is_active = 1 {where_clause}''',
        params, one=True
    )
    total = count['c']
    course_total = query_db(
        slug,
        '''SELECT COUNT(*) AS c FROM students
           WHERE course_id = ? AND is_active = 1''',
        [course['id']], one=True,
    )['c']
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    offset = (page - 1) * per_page

    rows = query_db(slug,
        f'''WITH presentation_counts AS (
                SELECT student_id, COUNT(*) AS presentation_count
                FROM presentation_participants
                WHERE course_id = ?
                  AND typeof(week_num) = 'integer'
                  AND week_num > 0
                  AND popping_version_compatible(data_version, ?) = 1
                GROUP BY student_id
            ),
            challenger_counts AS (
                SELECT challenger_id AS student_id,
                       COUNT(*) AS challenger_count
                FROM challenge_rounds
                WHERE course_id = ?
                  AND typeof(week_num) = 'integer'
                  AND week_num > 0
                  AND popping_version_compatible(data_version, ?) = 1
                GROUP BY challenger_id
            )
            SELECT s.id, s.student_id, s.name, s.display_name, s.team_id,
                   t.name as team_name, t.color as team_color,
                   s.last_login_at, s.last_active_at, s.last_team_joined_at,
                   COALESCE(presentation.presentation_count, 0)
                       AS presentation_count,
                   COALESCE(challenger.challenger_count, 0)
                       AS challenger_count
            FROM students s
            LEFT JOIN teams t ON s.team_id = t.id
            LEFT JOIN presentation_counts presentation
              ON presentation.student_id = s.id
            LEFT JOIN challenger_counts challenger
              ON challenger.student_id = s.id
            WHERE s.course_id = ? AND s.is_active = 1 {where_clause}
            ORDER BY {sort_sql} {order_sql},
                     s.student_id COLLATE NOCASE ASC, s.id ASC
            LIMIT ? OFFSET ?''',
        [course['id'], SCHEMA_VERSION, course['id'], SCHEMA_VERSION]
        + params + [per_page, offset]
    )

    students = []
    for r in rows:
        d = dict(r)
        d['is_online'] = _student_is_online(
            r['last_active_at'],
            now=presence_checked_at,
        )
        d['roster_name'] = r['name']
        students.append(d)

    max_teams = get_max_teams(slug, course['id'])
    teams = query_db(slug,
        'SELECT id, name, color FROM teams WHERE course_id = ? ORDER BY id LIMIT ?',
        [course['id'], max_teams]
    )

    return jsonify({
        'students': students,
        'teams': [dict(t) for t in teams],
        'page': page,
        'total_pages': total_pages,
        'total': total,
        'course_total': course_total,
        'sort': sort_col,
        'order': order
    })


@app.route('/api/add_student', methods=['POST'])
@instructor_login_required
def add_student():
    slug = session['slug']
    if is_demo_instance_slug(slug):
        return jsonify({'error': 'The demo roster is fixed at two students'}), 403
    data = request.get_json(silent=True) or {}
    student_id = data.get('student_id', '').strip()
    name = data.get('name', '').strip()
    pin = data.get('pin', '').strip()
    if not student_id or not pin:
        return jsonify({'error': 'Student ID and PIN are required'}), 400
    if len(student_id) > 100 or len(name) > 200:
        return jsonify({'error': 'Student ID or name is too long'}), 400
    if not re.fullmatch(r'[0-9]{4}', pin):
        return jsonify({'error': 'PIN must be exactly 4 digits'}), 400
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        course = db.execute('SELECT id FROM courses LIMIT 1').fetchone()
        state = db.execute(
            '''SELECT phase, session_key, roster_version
               FROM course_state WHERE course_id = ?''',
            [course['id']]
        ).fetchone()
        guard = _expected_roster_state_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        if not state or state['phase'] != 'setup':
            db.rollback()
            return jsonify({'error': 'Students can only be changed during setup'}), 409
        existing_rows = db.execute(
            '''SELECT id, student_id, name, pin, is_active FROM students
               WHERE course_id = ?
               ORDER BY is_active DESC, id''',
            [course['id']]
        ).fetchall()
        existing_by_sid = _roster_rows_by_casefolded_id(existing_rows)
        if existing_by_sid is None:
            db.rollback()
            return jsonify({'error': _ROSTER_CASE_COLLISION_MESSAGE}), 409
        student_id_key = student_id.casefold()
        existing = existing_by_sid.get(student_id_key)
        structural_change = not existing or not existing['is_active']
        if structural_change:
            freeze_guard = _session_roster_mutation_guard(
                db, course['id'], state
            )
            if freeze_guard:
                db.rollback()
                return jsonify({'error': freeze_guard[0]}), freeze_guard[1]

        normalized_name = (
            existing['name'] if existing and not name else (name or None)
        )
        pin_changed = bool(existing and existing['pin'] != pin)
        changed = structural_change or (
            (existing['name'] or None) != normalized_name
            or existing['pin'] != pin
        )
        if existing and changed:
            if existing['is_active']:
                db.execute(
                    'UPDATE students SET name = ?, pin = ? WHERE id = ?',
                    [normalized_name, pin, existing['id']]
                )
            else:
                db.execute(
                    '''UPDATE students
                       SET name = ?,
                           pin = ?,
                           is_active = 1,
                           last_team_id = COALESCE(last_team_id, team_id),
                           team_id = NULL,
                           last_active_at = NULL
                       WHERE id = ?''',
                    [normalized_name, pin, existing['id']]
                )
            result = {
                'success': True,
                'updated': bool(existing['is_active']),
                'reactivated': not bool(existing['is_active']),
            }
        elif existing:
            result = {'success': True, 'updated': True}
        else:
            db.execute(
                '''INSERT INTO students (course_id, student_id, name, pin)
                   VALUES (?, ?, ?, ?)''',
                [course['id'], student_id, normalized_name, pin]
            )
            result = {'success': True, 'added': True}
        result['changed'] = changed
        result['pin_changed'] = pin_changed
        if changed:
            roster_version = _bump_roster_version(
                slug, course['id'], db=db
            )
        else:
            roster_version = _current_roster_version(db, course['id'])
        result['roster_version'] = roster_version
        db.commit()
        return jsonify(result)
    except Exception:
        db.rollback()
        raise


@app.route('/api/assign_student', methods=['POST'])
@instructor_login_required
def assign_student():
    slug = session['slug']
    data = request.get_json(silent=True) or {}
    student_id = data.get('student_id')  # DB row id
    team_id = data.get('team_id')  # None or '' to unassign
    if not student_id:
        return jsonify({'error': 'Student ID required'}), 400
    try:
        student_id = int(student_id)
        team_id = int(team_id) if team_id not in (None, '') else None
    except (TypeError, ValueError):
        return jsonify({'error': 'Invalid student or team'}), 400

    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        course = db.execute('SELECT id FROM courses LIMIT 1').fetchone()
        state = db.execute(
            '''SELECT phase, session_key, roster_version, max_teams,
                      max_members_per_team FROM course_state
               WHERE course_id = ?''',
            [course['id']]
        ).fetchone()
        guard = _expected_roster_state_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        if not state or state['phase'] != 'setup':
            db.rollback()
            return jsonify({'error': 'Teams can only be changed during setup'}), 409
        student = db.execute(
            '''SELECT id, team_id FROM students
               WHERE id = ? AND course_id = ? AND is_active = 1''',
            [student_id, course['id']]
        ).fetchone()
        if not student:
            db.rollback()
            return jsonify({'error': 'Student not found'}), 404
        if student['team_id'] == team_id:
            roster_version = _current_roster_version(db, course['id'])
            db.commit()
            return jsonify({'success': True, 'roster_version': roster_version})

        if team_id is not None:
            visible = db.execute(
                'SELECT id FROM teams WHERE course_id = ? ORDER BY id LIMIT ?',
                [course['id'], state['max_teams'] or 6]
            ).fetchall()
            if team_id not in {team['id'] for team in visible}:
                db.rollback()
                return jsonify({'error': 'Team is not available'}), 404
            member_count = db.execute(
                '''SELECT COUNT(*) AS c FROM students
                   WHERE course_id = ? AND team_id = ? AND is_active = 1''',
                [course['id'], team_id]
            ).fetchone()['c']
            if member_count >= (state['max_members_per_team'] or 10):
                db.rollback()
                return jsonify({'error': 'That team is full'}), 409

        freeze_guard = _session_roster_mutation_guard(
            db, course['id'], state
        )
        if freeze_guard:
            db.rollback()
            return jsonify({'error': freeze_guard[0]}), freeze_guard[1]
        db.execute(
            '''UPDATE students
               SET last_team_id = CASE
                       WHEN ? IS NULL THEN COALESCE(last_team_id, team_id)
                       ELSE ?
                   END,
                   team_id = ?,
                   last_team_joined_at = CASE
                       WHEN ? IS NULL THEN last_team_joined_at
                       ELSE CURRENT_TIMESTAMP
                   END
               WHERE id = ?''',
            [team_id, team_id, team_id, team_id, student_id]
        )
        roster_version = _bump_roster_version(slug, course['id'], db=db)
        db.commit()
        return jsonify({'success': True, 'roster_version': roster_version})
    except Exception:
        db.rollback()
        raise


@app.route('/api/remove_student/<int:student_db_id>', methods=['DELETE'])
@instructor_login_required
def remove_student(student_db_id):
    slug = session['slug']
    if is_demo_instance_slug(slug):
        return jsonify({'error': 'The demo roster is fixed at two students'}), 403
    data = request.get_json(silent=True) or {}
    ensure_schema(slug)
    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        course = db.execute('SELECT id FROM courses LIMIT 1').fetchone()
        state = db.execute(
            '''SELECT phase, session_key, roster_version
               FROM course_state WHERE course_id = ?''',
            [course['id']]
        ).fetchone()
        guard = _expected_roster_state_guard(data, state)
        if guard:
            db.rollback()
            return jsonify({'error': guard[0]}), guard[1]
        if not state or state['phase'] != 'setup':
            db.rollback()
            return jsonify({'error': 'Students can only be changed during setup'}), 409
        exists = db.execute(
            '''SELECT id FROM students
               WHERE id = ? AND course_id = ? AND is_active = 1''',
            [student_db_id, course['id']]
        ).fetchone()
        if not exists:
            db.rollback()
            return jsonify({'error': 'Student not found'}), 404
        freeze_guard = _session_roster_mutation_guard(
            db, course['id'], state
        )
        if freeze_guard:
            db.rollback()
            return jsonify({'error': freeze_guard[0]}), freeze_guard[1]
        _archive_students(
            slug, [student_db_id], bump_roster=True, db=db, commit=False
        )
        roster_version = _current_roster_version(db, course['id'])
        db.commit()
        return jsonify({'success': True, 'roster_version': roster_version})
    except Exception:
        db.rollback()
        raise


@app.route('/api/reset_data', methods=['POST'])
@instructor_login_required
def reset_data():
    slug = session['slug']
    if is_demo_instance_slug(slug):
        return jsonify({'error': 'Use Reset in the private demo banner'}), 403
    data = request.get_json(silent=True) or {}
    if data.get('confirm_slug') != slug:
        return jsonify({'error': 'Type the course slug to confirm reset'}), 400
    ensure_schema(slug)
    initial_state = query_db(slug, 'SELECT * FROM course_state LIMIT 1', one=True)
    guard = _expected_state_guard(data, initial_state)
    if guard:
        return jsonify({'error': guard[0]}), guard[1]
    if initial_state['phase'] not in ('setup', 'ended'):
        return jsonify({'error': 'Data can only be reset during Setup or after End Session'}), 409

    # Snapshot the database *before* acquiring the write lock so the backup's
    # own SQLite connection doesn't compete with the held BEGIN IMMEDIATE.
    try:
        backup_name = _create_reset_backup(slug, prune=False)
    except Exception:
        return jsonify({'error': 'Could not create reset backup; try again'}), 500

    db = get_db(slug)
    db.execute('BEGIN IMMEDIATE')
    try:
        course = db.execute('SELECT id FROM courses LIMIT 1').fetchone()
        course_id = course['id']
        state = db.execute(
            'SELECT * FROM course_state WHERE course_id = ?', [course_id]
        ).fetchone()
        versions_changed = (
            not state
            or (state['state_version'] or 0)
            != (initial_state['state_version'] or 0)
            or (state['roster_version'] or 0)
            != (initial_state['roster_version'] or 0)
        )
        if versions_changed:
            db.rollback()
            _discard_reset_backup(slug, backup_name)
            return jsonify({
                'error': (
                    'Course data changed while the backup was being created. '
                    'Review the current state and try reset again.'
                )
            }), 409
        guard = _expected_state_guard(data, state)
        if guard:
            db.rollback()
            _discard_reset_backup(slug, backup_name)
            return jsonify({'error': guard[0]}), guard[1]
        if state['phase'] not in ('setup', 'ended'):
            db.rollback()
            _discard_reset_backup(slug, backup_name)
            return jsonify({'error': 'Data can only be reset during Setup or after End Session'}), 409
        optional_tables = {
            row['name'] for row in db.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table' "
                "AND name IN ('peer_reviews', 'discussion_responses', "
                "'discussion_selections')"
            ).fetchall()
        }
        if 'peer_reviews' in optional_tables:
            db.execute(
                'DELETE FROM peer_reviews WHERE course_id = ?', [course_id]
            )
        db.execute('DELETE FROM teammate_thumbs WHERE course_id = ?', [course_id])
        db.execute('DELETE FROM presentation_ratings WHERE course_id = ?', [course_id])
        db.execute('DELETE FROM hidden_discussion_questions WHERE course_id = ?', [course_id])
        db.execute('DELETE FROM challenge_hands WHERE course_id = ?', [course_id])
        db.execute('DELETE FROM challenge_rounds WHERE course_id = ?', [course_id])
        db.execute('DELETE FROM challenge_ratings WHERE course_id = ?', [course_id])
        db.execute(
            'DELETE FROM presentation_participants WHERE course_id = ?', [course_id])
        if 'discussion_responses' in optional_tables:
            db.execute('DELETE FROM discussion_responses WHERE course_id = ?', [course_id])
        if 'discussion_selections' in optional_tables:
            db.execute('DELETE FROM discussion_selections WHERE course_id = ?', [course_id])
        db.execute(
            '''UPDATE students
               SET team_id = NULL, last_login_at = NULL, last_active_at = NULL,
                   last_team_joined_at = NULL, last_team_id = NULL
               WHERE course_id = ?''',
            [course_id]
        )
        db.execute(
            '''UPDATE course_state SET
               phase = 'setup',
               active_team_id = NULL,
               active_question_id = NULL,
               current_question = NULL,
               current_discussion_key = NULL,
               current_discussion_source_key = NULL,
               current_discussion_title = NULL,
               current_discussion_content = NULL,
               presentation_started_at = NULL,
               presentation_created_at = NULL,
               presentation_time_cap = 300,
               presentation_remaining = NULL,
               poll_active = 0,
               poll_question_key = NULL,
               poll_started_at = NULL,
               poll_closed_at = NULL,
               challenge_ratings_closed_at = NULL,
               presentation_history = '[]',
               active_challenges_json = '[]',
               teams_locked = 0,
               session_started_at = NULL,
               session_key = COALESCE(session_key, 0) + 1,
               roster_version = COALESCE(roster_version, 0) + 1
               WHERE course_id = ?''',
            [course_id]
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    _prune_reset_backups(slug)
    return jsonify({'success': True, 'backup': backup_name})


def _exported_at_utc():
    """Return one machine-readable UTC timestamp for an export artifact."""
    return datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')


def _export_data_version(value):
    """Use public version notation while preserving malformed legacy values."""
    try:
        return public_version(value)
    except (TypeError, ValueError):
        return '' if value is None else str(value)


def _legacy_data_reason(week_num, data_version):
    """Classify why a durable feedback row belongs in the legacy export."""
    if _positive_lecture_week(week_num) is None:
        return 'unknown_week'
    try:
        parse_version(data_version)
    except (TypeError, ValueError):
        return 'malformed_data_version'
    return 'incompatible_data_version'

@app.route('/export/<slug>/legacy-feedback.csv')
@instructor_login_required
def export_legacy_feedback(slug):
    """Download unknown-week or incompatible durable feedback as CSV."""
    if session.get('slug') != slug:
        return 'Forbidden', 403

    db = None
    snapshot_open = False
    try:
        ensure_schema(slug)
        db = get_db(slug)
        db.execute('BEGIN')
        snapshot_open = True
        # This phase read establishes the snapshot used by every feedback
        # query, so a concurrent reset cannot mix pre- and post-reset rows.
        course = db.execute(
            '''SELECT c.id, c.code, s.phase, s.presentation_history
               FROM courses c
               LEFT JOIN course_state s ON s.course_id = c.id
               WHERE c.slug = ?''',
            [slug],
        ).fetchone()
        if not course:
            db.rollback()
            snapshot_open = False
            return 'Course not found', 404
        course = _course_with_config_metadata(
            course, _course_availability(slug).get('config')
        )
        if course['phase'] not in ('setup', 'ended'):
            db.rollback()
            snapshot_open = False
            return (
                'Data can only be exported during Setup or after End Session.',
                409,
            )
        legacy_where = (
            'p.course_id = ? AND (p.week_num IS NULL OR '
            "typeof(p.week_num) != 'integer' OR p.week_num <= 0 OR "
            'popping_version_compatible(p.data_version, ?) = 0)'
        )
        legacy_row_count = sum(
            db.execute(
                f'SELECT COUNT(*) FROM {table} p WHERE {legacy_where}',
                [course['id'], SCHEMA_VERSION],
            ).fetchone()[0]
            for table in (
                'teammate_thumbs', 'presentation_ratings',
                'presentation_participants', 'challenge_rounds',
                'challenge_ratings',
            )
        )
        raw_history = course['presentation_history']
        if (legacy_row_count > MAX_EXPORT_ROWS
                or len(str(raw_history or '').encode('utf-8')) > MAX_EXPORT_BYTES):
            db.rollback()
            snapshot_open = False
            return 'The legacy export exceeds the safety limit.', 413
        thumb_rows = db.execute(
            f'''SELECT g.student_id AS grader_id,
                       CASE
                         WHEN NULLIF(TRIM(g.display_name), '') IS NOT NULL
                           THEN g.display_name
                         WHEN NULLIF(TRIM(g.name), '') IS NOT NULL THEN g.name
                         ELSE g.student_id END AS grader_name,
                       r.student_id AS recipient_id,
                       CASE
                         WHEN NULLIF(TRIM(r.display_name), '') IS NOT NULL
                           THEN r.display_name
                         WHEN NULLIF(TRIM(r.name), '') IS NOT NULL THEN r.name
                         ELSE r.student_id END AS recipient_name,
                       p.grader_team_id, p.grader_team_name,
                       p.recipient_team_id, p.recipient_team_name,
                       p.session_key, p.week_num, p.question_key,
                       p.source_question_key, p.question_title, p.created_at,
                       p.data_version
                FROM teammate_thumbs p
                JOIN students g ON g.id = p.grader_id
                JOIN students r ON r.id = p.recipient_id
                WHERE {legacy_where}
                ORDER BY p.created_at, p.id''',
            [course['id'], SCHEMA_VERSION],
        ).fetchall()
        rating_rows = db.execute(
            f'''SELECT s.student_id AS grader_id,
                       CASE
                         WHEN NULLIF(TRIM(s.display_name), '') IS NOT NULL
                           THEN s.display_name
                         WHEN NULLIF(TRIM(s.name), '') IS NOT NULL THEN s.name
                         ELSE s.student_id END AS grader_name,
                       p.rater_team_id AS grader_team_id,
                       p.rater_team_name AS grader_team_name,
                       p.presenting_team_id, p.presenting_team_name,
                       p.session_key, p.week_num, p.question_key,
                       p.question_title, p.q1_developed, p.q2_easy,
                       p.created_at, p.data_version
                FROM presentation_ratings p
                JOIN students s ON s.id = p.student_id
                WHERE {legacy_where}
                ORDER BY p.created_at, p.id''',
            [course['id'], SCHEMA_VERSION],
        ).fetchall()
        participant_rows = db.execute(
            f'''SELECT p.session_key, p.week_num, p.presentation_key,
                       p.student_identifier,
                       CASE WHEN NULLIF(TRIM(p.student_name), '') IS NOT NULL
                              THEN p.student_name
                            ELSE p.student_identifier END AS student_name,
                       p.team_id, p.team_name, p.created_at, p.data_version
                FROM presentation_participants p
                WHERE {legacy_where}
                ORDER BY p.created_at, p.id''',
            [course['id'], SCHEMA_VERSION],
        ).fetchall()
        challenge_round_rows = db.execute(
            f'''SELECT p.session_key, p.week_num, p.challenge_key,
                       p.presentation_key, p.challenge_num,
                       challenger.student_id AS challenger_id,
                       CASE WHEN NULLIF(TRIM(p.challenger_name), '') IS NOT NULL
                              THEN p.challenger_name
                            ELSE challenger.student_id END AS challenger_name,
                       p.challenger_team_id, p.challenger_team_name,
                       p.presenting_team_id, p.presenting_team_name,
                       p.question_title, p.created_at, p.data_version
                FROM challenge_rounds p
                LEFT JOIN students challenger ON challenger.id = p.challenger_id
                WHERE {legacy_where}
                ORDER BY p.created_at, p.id''',
            [course['id'], SCHEMA_VERSION],
        ).fetchall()
        challenge_rating_rows = db.execute(
            f'''SELECT p.session_key, p.week_num, p.challenge_key,
                       p.presentation_key, round.challenge_num,
                       challenger.student_id AS challenger_id,
                       CASE
                         WHEN NULLIF(TRIM(round.challenger_name), '') IS NOT NULL
                           THEN round.challenger_name
                         WHEN NULLIF(TRIM(p.challenger_name), '') IS NOT NULL
                           THEN p.challenger_name
                         ELSE challenger.student_id END AS challenger_name,
                       COALESCE(round.challenger_team_id,
                                p.challenger_team_id) AS challenger_team_id,
                       COALESCE(round.challenger_team_name,
                                p.challenger_team_name) AS challenger_team_name,
                       round.presenting_team_id, round.presenting_team_name,
                       round.question_title,
                       rater.student_id AS rater_id,
                       CASE
                         WHEN NULLIF(TRIM(p.rater_name), '') IS NOT NULL
                           THEN p.rater_name
                         WHEN NULLIF(TRIM(rater.display_name), '') IS NOT NULL
                           THEN rater.display_name
                         WHEN NULLIF(TRIM(rater.name), '') IS NOT NULL
                           THEN rater.name
                         ELSE rater.student_id END AS rater_name,
                       p.rater_team_id, p.rater_team_name, p.score,
                       p.created_at, p.data_version
                FROM challenge_ratings p
                JOIN students rater ON rater.id = p.rater_id
                LEFT JOIN students challenger ON challenger.id = p.challenger_id
                LEFT JOIN challenge_rounds round
                  ON round.course_id = p.course_id
                 AND round.challenge_key = p.challenge_key
                WHERE {legacy_where}
                ORDER BY p.created_at, p.id''',
            [course['id'], SCHEMA_VERSION],
        ).fetchall()
        question_weeks = {
            row['id']: (
                1 if row['week_num'] is None else row['week_num']
            )
            for row in db.execute(
                'SELECT id, week_num FROM questions WHERE course_id = ?',
                [course['id']],
            ).fetchall()
        }
        rating_weeks = _compatible_rating_weeks(db, course['id'])

        history_rows = []
        if raw_history and str(raw_history).strip() not in ('', '[]'):
            malformed_container = False
            try:
                history_items = json.loads(raw_history)
            except (TypeError, ValueError):
                history_items = [raw_history]
                malformed_container = True
            if not isinstance(history_items, list):
                history_items = [history_items]
                malformed_container = True
            for item in history_items:
                if malformed_container or not isinstance(item, dict):
                    history_rows.append({
                        'item': {},
                        'week_num': None,
                        'data_version': '',
                        'history_json': (
                            str(raw_history) if malformed_container
                            else json.dumps(item, ensure_ascii=False)
                        ),
                    })
                    continue
                data_version = _history_data_version(item)
                week_num = _resolve_history_week(
                    item,
                    question_weeks=question_weeks,
                    rating_weeks=rating_weeks,
                )
                if (week_num is not None
                        and _data_version_is_compatible(data_version)):
                    continue
                history_rows.append({
                    'item': item,
                    'week_num': week_num,
                    'data_version': data_version,
                    'history_json': json.dumps(
                        item, ensure_ascii=False, sort_keys=True
                    ),
                })
        if legacy_row_count + len(history_rows) > MAX_EXPORT_ROWS:
            db.rollback()
            snapshot_open = False
            return 'The legacy export exceeds the safety limit.', 413
        db.commit()
        snapshot_open = False
    except Exception:
        if db is not None and snapshot_open:
            db.rollback()
        raise

    exported_at = _exported_at_utc()
    website_version = public_version(APP_VERSION)
    schema_version = public_version(SCHEMA_VERSION)
    export_format_version = public_version(EXPORT_FORMAT_VERSION)
    output = io.StringIO(newline='')
    writer = csv.writer(output)

    def write_row(values):
        writer.writerow(_spreadsheet_safe_row(values))

    def metadata_values(row):
        return [
            _export_data_version(row['data_version']),
            _legacy_data_reason(row['week_num'], row['data_version']),
            website_version,
            schema_version,
            export_format_version,
            exported_at,
        ]

    legacy_headers = [
        'record_type',
        'grader_id', 'grader_name', 'grader_team_id', 'grader_team',
        'recipient_id', 'recipient_name', 'recipient_team_id',
        'recipient_team', 'presenting_team_id', 'presenting_team',
        'q1_developed', 'q2_easy', 'session_key', 'lecture_week',
        'question_key', 'source_question_key', 'question_title', 'time',
        'challenge_key', 'presentation_key', 'challenge_number',
        'challenger_id', 'challenger_name', 'challenger_team_id',
        'challenger_team', 'rater_id', 'rater_name', 'rater_team_id',
        'rater_team', 'score_1to5', 'participant_id', 'participant_name',
        'participant_team_id', 'participant_team',
        'data_version', 'legacy_reason',
        'exported_by_website_version', 'database_schema_version',
        'export_format_version', 'exported_at_utc', 'history_json',
    ]
    write_row(legacy_headers)
    for row in thumb_rows:
        write_row([
            'teammate_thumb',
            row['grader_id'], row['grader_name'], row['grader_team_id'],
            row['grader_team_name'], row['recipient_id'],
            row['recipient_name'], row['recipient_team_id'],
            row['recipient_team_name'], '', '', '', '', row['session_key'],
            row['week_num'] if row['week_num'] is not None else 'unknown',
            row['question_key'], row['source_question_key'],
            row['question_title'], row['created_at'],
            *([''] * 12), *([''] * 4), *metadata_values(row), '',
        ])
    for row in rating_rows:
        write_row([
            'presentation_rating',
            row['grader_id'], row['grader_name'], row['grader_team_id'],
            row['grader_team_name'], '', '', '', '',
            row['presenting_team_id'], row['presenting_team_name'],
            row['q1_developed'], row['q2_easy'], row['session_key'],
            row['week_num'] if row['week_num'] is not None else 'unknown',
            row['question_key'], '', row['question_title'], row['created_at'],
            *([''] * 12), *([''] * 4), *metadata_values(row), '',
        ])
    for row in participant_rows:
        values = {
            'record_type': 'presentation_participant',
            'presenting_team_id': row['team_id'],
            'presenting_team': row['team_name'],
            'session_key': row['session_key'],
            'lecture_week': (
                row['week_num']
                if row['week_num'] is not None else 'unknown'
            ),
            'question_key': row['presentation_key'],
            'time': row['created_at'],
            'presentation_key': row['presentation_key'],
            'participant_id': row['student_identifier'],
            'participant_name': row['student_name'],
            'participant_team_id': row['team_id'],
            'participant_team': row['team_name'],
            'data_version': _export_data_version(row['data_version']),
            'legacy_reason': _legacy_data_reason(
                row['week_num'], row['data_version']
            ),
            'exported_by_website_version': website_version,
            'database_schema_version': schema_version,
            'export_format_version': export_format_version,
            'exported_at_utc': exported_at,
        }
        write_row([values.get(header, '') for header in legacy_headers])
    for row in challenge_round_rows:
        write_row([
            'challenge_round', '', '', '', '', '', '', '', '',
            row['presenting_team_id'], row['presenting_team_name'], '', '',
            row['session_key'],
            row['week_num'] if row['week_num'] is not None else 'unknown',
            row['presentation_key'], '', row['question_title'],
            row['created_at'], row['challenge_key'], row['presentation_key'],
            row['challenge_num'], row['challenger_id'], row['challenger_name'],
            row['challenger_team_id'], row['challenger_team_name'],
            '', '', '', '', '', *([''] * 4),
            *metadata_values(row), '',
        ])
    for row in challenge_rating_rows:
        write_row([
            'challenge_rating', row['rater_id'], row['rater_name'],
            row['rater_team_id'], row['rater_team_name'], '', '', '', '',
            row['presenting_team_id'], row['presenting_team_name'], '', '',
            row['session_key'],
            row['week_num'] if row['week_num'] is not None else 'unknown',
            row['presentation_key'], '', row['question_title'],
            row['created_at'], row['challenge_key'], row['presentation_key'],
            row['challenge_num'], row['challenger_id'], row['challenger_name'],
            row['challenger_team_id'], row['challenger_team_name'],
            row['rater_id'], row['rater_name'], row['rater_team_id'],
            row['rater_team_name'], row['score'], *([''] * 4),
            *metadata_values(row), '',
        ])

    for history_row in history_rows:
        item = history_row['item']
        qkey = _history_presentation_key(item) if item else ''
        week_num = history_row['week_num']
        data_version = history_row['data_version']
        values = {
            'record_type': 'presentation_history',
            'presenting_team_id': item.get('team_id', ''),
            'presenting_team': item.get('team', ''),
            'session_key': item.get('session_key', ''),
            'lecture_week': (
                week_num if week_num is not None else 'unknown'
            ),
            'question_key': qkey,
            'question_title': item.get('title', ''),
            'time': item.get('ended_at') or item.get('started_at', ''),
            'presentation_key': qkey,
            'data_version': _export_data_version(data_version),
            'legacy_reason': _legacy_data_reason(week_num, data_version),
            'exported_by_website_version': website_version,
            'database_schema_version': schema_version,
            'export_format_version': export_format_version,
            'exported_at_utc': exported_at,
            'history_json': history_row['history_json'],
        }
        write_row([values.get(header, '') for header in legacy_headers])
    encoded_output = output.getvalue().encode('utf-8-sig')
    if len(encoded_output) > MAX_EXPORT_BYTES:
        return 'The legacy export exceeds the safety limit.', 413
    content = io.BytesIO(encoded_output)
    filename = (
        f"popping_{course['code'] or slug}_legacy_unknown_week_feedback.csv"
    )
    return send_file(
        content,
        mimetype='text/csv; charset=utf-8',
        as_attachment=True,
        download_name=filename,
    )

def _compatible_export_week_exists(db, course_id, week_num, state_row):
    '''Return whether a later week has current-series durable results.'''
    for table in (
            'teammate_thumbs', 'presentation_ratings',
            'presentation_participants', 'challenge_rounds',
            'challenge_ratings'):
        found = db.execute(
            f'''SELECT 1 FROM {table}
                WHERE course_id = ? AND week_num = ?
                  AND popping_version_compatible(data_version, ?) = 1
                LIMIT 1''',
            [course_id, week_num, SCHEMA_VERSION],
        ).fetchone()
        if found:
            return True

    raw_history = state_row['presentation_history'] if state_row else None
    if not raw_history:
        return False
    try:
        history = json.loads(raw_history)
    except (TypeError, ValueError):
        return False
    if not isinstance(history, list):
        return False
    question_weeks = {
        row['id']: (1 if row['week_num'] is None else row['week_num'])
        for row in db.execute(
            'SELECT id, week_num FROM questions WHERE course_id = ?',
            [course_id],
        ).fetchall()
    }
    rating_weeks = _compatible_rating_weeks(db, course_id)
    return any(
        _history_item_is_compatible(item)
        and _resolve_history_week(
            item, question_weeks=question_weeks,
            rating_weeks=rating_weeks,
        ) == week_num
        for item in history
    )


@app.route('/export/<slug>')
@instructor_login_required
def export_data(slug):
    if session.get('slug') != slug:
        flash('Unauthorized', 'error')
        return redirect(url_for('index'))
    if request.args.get('weeks', '').lower() == 'all':
        return ('Export one week at a time: use ?week=N for a specific '
                'lecture week.'), 400

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        import collections
    except ImportError:
        flash('Export library not available. Contact administrator.', 'error')
        return redirect(url_for('instructor_course', slug=slug))

    db = None
    snapshot_open = False
    try:
        ensure_schema(slug)
        db = get_db(slug)
        db.execute('BEGIN IMMEDIATE')
        snapshot_open = True
        course = query_db(slug, 'SELECT * FROM courses LIMIT 1', one=True)
        if not course:
            db.rollback()
            snapshot_open = False
            flash('Course not found.', 'error')
            return redirect(url_for('index'))
        course = _course_with_config_metadata(
            course, _course_availability(slug).get('config')
        )

        cid = course['id']
        state_row = query_db(
            slug,
            'SELECT * FROM course_state WHERE course_id = ?',
            [cid],
            one=True,
        )
        phase_error = _export_phase_error(db, cid)
        if phase_error:
            db.rollback()
            snapshot_open = False
            return phase_error, 409

        current_week = (
            state_row['discussion_week'] if state_row else None
        ) or 1
        max_teams = (state_row['max_teams'] if state_row else None) or 6

        # Optional ?week=N exports a previous lecture week with the exact
        # same workbook layout as the current week.
        requested_week = (request.args.get('week') or '').strip()
        if requested_week:
            try:
                requested_week_num = int(requested_week)
            except ValueError:
                db.rollback()
                snapshot_open = False
                return 'Invalid week parameter.', 400
            if requested_week_num < 1:
                db.rollback()
                snapshot_open = False
                return 'Week must be a positive integer.', 400
            if (requested_week_num > current_week
                    and not _compatible_export_week_exists(
                        db, cid, requested_week_num, state_row
                    )):
                db.rollback()
                snapshot_open = False
                return (
                    f'Week {requested_week_num} is later than the current week '
                    f'and has no compatible saved results.'
                ), 400
            current_week = requested_week_num

        export_weeks = [current_week]
        week_ph = ','.join('?' * len(export_weeks))

        asset_files = []
        asset_bytes = 0

        def add_asset(fpath, archive_name):
            nonlocal asset_bytes
            if not os.path.isfile(fpath):
                return
            remaining = MAX_EXPORT_BYTES - asset_bytes
            if remaining < 0:
                return
            try:
                with open(fpath, 'rb') as handle:
                    content = handle.read(remaining + 1)
            except FileNotFoundError:
                return
            asset_bytes += len(content)
            if len(content) > remaining:
                return
            asset_files.append((archive_name, content))

        class_dir = _course_class_dir(slug)
        for export_week in export_weeks:
            discussion_path = _resolve_week_question_path(slug, export_week)
            add_asset(
                discussion_path,
                f'questions/week-{export_week}-questions.md',
            )

        for export_week in export_weeks:
            appendix_path = os.path.join(
                config.DATA_DIR, slug, 'appendix',
                f'week-{export_week}-appendix.md',
            )
            if not os.path.isfile(appendix_path):
                appendix_path = os.path.join(
                    class_dir, f'week-{export_week}-appendix.md'
                )
            add_asset(
                appendix_path,
                f'appendix/week-{export_week}-appendix.md',
            )

        if asset_bytes > MAX_EXPORT_BYTES:
            db.rollback()
            snapshot_open = False
            flash('The export files exceed the 50 MB safety limit.', 'error')
            return redirect(url_for('instructor_course', slug=slug))

        row_counts = query_db(
            slug,
            f'''SELECT
                   (SELECT COUNT(*) FROM students WHERE course_id = ?) AS students,
                   (SELECT COUNT(*) FROM (
                        SELECT id FROM teams WHERE course_id = ?
                        ORDER BY id LIMIT ?
                    )) AS teams,
                   (SELECT COUNT(*) FROM teammate_thumbs
                    WHERE course_id = ? AND week_num IN ({week_ph})
                      AND popping_version_compatible(data_version, ?) = 1)
                       AS thumbs,
                   (SELECT COUNT(*) FROM presentation_ratings
                    WHERE course_id = ? AND week_num IN ({week_ph})
                      AND popping_version_compatible(data_version, ?) = 1)
                       AS ratings,
                   (SELECT COUNT(*) FROM presentation_participants
                    WHERE course_id = ? AND week_num IN ({week_ph})
                      AND popping_version_compatible(data_version, ?) = 1)
                       AS presentation_participants,
                   (SELECT COUNT(*) FROM challenge_rounds
                    WHERE course_id = ? AND week_num IN ({week_ph})
                      AND popping_version_compatible(data_version, ?) = 1)
                       AS challenge_rounds,
                   (SELECT COUNT(*) FROM challenge_ratings
                    WHERE course_id = ? AND week_num IN ({week_ph})
                      AND popping_version_compatible(data_version, ?) = 1)
                       AS challenge_ratings''',
            (
                [cid, cid, max_teams, cid] + export_weeks + [SCHEMA_VERSION]
                + [cid] + export_weeks + [SCHEMA_VERSION]
                + [cid] + export_weeks + [SCHEMA_VERSION]
                + [cid] + export_weeks + [SCHEMA_VERSION]
                + [cid] + export_weeks + [SCHEMA_VERSION]
            ),
            one=True
        )
        if sum(row_counts) + row_counts['students'] > MAX_EXPORT_ROWS:
            db.rollback()
            snapshot_open = False
            flash('The export is too large. Please contact the administrator.', 'error')
            return redirect(url_for('instructor_course', slug=slug))

        # Gather one compatibility-scoped snapshot for every derived and raw
        # workbook view.
        students = query_db(
            slug,
            '''SELECT s.*, t.name as team_name
               FROM students s LEFT JOIN teams t ON t.id = CASE
                   WHEN s.is_active = 1 THEN s.team_id
                   ELSE COALESCE(s.team_id, s.last_team_id) END
               WHERE s.course_id = ?
               ORDER BY COALESCE(NULLIF(TRIM(s.display_name), ''),
                                 NULLIF(TRIM(s.name), ''), s.student_id)''',
            [cid],
        )
        participation_counts = _participation_counts_by_student(
            db, cid, [student['id'] for student in students]
        )
        teams = query_db(
            slug,
            '''SELECT t.*,
                    (SELECT COUNT(*) FROM students s
                     WHERE s.team_id = t.id AND s.is_active = 1) as member_count
               FROM teams t WHERE t.course_id = ? ORDER BY t.id LIMIT ?''',
            [cid, max_teams],
        )
        peer_reviews = query_db(
            slug,
            f'''SELECT p.grader_id, p.recipient_id,
                       g.student_id as grader_sid,
                       CASE
                         WHEN NULLIF(TRIM(g.display_name), '') IS NOT NULL
                           THEN g.display_name
                         WHEN NULLIF(TRIM(g.name), '') IS NOT NULL THEN g.name
                         ELSE g.student_id END AS grader_name,
                       r.student_id as recipient_sid,
                       CASE
                         WHEN NULLIF(TRIM(r.display_name), '') IS NOT NULL
                           THEN r.display_name
                         WHEN NULLIF(TRIM(r.name), '') IS NOT NULL THEN r.name
                         ELSE r.student_id END AS recipient_name,
                      p.grader_team_id, p.grader_team_name,
                      p.recipient_team_id, p.recipient_team_name,
                      p.session_key, p.week_num, p.question_key,
                      p.source_question_key, p.question_title,
                      'overall' AS criterion, 1 AS score, p.created_at,
                      p.data_version
               FROM teammate_thumbs p
               JOIN students g ON p.grader_id = g.id
               JOIN students r ON p.recipient_id = r.id
               WHERE p.course_id = ? AND p.week_num IN ({week_ph})
                 AND popping_version_compatible(p.data_version, ?) = 1
               ORDER BY p.created_at''',
            [cid] + export_weeks + [SCHEMA_VERSION],
        )
        ratings = query_db(
            slug,
            f'''SELECT pr.question_key, pr.session_key, pr.week_num,
                      pr.presenting_team_id, pr.presenting_team_name,
                      pr.question_id, pr.question_title,
                      pr.rater_team_id, pr.rater_team_name,
                      pr.student_id as rater_db_id,
                       s.student_id as rater_sid,
                       CASE
                         WHEN NULLIF(TRIM(s.display_name), '') IS NOT NULL
                           THEN s.display_name
                         WHEN NULLIF(TRIM(s.name), '') IS NOT NULL THEN s.name
                         ELSE s.student_id END AS rater_name,
                      pr.q1_developed, pr.q2_easy, pr.created_at,
                      pr.data_version
               FROM presentation_ratings pr
               JOIN students s ON pr.student_id = s.id
               WHERE pr.course_id = ? AND pr.week_num IN ({week_ph})
                 AND popping_version_compatible(pr.data_version, ?) = 1
               ORDER BY pr.question_key, pr.created_at''',
            [cid] + export_weeks + [SCHEMA_VERSION],
        )
        presentation_participants = query_db(
            slug,
            f'''SELECT participant.session_key, participant.week_num,
                      participant.presentation_key,
                      participant.student_identifier,
                      CASE WHEN NULLIF(TRIM(participant.student_name), '') IS NOT NULL
                             THEN participant.student_name
                           ELSE participant.student_identifier END AS student_name,
                      participant.team_id, participant.team_name,
                      participant.created_at, participant.data_version
               FROM presentation_participants participant
               WHERE participant.course_id = ?
                 AND typeof(participant.week_num) = 'integer'
                 AND participant.week_num > 0
                 AND participant.week_num IN ({week_ph})
                 AND popping_version_compatible(
                         participant.data_version, ?) = 1
               ORDER BY participant.session_key,
                        participant.presentation_key, participant.id''',
            [cid] + export_weeks + [SCHEMA_VERSION],
        )
        challenge_rounds = query_db(
            slug,
            f'''SELECT ch.session_key, ch.week_num, ch.challenge_key,
                      ch.presentation_key, ch.challenge_num,
                      challenger.student_id AS challenger_sid,
                      CASE WHEN NULLIF(TRIM(ch.challenger_name), '') IS NOT NULL
                             THEN ch.challenger_name
                           ELSE challenger.student_id END AS challenger_name,
                      ch.challenger_team_id, ch.challenger_team_name,
                      ch.presenting_team_id, ch.presenting_team_name,
                      ch.question_id, ch.question_title,
                      COUNT(rating.id) AS rating_count,
                      ROUND(AVG(rating.score), 2) AS average_score,
                      ch.created_at, ch.data_version
               FROM challenge_rounds ch
               LEFT JOIN students challenger
                 ON challenger.id = ch.challenger_id
               LEFT JOIN challenge_ratings rating
                 ON rating.course_id = ch.course_id
                AND rating.challenge_key = ch.challenge_key
                AND rating.session_key = ch.session_key
                AND rating.presentation_key = ch.presentation_key
                AND rating.challenger_id = ch.challenger_id
                AND rating.week_num = ch.week_num
                AND typeof(rating.week_num) = 'integer'
                AND rating.week_num > 0
                AND popping_version_compatible(
                        rating.data_version, ?) = 1
               WHERE ch.course_id = ? AND ch.week_num IN ({week_ph})
                 AND typeof(ch.week_num) = 'integer'
                 AND ch.week_num > 0
                 AND popping_version_compatible(ch.data_version, ?) = 1
               GROUP BY ch.id
               ORDER BY ch.session_key, ch.presentation_key, ch.challenge_num''',
            [SCHEMA_VERSION, cid] + export_weeks + [SCHEMA_VERSION],
        )
        challenge_ratings = query_db(
            slug,
            f'''SELECT cr.challenge_key, cr.presentation_key, cr.session_key,
                      cr.week_num, ch.challenge_num,
                      challenger.student_id AS challenger_sid,
                      CASE
                        WHEN NULLIF(TRIM(ch.challenger_name), '') IS NOT NULL
                          THEN ch.challenger_name
                        WHEN NULLIF(TRIM(cr.challenger_name), '') IS NOT NULL
                          THEN cr.challenger_name
                        ELSE challenger.student_id END AS challenger_name,
                      COALESCE(ch.challenger_team_id, cr.challenger_team_id)
                          AS challenger_team_id,
                      COALESCE(ch.challenger_team_name,
                               cr.challenger_team_name)
                          AS challenger_team_name,
                      ch.presenting_team_id, ch.presenting_team_name,
                      ch.question_id, ch.question_title,
                      cr.rater_id, s.student_id AS rater_sid,
                      CASE
                        WHEN NULLIF(TRIM(cr.rater_name), '') IS NOT NULL
                          THEN cr.rater_name
                        WHEN NULLIF(TRIM(s.display_name), '') IS NOT NULL
                          THEN s.display_name
                        WHEN NULLIF(TRIM(s.name), '') IS NOT NULL THEN s.name
                        ELSE s.student_id END AS rater_name,
                      cr.rater_team_id, cr.rater_team_name,
                      cr.score, cr.created_at, cr.data_version
               FROM challenge_ratings cr
               JOIN students s ON cr.rater_id = s.id
               LEFT JOIN students challenger
                 ON challenger.id = cr.challenger_id
               LEFT JOIN challenge_rounds ch
                 ON ch.course_id = cr.course_id
                AND ch.challenge_key = cr.challenge_key
                AND ch.session_key = cr.session_key
                AND ch.presentation_key = cr.presentation_key
                AND ch.challenger_id = cr.challenger_id
                AND ch.week_num = cr.week_num
                AND typeof(ch.week_num) = 'integer'
                AND ch.week_num > 0
                AND popping_version_compatible(ch.data_version, ?) = 1
               WHERE cr.course_id = ? AND cr.week_num IN ({week_ph})
                 AND typeof(cr.week_num) = 'integer'
                 AND cr.week_num > 0
                 AND popping_version_compatible(cr.data_version, ?) = 1
               ORDER BY cr.challenge_key, cr.created_at''',
            [SCHEMA_VERSION, cid] + export_weeks + [SCHEMA_VERSION],
        )
        coursewide_participation_versions = query_db(
            slug,
            '''SELECT DISTINCT data_version
               FROM presentation_participants
               WHERE course_id = ?
                 AND typeof(week_num) = 'integer'
                 AND week_num > 0
                 AND popping_version_compatible(data_version, ?) = 1
               UNION
               SELECT DISTINCT data_version
               FROM challenge_rounds
               WHERE course_id = ?
                 AND typeof(week_num) = 'integer'
                 AND week_num > 0
                 AND popping_version_compatible(data_version, ?) = 1''',
            [cid, SCHEMA_VERSION, cid, SCHEMA_VERSION],
        )


        question_weeks = {
            row['id']: (
                1 if row['week_num'] is None else row['week_num']
            )
            for row in query_db(
                slug,
                'SELECT id, week_num FROM questions WHERE course_id = ?',
                [cid],
            )
        }

        # Map only compatible history entries in this week's export scope.
        key_to_team = {}
        history_data_versions = set()
        rating_weeks = _compatible_rating_weeks(db, cid)
        history_items = []
        if state_row and state_row['presentation_history']:
            try:
                history_items = json.loads(state_row['presentation_history'])
            except (TypeError, ValueError):
                history_items = []
            if not isinstance(history_items, list):
                history_items = []
        for h in history_items:
            if not _history_item_is_compatible(h):
                continue
            history_version = _history_data_version(h)
            qkey = _history_presentation_key(h)
            history_week = _resolve_history_week(
                h, question_weeks=question_weeks,
                rating_weeks=rating_weeks,
            )
            if history_week != current_week:
                continue
            key_to_team[qkey] = h.get('team', 'Unknown')
            history_data_versions.add(history_version)

        team_id_to_name = {t['id']: t['name'] for t in teams}
        included_data_versions = history_data_versions | {
            row['data_version']
            for rows in (
                peer_reviews, ratings, presentation_participants,
                challenge_rounds, challenge_ratings
            )
            for row in rows
        }
        included_data_versions.update(
            row['data_version'] for row in coursewide_participation_versions
        )
        included_data_versions = sorted(
            included_data_versions, key=parse_version
        )
        public_data_versions = [
            public_version(value) for value in included_data_versions
        ]
        exported_at = _exported_at_utc()
        data_compatibility = compatibility_label(SCHEMA_VERSION)
        # Release the snapshot lock after all database rows and files are captured.
        db.commit()
        snapshot_open = False

        # ── styles ──

        wb = Workbook()
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='13294B', end_color='13294B', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        bold_font = Font(bold=True)
        freeze_align = Alignment(vertical='top', wrap_text=True)

        def style_header(ws, headers):
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            ws.freeze_panes = 'A2'

        def set_col_widths(ws, headers, rows):
            """Compute column widths from actual content."""
            for col_idx, h in enumerate(headers, 1):
                max_len = len(str(h))
                for row in rows:
                    val = row[col_idx - 1] if col_idx <= len(row) else ''
                    if val is not None:
                        max_len = max(max_len, min(len(str(val)), 50))
                ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 3, 12)

        # ════════════════════════════════════════════════════════════════════
        # TAB 1: Summary — quick overview
        # ════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = 'Summary'

        info_rows = [
            ('Course', course['name']),
            ('Code', course['code'] or ''),
            ('Semester', course['semester'] or ''),
            ('Lecture Week', current_week),
            ('Export Scope', f'Week {current_week}'),
            ('Participation Roster Scope',
             'Course-wide compatible participation through export time'),
            ('Website Version', public_version(APP_VERSION)),
            ('Database Schema Version', public_version(SCHEMA_VERSION)),
            ('Export Format Version', public_version(EXPORT_FORMAT_VERSION)),
            ('Data Compatibility', data_compatibility),
            ('Data Versions Included', ', '.join(public_data_versions)),
            ('Exported At (UTC)', exported_at),
            ('Active Students', sum(1 for student in students if student['is_active'])),
            ('Archived Students', sum(1 for student in students if not student['is_active'])),
            ('Current Visible Teams', len(teams)),
            ('Week Peer Reviews (thumbs)', len(peer_reviews)),
            ('Week Presentation Ratings', len(ratings)),
            ('Week Presentation Participants',
             len(presentation_participants)),
            ('Week Challenge Rounds', len(challenge_rounds)),
            ('Week Challenge Ratings', len(challenge_ratings)),
        ]
        for r, (label, val) in enumerate(info_rows, 1):
            ws1.cell(row=r, column=1, value=label).font = bold_font
            ws1.cell(row=r, column=2, value=val)

        # Team leaderboard
        r = len(info_rows) + 2
        ws1.cell(row=r, column=1, value='Team Leaderboard').font = bold_font
        r += 1
        lb_headers = ['Rank', 'Team', 'Members', 'Presentations', 'Avg Developed (1-5)', 'Avg Easy (1-5)', 'Combined Avg']
        for col, h in enumerate(lb_headers, 1):
            cell = ws1.cell(row=r, column=col, value=h)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align
        r += 1

        # Build team scores — fix: use combined count for sort denominator
        team_scores = collections.defaultdict(lambda: {'dev': [], 'easy': []})
        visible_team_ids = {team['id'] for team in teams}
        visible_team_names = {team['name'] for team in teams}
        for rt in ratings:
            if rt['presenting_team_id'] in visible_team_ids:
                tname = team_id_to_name[rt['presenting_team_id']]
            else:
                tname = rt['presenting_team_name'] or key_to_team.get(
                    rt['question_key'], 'Unknown'
                )
                if tname not in visible_team_names:
                    continue
            if rt['q1_developed'] is not None:
                team_scores[tname]['dev'].append(rt['q1_developed'])
            if rt['q2_easy'] is not None:
                team_scores[tname]['easy'].append(rt['q2_easy'])

        def combined_fraction(sc):
            all_vals = sc['dev'] + sc['easy']
            return Fraction(sum(all_vals), len(all_vals)) if all_vals else Fraction(0, 1)

        def combined_avg(sc):
            return round(float(combined_fraction(sc)), 2)

        team_sorted = sorted(
            team_scores.items(),
            key=lambda item: (-combined_fraction(item[1]), item[0].casefold())
        )
        rank_by_team = {}
        prior_fraction = None
        for position, (tname, sc) in enumerate(team_sorted, 1):
            fraction = combined_fraction(sc)
            if prior_fraction is None or fraction != prior_fraction:
                rank = position
            rank_by_team[tname] = rank
            avg_d = round(sum(sc['dev']) / len(sc['dev']), 2) if sc['dev'] else ''
            avg_e = round(sum(sc['easy']) / len(sc['easy']), 2) if sc['easy'] else ''
            total = combined_avg(sc)
            member_cnt = next((t['member_count'] for t in teams if t['name'] == tname), '')
            pres_cnt = sum(1 for k, v in key_to_team.items() if v == tname)
            ws1.cell(row=r, column=1, value=rank)
            ws1.cell(row=r, column=2, value=tname)
            ws1.cell(row=r, column=3, value=member_cnt)
            ws1.cell(row=r, column=4, value=pres_cnt)
            ws1.cell(row=r, column=5, value=avg_d)
            ws1.cell(row=r, column=6, value=avg_e)
            ws1.cell(row=r, column=7, value=total if total else '')
            r += 1
            prior_fraction = fraction

        ws1.column_dimensions['A'].width = 28
        ws1.column_dimensions['B'].width = 22
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 16
        ws1.column_dimensions['E'].width = 20
        ws1.column_dimensions['F'].width = 18
        ws1.column_dimensions['G'].width = 14

        # ════════════════════════════════════════════════════════════════════
        # TAB 2: Students, one row per student, gradebook-ready
        # ════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet('Students')
        s_headers = [
            'student_id', 'roster_name', 'display_name', 'team', 'status',
            'course_presentation_team_turns', 'course_challenger_turns',
            'thumbs_given', 'thumbs_received',
            'presentation_ratings_given',
            'challenges_rated',
            'last_login', 'last_active',
        ]
        style_header(ws2, s_headers)

        # Pre-compute per-student counts (keyed by DB id)
        thumbs_given = collections.Counter(pr['grader_id'] for pr in peer_reviews)
        thumbs_recv = collections.Counter(pr['recipient_id'] for pr in peer_reviews)
        ratings_given = collections.Counter(rt['rater_db_id'] for rt in ratings)
        challenge_ratings_given = collections.Counter(cr['rater_id'] for cr in challenge_ratings)

        s_rows = []
        for stu in students:
            row = [
                stu['student_id'],
                stu['name'] or '',
                stu['display_name'] or '',
                stu['team_name'] or '',
                'active' if stu['is_active'] else 'archived',
                participation_counts.get(stu['id'], {}).get(
                    'presentation_count', 0),
                participation_counts.get(stu['id'], {}).get(
                    'challenger_count', 0),
                thumbs_given.get(stu['id'], 0),
                thumbs_recv.get(stu['id'], 0),
                ratings_given.get(stu['id'], 0),
                challenge_ratings_given.get(stu['id'], 0),
                stu['last_login_at'] or '',
                stu['last_active_at'] or '',
            ]
            s_rows.append(row)

        for i, row in enumerate(s_rows, 2):
            for col, val in enumerate(row, 1):
                ws2.cell(row=i, column=col, value=val)
        set_col_widths(ws2, s_headers, s_rows)

        # ════════════════════════════════════════════════════════════════════
        # TAB 3: Participation Roster, latest course-wide totals
        # ════════════════════════════════════════════════════════════════════
        ws_participation_roster = wb.create_sheet('Participation Roster')
        participation_roster_headers = [
            'student_id', 'roster_name', 'display_name', 'team', 'status',
            'course_presentation_team_turns', 'course_challenger_turns',
        ]
        style_header(
            ws_participation_roster, participation_roster_headers
        )
        participation_roster_rows = []
        for student in students:
            counts = participation_counts.get(student['id'], {})
            participation_roster_rows.append([
                student['student_id'],
                student['name'] or '',
                student['display_name'] or '',
                student['team_name'] or '',
                'active' if student['is_active'] else 'archived',
                counts.get('presentation_count', 0),
                counts.get('challenger_count', 0),
            ])
        for i, row in enumerate(participation_roster_rows, 2):
            for col, value in enumerate(row, 1):
                ws_participation_roster.cell(
                    row=i, column=col, value=value
                )
        set_col_widths(
            ws_participation_roster, participation_roster_headers,
            participation_roster_rows,
        )

        # ════════════════════════════════════════════════════════════════════
        # TAB 4: Teams, one row per team with aggregates
        # ════════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet('Teams')
        t_headers = ['team_id', 'team_name', 'rank', 'member_count', 'presentations',
                     'avg_developed', 'avg_easy', 'combined_avg']
        style_header(ws3, t_headers)

        t_rows = []
        for t in teams:
            tname = t['name']
            sc = team_scores.get(tname, {'dev': [], 'easy': []})
            avg_d = round(sum(sc['dev']) / len(sc['dev']), 2) if sc['dev'] else ''
            avg_e = round(sum(sc['easy']) / len(sc['easy']), 2) if sc['easy'] else ''
            total = combined_avg(sc)
            pres_cnt = sum(1 for k, v in key_to_team.items() if v == tname)
            t_rows.append([
                t['id'], tname, rank_by_team.get(tname, ''), t['member_count'], pres_cnt,
                avg_d, avg_e, total if total else '',
            ])
        for i, row in enumerate(t_rows, 2):
            for col, val in enumerate(row, 1):
                ws3.cell(row=i, column=col, value=val)
        set_col_widths(ws3, t_headers, t_rows)

        # ════════════════════════════════════════════════════════════════════
        # TAB 5: Peer Reviews, raw thumbs-up data
        # ════════════════════════════════════════════════════════════════════
        ws4 = wb.create_sheet('Peer Reviews')
        pr_headers = [
            'grader_id', 'grader_name', 'grader_team_id', 'grader_team',
            'recipient_id', 'recipient_name', 'recipient_team_id',
            'recipient_team', 'session_key', 'week', 'discussion_post_key',
            'source_question_key', 'question_title', 'criterion', 'score',
            'data_version', 'time',
        ]
        style_header(ws4, pr_headers)

        pr_rows = []
        for pr in peer_reviews:
            pr_rows.append([
                pr['grader_sid'], pr['grader_name'], pr['grader_team_id'],
                pr['grader_team_name'], pr['recipient_sid'],
                pr['recipient_name'], pr['recipient_team_id'],
                pr['recipient_team_name'], pr['session_key'],
                pr['week_num'], pr['question_key'], pr['source_question_key'],
                pr['question_title'],
                pr['criterion'], pr['score'],
                _export_data_version(pr['data_version']), pr['created_at'],
            ])
        for i, row in enumerate(pr_rows, 2):
            for col, val in enumerate(row, 1):
                ws4.cell(row=i, column=col, value=val)
        set_col_widths(ws4, pr_headers, pr_rows)

        # ════════════════════════════════════════════════════════════════════
        # TAB 6: Presentation Ratings, raw star ratings
        # ════════════════════════════════════════════════════════════════════
        ws5 = wb.create_sheet('Presentation Ratings')
        rt_headers = ['session_key', 'week', 'presentation_key', 'question_id',
                      'question_title', 'presenting_team_id', 'presenting_team',
                      'rater_id', 'rater_name', 'rater_team_id', 'rater_team',
                      'developed_1to5', 'easy_1to5', 'data_version', 'time']
        style_header(ws5, rt_headers)

        rt_rows = []
        for rt in ratings:
            rt_rows.append([
                rt['session_key'],
                rt['week_num'],
                rt['question_key'],
                rt['question_id'], rt['question_title'] or '',
                rt['presenting_team_id'],
                rt['presenting_team_name'] or key_to_team.get(
                    rt['question_key'], 'Unknown'
                ),
                rt['rater_sid'], rt['rater_name'],
                rt['rater_team_id'],
                rt['rater_team_name'] or (
                    team_id_to_name.get(rt['rater_team_id'], '')
                    if rt['rater_team_id'] else ''
                ),
                rt['q1_developed'], rt['q2_easy'],
                _export_data_version(rt['data_version']), rt['created_at'],
            ])
        for i, row in enumerate(rt_rows, 2):
            for col, val in enumerate(row, 1):
                ws5.cell(row=i, column=col, value=val)
        set_col_widths(ws5, rt_headers, rt_rows)

        # ════════════════════════════════════════════════════════════════════
        # TAB 7: Presentation Participants, finalized team membership
        # ════════════════════════════════════════════════════════════════════
        ws_participants = wb.create_sheet('Presentation Participants')
        participant_headers = [
            'session_key', 'week', 'presentation_key', 'participant_id',
            'participant_name', 'team_id', 'team_name', 'data_version', 'time',
        ]
        style_header(ws_participants, participant_headers)

        participant_rows = []
        for participant in presentation_participants:
            participant_rows.append([
                participant['session_key'],
                participant['week_num'],
                participant['presentation_key'],
                participant['student_identifier'],
                participant['student_name']
                or participant['student_identifier'],
                participant['team_id'],
                participant['team_name'] or '',
                _export_data_version(participant['data_version']),
                participant['created_at'],
            ])
        for i, row in enumerate(participant_rows, 2):
            for col, val in enumerate(row, 1):
                ws_participants.cell(row=i, column=col, value=val)
        set_col_widths(
            ws_participants, participant_headers, participant_rows
        )

        # ════════════════════════════════════════════════════════════════════
        # TABS 8-9: Challenge rounds and raw 1-5 ratings
        # ════════════════════════════════════════════════════════════════════
        ws6 = wb.create_sheet('Challenge Rounds')
        round_headers = [
            'session_key', 'week', 'challenge_key', 'presentation_key',
            'challenge_number', 'presenting_team_id', 'presenting_team',
            'question_id', 'question_title', 'challenger_id',
            'challenger_name', 'challenger_team_id', 'challenger_team',
            'ratings_submitted', 'average_score_1to5', 'data_version', 'time',
        ]
        style_header(ws6, round_headers)

        round_rows = []
        for challenge in challenge_rounds:
            round_rows.append([
                challenge['session_key'],
                challenge['week_num'],
                challenge['challenge_key'],
                challenge['presentation_key'],
                challenge['challenge_num'],
                challenge['presenting_team_id'],
                challenge['presenting_team_name'] or '',
                challenge['question_id'],
                challenge['question_title'] or '',
                challenge['challenger_sid'],
                challenge['challenger_name'] or '',
                challenge['challenger_team_id'],
                challenge['challenger_team_name'] or '',
                challenge['rating_count'],
                challenge['average_score'],
                _export_data_version(challenge['data_version']),
                challenge['created_at'],
            ])
        for i, row in enumerate(round_rows, 2):
            for col, val in enumerate(row, 1):
                ws6.cell(row=i, column=col, value=val)
        set_col_widths(ws6, round_headers, round_rows)

        ws7 = wb.create_sheet('Challenge Ratings')
        cr_headers = [
            'session_key', 'week', 'challenge_key', 'presentation_key',
            'challenge_number', 'presenting_team_id', 'presenting_team',
            'question_id', 'question_title', 'challenger_id',
            'challenger_name', 'challenger_team_id', 'challenger_team',
            'rater_id', 'rater_name', 'rater_team_id', 'rater_team',
            'score_1to5', 'data_version', 'time',
        ]
        style_header(ws7, cr_headers)

        cr_rows = []
        for cr in challenge_ratings:
            cr_rows.append([
                cr['session_key'],
                cr['week_num'],
                cr['challenge_key'],
                cr['presentation_key'],
                cr['challenge_num'],
                cr['presenting_team_id'],
                cr['presenting_team_name'] or '',
                cr['question_id'],
                cr['question_title'] or '',
                cr['challenger_sid'],
                cr['challenger_name'] or '',
                cr['challenger_team_id'],
                cr['challenger_team_name'] or '',
                cr['rater_sid'], cr['rater_name'] or '',
                cr['rater_team_id'],
                cr['rater_team_name'] or '',
                cr['score'], _export_data_version(cr['data_version']),
                cr['created_at'],
            ])
        for i, row in enumerate(cr_rows, 2):
            for col, val in enumerate(row, 1):
                ws7.cell(row=i, column=col, value=val)
        set_col_widths(ws7, cr_headers, cr_rows)

        # ── save ──
        # Save the workbook to an in-memory buffer
        xlsx_buf = BytesIO()
        for worksheet in wb.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.value = _spreadsheet_safe_value(cell.value)
        wb.save(xlsx_buf)
        xlsx_buf.seek(0)

        manifest = {
            'website_version': public_version(APP_VERSION),
            'database_schema_version': public_version(SCHEMA_VERSION),
            'export_format_version': public_version(EXPORT_FORMAT_VERSION),
            'data_compatibility': data_compatibility,
            'data_versions': public_data_versions,
            'exported_at_utc': exported_at,
        }
        manifest_bytes = (
            json.dumps(manifest, indent=2, sort_keys=True) + '\n'
        ).encode('utf-8')

        if (xlsx_buf.getbuffer().nbytes + asset_bytes + len(manifest_bytes)
                > MAX_EXPORT_BYTES):
            flash('The export files exceed the 50 MB safety limit.', 'error')
            return redirect(url_for('instructor_course', slug=slug))

        # Build a ZIP containing the Excel file + all question content files
        import zipfile
        zip_buf = BytesIO()
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Excel interaction data
            with zf.open('course_data.xlsx', 'w') as workbook_member:
                import shutil
                xlsx_buf.seek(0)
                shutil.copyfileobj(xlsx_buf, workbook_member, 1024 * 1024)

            zf.writestr('manifest.json', manifest_bytes)
            for archive_name, content in asset_files:
                zf.writestr(archive_name, content)

        zip_buf.seek(0)
        filename = (
            f"popping_{course['code'] or slug}_week_{current_week}_export.zip"
        )
        return send_file(
            zip_buf,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename,
        )

    except Exception:
        if snapshot_open and db is not None:
            db.rollback()
        app.logger.exception('Export failed for course %s', slug)
        flash('Export failed — please try again.', 'error')
        return redirect(url_for('instructor_course', slug=slug))
