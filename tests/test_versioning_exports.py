"""Regression tests for application, schema, data, and export versioning."""

from contextlib import contextmanager
import csv
from datetime import datetime, timezone
import io
import json
from pathlib import Path
import sqlite3
import sys
import uuid
import zipfile

import pytest


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import app as app_module  # noqa: E402
import config  # noqa: E402
import database  # noqa: E402
from versioning import (  # noqa: E402
    APP_VERSION,
    BASELINE_DATA_VERSION,
    BASELINE_SCHEMA_VERSION,
    EXPORT_FORMAT_VERSION,
    SCHEMA_VERSION,
    parse_version,
    public_version,
    sqlite_versions_compatible,
    versions_compatible,
)


SESSION_KEY = 11
BASELINE_FEEDBACK_TABLES = (
    "teammate_thumbs",
    "presentation_ratings",
    "challenge_rounds",
    "challenge_ratings",
)
FEEDBACK_TABLES = (*BASELINE_FEEDBACK_TABLES, "presentation_participants")


def test_v1_versions_and_public_form_are_aligned():
    assert APP_VERSION == "1.1.0"
    assert SCHEMA_VERSION == "1.1.0"
    assert EXPORT_FORMAT_VERSION == "1.1.0"
    assert BASELINE_SCHEMA_VERSION == "1.0.0"
    assert BASELINE_DATA_VERSION == "1.0.0"
    assert public_version() == "v1.1.0"
    assert public_version(SCHEMA_VERSION) == "v1.1.0"
    assert parse_version(SCHEMA_VERSION)[2] == 0


@pytest.mark.parametrize(
    "left,right,expected",
    (
        ("1.1.0", "1.1.99", True),
        ("1.0.3", "1.0.10", True),
        ("1.0.3", "1.1.0", False),
        ("1.10.0", "1.1.99", False),
        ("2.0.0", "1.0.0", False),
    ),
)
def test_compatibility_uses_numeric_major_and_minor(left, right, expected):
    assert versions_compatible(left, right) is expected


@pytest.mark.parametrize(
    "value",
    (None, "", "v1.0.0", " 1.0.0", "1.0.0 ", "01.0.0", "1.0"),
)
def test_malformed_internal_versions_fail_closed(value):
    with pytest.raises(ValueError):
        parse_version(value)
    assert sqlite_versions_compatible(value, SCHEMA_VERSION) == 0
    assert sqlite_versions_compatible(SCHEMA_VERSION, value) == 0


def test_fresh_schema_upgrades_baseline_and_all_versioned_tables():
    db = sqlite3.connect(":memory:")
    try:
        db.executescript(
            (PROJECT_ROOT / "popping.sql").read_text(encoding="utf-8")
        )
        assert db.execute(
            """SELECT schema_version, applied_by_app_version
               FROM schema_migrations ORDER BY id"""
        ).fetchall() == [(BASELINE_SCHEMA_VERSION, BASELINE_SCHEMA_VERSION)]
        for table in BASELINE_FEEDBACK_TABLES:
            columns = {
                row[1]: row for row in db.execute(f"PRAGMA table_info({table})")
            }
            assert columns["data_version"][4] == "'1.0.0'"

        database.upgrade_schema_connection(db)
        assert db.execute(
            """SELECT schema_version, applied_by_app_version
               FROM schema_migrations ORDER BY id"""
        ).fetchall() == [
            (BASELINE_SCHEMA_VERSION, BASELINE_SCHEMA_VERSION),
            (SCHEMA_VERSION, APP_VERSION),
        ]
        for table in FEEDBACK_TABLES:
            columns = {
                row[1]: row for row in db.execute(f"PRAGMA table_info({table})")
            }
            assert columns["data_version"][3] == 1
            assert columns["data_version"][4] is None
    finally:
        db.close()


@pytest.fixture
def versioned_course_env(tmp_path, monkeypatch):
    data_dir = tmp_path / "data"
    classes_dir = tmp_path / "classes"
    data_dir.mkdir()
    classes_dir.mkdir()
    slug = f"version_{uuid.uuid4().hex[:8]}"
    class_dir = classes_dir / slug
    class_dir.mkdir()
    (class_dir / "course.yaml").write_text(
        "\n".join(
            (
                f"slug: {slug}",
                "name: Versioning Test",
                "code: VERSION101",
                "semester: Test",
                "active: true",
                "",
            )
        ),
        encoding="utf-8",
    )
    (class_dir / "week-1-questions.md").write_text(
        "---\ntitle: Versioned Question\nid: versioned-question\n---\n\nDiscuss it.\n",
        encoding="utf-8",
    )

    course_dir = data_dir / slug
    course_dir.mkdir()
    db_path = course_dir / "popping.db"

    monkeypatch.setattr(config, "DATA_DIR", str(data_dir))
    monkeypatch.setattr(config, "CLASSES_DIR", str(classes_dir))
    monkeypatch.setattr(config, "CONFIG_DIR", str(classes_dir))
    monkeypatch.setitem(app_module.app.config, "TESTING", True)
    monkeypatch.setitem(
        app_module.app.config, "SECRET_KEY", "versioning-test-secret"
    )

    database.forget_schema(slug)
    app_module._clear_course_availability_cache(slug)

    db = sqlite3.connect(db_path)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys=ON")
    db.executescript(
        Path(config.DATABASE_SCHEMA).read_text(encoding="utf-8")
    )
    instructor_id = db.execute(
        """INSERT INTO instructors (username, name, pin)
           VALUES ('instructor', 'Test Instructor', '9999')"""
    ).lastrowid
    course_id = db.execute(
        """INSERT INTO courses
           (name, code, semester, slug, instructor_id, is_active)
           VALUES ('Versioning Test', 'VERSION101', 'Test', ?, ?, 1)""",
        (slug, instructor_id),
    ).lastrowid

    teams = {}
    for number in range(1, 4):
        name = f"Team {number}"
        teams[name] = db.execute(
            "INSERT INTO teams (course_id, name) VALUES (?, ?)",
            (course_id, name),
        ).lastrowid

    student_specs = (
        ("s1", "Alice", teams["Team 1"]),
        ("s2", "Bob", teams["Team 1"]),
        ("s3", "Cara", teams["Team 2"]),
        ("s4", "Dan", teams["Team 2"]),
        ("s5", "Eve", teams["Team 3"]),
    )
    students = {}
    for number, (student_id, name, team_id) in enumerate(student_specs, 1):
        students[student_id] = db.execute(
            """INSERT INTO students
               (course_id, student_id, name, pin, team_id)
               VALUES (?, ?, ?, ?, ?)""",
            (course_id, student_id, name, f"{number}{number}{number}{number}", team_id),
        ).lastrowid

    question_id = db.execute(
        """INSERT INTO questions
           (course_id, question_num, question_text, title, content, week_num,
            source_key)
           VALUES (?, 1, 'Discuss it.', 'Versioned Question', 'Discuss it.',
                   1, 'week-1-q-versioned-question')""",
        (course_id,),
    ).lastrowid
    db.execute(
        """INSERT INTO course_state
           (course_id, phase, max_teams, max_members_per_team,
            discussion_week, presentation_history, roster_version, session_key,
            active_challenges_json)
           VALUES (?, 'setup', 3, 10, 1, '[]', 0, ?, '[]')""",
        (course_id, SESSION_KEY),
    )
    db.commit()
    database.upgrade_schema_connection(db)
    db.commit()
    db.close()

    env = {
        "slug": slug,
        "db_path": db_path,
        "course_id": course_id,
        "instructor_id": instructor_id,
        "teams": teams,
        "students": students,
        "question_id": question_id,
    }
    yield env

    database.forget_schema(slug)
    app_module._clear_course_availability_cache(slug)


@contextmanager
def _connect(env):
    db = sqlite3.connect(env["db_path"])
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys=ON")
    try:
        yield db
    finally:
        db.close()


def _student_client(env, student_id):
    client = app_module.app.test_client()
    with client.session_transaction() as browser_session:
        browser_session["role"] = "student"
        browser_session["student_id"] = student_id
        browser_session["name"] = student_id
        browser_session["slug"] = env["slug"]
    return client


def _instructor_client(env):
    client = app_module.app.test_client()
    with client.session_transaction() as browser_session:
        browser_session["role"] = "instructor"
        browser_session["instructor_id"] = env["instructor_id"]
        browser_session["slug"] = env["slug"]
    return client


def _set_state(env, **fields):
    assignments = ", ".join(f"{name} = ?" for name in fields)
    with _connect(env) as db:
        db.execute(
            f"UPDATE course_state SET {assignments} WHERE course_id = ?",
            [*fields.values(), env["course_id"]],
        )
        db.commit()


def test_sqlite_compatibility_function_is_registered(versioned_course_env):
    with app_module.app.app_context():
        database.forget_schema(versioned_course_env["slug"])
        db = database.get_db(versioned_course_env["slug"])
        assert db.execute(
            "SELECT popping_version_compatible('1.1.9', '1.1.0')"
        ).fetchone()[0] == 1
        assert db.execute(
            "SELECT popping_version_compatible('broken', '1.1.0')"
        ).fetchone()[0] == 0


def test_readiness_requires_a_registered_schema_migration_path(
        versioned_course_env, monkeypatch):
    env = versioned_course_env
    calls = []

    def reject_schema_inspection(connection, allow_unversioned=True):
        calls.append(allow_unversioned)
        raise RuntimeError("Unsupported schema ledger")

    monkeypatch.setattr(
        app_module,
        "inspect_schema_version",
        reject_schema_inspection,
    )
    app_module._clear_course_availability_cache(env["slug"])

    assert app_module._course_availability(env["slug"])["status"] == "invalid"
    response = app_module.app.test_client().get("/healthz")
    assert response.status_code == 503
    assert response.get_json() == {"status": "unavailable"}
    assert calls == [True, True]


def test_healthz_accepts_versioned_schema_and_reports_versions(
        versioned_course_env):
    response = app_module.app.test_client().get("/healthz")

    assert response.status_code == 200
    assert response.get_json() == {
        "status": "ok",
        "courses_checked": 1,
        "website_version": public_version(APP_VERSION),
        "database_schema_version": public_version(SCHEMA_VERSION),
    }


def test_future_schema_fails_closed_in_health_and_course_availability(
        versioned_course_env):
    env = versioned_course_env
    with _connect(env) as db:
        db.execute(
            """INSERT INTO schema_migrations
               (schema_version, applied_by_app_version)
               VALUES ('1.2.0', '1.2.0')"""
        )
        db.commit()

    app_module._clear_course_availability_cache(env["slug"])
    assert app_module._course_availability(env["slug"])["status"] == "invalid"

    response = app_module.app.test_client().get("/healthz")
    assert response.status_code == 503
    assert response.get_json() == {"status": "unavailable"}


def test_feedback_write_paths_stamp_and_preserve_data_version(
        versioned_course_env):
    env = versioned_course_env
    _set_state(env, phase="discussion")
    teammate = _student_client(env, "s1")
    created_thumb = teammate.post(
        "/api/grade_peer",
        json={"recipient_id": "s2", "selected": True},
    )
    assert created_thumb.status_code == 200

    with _connect(env) as db:
        thumb = db.execute(
            "SELECT id, data_version FROM teammate_thumbs"
        ).fetchone()
        assert thumb["data_version"] == APP_VERSION
        db.execute(
            "UPDATE teammate_thumbs SET data_version = '1.1.7' WHERE id = ?",
            (thumb["id"],),
        )
        db.commit()
    repeated_thumb = teammate.post(
        "/api/grade_peer",
        json={"recipient_id": "s2", "selected": True},
    )
    assert repeated_thumb.status_code == 200

    now = datetime.now(timezone.utc).replace(tzinfo=None).strftime(
        "%Y-%m-%d %H:%M:%S.%f"
    )
    _set_state(
        env,
        phase="competition",
        active_team_id=env["teams"]["Team 1"],
        active_question_id=env["question_id"],
        current_question="Versioned Question",
        presentation_started_at=now,
        presentation_created_at=now,
        poll_active=1,
        poll_question_key="pres-versioned",
        poll_started_at=now,
        poll_closed_at=None,
        challenge_ratings_closed_at=None,
        active_challenges_json="[]",
    )
    rater = _student_client(env, "s3")
    created_rating = rater.post(
        "/api/submit_rating",
        json={
            "presentation_key": "pres-versioned",
            "q1_developed": 4,
            "q2_easy": 5,
        },
    )
    assert created_rating.status_code == 200
    with _connect(env) as db:
        rating = db.execute(
            "SELECT id, data_version FROM presentation_ratings"
        ).fetchone()
        assert rating["data_version"] == APP_VERSION
        db.execute(
            "UPDATE presentation_ratings SET data_version = '1.1.8' WHERE id = ?",
            (rating["id"],),
        )
        db.commit()
    repeated_rating = rater.post(
        "/api/submit_rating",
        json={
            "presentation_key": "pres-versioned",
            "q1_developed": 3,
            "q2_easy": 4,
        },
    )
    assert repeated_rating.status_code == 200

    _set_state(
        env,
        poll_active=0,
        poll_started_at=None,
        poll_closed_at=None,
        challenge_ratings_closed_at=None,
    )
    challenger = _student_client(env, "s5")
    raised = challenger.post(
        "/api/raise_hand", json={"presentation_key": "pres-versioned"}
    )
    assert raised.status_code == 200
    selected = _instructor_client(env).post(
        "/api/select_challenger",
        json={
            "presentation_key": "pres-versioned",
            "student_id": env["students"]["s5"],
        },
    )
    assert selected.status_code == 200
    challenge_key = selected.get_json()["challenge_key"]

    created_challenge_rating = rater.post(
        "/api/submit_challenge_rating",
        json={"challenge_key": challenge_key, "score": 5},
    )
    assert created_challenge_rating.status_code == 200
    with _connect(env) as db:
        round_row = db.execute(
            "SELECT data_version FROM challenge_rounds WHERE challenge_key = ?",
            (challenge_key,),
        ).fetchone()
        challenge_rating = db.execute(
            """SELECT id, data_version FROM challenge_ratings
               WHERE challenge_key = ?""",
            (challenge_key,),
        ).fetchone()
        assert round_row["data_version"] == APP_VERSION
        assert challenge_rating["data_version"] == APP_VERSION
        db.execute(
            "UPDATE challenge_ratings SET data_version = '1.1.9' WHERE id = ?",
            (challenge_rating["id"],),
        )
        db.commit()
    repeated_challenge_rating = rater.post(
        "/api/submit_challenge_rating",
        json={"challenge_key": challenge_key, "score": 4},
    )
    assert repeated_challenge_rating.status_code == 200

    with _connect(env) as db:
        assert db.execute(
            "SELECT data_version FROM teammate_thumbs"
        ).fetchone()[0] == "1.1.7"
        assert db.execute(
            "SELECT data_version FROM presentation_ratings"
        ).fetchone()[0] == "1.1.8"
        assert db.execute(
            "SELECT data_version FROM challenge_ratings"
        ).fetchone()[0] == "1.1.9"


def _seed_versioned_export_rows(env):
    cases = (
        ("compatible", "1.1.3", 1),
        ("unknown", "1.0.4", None),
        ("incompatible", "0.9.9", 1),
        ("malformed", "not-a-version", 1),
    )
    with _connect(env) as db:
        for number, (label, data_version, week_num) in enumerate(cases, 1):
            db.execute(
                """INSERT INTO teammate_thumbs
                   (course_id, session_key, week_num, question_key,
                    source_question_key, question_title, grader_id,
                    recipient_id, grader_team_id, grader_team_name,
                    recipient_team_id, recipient_team_name, data_version)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Team 1', ?, 'Team 1', ?)""",
                (
                    env["course_id"], SESSION_KEY, week_num,
                    f"thumb-{label}", f"thumb-{label}", label,
                    env["students"]["s1"], env["students"]["s2"],
                    env["teams"]["Team 1"], env["teams"]["Team 1"],
                    data_version,
                ),
            )
            db.execute(
                """INSERT INTO presentation_ratings
                   (course_id, student_id, question_key, session_key, week_num,
                    presenting_team_id, presenting_team_name, question_id,
                    question_title, rater_team_id, rater_team_name,
                    q1_developed, q2_easy, data_version)
                   VALUES (?, ?, ?, ?, ?, ?, 'Team 1', ?, ?, ?, 'Team 2',
                           4, 5, ?)""",
                (
                    env["course_id"], env["students"]["s3"],
                    f"presentation-{label}", SESSION_KEY, week_num,
                    env["teams"]["Team 1"],
                    env["question_id"] if week_num is not None else None,
                    label,
                    env["teams"]["Team 2"], data_version,
                ),
            )
            challenge_key = f"challenge-{label}"
            db.execute(
                """INSERT INTO challenge_rounds
                   (course_id, session_key, week_num, presentation_key,
                    challenge_key, challenge_num, challenger_id,
                    challenger_name, challenger_team_id, challenger_team_name,
                    presenting_team_id, presenting_team_name, question_id,
                    question_title, data_version)
                   VALUES (?, ?, ?, ?, ?, ?, ?, 'Eve', ?, 'Team 3', ?,
                           'Team 1', ?, ?, ?)""",
                (
                    env["course_id"], SESSION_KEY, week_num,
                    f"presentation-{label}", challenge_key, number,
                    env["students"]["s5"], env["teams"]["Team 3"],
                    env["teams"]["Team 1"], env["question_id"], label,
                    data_version,
                ),
            )
            db.execute(
                """INSERT INTO challenge_ratings
                   (course_id, session_key, week_num, challenge_key,
                    presentation_key, challenger_id, challenger_name,
                    challenger_team_id, challenger_team_name, rater_id,
                    rater_name, rater_team_id, rater_team_name, score,
                    data_version)
                   VALUES (?, ?, ?, ?, ?, ?, 'Eve', ?, 'Team 3', ?, 'Cara',
                           ?, 'Team 2', 4, ?)""",
                (
                    env["course_id"], SESSION_KEY, week_num, challenge_key,
                    f"presentation-{label}", env["students"]["s5"],
                    env["teams"]["Team 3"], env["students"]["s3"],
                    env["teams"]["Team 2"], data_version,
                ),
            )
            db.execute(
                """INSERT INTO presentation_participants
                   (course_id, session_key, week_num, presentation_key,
                    student_id, student_identifier, student_name,
                    team_id, team_name, data_version)
                   VALUES (?, ?, ?, ?, ?, 's1', 'Alice', ?, 'Team 1', ?)""",
                (
                    env["course_id"], SESSION_KEY, week_num,
                    f"presentation-{label}", env["students"]["s1"],
                    env["teams"]["Team 1"], data_version,
                ),
            )
        db.commit()
    _set_state(env, phase="ended", discussion_week=1)


def _workbook_rows(workbook, sheet_name):
    values = list(workbook[sheet_name].iter_rows(values_only=True))
    headers = values[0]
    return [dict(zip(headers, row)) for row in values[1:]]


def _assert_utc_timestamp(value):
    assert isinstance(value, str) and value.endswith("Z")
    datetime.fromisoformat(value[:-1] + "+00:00")


def test_weekly_export_routes_only_compatible_known_week_rows_and_versions(
        versioned_course_env):
    from openpyxl import load_workbook

    env = versioned_course_env
    _seed_versioned_export_rows(env)
    with _connect(env) as db:
        db.execute(
            """UPDATE students
               SET is_active = 0, last_team_id = team_id, team_id = NULL
               WHERE id = ?""",
            (env["students"]["s2"],),
        )
        db.execute(
            """INSERT INTO presentation_participants
               (course_id, session_key, week_num, presentation_key,
                student_id, student_identifier, student_name,
                team_id, team_name, data_version)
               VALUES (?, ?, 2, 'presentation-prior-week', ?, 's1', 'Alice',
                       ?, 'Team 1', '1.1.3')""",
            (
                env["course_id"], SESSION_KEY + 1,
                env["students"]["s1"], env["teams"]["Team 1"],
            ),
        )
        db.execute(
            """INSERT INTO challenge_rounds
               (course_id, session_key, week_num, presentation_key,
                challenge_key, challenge_num, challenger_id,
                challenger_name, challenger_team_id, challenger_team_name,
                presenting_team_id, presenting_team_name, question_id,
                question_title, data_version)
               VALUES (?, ?, 2, 'presentation-prior-week',
                       'challenge-prior-week', 1, ?, 'Bob', ?, 'Team 1',
                       ?, 'Team 2', ?, 'Prior week', '1.1.3')""",
            (
                env["course_id"], SESSION_KEY + 1,
                env["students"]["s2"], env["teams"]["Team 1"],
                env["teams"]["Team 2"], env["question_id"],
            ),
        )
        db.commit()
    response = _instructor_client(env).get(f"/export/{env['slug']}")

    assert response.status_code == 200
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        manifest = json.loads(archive.read("manifest.json"))
        workbook = load_workbook(
            io.BytesIO(archive.read("course_data.xlsx")), read_only=True
        )

    assert manifest["website_version"] == public_version(APP_VERSION)
    assert manifest["database_schema_version"] == public_version(SCHEMA_VERSION)
    assert manifest["export_format_version"] == public_version(
        EXPORT_FORMAT_VERSION
    )
    assert manifest["data_compatibility"] == "v1.1.x"
    assert manifest["data_versions"] == ["v1.1.3"]
    _assert_utc_timestamp(manifest["exported_at_utc"])

    summary = {
        row[0]: row[1]
        for row in workbook["Summary"].iter_rows(values_only=True)
        if row[0]
    }
    assert summary["Website Version"] == "v1.1.0"
    assert summary["Database Schema Version"] == "v1.1.0"
    assert summary["Export Format Version"] == "v1.1.0"
    assert summary["Data Compatibility"] == "v1.1.x"
    assert summary["Data Versions Included"] == "v1.1.3"
    assert summary["Participation Roster Scope"] == (
        "Course-wide compatible participation through export time"
    )
    _assert_utc_timestamp(summary["Exported At (UTC)"])
    assert summary["Week Peer Reviews (thumbs)"] == 1
    assert summary["Week Presentation Ratings"] == 1
    assert summary["Week Presentation Participants"] == 1
    assert summary["Week Challenge Rounds"] == 1
    assert summary["Week Challenge Ratings"] == 1

    peer_rows = _workbook_rows(workbook, "Peer Reviews")
    presentation_rows = _workbook_rows(workbook, "Presentation Ratings")
    round_rows = _workbook_rows(workbook, "Challenge Rounds")
    challenge_rating_rows = _workbook_rows(workbook, "Challenge Ratings")
    participant_rows = _workbook_rows(workbook, "Presentation Participants")
    roster_rows = _workbook_rows(workbook, "Participation Roster")
    student_rows = _workbook_rows(workbook, "Students")
    assert [row["discussion_post_key"] for row in peer_rows] == [
        "thumb-compatible"
    ]
    assert [row["presentation_key"] for row in presentation_rows] == [
        "presentation-compatible"
    ]
    assert [row["challenge_key"] for row in round_rows] == [
        "challenge-compatible"
    ]
    assert [row["challenge_key"] for row in challenge_rating_rows] == [
        "challenge-compatible"
    ]
    assert participant_rows == [{
        "session_key": SESSION_KEY,
        "week": 1,
        "presentation_key": "presentation-compatible",
        "participant_id": "s1",
        "participant_name": "Alice",
        "team_id": env["teams"]["Team 1"],
        "team_name": "Team 1",
        "data_version": "v1.1.3",
        "time": participant_rows[0]["time"],
    }]
    alice = next(row for row in student_rows if row["student_id"] == "s1")
    assert alice["course_presentation_team_turns"] == 2
    assert alice["course_challenger_turns"] == 0
    assert next(
        workbook["Participation Roster"].iter_rows(values_only=True)
    ) == (
        "student_id", "name", "team", "status",
        "course_presentation_team_turns", "course_challenger_turns",
    )
    assert len(roster_rows) == 5
    assert {row["student_id"] for row in roster_rows} == {
        "s1", "s2", "s3", "s4", "s5",
    }
    roster_alice = next(
        row for row in roster_rows if row["student_id"] == "s1"
    )
    assert roster_alice == {
        "student_id": "s1",
        "name": "Alice",
        "team": "Team 1",
        "status": "active",
        "course_presentation_team_turns": 2,
        "course_challenger_turns": 0,
    }
    roster_bob = next(
        row for row in roster_rows if row["student_id"] == "s2"
    )
    assert roster_bob == {
        "student_id": "s2",
        "name": "Bob",
        "team": "Team 1",
        "status": "archived",
        "course_presentation_team_turns": 0,
        "course_challenger_turns": 1,
    }
    for rows in (
        peer_rows, presentation_rows, round_rows, challenge_rating_rows
    ):
        assert [row["data_version"] for row in rows] == ["v1.1.3"]


def test_participants_export_and_roster_counts_agree_on_invalid_weeks(
        versioned_course_env):
    """Hand-corrupted week rows vanish from BOTH export and roster counts.

    The dashboard counting query and the participants export must apply the
    same week-validity rule so they can never disagree about a corrupted row.
    """
    from openpyxl import load_workbook

    env = versioned_course_env
    _seed_versioned_export_rows(env)
    with _connect(env) as db:
        for bad_week in (0, -1):
            db.execute(
                """INSERT INTO presentation_participants
                   (course_id, session_key, week_num, presentation_key,
                    student_id, student_identifier, student_name,
                    team_id, team_name, data_version)
                   VALUES (?, ?, ?, ?, ?, 's1', 'Alice', ?, 'Team 1',
                           '1.1.3')""",
                (
                    env["course_id"], SESSION_KEY, bad_week,
                    f"presentation-bad-week-{bad_week}",
                    env["students"]["s1"], env["teams"]["Team 1"],
                ),
            )
        db.commit()

    response = _instructor_client(env).get(f"/export/{env['slug']}")

    assert response.status_code == 200
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        workbook = load_workbook(
            io.BytesIO(archive.read("course_data.xlsx")), read_only=True
        )
    participant_rows = _workbook_rows(workbook, "Presentation Participants")
    assert [row["presentation_key"] for row in participant_rows] == [
        "presentation-compatible"
    ]
    roster_rows = _workbook_rows(workbook, "Participation Roster")
    roster_alice = next(
        row for row in roster_rows if row["student_id"] == "s1"
    )
    assert roster_alice["course_presentation_team_turns"] == 1


def test_export_previous_week_uses_same_layout(versioned_course_env):
    """?week=N exports that week's rows with the identical workbook layout."""
    from openpyxl import load_workbook

    env = versioned_course_env
    _seed_versioned_export_rows(env)
    with _connect(env) as db:
        db.execute(
            """INSERT INTO presentation_participants
               (course_id, session_key, week_num, presentation_key,
                student_id, student_identifier, student_name,
                team_id, team_name, data_version)
               VALUES (?, ?, 2, 'presentation-week-2', ?, 's2', 'Bob',
                       ?, 'Team 1', '1.1.3')""",
            (
                env["course_id"], SESSION_KEY + 1,
                env["students"]["s2"], env["teams"]["Team 1"],
            ),
        )
        db.commit()
    _set_state(env, phase="ended", discussion_week=2)
    client = _instructor_client(env)

    def export_workbook(response):
        assert response.status_code == 200
        with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
            return load_workbook(
                io.BytesIO(archive.read("course_data.xlsx")), read_only=True
            )

    # Default export: current week (2) only.
    response = client.get(f"/export/{env['slug']}")
    assert "week_2" in response.headers["Content-Disposition"]
    workbook = export_workbook(response)
    participant_rows = _workbook_rows(workbook, "Presentation Participants")
    assert [row["presentation_key"] for row in participant_rows] == [
        "presentation-week-2"
    ]

    # Explicit previous week: same layout, week-1 rows, week-1 labels.
    response = client.get(f"/export/{env['slug']}?week=1")
    assert "week_1" in response.headers["Content-Disposition"]
    workbook = export_workbook(response)
    participant_rows = _workbook_rows(workbook, "Presentation Participants")
    assert [row["presentation_key"] for row in participant_rows] == [
        "presentation-compatible"
    ]
    summary = {
        row[0]: row[1]
        for row in workbook["Summary"].iter_rows(values_only=True)
        if row[0]
    }
    assert summary["Lecture Week"] == 1
    assert summary["Export Scope"] == "Week 1"
    # The course-wide roster still pools all weeks in a week-1 export.
    roster_rows = _workbook_rows(workbook, "Participation Roster")
    roster_bob = next(
        row for row in roster_rows if row["student_id"] == "s2"
    )
    assert roster_bob["course_presentation_team_turns"] == 1


def test_export_week_parameter_validation(versioned_course_env):
    env = versioned_course_env
    _seed_versioned_export_rows(env)
    _set_state(env, phase="ended", discussion_week=2)
    client = _instructor_client(env)

    for bad in ("0", "-1", "abc", "3", "2.5"):
        response = client.get(f"/export/{env['slug']}?week={bad}")
        assert response.status_code == 400, f"week={bad!r} must be rejected"
    response = client.get(f"/export/{env['slug']}?weeks=all")
    assert response.status_code == 400
    # An empty week parameter behaves like no parameter (current week).
    assert client.get(f"/export/{env['slug']}?week=").status_code == 200
    # Boundary values are accepted.
    assert client.get(f"/export/{env['slug']}?week=1").status_code == 200
    assert client.get(f"/export/{env['slug']}?week=2").status_code == 200


def test_tools_menu_lists_downloadable_weeks(versioned_course_env):
    env = versioned_course_env
    _set_state(env, phase="ended", discussion_week=3)
    response = _instructor_client(env).get(f"/instructor/{env['slug']}")

    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert "Download Results" in html
    assert "Current Week (3)" in html
    assert f"/export/{env['slug']}?week=2" in html
    assert f"/export/{env['slug']}?week=1" in html


def test_legacy_export_routes_unknown_incompatible_and_malformed_four_types(
        versioned_course_env):
    env = versioned_course_env
    _seed_versioned_export_rows(env)
    response = _instructor_client(env).get(
        f"/export/{env['slug']}/legacy-feedback.csv"
    )

    assert response.status_code == 200
    rows = list(csv.DictReader(io.StringIO(
        response.data.decode("utf-8-sig")
    )))
    assert len(rows) == 15
    assert {row["record_type"] for row in rows} == {
        "teammate_thumb",
        "presentation_rating",
        "challenge_round",
        "challenge_rating",
        "presentation_participant",
    }

    def row_key(row):
        if row["record_type"] == "teammate_thumb":
            return row["question_key"].removeprefix("thumb-")
        if row["record_type"] == "presentation_rating":
            return row["question_key"].removeprefix("presentation-")
        if row["record_type"] == "presentation_participant":
            return row["presentation_key"].removeprefix("presentation-")
        return row["challenge_key"].removeprefix("challenge-")

    by_label = {}
    for row in rows:
        by_label.setdefault(row_key(row), []).append(row)
        assert row["exported_by_website_version"] == "v1.1.0"
        assert row["database_schema_version"] == "v1.1.0"
        assert row["export_format_version"] == "v1.1.0"
        _assert_utc_timestamp(row["exported_at_utc"])

    assert set(by_label) == {"unknown", "incompatible", "malformed"}
    assert {row["legacy_reason"] for row in by_label["unknown"]} == {
        "unknown_week"
    }
    assert {row["data_version"] for row in by_label["unknown"]} == {
        "v1.0.4"
    }
    assert {row["legacy_reason"] for row in by_label["incompatible"]} == {
        "incompatible_data_version"
    }
    assert {row["data_version"] for row in by_label["incompatible"]} == {
        "v0.9.9"
    }
    assert {row["legacy_reason"] for row in by_label["malformed"]} == {
        "malformed_data_version"
    }
    assert {row["data_version"] for row in by_label["malformed"]} == {
        "not-a-version"
    }
    assert not any(row_key(row) == "compatible" for row in rows)


def test_unknown_week_reason_precedes_malformed_version(versioned_course_env):
    env = versioned_course_env
    with _connect(env) as db:
        db.execute(
            """INSERT INTO teammate_thumbs
               (course_id, session_key, week_num, question_key,
                source_question_key, grader_id, recipient_id, data_version)
               VALUES (?, ?, NULL, 'both-invalid', 'both-invalid', ?, ?,
                       'not-a-version')""",
            (
                env["course_id"], SESSION_KEY,
                env["students"]["s1"], env["students"]["s2"],
            ),
        )
        db.execute("UPDATE course_state SET phase = 'ended'")
        db.commit()

    response = _instructor_client(env).get(
        f"/export/{env['slug']}/legacy-feedback.csv"
    )
    rows = list(csv.DictReader(io.StringIO(
        response.data.decode("utf-8-sig")
    )))
    assert len(rows) == 1
    assert rows[0]["legacy_reason"] == "unknown_week"



def test_ledger_present_database_with_missing_data_version_schema_fails_ready(
        versioned_course_env):
    env = versioned_course_env
    with _connect(env) as db:
        db.execute("ALTER TABLE challenge_ratings DROP COLUMN data_version")
        db.commit()

    app_module._clear_course_availability_cache(env["slug"])
    assert app_module._course_availability(env["slug"])["status"] == "invalid"

    response = app_module.app.test_client().get("/healthz")
    assert response.status_code == 503
    assert response.get_json() == {"status": "unavailable"}


def test_history_current_and_legacy_routes_share_week_inference(
        versioned_course_env):
    from openpyxl import load_workbook

    env = versioned_course_env
    history = [
        {
            "presentation_key": "history-question",
            "session_key": SESSION_KEY,
            "data_version": "1.1.0",
            "question_id": env["question_id"],
            "team_id": env["teams"]["Team 1"],
            "team": "Team 1",
            "title": "Inferred from question",
            "started_at": "2026-01-01 10:00:00",
        },
        {
            "presentation_key": "history-rating",
            "session_key": SESSION_KEY,
            "data_version": "1.1.4",
            "team_id": env["teams"]["Team 1"],
            "team": "Team 1",
            "title": "Inferred from rating",
            "started_at": "2026-01-01 10:05:00",
        },
        {
            "presentation_key": "history-unknown",
            "session_key": SESSION_KEY,
            "data_version": "1.1.5",
            "team": "Team 2",
            "title": "Unknown week",
        },
        {
            "presentation_key": "history-old",
            "session_key": SESSION_KEY,
            "week_num": 1,
            "data_version": "0.9.9",
            "team": "Team 2",
            "title": "Old series",
        },
        {
            "presentation_key": "history-malformed",
            "session_key": SESSION_KEY,
            "week_num": 1,
            "data_version": "bad-version",
            "team": "Team 2",
            "title": "Malformed version",
        },
    ]
    with _connect(env) as db:
        db.execute(
            """INSERT INTO presentation_ratings
               (course_id, student_id, question_key, session_key, week_num,
                presenting_team_id, presenting_team_name, question_title,
                rater_team_id, rater_team_name, q1_developed, q2_easy,
                data_version)
               VALUES (?, ?, 'history-rating', ?, 1, ?, 'Team 1',
                       'Inferred from rating', ?, 'Team 2', 4, 4, '1.1.6')""",
            (
                env["course_id"], env["students"]["s3"], SESSION_KEY,
                env["teams"]["Team 1"], env["teams"]["Team 2"],
            ),
        )
        db.execute(
            """UPDATE course_state
               SET phase = 'ended', discussion_week = 1,
                   presentation_history = ?
               WHERE course_id = ?""",
            (json.dumps(history), env["course_id"]),
        )
        db.commit()

    current = _instructor_client(env).get(f"/export/{env['slug']}")
    assert current.status_code == 200
    with zipfile.ZipFile(io.BytesIO(current.data)) as archive:
        manifest = json.loads(archive.read("manifest.json"))
        workbook = load_workbook(
            io.BytesIO(archive.read("course_data.xlsx")), read_only=True
        )
    assert manifest["data_versions"] == ["v1.1.0", "v1.1.4", "v1.1.6"]
    team_rows = _workbook_rows(workbook, "Teams")
    team_1 = next(row for row in team_rows if row["team_name"] == "Team 1")
    assert team_1["presentations"] == 2

    legacy = _instructor_client(env).get(
        f"/export/{env['slug']}/legacy-feedback.csv"
    )
    assert legacy.status_code == 200
    rows = list(csv.DictReader(io.StringIO(
        legacy.data.decode("utf-8-sig")
    )))
    history_rows = [
        row for row in rows if row["record_type"] == "presentation_history"
    ]
    by_key = {row["presentation_key"]: row for row in history_rows}
    assert set(by_key) == {
        "history-unknown", "history-old", "history-malformed"
    }
    assert by_key["history-unknown"]["legacy_reason"] == "unknown_week"
    assert by_key["history-old"]["legacy_reason"] == (
        "incompatible_data_version"
    )
    assert by_key["history-malformed"]["legacy_reason"] == (
        "malformed_data_version"
    )
    assert "history-question" not in by_key
    assert "history-rating" not in by_key
    assert json.loads(by_key["history-old"]["history_json"])[
        "data_version"
    ] == "0.9.9"


def test_history_with_zero_week_question_is_legacy_only(
        versioned_course_env):
    from openpyxl import load_workbook

    env = versioned_course_env
    presentation_key = "history-zero-question-week"
    history = [{
        "presentation_key": presentation_key,
        "session_key": SESSION_KEY,
        "question_id": env["question_id"],
        "data_version": "1.1.4",
        "team_id": env["teams"]["Team 1"],
        "team": "Team 1",
        "title": "Invalid zero-week question",
    }]
    with _connect(env) as db:
        db.execute(
            "UPDATE questions SET week_num = 0 WHERE id = ?",
            (env["question_id"],),
        )
        db.execute(
            """UPDATE course_state
               SET phase = 'ended', discussion_week = 1,
                   presentation_history = ?
               WHERE course_id = ?""",
            (json.dumps(history), env["course_id"]),
        )
        db.commit()

    current = _instructor_client(env).get(f"/export/{env['slug']}")
    assert current.status_code == 200
    with zipfile.ZipFile(io.BytesIO(current.data)) as archive:
        workbook = load_workbook(
            io.BytesIO(archive.read("course_data.xlsx")), read_only=True
        )
    team_rows = _workbook_rows(workbook, "Teams")
    team_1 = next(row for row in team_rows if row["team_name"] == "Team 1")
    assert team_1["presentations"] == 0

    legacy = _instructor_client(env).get(
        f"/export/{env['slug']}/legacy-feedback.csv"
    )
    assert legacy.status_code == 200
    rows = list(csv.DictReader(io.StringIO(
        legacy.data.decode("utf-8-sig")
    )))
    history_rows = [
        row for row in rows
        if row["record_type"] == "presentation_history"
    ]
    assert len(history_rows) == 1
    assert history_rows[0]["presentation_key"] == presentation_key
    assert history_rows[0]["lecture_week"] == "unknown"
    assert history_rows[0]["legacy_reason"] == "unknown_week"
    assert history_rows[0]["data_version"] == "v1.1.4"

def test_nonpositive_and_noninteger_weeks_route_as_unknown(
        versioned_course_env):
    assert app_module._resolve_history_week({"week_num": 0}) is None
    assert app_module._resolve_history_week({"week_num": -2}) is None
    assert app_module._resolve_history_week({"week_num": "not-a-week"}) is None
    assert app_module._resolve_history_week({"week_num": 1.5}) is None
    assert app_module._resolve_history_week({"week_num": 1.0}) is None
    assert app_module._resolve_history_week({"week_num": "1"}) is None
    assert app_module._resolve_history_week({"week_num": True}) is None
    assert app_module._legacy_data_reason(0, "1.0.0") == "unknown_week"
    assert app_module._legacy_data_reason(-2, "1.0.0") == "unknown_week"
    assert app_module._legacy_data_reason(
        "not-a-week", "1.0.0"
    ) == "unknown_week"
    for malformed_week in (1.5, 1.0, "1", True):
        assert app_module._legacy_data_reason(
            malformed_week, "1.0.0"
        ) == "unknown_week"

    env = versioned_course_env
    with _connect(env) as db:
        for label, week_num in (
            ("zero-week", 0),
            ("negative-week", -2),
            ("fractional-week", 1.5),
            ("text-week", "not-a-week"),
        ):
            db.execute(
                """INSERT INTO teammate_thumbs
                   (course_id, session_key, week_num, question_key,
                    source_question_key, grader_id, recipient_id,
                    data_version)
                   VALUES (?, ?, ?, ?, ?, ?, ?, '1.0.0')""",
                (
                    env["course_id"], SESSION_KEY, week_num, label, label,
                    env["students"]["s1"], env["students"]["s2"],
                ),
            )
        db.execute("UPDATE course_state SET phase = 'ended'")
        db.commit()

    response = _instructor_client(env).get(
        f"/export/{env['slug']}/legacy-feedback.csv"
    )
    rows = list(csv.DictReader(io.StringIO(
        response.data.decode("utf-8-sig")
    )))
    assert {row["question_key"] for row in rows} == {
        "zero-week", "negative-week", "fractional-week", "text-week"
    }
    assert {row["legacy_reason"] for row in rows} == {"unknown_week"}


def test_invalid_participation_weeks_are_legacy_only_everywhere(
        versioned_course_env):
    from openpyxl import load_workbook

    env = versioned_course_env
    invalid_weeks = (
        ("null-week", None),
        ("zero-week", 0),
        ("negative-week", -2),
        ("text-week", "not-a-week"),
        ("real-week", 1.5),
    )
    with _connect(env) as db:
        for number, (label, week_num) in enumerate(invalid_weeks, 1):
            db.execute(
                """INSERT INTO presentation_participants
                   (course_id, session_key, week_num, presentation_key,
                    student_id, student_identifier, data_version)
                   VALUES (?, ?, ?, ?, ?, 's1', '1.1.9')""",
                (
                    env["course_id"], SESSION_KEY, week_num,
                    f"presentation-{label}", env["students"]["s1"],
                ),
            )
            db.execute(
                """INSERT INTO challenge_rounds
                   (course_id, session_key, week_num, presentation_key,
                    challenge_key, challenge_num, challenger_id, data_version)
                   VALUES (?, ?, ?, ?, ?, ?, ?, '1.1.9')""",
                (
                    env["course_id"], SESSION_KEY, week_num,
                    f"presentation-{label}", f"challenge-{label}", number,
                    env["students"]["s5"],
                ),
            )
        db.commit()

    client = _instructor_client(env)
    teams_response = client.get("/api/teams")
    assert teams_response.status_code == 200
    team_members = {
        member["student_id"]: member
        for team in teams_response.get_json()
        for member in team["members"]
    }
    assert team_members["s1"]["presentation_count"] == 0
    assert team_members["s5"]["challenger_count"] == 0

    students_response = client.get("/api/students?per_page=100")
    assert students_response.status_code == 200
    student_rows = {
        row["student_id"]: row
        for row in students_response.get_json()["students"]
    }
    assert student_rows["s1"]["presentation_count"] == 0
    assert student_rows["s5"]["challenger_count"] == 0

    _set_state(env, phase="ended")
    current = client.get(f"/export/{env['slug']}")
    assert current.status_code == 200
    with zipfile.ZipFile(io.BytesIO(current.data)) as archive:
        manifest = json.loads(archive.read("manifest.json"))
        workbook = load_workbook(
            io.BytesIO(archive.read("course_data.xlsx")), read_only=True
        )

    assert "v1.1.9" not in manifest["data_versions"]
    assert _workbook_rows(workbook, "Presentation Participants") == []
    assert _workbook_rows(workbook, "Challenge Rounds") == []
    roster_rows = {
        row["student_id"]: row
        for row in _workbook_rows(workbook, "Participation Roster")
    }
    assert roster_rows["s1"]["course_presentation_team_turns"] == 0
    assert roster_rows["s5"]["course_challenger_turns"] == 0

    legacy = client.get(f"/export/{env['slug']}/legacy-feedback.csv")
    assert legacy.status_code == 200
    legacy_rows = list(csv.DictReader(io.StringIO(
        legacy.data.decode("utf-8-sig")
    )))
    assert len(legacy_rows) == 10
    assert {row["record_type"] for row in legacy_rows} == {
        "presentation_participant", "challenge_round",
    }
    assert {row["legacy_reason"] for row in legacy_rows} == {"unknown_week"}
    expected_labels = {label for label, _week_num in invalid_weeks}
    assert {
        row["presentation_key"].removeprefix("presentation-")
        for row in legacy_rows
        if row["record_type"] == "presentation_participant"
    } == expected_labels
    assert {
        row["challenge_key"].removeprefix("challenge-")
        for row in legacy_rows
        if row["record_type"] == "challenge_round"
    } == expected_labels


def test_live_counts_and_saved_controls_ignore_incompatible_rows(
        versioned_course_env):
    env = versioned_course_env
    _set_state(env, phase="discussion")
    with _connect(env) as db:
        for grader, recipient, version in (
            ("s1", "s2", "1.1.7"),
            ("s2", "s1", "0.9.9"),
        ):
            db.execute(
                """INSERT INTO teammate_thumbs
                   (course_id, session_key, week_num, question_key,
                    source_question_key, question_title, grader_id,
                    recipient_id, grader_team_id, grader_team_name,
                    recipient_team_id, recipient_team_name, data_version)
                   VALUES (?, ?, 1, 'discussion', 'discussion', 'Discussion',
                           ?, ?, ?, 'Team 1', ?, 'Team 1', ?)""",
                (
                    env["course_id"], SESSION_KEY,
                    env["students"][grader], env["students"][recipient],
                    env["teams"]["Team 1"], env["teams"]["Team 1"], version,
                ),
            )
        db.commit()

    discussion_state = _instructor_client(env).get("/api/state").get_json()
    team_1 = next(
        row for row in discussion_state["thumb_team_progress"]
        if row["team_name"] == "Team 1"
    )
    assert team_1["participant_count"] == 1
    assert team_1["thumb_count"] == 1
    assert _student_client(env, "s1").get("/api/my_responses").get_json()[
        "thumb_recipient_ids"
    ] == ["s2"]
    assert _student_client(env, "s2").get("/api/my_responses").get_json()[
        "thumb_recipient_ids"
    ] == []

    now = datetime.now(timezone.utc).replace(tzinfo=None).strftime(
        "%Y-%m-%d %H:%M:%S.%f"
    )
    presentation_key = "live-compatible-presentation"
    challenge_key = "live-compatible-challenge"
    _set_state(
        env,
        phase="competition",
        active_team_id=env["teams"]["Team 1"],
        active_question_id=env["question_id"],
        current_question="Versioned Question",
        presentation_started_at=now,
        presentation_created_at=now,
        poll_active=1,
        poll_question_key=presentation_key,
        poll_started_at=now,
        poll_closed_at=None,
        active_challenges_json=json.dumps([{
            "challenge_key": challenge_key,
            "challenge_num": 1,
            "challenger_id": env["students"]["s5"],
            "challenger_name": "Eve",
            "challenger_team_id": env["teams"]["Team 3"],
            "challenger_team_name": "Team 3",
        }]),
    )
    with _connect(env) as db:
        for student_id, version, score in (
            ("s3", "1.1.8", 3),
            ("s4", "0.9.9", 5),
        ):
            db.execute(
                """INSERT INTO presentation_ratings
                   (course_id, student_id, question_key, session_key, week_num,
                    presenting_team_id, presenting_team_name, question_id,
                    question_title, rater_team_id, rater_team_name,
                    q1_developed, q2_easy, data_version)
                   VALUES (?, ?, ?, ?, 1, ?, 'Team 1', ?,
                           'Versioned Question', ?, 'Team 2', ?, ?, ?)""",
                (
                    env["course_id"], env["students"][student_id],
                    presentation_key, SESSION_KEY,
                    env["teams"]["Team 1"], env["question_id"],
                    env["teams"]["Team 2"], score, score, version,
                ),
            )
        db.execute(
            """INSERT INTO challenge_rounds
               (course_id, session_key, week_num, presentation_key,
                challenge_key, challenge_num, challenger_id, challenger_name,
                challenger_team_id, challenger_team_name, presenting_team_id,
                presenting_team_name, question_id, question_title,
                data_version)
               VALUES (?, ?, 1, ?, ?, 1, ?, 'Eve', ?, 'Team 3', ?,
                       'Team 1', ?, 'Versioned Question', '1.1.8')""",
            (
                env["course_id"], SESSION_KEY, presentation_key,
                challenge_key, env["students"]["s5"],
                env["teams"]["Team 3"], env["teams"]["Team 1"],
                env["question_id"],
            ),
        )
        for student_id, version, score in (
            ("s3", "1.1.8", 3),
            ("s4", "0.9.9", 5),
        ):
            db.execute(
                """INSERT INTO challenge_ratings
                   (course_id, session_key, week_num, challenge_key,
                    presentation_key, challenger_id, challenger_name,
                    challenger_team_id, challenger_team_name, rater_id,
                    rater_name, rater_team_id, rater_team_name, score,
                    data_version)
                   VALUES (?, ?, 1, ?, ?, ?, 'Eve', ?, 'Team 3', ?, ?,
                           ?, 'Team 2', ?, ?)""",
                (
                    env["course_id"], SESSION_KEY, challenge_key,
                    presentation_key, env["students"]["s5"],
                    env["teams"]["Team 3"], env["students"][student_id],
                    student_id, env["teams"]["Team 2"], score, version,
                ),
            )
        db.commit()

    live_state = _instructor_client(env).get("/api/state").get_json()
    assert live_state["poll_count"] == 1
    assert live_state["challenge_rating_summaries"][challenge_key][
        "submitted_count"
    ] == 1

    compatible_controls = _student_client(env, "s3").get(
        "/api/my_responses"
    ).get_json()
    assert compatible_controls["rating"] == {
        "q1_developed": 3,
        "q2_easy": 3,
    }
    assert compatible_controls["challenge_ratings"] == {challenge_key: 3}

    legacy_controls = _student_client(env, "s4").get(
        "/api/my_responses"
    ).get_json()
    assert legacy_controls["rating"] is None
    assert legacy_controls["challenge_ratings"] == {}

def test_end_summary_rankings_and_live_history_ignore_incompatible_rows(
        versioned_course_env):
    env = versioned_course_env
    history = [
        {
            "presentation_key": "rank-current",
            "session_key": SESSION_KEY,
            "week_num": 1,
            "data_version": "1.1.8",
            "team_id": env["teams"]["Team 1"],
            "team": "Team 1",
        },
        {
            "presentation_key": "rank-baseline",
            "session_key": SESSION_KEY,
            "week_num": 1,
            "team_id": env["teams"]["Team 3"],
            "team": "Team 3",
        },
        {
            "presentation_key": "rank-old",
            "session_key": SESSION_KEY,
            "week_num": 1,
            "data_version": "0.9.9",
            "team_id": env["teams"]["Team 2"],
            "team": "Team 2",
        },
        {
            "presentation_key": "rank-malformed",
            "session_key": SESSION_KEY,
            "week_num": 1,
            "data_version": "bad-version",
            "team": "Team 2",
        },
        "malformed-history-item",
    ]
    with _connect(env) as db:
        for key, student_id, team_name, team_id, score, version in (
            ("rank-current", "s3", "Team 1", env["teams"]["Team 1"], 2,
             "1.1.8"),
            ("rank-old", "s4", "Team 2", env["teams"]["Team 2"], 5,
             "0.9.9"),
        ):
            db.execute(
                """INSERT INTO presentation_ratings
                   (course_id, student_id, question_key, session_key, week_num,
                    presenting_team_id, presenting_team_name, rater_team_id,
                    rater_team_name, q1_developed, q2_easy, data_version)
                   VALUES (?, ?, ?, ?, 1, ?, ?, ?, 'Team 2', ?, ?, ?)""",
                (
                    env["course_id"], env["students"][student_id], key,
                    SESSION_KEY, team_id, team_name,
                    env["teams"]["Team 2"], score, score, version,
                ),
            )
        for key, student_id, challenger_id, challenger_name, score, version in (
            ("rank-ch-current", "s3", "s5", "Eve", 2, "1.1.8"),
            ("rank-ch-old", "s4", "s1", "Alice", 5, "0.9.9"),
        ):
            db.execute(
                """INSERT INTO challenge_ratings
                   (course_id, session_key, week_num, challenge_key,
                    presentation_key, challenger_id, challenger_name,
                    rater_id, rater_name, rater_team_id, rater_team_name,
                    score, data_version)
                   VALUES (?, ?, 1, ?, ?, ?, ?, ?, ?, ?, 'Team 2', ?, ?)""",
                (
                    env["course_id"], SESSION_KEY, key, key,
                    env["students"][challenger_id], challenger_name,
                    env["students"][student_id], student_id,
                    env["teams"]["Team 2"], score, version,
                ),
            )
        db.execute(
            """UPDATE course_state
               SET phase = 'ended', presentation_history = ?
               WHERE course_id = ?""",
            (json.dumps(history), env["course_id"]),
        )
        db.commit()

    instructor_state = _instructor_client(env).get("/api/state").get_json()
    assert instructor_state["completed_presentation_count"] == 1
    assert {
        item["presentation_key"]
        for item in instructor_state["presentation_history"]
    } == {"rank-current"}

    instructor_html = _instructor_client(env).get(
        f"/instructor/{env['slug']}"
    ).get_data(as_text=True)
    assert (
        "Students who submitted at least one response this session:</strong> 1"
        in instructor_html
    )
    assert "#1 Team 1:" in instructor_html
    assert "#1 Team 2:" not in instructor_html
    assert "#1 Eve:" in instructor_html
    assert "#1 Alice:" not in instructor_html

    student_results = _student_client(env, "s1").get("/api/poll").get_json()
    assert student_results["top_teams"] == [{"name": "Team 1", "rank": 1}]
    assert student_results["top_challengers"] == [
        {"name": "Eve", "rank": 1}
    ]

def test_integrity_guards_still_detect_incompatible_feedback_and_history(
        versioned_course_env):
    env = versioned_course_env
    incompatible_history = json.dumps([{
        "presentation_key": "guard-old",
        "session_key": SESSION_KEY,
        "week_num": 1,
        "data_version": "0.9.9",
    }])
    with _connect(env) as db:
        db.execute(
            """INSERT INTO teammate_thumbs
               (course_id, session_key, week_num, question_key,
                source_question_key, grader_id, recipient_id, data_version)
               VALUES (?, ?, 1, 'guard-old', 'guard-old', ?, ?, '0.9.9')""",
            (
                env["course_id"], SESSION_KEY,
                env["students"]["s1"], env["students"]["s2"],
            ),
        )
        db.commit()
        assert app_module._current_session_has_durable_activity(
            db, env["course_id"], SESSION_KEY
        )
        assert app_module._session_has_week_scoped_activity(
            db, env["course_id"], SESSION_KEY, "[]"
        )

        db.execute("DELETE FROM teammate_thumbs")
        db.execute(
            "UPDATE course_state SET presentation_history = ?",
            (incompatible_history,),
        )
        db.commit()
        assert app_module._current_session_has_durable_activity(
            db, env["course_id"], SESSION_KEY
        )
        assert app_module._session_has_week_scoped_activity(
            db, env["course_id"], SESSION_KEY, incompatible_history
        )


def test_current_challenge_rating_export_does_not_borrow_legacy_round_metadata(
        versioned_course_env):
    from openpyxl import load_workbook

    env = versioned_course_env
    challenge_key = "mixed-parent-version"
    with _connect(env) as db:
        db.execute(
            """INSERT INTO challenge_rounds
               (course_id, session_key, week_num, presentation_key,
                challenge_key, challenge_num, challenger_id, challenger_name,
                challenger_team_id, challenger_team_name, presenting_team_id,
                presenting_team_name, question_id, question_title,
                data_version)
               VALUES (?, ?, 1, 'mixed-presentation', ?, 4, ?, 'Eve', ?,
                       'Team 3', ?, 'Legacy Team', ?, 'Legacy Question',
                       '0.9.9')""",
            (
                env["course_id"], SESSION_KEY, challenge_key,
                env["students"]["s5"], env["teams"]["Team 3"],
                env["teams"]["Team 1"], env["question_id"],
            ),
        )
        db.execute(
            """INSERT INTO challenge_ratings
               (course_id, session_key, week_num, challenge_key,
                presentation_key, challenger_id, challenger_name,
                challenger_team_id, challenger_team_name, rater_id,
                rater_name, rater_team_id, rater_team_name, score,
                data_version)
               VALUES (?, ?, 1, ?, 'mixed-presentation', ?, 'Eve', ?,
                       'Team 3', ?, 'Cara', ?, 'Team 2', 4, '1.1.7')""",
            (
                env["course_id"], SESSION_KEY, challenge_key,
                env["students"]["s5"], env["teams"]["Team 3"],
                env["students"]["s3"], env["teams"]["Team 2"],
            ),
        )
        db.execute("UPDATE course_state SET phase = 'ended'")
        db.commit()

    response = _instructor_client(env).get(f"/export/{env['slug']}")
    assert response.status_code == 200
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        workbook = load_workbook(
            io.BytesIO(archive.read("course_data.xlsx")), read_only=True
        )
    assert _workbook_rows(workbook, "Challenge Rounds") == []
    rating_rows = _workbook_rows(workbook, "Challenge Ratings")
    assert len(rating_rows) == 1
    assert rating_rows[0]["challenge_key"] == challenge_key
    assert rating_rows[0]["challenge_number"] is None
    assert rating_rows[0]["presenting_team"] is None
    assert rating_rows[0]["question_title"] is None


def test_challenge_export_joins_require_matching_snapshot_identity(
        versioned_course_env):
    from openpyxl import load_workbook

    env = versioned_course_env
    scoped_key = "scoped-challenge"
    legacy_parent_key = "legacy-parent-current-rating"
    scoped_presentation = "scoped-presentation"

    with _connect(env) as db:
        def add_round(key, *, week, presentation, number, title):
            db.execute(
                """INSERT INTO challenge_rounds
                   (course_id, session_key, week_num, presentation_key,
                    challenge_key, challenge_num, challenger_id,
                    challenger_name, challenger_team_id, challenger_team_name,
                    presenting_team_id, presenting_team_name, question_id,
                    question_title, data_version)
                   VALUES (?, ?, ?, ?, ?, ?, ?, 'Eve', ?, 'Team 3', ?,
                           'Team 1', ?, ?, '1.1.7')""",
                (
                    env["course_id"], SESSION_KEY, week, presentation, key,
                    number, env["students"]["s5"], env["teams"]["Team 3"],
                    env["teams"]["Team 1"], env["question_id"], title,
                ),
            )

        def add_rating(
                key, *, rater, score, week=1, session_key=SESSION_KEY,
                presentation=scoped_presentation, challenger="s5"):
            db.execute(
                """INSERT INTO challenge_ratings
                   (course_id, session_key, week_num, challenge_key,
                    presentation_key, challenger_id, rater_id, score,
                    data_version)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, '1.1.8')""",
                (
                    env["course_id"], session_key, week, key, presentation,
                    env["students"][challenger], env["students"][rater], score,
                ),
            )

        add_round(
            scoped_key, week=1, presentation=scoped_presentation, number=7,
            title="Scoped Question",
        )
        add_rating(scoped_key, rater="s1", score=4)
        add_rating(scoped_key, rater="s2", score=5, week=0)
        add_rating(
            scoped_key, rater="s3", score=1,
            session_key=SESSION_KEY + 1,
        )
        add_rating(
            scoped_key, rater="s4", score=2,
            presentation="other-presentation",
        )
        add_rating(
            scoped_key, rater="s5", score=3, challenger="s4",
        )

        add_round(
            legacy_parent_key, week=0, presentation="legacy-presentation",
            number=8, title="Legacy Question",
        )
        add_rating(
            legacy_parent_key, rater="s2", score=5,
            presentation="legacy-presentation",
        )
        db.execute("UPDATE course_state SET phase = 'ended'")
        db.commit()

    client = _instructor_client(env)
    response = client.get(f"/export/{env['slug']}")
    assert response.status_code == 200
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        workbook = load_workbook(
            io.BytesIO(archive.read("course_data.xlsx")), read_only=True
        )

    round_rows = _workbook_rows(workbook, "Challenge Rounds")
    assert len(round_rows) == 1
    assert round_rows[0]["challenge_key"] == scoped_key
    assert round_rows[0]["ratings_submitted"] == 1
    assert round_rows[0]["average_score_1to5"] == 4

    rating_rows = _workbook_rows(workbook, "Challenge Ratings")
    scoped_rows = {
        row["rater_id"]: row
        for row in rating_rows
        if row["challenge_key"] == scoped_key
    }
    assert set(scoped_rows) == {"s1", "s3", "s4", "s5"}
    assert scoped_rows["s1"]["challenge_number"] == 7
    assert scoped_rows["s1"]["presenting_team"] == "Team 1"
    assert scoped_rows["s1"]["question_title"] == "Scoped Question"
    for rater_id in ("s3", "s4", "s5"):
        assert scoped_rows[rater_id]["challenge_number"] is None
        assert scoped_rows[rater_id]["presenting_team"] is None
        assert scoped_rows[rater_id]["question_title"] is None

    legacy_parent_rating = next(
        row for row in rating_rows
        if row["challenge_key"] == legacy_parent_key
    )
    assert legacy_parent_rating["challenge_number"] is None
    assert legacy_parent_rating["presenting_team"] is None
    assert legacy_parent_rating["question_title"] is None

    legacy = client.get(f"/export/{env['slug']}/legacy-feedback.csv")
    assert legacy.status_code == 200
    legacy_rows = list(csv.DictReader(io.StringIO(
        legacy.data.decode("utf-8-sig")
    )))
    assert {
        (row["record_type"], row["challenge_key"], row["legacy_reason"])
        for row in legacy_rows
        if row["challenge_key"] in {scoped_key, legacy_parent_key}
    } == {
        ("challenge_rating", scoped_key, "unknown_week"),
        ("challenge_round", legacy_parent_key, "unknown_week"),
    }
