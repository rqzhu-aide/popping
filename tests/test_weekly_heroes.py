"""Regression coverage for durable Weekly Hero summaries and badges."""

from contextlib import contextmanager
import io
import json
from pathlib import Path
import shutil
import sqlite3
import subprocess
import sys
import uuid
import zipfile

import pytest


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import app as app_module  # noqa: E402
import config  # noqa: E402
import database  # noqa: E402
from versioning import (  # noqa: E402
    APP_VERSION,
    BASELINE_SCHEMA_VERSION,
    SCHEMA_VERSION,
    SCHEMA_VERSION_HISTORY,
)


SESSION_KEY = 7
V1_1 = "1.1.0"
V1_2 = "1.2.0"
JS_HARNESS = (
    Path(__file__).resolve().parent / "js" / "weekly_hero_badges_harness.js"
)


def _baseline_connection(path=":memory:"):
    db = sqlite3.connect(path)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys=ON")
    db.executescript(
        (PROJECT_ROOT / "popping.sql").read_text(encoding="utf-8")
    )
    return db


def _upgrade_through(db, target):
    current = BASELINE_SCHEMA_VERSION
    for next_version in SCHEMA_VERSION_HISTORY[1:]:
        migration = database._SCHEMA_MIGRATIONS[(current, next_version)]
        migration(db)
        db.execute(
            """INSERT INTO schema_migrations
               (schema_version, applied_by_app_version) VALUES (?, ?)""",
            [next_version, APP_VERSION],
        )
        current = next_version
        if current == target:
            break
    db.commit()


@pytest.fixture
def hero_env(tmp_path, monkeypatch):
    """Create a current-schema course with stable team/student identities."""
    data_dir = tmp_path / "data"
    classes_dir = tmp_path / "classes"
    data_dir.mkdir()
    classes_dir.mkdir()
    slug = f"heroes_{uuid.uuid4().hex[:8]}"
    class_dir = classes_dir / slug
    class_dir.mkdir()
    (class_dir / "course.yaml").write_text(
        "\n".join((
            f"slug: {slug}",
            "name: Weekly Hero Test",
            "code: HERO101",
            "semester: Test",
            "active: true",
            "",
        )),
        encoding="utf-8",
    )
    (class_dir / "week-02-questions.md").write_text(
        "---\ntitle: Hero question\nid: hero-question\n---\n\nDiscuss.\n",
        encoding="utf-8",
    )
    course_dir = data_dir / slug
    course_dir.mkdir()
    db_path = course_dir / "popping.db"

    monkeypatch.setattr(config, "DATA_DIR", str(data_dir))
    monkeypatch.setattr(config, "CLASSES_DIR", str(classes_dir))
    monkeypatch.setattr(config, "CONFIG_DIR", str(classes_dir))
    monkeypatch.setitem(app_module.app.config, "TESTING", True)
    monkeypatch.setitem(
        app_module.app.config, "SECRET_KEY", "weekly-hero-test-key"
    )
    database.forget_schema(slug)
    app_module._clear_course_availability_cache(slug)

    db = _baseline_connection(db_path)
    instructor_id = db.execute(
        """INSERT INTO instructors (username, name, pin)
           VALUES ('instructor', 'Hero Instructor', '9999')"""
    ).lastrowid
    course_id = db.execute(
        """INSERT INTO courses
           (name, code, semester, slug, instructor_id, is_active)
           VALUES ('Weekly Hero Test', 'HERO101', 'Test', ?, ?, 1)""",
        [slug, instructor_id],
    ).lastrowid
    teams = {}
    for number in range(1, 6):
        name = f"Team {number}"
        teams[name] = db.execute(
            "INSERT INTO teams (course_id, name) VALUES (?, ?)",
            [course_id, name],
        ).lastrowid

    # s1 is currently on Team 5 but historically presented for Team 1.
    student_specs = (
        ("s1", "Alice Historical", "1111", teams["Team 5"]),
        ("s2", "Bob", "2222", teams["Team 2"]),
        ("s3", "Cara", "3333", teams["Team 3"]),
        ("s4", "Dan", "4444", teams["Team 4"]),
        ("s5", "Eve Current Team One", "5555", teams["Team 1"]),
        ("s6", "Faye", "6666", teams["Team 5"]),
        ("s7", "Gus", "7777", teams["Team 5"]),
        ("s8", "Hana", "8888", teams["Team 5"]),
    )
    students = {}
    for student_id, name, pin, team_id in student_specs:
        students[student_id] = db.execute(
            """INSERT INTO students
               (course_id, student_id, name, pin, team_id)
               VALUES (?, ?, ?, ?, ?)""",
            [course_id, student_id, name, pin, team_id],
        ).lastrowid
    question_id = db.execute(
        """INSERT INTO questions
           (course_id, question_num, question_text, title, content, week_num,
            source_key)
           VALUES (?, 1, 'Discuss.', 'Hero question', 'Discuss.', 2,
                   'week-2-q-hero-question')""",
        [course_id],
    ).lastrowid
    db.execute(
        """INSERT INTO course_state
           (course_id, phase, max_teams, max_members_per_team,
            discussion_week, presentation_history, roster_version,
            session_key, active_challenges_json)
           VALUES (?, 'setup', 5, 10, 2, '[]', 0, ?, '[]')""",
        [course_id, SESSION_KEY],
    )
    db.commit()
    database.upgrade_schema_connection(db)
    db.commit()
    db.close()

    env = {
        "slug": slug,
        "data_dir": data_dir,
        "db_path": db_path,
        "course_id": course_id,
        "instructor_id": instructor_id,
        "teams": teams,
        "students": students,
        "question_id": question_id,
    }
    yield env

    database.forget_schema(slug)
    app_module._clear_course_availability_cache(slug)


@contextmanager
def _connect(env):
    db = sqlite3.connect(env["db_path"])
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys=ON")
    try:
        yield db
    finally:
        db.close()


def _instructor_client(env):
    client = app_module.app.test_client()
    with client.session_transaction() as browser_session:
        browser_session["role"] = "instructor"
        browser_session["instructor_id"] = env["instructor_id"]
        browser_session["slug"] = env["slug"]
        browser_session["instructor_auth_token"] = (
            app_module._instructor_session_token(
                env["slug"], env["instructor_id"], "9999"
            )
        )
    return client


def _student_client(env, student_identifier="s1"):
    with _connect(env) as db:
        student = db.execute(
            """SELECT id, student_id, pin FROM students
               WHERE course_id = ? AND student_id = ?""",
            [env["course_id"], student_identifier],
        ).fetchone()
    client = app_module.app.test_client()
    with client.session_transaction() as browser_session:
        browser_session["role"] = "student"
        browser_session["student_id"] = student["student_id"]
        browser_session["name"] = student["student_id"]
        browser_session["slug"] = env["slug"]
        browser_session["student_auth_token"] = (
            app_module._student_session_token(
                env["slug"], student["id"], student["student_id"],
                student["pin"],
            )
        )
    return client


def _set_state(env, **fields):
    assert fields
    assignments = ", ".join(f"{column} = ?" for column in fields)
    with _connect(env) as db:
        db.execute(
            f"UPDATE course_state SET {assignments} WHERE course_id = ?",
            [*fields.values(), env["course_id"]],
        )
        db.commit()


def _seed_ranked_week(env, *, week=1, data_version=SCHEMA_VERSION):
    """Seed four team results and three challenger results for one week."""
    team_scores = {
        "Team 1": 5,
        "Team 2": 4,
        "Team 3": 3,
        "Team 4": 3,
    }
    historical_members = {
        "Team 1": "s1",
        "Team 2": "s2",
        "Team 3": "s3",
        "Team 4": "s4",
    }
    with _connect(env) as db:
        for number, (team_name, score) in enumerate(team_scores.items(), 1):
            presentation_key = f"week-{week}-team-{number}"
            team_id = env["teams"][team_name]
            db.execute(
                """INSERT INTO presentation_ratings
                   (course_id, student_id, question_key, session_key,
                    week_num, presenting_team_id, presenting_team_name,
                    rater_team_id, rater_team_name, data_version,
                    q1_developed, q2_easy)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Team 5', ?, ?, ?)""",
                [
                    env["course_id"], env["students"]["s8"],
                    presentation_key, SESSION_KEY, week, team_id, team_name,
                    env["teams"]["Team 5"], data_version, score, score,
                ],
            )
            member_identifier = historical_members[team_name]
            db.execute(
                """INSERT INTO presentation_participants
                   (course_id, session_key, week_num, presentation_key,
                    student_id, student_identifier, student_name,
                    team_id, team_name, data_version)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                [
                    env["course_id"], SESSION_KEY, week, presentation_key,
                    env["students"][member_identifier], member_identifier,
                    f"Historical {member_identifier}", team_id, team_name,
                    data_version,
                ],
            )

        for number, (challenger, score) in enumerate(
                (("s6", 5), ("s7", 5), ("s8", 4)), 1):
            db.execute(
                """INSERT INTO challenge_ratings
                   (course_id, session_key, week_num, challenge_key,
                    presentation_key, challenger_id, challenger_name,
                    challenger_team_id, challenger_team_name, rater_id,
                    rater_name, rater_team_id, rater_team_name,
                    data_version, score)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Team 5', ?, 'Bob', ?,
                           'Team 2', ?, ?)""",
                [
                    env["course_id"], SESSION_KEY, week,
                    f"week-{week}-challenge-{number}",
                    f"week-{week}-team-1", env["students"][challenger],
                    f"Challenge {challenger}", env["teams"]["Team 5"],
                    env["students"]["s2"], env["teams"]["Team 2"],
                    data_version, score,
                ],
            )
        db.commit()


def _insert_summary(db, env, week):
    return db.execute(
        """INSERT INTO weekly_hero_summaries
           (course_id, week_num, calculation_version,
            source_schema_version, source_data_versions,
            source_fingerprint, source_presentation_rating_count,
            source_challenge_rating_count, source_participant_count,
            source_history_item_count, data_version)
           VALUES (?, ?, ?, ?, '[]', ?, 0, 0, 0, 0, ?)""",
        [
            env["course_id"], week,
            database.WEEKLY_HERO_CALCULATION_VERSION, SCHEMA_VERSION,
            f"{week:064x}", APP_VERSION,
        ],
    ).lastrowid


def _insert_award(
        db, env, summary_id, *, award_type, student="s1", rank=1,
        result_suffix=None):
    result_suffix = result_suffix or f"{award_type}-{student}"
    if award_type == "bolt":
        result_id = db.execute(
            """INSERT INTO weekly_hero_results
               (summary_id, result_key, category, award_type, rank,
                score_sum, score_count, rating_count, team_id, team_name,
                challenger_id, challenger_identifier, challenger_name)
               VALUES (?, ?, 'challenger', 'bolt', 1, 5, 1, 1,
                       NULL, NULL, ?, ?, ?)""",
            [
                summary_id, f"challenger:{result_suffix}",
                env["students"][student], student, f"Hero {student}",
            ],
        ).lastrowid
        team_id = None
        team_name = None
    else:
        team_name = f"Team {rank}"
        team_id = env["teams"][team_name]
        result_id = db.execute(
            """INSERT INTO weekly_hero_results
               (summary_id, result_key, category, award_type, rank,
                score_sum, score_count, rating_count,
                developed_score_sum, developed_score_count,
                easy_score_sum, easy_score_count, team_id, team_name,
                challenger_id, challenger_identifier, challenger_name)
               VALUES (?, ?, 'team', ?, ?, 10, 2, 1, 5, 1, 5, 1, ?, ?,
                       NULL, NULL, NULL)""",
            [
                summary_id, f"team:{result_suffix}", award_type, rank,
                team_id, team_name,
            ],
        ).lastrowid
    db.execute(
        """INSERT INTO weekly_hero_recipients
           (result_id, recipient_key, student_id, student_identifier,
            student_name, team_id, team_name)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        [
            result_id, f"student:{student}", env["students"][student],
            student, f"Hero {student}", team_id, team_name,
        ],
    )
    return result_id


def _raw_source_snapshot(db, course_id, week):
    snapshot = {}
    for table in (
            "presentation_ratings", "challenge_ratings",
            "presentation_participants"):
        rows = db.execute(
            f"SELECT * FROM {table} WHERE course_id = ? AND week_num = ? "
            "ORDER BY id",
            [course_id, week],
        ).fetchall()
        snapshot[table] = [tuple(row) for row in rows]
    snapshot["presentation_history"] = db.execute(
        "SELECT presentation_history FROM course_state WHERE course_id = ?",
        [course_id],
    ).fetchone()[0]
    return snapshot


def test_v1_2_to_v1_3_migration_adds_and_validates_weekly_hero_tables():
    db = _baseline_connection()
    try:
        _upgrade_through(db, V1_2)
        assert database.inspect_schema_version(db) == V1_2
        assert not db.execute(
            """SELECT 1 FROM sqlite_master
               WHERE type = 'table' AND name = 'weekly_hero_summaries'"""
        ).fetchone()

        assert database.upgrade_schema_connection(db) == SCHEMA_VERSION
        assert database.validate_current_schema(db) == SCHEMA_VERSION
        assert [
            row["schema_version"] for row in db.execute(
                "SELECT schema_version FROM schema_migrations ORDER BY id"
            )
        ] == list(SCHEMA_VERSION_HISTORY)
        tables = {
            row[0] for row in db.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            )
        }
        assert {
            "weekly_hero_summaries",
            "weekly_hero_results",
            "weekly_hero_recipients",
        }.issubset(tables)

        db.execute("DROP TABLE weekly_hero_recipients")
        with pytest.raises(
            RuntimeError,
            match="weekly_hero_recipients is missing required column",
        ):
            database.validate_current_schema(db)
    finally:
        db.close()


def test_explicit_v1_2_backfill_preserves_every_source_row(hero_env):
    _seed_ranked_week(hero_env, data_version="1.2.7")
    with _connect(hero_env) as db:
        before = _raw_source_snapshot(db, hero_env["course_id"], 1)
        assert database.detect_weekly_hero_source_schema_version(
            db, hero_env["course_id"], 1
        ) == V1_2
        preview = database.calculate_weekly_hero_preview(
            db,
            hero_env["course_id"],
            1,
            source_schema_version=V1_2,
        )
        assert preview["source_data_versions"] == ["1.2.7"]

        outcome = database.save_weekly_hero_summary(db, preview)
        db.commit()

        assert outcome["status"] == "created"
        assert _raw_source_snapshot(db, hero_env["course_id"], 1) == before
        assert tuple(db.execute(
            """SELECT source_schema_version, data_version
               FROM weekly_hero_summaries"""
        ).fetchone()) == (V1_2, APP_VERSION)


def test_weekly_preview_uses_competition_ranks_and_historical_members(
        hero_env):
    _seed_ranked_week(hero_env)
    with _connect(hero_env) as db:
        preview = database.calculate_weekly_hero_preview(
            db, hero_env["course_id"], 1
        )

    teams = [
        result for result in preview["results"]
        if result["category"] == "team"
    ]
    assert [
        (result["team_name"], result["rank"], result["award_type"])
        for result in teams
    ] == [
        ("Team 1", 1, "gold"),
        ("Team 2", 2, "silver"),
        ("Team 3", 3, "bronze"),
        ("Team 4", 3, "bronze"),
    ]
    gold = teams[0]
    assert [recipient["student_identifier"] for recipient in gold["recipients"]] == [
        "s1"
    ]
    assert gold["recipients"][0]["team_name"] == "Team 1"
    # s5 is currently on Team 1, but did not belong to the historical
    # presentation snapshot and therefore must not receive the medal.
    assert all(
        recipient["student_identifier"] != "s5"
        for result in teams for recipient in result["recipients"]
    )


def test_weekly_preview_awards_all_first_place_challenger_ties(hero_env):
    _seed_ranked_week(hero_env)
    with _connect(hero_env) as db:
        preview = database.calculate_weekly_hero_preview(
            db, hero_env["course_id"], 1
        )

    challengers = [
        result for result in preview["results"]
        if result["category"] == "challenger"
    ]
    assert [
        (result["challenger_identifier"], result["rank"],
         result["award_type"])
        for result in challengers
    ] == [("s6", 1, "bolt"), ("s7", 1, "bolt")]


def test_fingerprint_ignores_profile_names_shadowed_by_recorded_name(
        hero_env):
    _seed_ranked_week(hero_env)
    with _connect(hero_env) as db:
        original = database.calculate_weekly_hero_preview(
            db, hero_env["course_id"], 1
        )
        created = database.save_weekly_hero_summary(db, original)

        db.execute(
            """UPDATE students
               SET name = 'Updated uploaded name',
                   display_name = 'Updated display name'
               WHERE id = ?""",
            [hero_env["students"]["s6"]],
        )
        after_profile_edit = database.calculate_weekly_hero_preview(
            db, hero_env["course_id"], 1
        )
        repeated = database.save_weekly_hero_summary(db, original)

    assert after_profile_edit["source_fingerprint"] == (
        original["source_fingerprint"]
    )
    assert after_profile_edit["results"] == original["results"]
    assert repeated["status"] == "unchanged"
    assert repeated["summary_id"] == created["summary_id"]


def test_fingerprint_tracks_challenger_profile_name_only_when_used(
        hero_env):
    _seed_ranked_week(hero_env)
    challenger_id = hero_env["students"]["s6"]
    with _connect(hero_env) as db:
        db.execute(
            """UPDATE challenge_ratings SET challenger_name = NULL
               WHERE challenger_id = ?""",
            [challenger_id],
        )
        db.execute(
            """UPDATE students
               SET name = 'Uploaded fallback', display_name = 'First choice'
               WHERE id = ?""",
            [challenger_id],
        )
        original = database.calculate_weekly_hero_preview(
            db, hero_env["course_id"], 1
        )

        db.execute(
            "UPDATE students SET name = 'Unused uploaded edit' WHERE id = ?",
            [challenger_id],
        )
        unused_name_edit = database.calculate_weekly_hero_preview(
            db, hero_env["course_id"], 1
        )
        assert unused_name_edit["source_fingerprint"] == (
            original["source_fingerprint"]
        )

        db.execute(
            "UPDATE students SET display_name = 'Second choice' WHERE id = ?",
            [challenger_id],
        )
        changed = database.calculate_weekly_hero_preview(
            db, hero_env["course_id"], 1
        )

    assert changed["source_fingerprint"] != original["source_fingerprint"]
    challenger = next(
        result for result in changed["results"]
        if result["category"] == "challenger"
        and result["challenger_id"] == challenger_id
    )
    assert challenger["challenger_name"] == "Second choice"
    assert challenger["recipients"][0]["student_name"] == "Second choice"


def test_fingerprint_ignores_fallback_profile_of_non_awarded_challenger(
        hero_env):
    _seed_ranked_week(hero_env)
    challenger_id = hero_env["students"]["s8"]
    with _connect(hero_env) as db:
        db.execute(
            """UPDATE challenge_ratings SET challenger_name = NULL
               WHERE challenger_id = ?""",
            [challenger_id],
        )
        db.execute(
            "UPDATE students SET display_name = 'First choice' WHERE id = ?",
            [challenger_id],
        )
        original = database.calculate_weekly_hero_preview(
            db, hero_env["course_id"], 1
        )

        db.execute(
            "UPDATE students SET display_name = 'Second choice' WHERE id = ?",
            [challenger_id],
        )
        changed = database.calculate_weekly_hero_preview(
            db, hero_env["course_id"], 1
        )

    assert changed["results"] == original["results"]
    assert changed["source_fingerprint"] == original["source_fingerprint"]


def test_fingerprint_tracks_stable_challenge_evidence_key(hero_env):
    _seed_ranked_week(hero_env)
    with _connect(hero_env) as db:
        original = database.calculate_weekly_hero_preview(
            db, hero_env["course_id"], 1
        )
        db.execute(
            """UPDATE challenge_ratings
               SET challenge_key = challenge_key || '-corrected'
               WHERE challenger_id = ?""",
            [hero_env["students"]["s8"]],
        )
        changed = database.calculate_weekly_hero_preview(
            db, hero_env["course_id"], 1
        )

    assert changed["results"] == original["results"]
    assert changed["source_fingerprint"] != original["source_fingerprint"]


def test_summary_save_is_idempotent_and_replace_is_explicit(hero_env):
    _seed_ranked_week(hero_env)
    with _connect(hero_env) as db:
        original = database.calculate_weekly_hero_preview(
            db, hero_env["course_id"], 1
        )
        created = database.save_weekly_hero_summary(db, original)
        repeated = database.save_weekly_hero_summary(db, original)
        assert created["status"] == "created"
        assert repeated == {
            "status": "unchanged",
            "summary_id": created["summary_id"],
            "preview": repeated["preview"],
        }
        assert db.execute(
            "SELECT COUNT(*) FROM weekly_hero_summaries"
        ).fetchone()[0] == 1

        db.execute(
            """UPDATE presentation_ratings SET q1_developed = 4
               WHERE question_key = 'week-1-team-1'"""
        )
        changed = database.calculate_weekly_hero_preview(
            db, hero_env["course_id"], 1
        )
        assert changed["source_fingerprint"] != original["source_fingerprint"]
        with pytest.raises(
            RuntimeError, match="explicitly allow replacement"
        ):
            database.save_weekly_hero_summary(db, changed)

        replaced = database.save_weekly_hero_summary(
            db, changed, replace=True
        )
        db.commit()
        assert replaced["status"] == "replaced"
        assert replaced["summary_id"] != created["summary_id"]
        assert db.execute(
            "SELECT COUNT(*) FROM weekly_hero_summaries"
        ).fetchone()[0] == 1
        assert not db.execute(
            "SELECT 1 FROM weekly_hero_results WHERE summary_id = ?",
            [created["summary_id"]],
        ).fetchone()


def test_missing_awarded_team_participant_coverage_refuses_save(hero_env):
    _seed_ranked_week(hero_env)
    with _connect(hero_env) as db:
        db.execute(
            """DELETE FROM presentation_participants
               WHERE presentation_key = 'week-1-team-1'"""
        )
        preview = database.calculate_weekly_hero_preview(
            db, hero_env["course_id"], 1
        )
        assert preview["recipient_coverage_complete"] is False
        assert preview["missing_participant_presentations"] == [{
            "result_key": f"team-id:{hero_env['teams']['Team 1']}",
            "team_id": hero_env["teams"]["Team 1"],
            "team_name": "Team 1",
            "presentation_key": "week-1-team-1",
        }]
        with pytest.raises(RuntimeError, match="lack participant snapshots"):
            database.save_weekly_hero_summary(db, preview)
        assert db.execute(
            "SELECT COUNT(*) FROM weekly_hero_summaries"
        ).fetchone()[0] == 0


def test_badge_counts_respect_prior_week_and_student_filters(hero_env):
    with _connect(hero_env) as db:
        week_1 = _insert_summary(db, hero_env, 1)
        _insert_award(db, hero_env, week_1, award_type="gold", student="s1")
        _insert_award(db, hero_env, week_1, award_type="bolt", student="s1")
        _insert_award(
            db, hero_env, week_1, award_type="bronze", student="s2", rank=3
        )
        week_2 = _insert_summary(db, hero_env, 2)
        _insert_award(
            db, hero_env, week_2, award_type="silver", student="s1", rank=2
        )
        week_3 = _insert_summary(db, hero_env, 3)
        _insert_award(db, hero_env, week_3, award_type="gold", student="s1")
        db.commit()

        assert database.get_weekly_hero_badge_counts(
            db,
            hero_env["course_id"],
            before_week=3,
            student_ids=[hero_env["students"]["s1"]],
        ) == {
            hero_env["students"]["s1"]: {
                "bolt": 1,
                "gold": 1,
                "silver": 1,
            }
        }
        assert database.get_weekly_hero_badge_counts(
            db,
            hero_env["course_id"],
            before_week=2,
            student_ids=[hero_env["students"]["s2"]],
        ) == {hero_env["students"]["s2"]: {"bronze": 1}}
        assert database.get_weekly_hero_badge_counts(
            db,
            hero_env["course_id"],
            before_week=1,
            student_ids=[hero_env["students"]["s1"]],
        ) == {}
        assert database.get_weekly_hero_badge_counts(
            db, hero_env["course_id"], student_ids=[]
        ) == {}


def test_end_session_refuses_mixed_sources_and_preserves_v1_2_summary(
        hero_env):
    _seed_ranked_week(hero_env, data_version="1.2.7")
    with _connect(hero_env) as db:
        legacy_preview = database.calculate_weekly_hero_preview(
            db,
            hero_env["course_id"],
            1,
            source_schema_version=V1_2,
        )
        saved = database.save_weekly_hero_summary(db, legacy_preview)
        db.commit()
        summary_before = dict(db.execute(
            "SELECT * FROM weekly_hero_summaries WHERE id = ?",
            [saved["summary_id"]],
        ).fetchone())
        results_before = [
            tuple(row) for row in db.execute(
                """SELECT * FROM weekly_hero_results
                   WHERE summary_id = ? ORDER BY id""",
                [saved["summary_id"]],
            )
        ]
        recipients_before = [
            tuple(row) for row in db.execute(
                """SELECT recipient.* FROM weekly_hero_recipients recipient
                   JOIN weekly_hero_results result
                     ON result.id = recipient.result_id
                   WHERE result.summary_id = ? ORDER BY recipient.id""",
                [saved["summary_id"]],
            )
        ]

        db.execute(
            """INSERT INTO presentation_ratings
               (course_id, student_id, question_key, session_key, week_num,
                presenting_team_id, presenting_team_name, rater_team_id,
                rater_team_name, data_version, q1_developed, q2_easy)
               VALUES (?, ?, 'week-1-current-series', ?, 1, ?, 'Team 5', ?,
                       'Team 2', ?, 2, 2)""",
            [
                hero_env["course_id"], hero_env["students"]["s2"],
                SESSION_KEY, hero_env["teams"]["Team 5"],
                hero_env["teams"]["Team 2"], SCHEMA_VERSION,
            ],
        )
        db.execute(
            """INSERT INTO presentation_participants
               (course_id, session_key, week_num, presentation_key,
                student_id, student_identifier, student_name,
                team_id, team_name, data_version)
               VALUES (?, ?, 1, 'week-1-current-series', ?, 's5',
                       'Current s5', ?, 'Team 5', ?)""",
            [
                hero_env["course_id"], SESSION_KEY,
                hero_env["students"]["s5"],
                hero_env["teams"]["Team 5"], SCHEMA_VERSION,
            ],
        )
        db.execute(
            """UPDATE course_state
               SET phase = 'competition', discussion_week = 1
               WHERE course_id = ?""",
            [hero_env["course_id"]],
        )
        db.commit()

    state_before = None
    with _connect(hero_env) as db:
        state_before = dict(db.execute(
            """SELECT phase, session_key, roster_version
               FROM course_state WHERE course_id = ?""",
            [hero_env["course_id"]],
        ).fetchone())

    response = _instructor_client(hero_env).post(
        "/api/set_phase",
        json={
            "phase": "ended",
            "expected_phase": "competition",
            "expected_session_key": SESSION_KEY,
            "presentation_key": "",
            "confirm_end_session": True,
        },
    )

    assert response.status_code == 409
    message = response.get_json()["error"].casefold()
    assert "weekly hero" in message
    assert any(fragment in message for fragment in (
        "mixed", "multiple compatibility", "more than one data version",
    ))
    assert "backfill" in message
    with _connect(hero_env) as db:
        state_after = dict(db.execute(
            """SELECT phase, session_key, roster_version
               FROM course_state WHERE course_id = ?""",
            [hero_env["course_id"]],
        ).fetchone())
        summary_after = dict(db.execute(
            "SELECT * FROM weekly_hero_summaries WHERE id = ?",
            [saved["summary_id"]],
        ).fetchone())
        results_after = [
            tuple(row) for row in db.execute(
                """SELECT * FROM weekly_hero_results
                   WHERE summary_id = ? ORDER BY id""",
                [saved["summary_id"]],
            )
        ]
        recipients_after = [
            tuple(row) for row in db.execute(
                """SELECT recipient.* FROM weekly_hero_recipients recipient
                   JOIN weekly_hero_results result
                     ON result.id = recipient.result_id
                   WHERE result.summary_id = ? ORDER BY recipient.id""",
                [saved["summary_id"]],
            )
        ]
        assert db.execute(
            "SELECT COUNT(*) FROM weekly_hero_summaries"
        ).fetchone()[0] == 1
    assert state_after == state_before
    assert summary_after == summary_before
    assert results_after == results_before
    assert recipients_after == recipients_before


def test_end_session_rolls_back_summary_if_transition_fails(
        hero_env, monkeypatch):
    _set_state(hero_env, phase="competition")
    original_save = database.save_weekly_hero_summary

    def save_then_fail(db, preview, replace=False):
        original_save(db, preview, replace=replace)
        raise RuntimeError("forced post-summary transition failure")

    monkeypatch.setattr(app_module, "save_weekly_hero_summary", save_then_fail)
    response = _instructor_client(hero_env).post(
        "/api/set_phase",
        json={
            "phase": "ended",
            "expected_phase": "competition",
            "expected_session_key": SESSION_KEY,
            "presentation_key": "",
            "confirm_end_session": True,
        },
    )
    assert response.status_code == 500
    assert response.get_json() == {
        "error": "Something went wrong. Please try again."
    }

    with _connect(hero_env) as db:
        state = db.execute(
            "SELECT phase, roster_version FROM course_state"
        ).fetchone()
        assert dict(state) == {"phase": "competition", "roster_version": 0}
        assert db.execute(
            "SELECT COUNT(*) FROM weekly_hero_summaries"
        ).fetchone()[0] == 0


def test_reset_course_data_deletes_summary_and_cascaded_awards(
        hero_env, monkeypatch):
    with _connect(hero_env) as db:
        summary = _insert_summary(db, hero_env, 1)
        _insert_award(db, hero_env, summary, award_type="gold", student="s1")
        db.commit()
    monkeypatch.setattr(
        app_module, "_create_reset_backup", lambda _slug, **_kwargs: "backup.db"
    )
    monkeypatch.setattr(app_module, "_prune_reset_backups", lambda _slug: None)

    response = _instructor_client(hero_env).post(
        "/api/reset_data",
        json={
            "confirm_slug": hero_env["slug"],
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
            "presentation_key": "",
        },
    )

    assert response.status_code == 200
    with _connect(hero_env) as db:
        for table in (
                "weekly_hero_summaries", "weekly_hero_results",
                "weekly_hero_recipients"):
            assert db.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0] == 0


def test_weekly_export_has_exact_weekly_hero_tab_without_pin(hero_env):
    from openpyxl import load_workbook

    with _connect(hero_env) as db:
        summary = _insert_summary(db, hero_env, 1)
        _insert_award(db, hero_env, summary, award_type="gold", student="s1")
        db.commit()

    response = _instructor_client(hero_env).get(
        f"/export/{hero_env['slug']}", query_string={"week": 1}
    )

    assert response.status_code == 200
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        workbook = load_workbook(
            io.BytesIO(archive.read("course_data.xlsx")), read_only=True
        )
    assert workbook.sheetnames.count("Weekly Hero") == 1
    assert workbook.sheetnames[1] == "Weekly Hero"
    sheet = workbook["Weekly Hero"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "week", "achievement", "rank", "team_id", "team_name",
        "student_id", "student_name_at_award", "average_score",
        "ratings_submitted", "score_sum", "score_count",
        "calculation_version", "source_schema_version",
        "source_data_versions", "source_fingerprint", "data_version",
        "calculated_at",
    )
    assert rows[1][0:7] == (
        1, "Gold Team", 1, hero_env["teams"]["Team 1"], "Team 1",
        "s1", "Hero s1",
    )
    assert all("pin" not in str(header).casefold() for header in rows[0])
    known_pins = {"1111", "2222", "3333", "4444", "5555", "6666",
                  "7777", "8888", "9999"}
    assert not any(
        str(value) in known_pins
        for row in rows for value in row if value is not None
    )


def test_badges_are_instructor_only_and_competition_only(hero_env):
    with _connect(hero_env) as db:
        summary = _insert_summary(db, hero_env, 1)
        _insert_award(db, hero_env, summary, award_type="gold", student="s1")
        _insert_award(db, hero_env, summary, award_type="bolt", student="s6")
        db.commit()

    instructor = _instructor_client(hero_env)
    for phase in ("setup", "discussion"):
        _set_state(hero_env, phase=phase)
        teams = instructor.get("/api/teams").get_json()
        assert all(
            "hero_badges" not in member
            for team in teams for member in team["members"]
        )
        page = instructor.get(f"/instructor/{hero_env['slug']}")
        assert page.status_code == 200
        assert b'data-role="weekly-hero-badges"' not in page.data

    active_challenge = {
        "challenge_key": "week-2-live-challenge",
        "challenge_num": 1,
        "challenger_id": hero_env["students"]["s6"],
        "challenger_identifier": "s6",
        "challenger_name": "Faye",
        "challenger_team_id": hero_env["teams"]["Team 5"],
        "challenger_team_name": "Team 5",
    }
    _set_state(
        hero_env,
        phase="competition",
        active_team_id=hero_env["teams"]["Team 2"],
        active_question_id=hero_env["question_id"],
        poll_question_key="week-2-live",
        active_challenges_json=json.dumps([active_challenge]),
    )
    instructor_teams = instructor.get("/api/teams").get_json()
    member_by_identifier = {
        member["student_id"]: member
        for team in instructor_teams for member in team["members"]
    }
    assert member_by_identifier["s1"]["hero_badges"] == {"gold": 1}
    assert member_by_identifier["s2"]["hero_badges"] == {}

    instructor_state = instructor.get("/api/state").get_json()
    assert instructor_state["active_challenges"][0]["hero_badges"] == {
        "bolt": 1
    }
    page = instructor.get(f"/instructor/{hero_env['slug']}")
    assert page.status_code == 200
    assert b"Gold team award: 1 week" in page.data

    student = _student_client(hero_env, "s2")
    student_teams = student.get("/api/teams").get_json()
    assert all(
        "hero_badges" not in member
        for team in student_teams for member in team["members"]
    )
    student_state = student.get("/api/state").get_json()
    assert "hero_badges" not in json.dumps(student_state)


def test_badge_markup_is_scoped_to_selected_team_and_active_challenger():
    template = (PROJECT_ROOT / "templates" / "instructor.html").read_text(
        encoding="utf-8"
    )
    script = (PROJECT_ROOT / "static" / "js" / "app.js").read_text(
        encoding="utf-8"
    )
    competition = template.index("{% elif state.phase == 'competition' %}")
    ended = template.index("<!-- ===== ENDED PHASE ===== -->", competition)
    call = "{{ weekly_hero_badges(member) }}"
    assert template.count(call) == 1
    assert competition < template.index(call) < ended
    team_section = template.index(
        'class="team-participation-team"', competition
    )
    assert "hidden" in template[team_section:team_section + 180]

    hands_start = script.index("// Render raised hands")
    active_start = script.index("// Render active challengers", hands_start)
    assert "weeklyHeroBadgesHtml" not in script[hands_start:active_start]
    active_end = script.index("window.selectChallenger", active_start)
    assert "weeklyHeroBadgesHtml(ch)" in script[active_start:active_end]


def test_weekly_hero_badge_javascript_harness():
    node = shutil.which("node")
    if node is None:
        pytest.skip("Node.js is required for the Weekly Hero badge harness")
    result = subprocess.run(
        [node, str(JS_HARNESS)],
        capture_output=True,
        text=True,
        timeout=30,
        cwd=PROJECT_ROOT,
    )
    assert result.returncode == 0, result.stdout + result.stderr
