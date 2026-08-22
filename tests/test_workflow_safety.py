"""Focused regression tests for destructive and live classroom workflows.

These tests intentionally exercise public Flask routes where possible. They use
an isolated SQLite database for every test and never touch checked-in course
data.
"""

from contextlib import contextmanager
from concurrent.futures import ThreadPoolExecutor
import csv
from datetime import datetime, timedelta
import hashlib
import io
import json
from pathlib import Path
import sqlite3
import sys
import uuid
import zipfile

import pytest


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import app as app_module  # noqa: E402
import config  # noqa: E402
import database  # noqa: E402


SESSION_KEY = 7


@pytest.fixture
def course_env(tmp_path, monkeypatch):
    """Create a complete course in a per-test data directory."""
    data_dir = tmp_path / "data"
    classes_dir = tmp_path / "classes"
    data_dir.mkdir()
    classes_dir.mkdir()
    slug = f"safety_{uuid.uuid4().hex[:8]}"
    class_dir = classes_dir / slug
    class_dir.mkdir()
    (class_dir / "course.yaml").write_text(
        "\n".join(
            (
                f"slug: {slug}",
                "name: Workflow Safety",
                "code: SAFE101",
                "semester: Test",
                "active: true",
                "",
            )
        ),
        encoding="utf-8",
    )
    course_dir = data_dir / slug
    course_dir.mkdir()
    db_path = course_dir / "popping.db"

    monkeypatch.setattr(config, "DATA_DIR", str(data_dir))
    monkeypatch.setattr(config, "CLASSES_DIR", str(classes_dir))
    monkeypatch.setattr(config, "CONFIG_DIR", str(classes_dir))
    monkeypatch.setitem(app_module.app.config, "TESTING", True)
    monkeypatch.setitem(app_module.app.config, "SECRET_KEY", "workflow-test-key")
    monkeypatch.setitem(
        app_module.app.config, "MAX_CONTENT_LENGTH", 2 * 1024 * 1024
    )

    database._schema_checked.discard(slug)
    question_cache = getattr(app_module, "_question_html_cache", None)
    if isinstance(question_cache, dict):
        question_cache.clear()

    db = sqlite3.connect(db_path)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys=ON")
    db.executescript(Path(config.DATABASE_SCHEMA).read_text(encoding="utf-8"))

    instructor_id = db.execute(
        "INSERT INTO instructors (username, name, pin) VALUES (?, ?, ?)",
        ("instructor", "Test Instructor", "9999"),
    ).lastrowid
    course_id = db.execute(
        """INSERT INTO courses (name, code, semester, slug, instructor_id)
           VALUES (?, ?, ?, ?, ?)""",
        ("Workflow Safety", "SAFE101", "Test", slug, instructor_id),
    ).lastrowid

    teams = {}
    for number in range(1, 5):
        name = f"Team {number}"
        teams[name] = db.execute(
            "INSERT INTO teams (course_id, name) VALUES (?, ?)",
            (course_id, name),
        ).lastrowid

    student_specs = (
        ("s1", "Alice", "1111", teams["Team 1"]),
        ("s2", "Bob", "2222", teams["Team 1"]),
        ("s3", "Unassigned", "3333", None),
        ("s4", "Dana", "4444", teams["Team 2"]),
    )
    students = {}
    for student_id, name, pin, team_id in student_specs:
        students[student_id] = db.execute(
            """INSERT INTO students
               (course_id, student_id, name, pin, team_id)
               VALUES (?, ?, ?, ?, ?)""",
            (course_id, student_id, name, pin, team_id),
        ).lastrowid

    question_id = db.execute(
        """INSERT INTO questions
           (course_id, question_num, question_text, title, content, week_num,
            source_key)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (
            course_id,
            1,
            "Explain the result.",
            "Question One",
            "Explain the result.",
            1,
            "week-1-question-1",
        ),
    ).lastrowid
    db.execute(
        """INSERT INTO course_state
           (course_id, phase, max_teams, max_members_per_team,
            discussion_week, presentation_history, roster_version, session_key)
           VALUES (?, 'setup', 4, 10, 1, '[]', 0, ?)""",
        (course_id, SESSION_KEY),
    )
    db.commit()
    db.execute("BEGIN IMMEDIATE")
    database.migrate_schema_connection(db)
    db.commit()
    db.close()

    env = {
        "slug": slug,
        "data_dir": data_dir,
        "db_path": db_path,
        "course_id": course_id,
        "instructor_id": instructor_id,
        "teams": teams,
        "students": students,
        "question_id": question_id,
    }
    yield env

    database._schema_checked.discard(slug)


@contextmanager
def _connect(env):
    db = sqlite3.connect(env["db_path"])
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys=ON")
    try:
        yield db
    finally:
        db.close()


def _instructor_client(env):
    client = app_module.app.test_client()
    with client.session_transaction() as flask_session:
        flask_session["role"] = "instructor"
        flask_session["instructor_id"] = env["instructor_id"]
        flask_session["slug"] = env["slug"]
    return client


def _student_client(env, student_id="s1"):
    client = app_module.app.test_client()
    with client.session_transaction() as flask_session:
        flask_session["role"] = "student"
        flask_session["student_id"] = student_id
        flask_session["name"] = student_id
        flask_session["slug"] = env["slug"]
    return client


def _setup_payload(roster_version=0, **extra):
    return {
        "expected_phase": "setup",
        "expected_session_key": SESSION_KEY,
        "expected_roster_version": roster_version,
        **extra,
    }


def _assign_all_active_students(env):
    """Assign active unassigned students and return the new roster version."""
    with _connect(env) as db:
        db.execute(
            '''UPDATE students SET team_id = ?
               WHERE course_id = ? AND is_active = 1 AND team_id IS NULL''',
            [env["teams"]["Team 2"], env["course_id"]],
        )
        db.execute(
            '''UPDATE course_state
               SET roster_version = COALESCE(roster_version, 0) + 1
               WHERE course_id = ?''',
            [env["course_id"]],
        )
        version = db.execute(
            "SELECT roster_version FROM course_state WHERE course_id = ?",
            [env["course_id"]],
        ).fetchone()["roster_version"]
        db.commit()
    return version


def _set_state(env, **fields):
    allowed = {
        "phase",
        "discussion_week",
        "session_key",
        "active_team_id",
        "active_question_id",
        "current_question",
        "presentation_started_at",
        "presentation_created_at",
        "presentation_time_cap",
        "presentation_remaining",
        "poll_active",
        "poll_question_key",
        "poll_started_at",
        "poll_closed_at",
        "challenge_ratings_closed_at",
        "presentation_history",
        "session_started_at",
        "current_discussion_key",
        "current_discussion_source_key",
        "current_discussion_title",
        "current_discussion_content",
    }
    assert fields and set(fields).issubset(allowed)
    assignments = ", ".join(f"{name} = ?" for name in fields)
    with _connect(env) as db:
        db.execute(
            f"UPDATE course_state SET {assignments} WHERE course_id = ?",
            [*fields.values(), env["course_id"]],
        )
        db.commit()


def _state_row(env):
    with _connect(env) as db:
        row = db.execute(
            "SELECT * FROM course_state WHERE course_id = ?",
            (env["course_id"],),
        ).fetchone()
        return dict(row)


PRESENTATION_SNAPSHOT_FIELDS = (
    "phase",
    "session_key",
    "active_team_id",
    "active_question_id",
    "current_question",
    "presentation_started_at",
    "presentation_created_at",
    "presentation_time_cap",
    "presentation_remaining",
    "poll_active",
    "poll_question_key",
    "poll_started_at",
    "presentation_history",
    "poll_closed_at",
    "challenge_ratings_closed_at",
)


def _presentation_snapshot(env):
    state = _state_row(env)
    return {name: state[name] for name in PRESENTATION_SNAPSHOT_FIELDS}


def _activate_presentation(env, **overrides):
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S.%f")
    fields = {
        "phase": "competition",
        "active_team_id": env["teams"]["Team 1"],
        "active_question_id": env["question_id"],
        "current_question": "Question One",
        "presentation_started_at": now,
        "presentation_created_at": now,
        "presentation_time_cap": 300,
        "presentation_remaining": None,
        "poll_active": 0,
        "poll_question_key": "pres-current",
        "poll_started_at": None,
        "poll_closed_at": None,
        "challenge_ratings_closed_at": None,
        "presentation_history": "[]",
    }
    fields.update(overrides)
    _set_state(env, **fields)


def _seed_response_history(env, session_key=SESSION_KEY):
    with _connect(env) as db:
        db.execute(
            f"""INSERT INTO teammate_thumbs
               (data_version, course_id, session_key, week_num, question_key,
                grader_id, recipient_id)
               VALUES ('{app_module.APP_VERSION}', ?, ?, ?, ?, ?, ?)""",
            (
                env["course_id"],
                session_key,
                1,
                "discussion-1",
                env["students"]["s2"],
                env["students"]["s1"],
            ),
        )
        db.execute(
            f"""INSERT INTO presentation_ratings
               (data_version, course_id, student_id, question_key, session_key, week_num,
                presenting_team_id, presenting_team_name, question_id,
                question_title, rater_team_id, rater_team_name,
                q1_developed, q2_easy)
               VALUES ('{app_module.APP_VERSION}', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                env["course_id"],
                env["students"]["s1"],
                "presentation-1",
                session_key,
                1,
                env["teams"]["Team 2"],
                "Team 2",
                env["question_id"],
                "Question One",
                env["teams"]["Team 1"],
                "Team 1",
                4,
                5,
            ),
        )
        db.commit()


def _history_counts(env):
    with _connect(env) as db:
        return {
            "thumbs": db.execute(
                "SELECT COUNT(*) FROM teammate_thumbs"
            ).fetchone()[0],
            "ratings": db.execute(
                "SELECT COUNT(*) FROM presentation_ratings"
            ).fetchone()[0],
        }


@pytest.mark.parametrize(
    ("route", "state_overrides"),
    (
        ("/api/stop_presentation", {}),
        (
            "/api/resume_presentation",
            {"presentation_started_at": None, "presentation_remaining": 120},
        ),
        ("/api/reset_presentation_timer", {}),
        ("/api/next_presentation", {}),
        ("/api/start_poll", {"poll_active": 0, "poll_started_at": None}),
        (
            "/api/stop_poll",
            {
                "poll_active": 1,
                "poll_started_at": datetime.utcnow().strftime(
                    "%Y-%m-%d %H:%M:%S"
                ),
            },
        ),
    ),
)
def test_stale_presentation_token_is_rejected_without_mutation(
    course_env, route, state_overrides
):
    _activate_presentation(course_env, **state_overrides)
    before = _presentation_snapshot(course_env)

    response = _instructor_client(course_env).post(
        route, json={"presentation_key": "pres-stale"}
    )

    assert response.status_code == 409
    assert _presentation_snapshot(course_env) == before


def test_stale_presentation_token_blocks_phase_exit(course_env):
    _activate_presentation(course_env)
    before = _presentation_snapshot(course_env)

    response = _instructor_client(course_env).post(
        "/api/set_phase",
        json={
            "phase": "ended",
            "expected_phase": "competition",
            "expected_session_key": SESSION_KEY,
            "presentation_key": "pres-stale",
            "confirm_end_session": True,
        },
    )

    assert response.status_code == 409
    assert _presentation_snapshot(course_env) == before


def test_phase_navigation_does_not_split_current_session(course_env):
    client = _instructor_client(course_env)
    roster_version = _assign_all_active_students(course_env)
    transitions = (
        ("setup", "discussion"),
        ("discussion", "setup"),
        ("setup", "discussion"),
        ("discussion", "competition"),
        ("competition", "discussion"),
        ("discussion", "ended"),
    )

    for current_phase, next_phase in transitions:
        response = client.post(
            "/api/set_phase",
            json={
                "phase": next_phase,
                "expected_phase": current_phase,
                "expected_session_key": SESSION_KEY,
                "expected_roster_version": roster_version,
                "confirm_end_session": next_phase == "ended",
            },
        )
        assert response.status_code == 200
        assert response.get_json()["session_key"] == SESSION_KEY
        state = _state_row(course_env)
        assert state["phase"] == next_phase
        assert state["session_key"] == SESSION_KEY


def test_end_session_requires_explicit_confirmation(course_env):
    _set_state(course_env, phase="discussion")
    before = _state_row(course_env)

    response = _instructor_client(course_env).post(
        "/api/set_phase",
        json={
            "phase": "ended",
            "expected_phase": "discussion",
            "expected_session_key": SESSION_KEY,
        },
    )

    assert response.status_code == 409
    assert response.get_json()["requires_confirmation"] is True
    assert _state_row(course_env) == before


def test_setup_phase_exit_blocks_unassigned_students_without_mutation(
    course_env,
):
    before = _state_row(course_env)

    response = _instructor_client(course_env).post(
        "/api/set_phase",
        json=_setup_payload(phase="discussion"),
    )

    assert response.status_code == 409
    data = response.get_json()
    assert "not assigned to a team" in data["error"].lower()
    assert data["unassigned_count"] == 1
    assert _state_row(course_env) == before


def test_unassigned_phase_block_cannot_be_bypassed_by_confirmation(course_env):
    before = _state_row(course_env)

    response = _instructor_client(course_env).post(
        "/api/set_phase",
        json=_setup_payload(
            phase="discussion",
            confirm_unassigned_students=True,
        ),
    )

    assert response.status_code == 409
    data = response.get_json()
    assert "not assigned to a team" in data["error"].lower()
    assert data["unassigned_count"] == 1
    assert _state_row(course_env) == before


def test_setup_phase_exit_needs_no_confirmation_when_everyone_is_assigned(
    course_env,
):
    roster_version = _assign_all_active_students(course_env)

    response = _instructor_client(course_env).post(
        "/api/set_phase",
        json=_setup_payload(
            roster_version=roster_version,
            phase="competition",
        ),
    )

    assert response.status_code == 200
    assert response.get_json()["phase"] == "competition"
    assert _state_row(course_env)["phase"] == "competition"


def test_unassigned_phase_exit_rejects_a_stale_roster(course_env):
    instructor = _instructor_client(course_env)
    payload = _setup_payload(phase="discussion")
    blocked = instructor.post("/api/set_phase", json=payload)
    assert blocked.status_code == 409
    assert blocked.get_json()["unassigned_count"] == 1

    roster_change = _student_client(course_env).post(
        "/api/join_team",
        json={"team_id": 0},
    )
    assert roster_change.status_code == 200

    retried = instructor.post("/api/set_phase", json=payload)

    assert retried.status_code == 409
    assert "roster changed" in retried.get_json()["error"].lower()
    state = _state_row(course_env)
    assert state["phase"] == "setup"
    assert state["roster_version"] == 1


def test_next_presentation_waits_for_open_poll(course_env, monkeypatch):
    clock = {"now": datetime(2026, 8, 13, 12, 0, 0)}
    monkeypatch.setattr(app_module, "_utcnow", lambda: clock["now"])
    _set_state(course_env, discussion_week=2)
    _activate_presentation(
        course_env,
        poll_active=1,
        poll_started_at=clock["now"].strftime("%Y-%m-%d %H:%M:%S.%f"),
    )
    client = _instructor_client(course_env)
    payload = {
        "presentation_key": "pres-current",
        "expected_phase": "competition",
        "expected_session_key": SESSION_KEY,
    }
    before = _presentation_snapshot(course_env)

    blocked = client.post("/api/next_presentation", json=payload)

    assert blocked.status_code == 409
    assert "active rating poll" in blocked.get_json()["error"]
    assert _presentation_snapshot(course_env) == before

    stopped = client.post("/api/stop_poll", json=payload)
    assert stopped.status_code == 200
    assert stopped.get_json()["ratings_settling"] is True
    assert stopped.get_json()["ratings_settling_remaining"] == 3
    cutoff = _state_row(course_env)["poll_closed_at"]

    settling = client.post("/api/next_presentation", json=payload)
    assert settling.status_code == 409
    assert settling.get_json()["ratings_settling"] is True
    assert _state_row(course_env)["poll_closed_at"] == cutoff

    clock["now"] += timedelta(seconds=app_module.POLL_SUBMISSION_GRACE_SECONDS)
    finished = client.post("/api/next_presentation", json=payload)
    assert finished.status_code == 200
    state = _state_row(course_env)
    assert state["active_team_id"] is None
    history = json.loads(state["presentation_history"])
    assert history[-1]["presentation_key"] == "pres-current"
    assert history[-1]["week_num"] == 2

    repeated = client.post("/api/next_presentation", json=payload)
    assert repeated.status_code == 200
    assert repeated.get_json()["already_finished"] is True


@pytest.mark.parametrize(
    ("route", "payload"),
    (
        (
            "/api/next_presentation",
            {"presentation_key": "pres-current"},
        ),
        (
            "/api/start_poll",
            {"presentation_key": "pres-current"},
        ),
        (
            "/api/cancel_presentation",
            {"presentation_key": "pres-current"},
        ),
        (
            "/api/set_phase",
            {
                "phase": "ended",
                "expected_phase": "competition",
                "expected_session_key": SESSION_KEY,
                "presentation_key": "pres-current",
                "confirm_end_session": True,
            },
        ),
    ),
)
def test_instructor_mutations_wait_for_rating_grace(
        course_env, monkeypatch, route, payload):
    poll_started = datetime(2026, 7, 25, 12, 0, 0)
    request_arrived = poll_started + timedelta(
        seconds=app_module.POLL_DURATION + 1
    )
    _activate_presentation(
        course_env,
        poll_active=1,
        poll_started_at=poll_started.strftime("%Y-%m-%d %H:%M:%S.%f"),
    )
    before = _presentation_snapshot(course_env)
    monkeypatch.setattr(app_module, "_utcnow", lambda: request_arrived)

    response = _instructor_client(course_env).post(route, json=payload)

    assert response.status_code == 409
    assert response.get_json()["error"] == app_module.POLL_SETTLING_MESSAGE
    assert _presentation_snapshot(course_env) == before


def test_next_presentation_can_finish_after_rating_grace(
        course_env, monkeypatch):
    poll_started = datetime(2026, 7, 25, 12, 0, 0)
    request_arrived = poll_started + timedelta(
        seconds=(
            app_module.POLL_DURATION
            + app_module.POLL_SUBMISSION_GRACE_SECONDS
        )
    )
    _activate_presentation(
        course_env,
        poll_active=1,
        poll_started_at=poll_started.strftime("%Y-%m-%d %H:%M:%S.%f"),
    )
    monkeypatch.setattr(app_module, "_utcnow", lambda: request_arrived)

    response = _instructor_client(course_env).post(
        "/api/next_presentation",
        json={"presentation_key": "pres-current"},
    )

    assert response.status_code == 200
    state = _state_row(course_env)
    assert state["active_team_id"] is None
    history = json.loads(state["presentation_history"])
    assert history[-1]["presentation_key"] == "pres-current"


def test_start_poll_preserves_fractional_start_time(
        course_env, monkeypatch):
    clock = {"now": datetime(2026, 7, 25, 12, 0, 0, 900000)}
    _activate_presentation(course_env)
    monkeypatch.setattr(app_module, "_utcnow", lambda: clock["now"])
    client = _instructor_client(course_env)

    started = client.post(
        "/api/start_poll",
        json={"presentation_key": "pres-current"},
    )

    expected = "2026-07-25 12:00:00.900000"
    assert started.status_code == 200
    assert started.get_json()["poll_started_at"] == expected
    assert _state_row(course_env)["poll_started_at"] == expected

    clock["now"] += timedelta(seconds=39, milliseconds=500)
    live_state = client.get("/api/poll").get_json()["state"]
    assert live_state["poll_active"] is True
    assert live_state["poll_remaining"] == 1

    clock["now"] += timedelta(milliseconds=500)
    closed_state = client.get("/api/poll").get_json()["state"]
    assert closed_state["poll_active"] is False
    assert closed_state["poll_remaining"] == 0


def test_leaving_ended_starts_exactly_one_new_session(course_env):
    _set_state(course_env, phase="ended")
    client = _instructor_client(course_env)
    roster_version = _assign_all_active_students(course_env)

    first = client.post(
        "/api/set_phase",
        json={
            "phase": "setup",
            "expected_phase": "ended",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert first.status_code == 200
    assert first.get_json()["session_key"] == SESSION_KEY + 1

    second = client.post(
        "/api/set_phase",
        json={
            "phase": "discussion",
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY + 1,
            "expected_roster_version": roster_version,
        },
    )
    assert second.status_code == 200
    assert second.get_json()["session_key"] == SESSION_KEY + 1
    assert _state_row(course_env)["session_key"] == SESSION_KEY + 1


def test_archiving_preserves_history_and_reactivation_reuses_identity(course_env):
    # Historical responses from an earlier session must not prevent roster
    # maintenance, and their student foreign keys must remain stable.
    _seed_response_history(course_env, session_key=SESSION_KEY - 1)
    original_student_db_id = course_env["students"]["s1"]
    history_before = _history_counts(course_env)
    client = _instructor_client(course_env)

    archived = client.delete(
        f"/api/remove_student/{original_student_db_id}",
        json=_setup_payload(),
    )
    assert archived.status_code == 200
    with _connect(course_env) as db:
        student = db.execute(
            """SELECT id, is_active, team_id, last_team_id
               FROM students WHERE student_id = 's1'"""
        ).fetchone()
    assert student["id"] == original_student_db_id
    assert student["is_active"] == 0
    assert student["team_id"] is None
    assert student["last_team_id"] == course_env["teams"]["Team 1"]
    assert _history_counts(course_env) == history_before

    restored = client.post(
        "/api/add_student",
        json=_setup_payload(
            roster_version=1,
            student_id="S1",
            name="Alice Restored",
            pin="5555",
        ),
    )
    assert restored.status_code == 200
    assert restored.get_json()["reactivated"] is True
    with _connect(course_env) as db:
        student = db.execute(
            """SELECT id, student_id, name, pin, is_active
               FROM students WHERE student_id = 's1'"""
        ).fetchone()
    assert student["id"] == original_student_db_id
    assert student["student_id"] == "s1"
    assert student["name"] == "Alice Restored"
    assert student["pin"] == "5555"
    assert student["is_active"] == 1
    assert _history_counts(course_env) == history_before


def test_setup_roster_mutation_rejects_stale_version_without_changes(course_env):
    client = _instructor_client(course_env)

    stale = client.post(
        "/api/set_max_teams",
        json=_setup_payload(roster_version=9, max_teams=3),
    )

    assert stale.status_code == 409
    with _connect(course_env) as db:
        state = db.execute(
            "SELECT max_teams, roster_version FROM course_state"
        ).fetchone()
    assert dict(state) == {"max_teams": 4, "roster_version": 0}

    current = client.post(
        "/api/set_max_teams",
        json=_setup_payload(max_teams=3),
    )
    assert current.status_code == 200
    with _connect(course_env) as db:
        state = db.execute(
            "SELECT max_teams, roster_version FROM course_state"
        ).fetchone()
    assert dict(state) == {"max_teams": 3, "roster_version": 1}


def test_instructor_unassignment_preserves_last_team(course_env):
    client = _instructor_client(course_env)
    student_id = course_env["students"]["s1"]
    team_1 = course_env["teams"]["Team 1"]
    team_2 = course_env["teams"]["Team 2"]

    unassigned = client.post(
        "/api/assign_student",
        json=_setup_payload(student_id=student_id, team_id=None),
    )
    assert unassigned.status_code == 200
    with _connect(course_env) as db:
        student = db.execute(
            "SELECT team_id, last_team_id FROM students WHERE id = ?",
            (student_id,),
        ).fetchone()
    assert student["team_id"] is None
    assert student["last_team_id"] == team_1

    stale = client.post(
        "/api/assign_student",
        json=_setup_payload(
            roster_version=0,
            student_id=student_id,
            team_id=team_2,
        ),
    )
    assert stale.status_code == 409

    reassigned = client.post(
        "/api/assign_student",
        json=_setup_payload(
            roster_version=1,
            student_id=student_id,
            team_id=team_2,
        ),
    )
    assert reassigned.status_code == 200
    with _connect(course_env) as db:
        student = db.execute(
            "SELECT team_id, last_team_id FROM students WHERE id = ?",
            (student_id,),
        ).fetchone()
    assert student["team_id"] == team_2
    assert student["last_team_id"] == team_2


def test_non_roster_setup_control_rejects_stale_session(course_env):
    response = _instructor_client(course_env).post(
        "/api/toggle_lock_teams",
        json={
            "locked": True,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY - 1,
        },
    )

    assert response.status_code == 409
    with _connect(course_env) as db:
        locked = db.execute(
            "SELECT teams_locked FROM course_state"
        ).fetchone()[0]
    assert locked == 0


def _reset_payload(env, expected_phase):
    return {
        "confirm_slug": env["slug"],
        "expected_phase": expected_phase,
        "expected_session_key": SESSION_KEY,
    }


def test_reset_rejects_live_phase_without_side_effects(course_env):
    _set_state(course_env, phase="discussion")
    _seed_response_history(course_env)
    history_before = _history_counts(course_env)
    state_before = _state_row(course_env)

    response = _instructor_client(course_env).post(
        "/api/reset_data", json=_reset_payload(course_env, "discussion")
    )

    assert response.status_code == 409
    assert _history_counts(course_env) == history_before
    assert _state_row(course_env) == state_before
    assert not (
        Path(course_env["data_dir"])
        / course_env["slug"]
        / "reset-backups"
    ).exists()


def test_reset_rejects_stale_state_without_side_effects(course_env):
    _set_state(course_env, phase="ended")
    _seed_response_history(course_env)
    history_before = _history_counts(course_env)
    state_before = _state_row(course_env)

    response = _instructor_client(course_env).post(
        "/api/reset_data", json=_reset_payload(course_env, "setup")
    )

    assert response.status_code == 409
    assert _history_counts(course_env) == history_before
    assert _state_row(course_env) == state_before
    assert not (
        Path(course_env["data_dir"])
        / course_env["slug"]
        / "reset-backups"
    ).exists()


def test_reset_creates_recoverable_backup_before_clearing(course_env):
    _set_state(course_env, phase="ended")
    _seed_response_history(course_env)

    response = _instructor_client(course_env).post(
        "/api/reset_data", json=_reset_payload(course_env, "ended")
    )

    assert response.status_code == 200
    backup_dir = (
        Path(course_env["data_dir"])
        / course_env["slug"]
        / "reset-backups"
    )
    backup_files = list(backup_dir.glob("popping-before-reset-*.db"))
    assert len(backup_files) == 1
    assert response.get_json()["backup"] == backup_files[0].name

    backup = sqlite3.connect(backup_files[0])
    backup.row_factory = sqlite3.Row
    try:
        assert backup.execute("PRAGMA integrity_check").fetchone()[0] == "ok"
        backup_state = backup.execute(
            "SELECT phase, session_key FROM course_state"
        ).fetchone()
        assert dict(backup_state) == {
            "phase": "ended",
            "session_key": SESSION_KEY,
        }
        assert backup.execute(
            "SELECT COUNT(*) FROM teammate_thumbs"
        ).fetchone()[0] == 1
        assert backup.execute(
            "SELECT COUNT(*) FROM presentation_ratings"
        ).fetchone()[0] == 1
    finally:
        backup.close()

    assert _history_counts(course_env) == {"thumbs": 0, "ratings": 0}
    live_state = _state_row(course_env)
    assert live_state["phase"] == "setup"
    assert live_state["session_key"] == SESSION_KEY + 1


def test_reset_rejects_change_during_uncontended_backup(course_env, monkeypatch):
    """A write during backup must abort reset instead of being lost."""
    _set_state(course_env, phase="ended")
    original_backup = app_module._create_reset_backup
    lock_observed = []

    def backup_uncontended(slug, *, prune=True):
        # If the write lock were held during backup, this second connection
        # would get SQLITE_BUSY.  With the backup before the lock, it succeeds.
        contender = sqlite3.connect(course_env["db_path"], timeout=0.05)
        try:
            contender.execute(
                "UPDATE course_state SET roster_version = roster_version + 1"
            )
            contender.commit()
        except sqlite3.OperationalError as exc:
            lock_observed.append(True)
        else:
            lock_observed.append(False)
        finally:
            contender.close()
        return original_backup(slug, prune=prune)

    monkeypatch.setattr(
        app_module, "_create_reset_backup", backup_uncontended
    )
    response = _instructor_client(course_env).post(
        "/api/reset_data", json=_reset_payload(course_env, "ended")
    )

    assert response.status_code == 409
    assert b"Course data changed while the backup was being created" in response.data
    assert lock_observed == [False]
    backup_dir = (
        Path(course_env["data_dir"])
        / course_env["slug"]
        / "reset-backups"
    )
    assert list(backup_dir.glob("popping-before-reset-*.db")) == []
    state = _state_row(course_env)
    assert state["phase"] == "ended"
    assert state["session_key"] == SESSION_KEY


def test_aborted_reset_preserves_three_prior_backups(
        course_env, monkeypatch):
    _set_state(course_env, phase="ended")
    original_backup = app_module._create_reset_backup
    clock = {"tick": 0}

    def unique_now():
        clock["tick"] += 1
        return datetime(2026, 8, 14) + timedelta(
            microseconds=clock["tick"]
        )

    monkeypatch.setattr(app_module, "_utcnow", unique_now)
    for _ in range(3):
        original_backup(course_env["slug"])
    backup_dir = (
        Path(course_env["data_dir"])
        / course_env["slug"]
        / "reset-backups"
    )
    prior_names = {
        path.name for path in backup_dir.glob("popping-before-reset-*.db")
    }
    assert len(prior_names) == 3

    def backup_after_change(slug, *, prune=True):
        contender = sqlite3.connect(course_env["db_path"])
        try:
            contender.execute(
                "UPDATE course_state SET roster_version = roster_version + 1"
            )
            contender.commit()
        finally:
            contender.close()
        return original_backup(slug, prune=prune)

    monkeypatch.setattr(
        app_module, "_create_reset_backup", backup_after_change
    )
    response = _instructor_client(course_env).post(
        "/api/reset_data", json=_reset_payload(course_env, "ended")
    )

    assert response.status_code == 409
    remaining_names = {
        path.name for path in backup_dir.glob("popping-before-reset-*.db")
    }
    assert remaining_names == prior_names
    assert _state_row(course_env)["phase"] == "ended"


def test_reset_restores_hidden_discussion_questions(course_env):
    with _connect(course_env) as db:
        db.execute(
            """INSERT INTO hidden_discussion_questions
               (course_id, week_num, question_key)
               VALUES (?, 1, 'discussion-1')""",
            (course_env["course_id"],),
        )
        db.commit()
    _set_state(course_env, phase="ended")

    response = _instructor_client(course_env).post(
        "/api/reset_data", json=_reset_payload(course_env, "ended")
    )

    assert response.status_code == 200
    with _connect(course_env) as db:
        assert db.execute(
            "SELECT COUNT(*) FROM hidden_discussion_questions"
        ).fetchone()[0] == 0


def test_reset_succeeds_without_optional_peer_reviews_table(course_env):
    with _connect(course_env) as db:
        db.execute("DROP TABLE peer_reviews")
        db.commit()
    _set_state(course_env, phase="ended")

    response = _instructor_client(course_env).post(
        "/api/reset_data", json=_reset_payload(course_env, "ended")
    )

    assert response.status_code == 200
    backup_path = (
        Path(course_env["data_dir"])
        / course_env["slug"]
        / "reset-backups"
        / response.get_json()["backup"]
    )
    assert backup_path.is_file()
    assert _state_row(course_env)["phase"] == "setup"
    assert _state_row(course_env)["session_key"] == SESSION_KEY + 1


def test_unassigned_student_sees_discussion_questions_container(course_env):
    _set_state(course_env, phase="discussion")

    response = _student_client(course_env, "s3").get("/dashboard")

    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert 'id="student-disc-questions-list"' in html


def test_student_thumb_is_session_scoped_and_phase_gated(course_env):
    """Thumbs recognize a teammate for the whole discussion phase: one per
    teammate per session, recorded only during discussion, with no question key."""
    client = _student_client(course_env, "s1")

    _set_state(course_env, phase="setup")
    rejected = client.post(
        "/api/grade_peer", json={"recipient_id": "s2", "selected": True}
    )
    assert rejected.status_code == 403
    assert _history_counts(course_env)["thumbs"] == 0

    _set_state(course_env, phase="discussion")
    # No question is ever posted in the new model: all week questions are
    # visible at once and thumbs are phase-scoped.
    assert _state_row(course_env)["current_discussion_key"] is None
    saved = client.post(
        "/api/grade_peer", json={"recipient_id": "s2", "selected": True}
    )
    assert saved.status_code == 200
    assert _history_counts(course_env)["thumbs"] == 1
    with _connect(course_env) as db:
        key = db.execute("SELECT question_key FROM teammate_thumbs").fetchone()[0]
    assert key == "discussion"

    # Session-scoped: repeating the same thumb is idempotent (no duplicate).
    again = client.post(
        "/api/grade_peer", json={"recipient_id": "s2", "selected": True}
    )
    assert again.status_code == 200
    assert _history_counts(course_env)["thumbs"] == 1


def test_student_rating_records_selected_week(course_env):
    with _connect(course_env) as db:
        question_id = db.execute(
            """INSERT INTO questions
               (course_id, question_num, question_text, title, week_num,
                source_key)
               VALUES (?, 1, 'Week 2 question', 'Week 2 question', 2,
                       'presentation:2:1')""",
            (course_env["course_id"],),
        ).lastrowid
        db.execute(
            "UPDATE course_state SET discussion_week = 2"
        )
        db.commit()
    _activate_presentation(
        course_env,
        active_question_id=question_id,
        poll_active=1,
        poll_started_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
    )

    response = _student_client(course_env, "s4").post(
        "/api/submit_rating",
        json={
            "presentation_key": "pres-current",
            "q1_developed": 4,
            "q2_easy": 5,
        },
    )

    assert response.status_code == 200
    with _connect(course_env) as db:
        saved_week = db.execute(
            "SELECT week_num FROM presentation_ratings"
        ).fetchone()[0]
    assert saved_week == 2


def test_sixty_students_can_submit_ratings_concurrently(course_env):
    student_ids = [f"load-{index:02d}" for index in range(1, 61)]
    with _connect(course_env) as db:
        for student_id in student_ids:
            db.execute(
                """INSERT INTO students
                   (course_id, student_id, name, pin, team_id)
                   VALUES (?, ?, ?, '1111', ?)""",
                (
                    course_env["course_id"],
                    student_id,
                    student_id,
                    course_env["teams"]["Team 2"],
                ),
            )
        db.commit()
    _activate_presentation(
        course_env,
        poll_active=1,
        poll_started_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
    )

    clients = []
    for student_id in student_ids:
        client = _student_client(course_env, student_id)
        with client.session_transaction() as flask_session:
            flask_session["activity_session_key"] = SESSION_KEY
            flask_session["last_active_synced_at"] = datetime.utcnow().isoformat()
        clients.append(client)

    def submit(client):
        response = client.post(
            "/api/submit_rating",
            json={
                "presentation_key": "pres-current",
                "q1_developed": 4,
                "q2_easy": 5,
            },
        )
        return response.status_code

    with ThreadPoolExecutor(max_workers=30) as pool:
        statuses = list(pool.map(submit, clients))

    assert statuses == [200] * len(student_ids)
    with _connect(course_env) as db:
        assert db.execute(
            """SELECT COUNT(*) FROM presentation_ratings
               WHERE course_id = ? AND question_key = 'pres-current'""",
            (course_env["course_id"],),
        ).fetchone()[0] == len(student_ids)


def test_rating_deadline_uses_arrival_time_before_write_lock(
        course_env, monkeypatch):
    poll_started = datetime(2026, 7, 25, 12, 0, 0)
    clock = {"now": poll_started + timedelta(seconds=42)}
    _activate_presentation(
        course_env,
        poll_active=1,
        poll_started_at=poll_started.strftime("%Y-%m-%d %H:%M:%S"),
    )
    client = _student_client(course_env, "s4")
    with client.session_transaction() as flask_session:
        flask_session["activity_session_key"] = SESSION_KEY
        flask_session["last_active_synced_at"] = clock["now"].isoformat()

    monkeypatch.setattr(app_module, "_utcnow", lambda: clock["now"])
    visible_state = client.get("/api/poll").get_json()["state"]
    assert visible_state["poll_active"] is False
    assert visible_state["poll_remaining"] == 0

    original_get_db = app_module.get_db

    class DelayedWriteConnection:
        def __init__(self, connection):
            self._connection = connection

        def execute(self, sql, *args, **kwargs):
            if sql == "BEGIN IMMEDIATE":
                clock["now"] = poll_started + timedelta(seconds=44)
            return self._connection.execute(sql, *args, **kwargs)

        def __getattr__(self, name):
            return getattr(self._connection, name)

    monkeypatch.setattr(
        app_module,
        "get_db",
        lambda slug: DelayedWriteConnection(original_get_db(slug)),
    )

    response = client.post(
        "/api/submit_rating",
        json={
            "presentation_key": "pres-current",
            "q1_developed": 4,
            "q2_easy": 5,
        },
    )

    assert clock["now"] == poll_started + timedelta(seconds=44)
    assert response.status_code == 200


def test_rating_grace_rejects_arrival_at_exact_end(
        course_env, monkeypatch):
    poll_started = datetime(2026, 7, 25, 12, 0, 0)
    request_arrived = poll_started + timedelta(
        seconds=(
            app_module.POLL_DURATION
            + app_module.POLL_SUBMISSION_GRACE_SECONDS
        )
    )
    _activate_presentation(
        course_env,
        poll_active=1,
        poll_started_at=poll_started.strftime("%Y-%m-%d %H:%M:%S.%f"),
    )
    client = _student_client(course_env, "s4")
    with client.session_transaction() as flask_session:
        flask_session["activity_session_key"] = SESSION_KEY
        flask_session["last_active_synced_at"] = request_arrived.isoformat()
    monkeypatch.setattr(app_module, "_utcnow", lambda: request_arrived)

    response = client.post(
        "/api/submit_rating",
        json={
            "presentation_key": "pres-current",
            "q1_developed": 4,
            "q2_easy": 5,
        },
    )

    assert response.status_code == 403
    assert response.get_json()["error"] == "The rating poll is closed"
    with _connect(course_env) as db:
        assert db.execute(
            "SELECT COUNT(*) FROM presentation_ratings"
        ).fetchone()[0] == 0


def test_rating_arrival_before_poll_start_is_rejected(
        course_env, monkeypatch):
    request_arrived = datetime(2026, 7, 25, 12, 0, 0)
    poll_started = request_arrived + timedelta(seconds=1)
    _activate_presentation(
        course_env,
        poll_active=1,
        poll_started_at=poll_started.strftime("%Y-%m-%d %H:%M:%S"),
    )
    client = _student_client(course_env, "s4")
    with client.session_transaction() as flask_session:
        flask_session["activity_session_key"] = SESSION_KEY
        flask_session["last_active_synced_at"] = request_arrived.isoformat()
    monkeypatch.setattr(app_module, "_utcnow", lambda: request_arrived)

    response = client.post(
        "/api/submit_rating",
        json={
            "presentation_key": "pres-current",
            "q1_developed": 4,
            "q2_easy": 5,
        },
    )

    assert response.status_code == 403
    assert response.get_json()["error"] == "The rating poll is closed"


def test_course_database_busy_timeout_is_eight_seconds(course_env):
    with app_module.app.app_context():
        db = database.get_db(course_env["slug"])
        busy_timeout = db.execute("PRAGMA busy_timeout").fetchone()[0]

    assert busy_timeout == 8000


def test_locked_rating_returns_retryable_json_without_late_commit(
        course_env, monkeypatch):
    _activate_presentation(
        course_env,
        poll_active=1,
        poll_started_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S.%f"),
    )
    client = _student_client(course_env, "s4")
    with client.session_transaction() as flask_session:
        flask_session["activity_session_key"] = SESSION_KEY
        flask_session["last_active_synced_at"] = datetime.utcnow().isoformat()

    with app_module.app.app_context():
        database.ensure_schema(course_env["slug"])
    assert client.get("/api/poll").status_code == 200

    monkeypatch.setattr(database, "SQLITE_BUSY_TIMEOUT_SECONDS", 0.05)
    monkeypatch.setattr(database, "SQLITE_BUSY_TIMEOUT_MS", 50)
    blocker = sqlite3.connect(
        course_env["db_path"], timeout=0.1, isolation_level=None
    )
    try:
        blocker.execute("BEGIN IMMEDIATE")
        response = client.post(
            "/api/submit_rating",
            json={
                "presentation_key": "pres-current",
                "q1_developed": 4,
                "q2_easy": 5,
            },
        )
    finally:
        blocker.rollback()
        blocker.close()

    assert response.status_code == 503
    assert response.is_json
    assert response.get_json()["retry_after"] == 2
    assert response.headers["Retry-After"] == "2"
    with _connect(course_env) as db:
        assert db.execute(
            "SELECT COUNT(*) FROM presentation_ratings"
        ).fetchone()[0] == 0

    retried = client.post(
        "/api/submit_rating",
        json={
            "presentation_key": "pres-current",
            "q1_developed": 4,
            "q2_easy": 5,
        },
    )

    assert retried.status_code == 200
    with _connect(course_env) as db:
        assert db.execute(
            "SELECT COUNT(*) FROM presentation_ratings"
        ).fetchone()[0] == 1


def test_non_lock_operational_error_remains_json_500(
        course_env, monkeypatch):
    def fail_query(*_args, **_kwargs):
        raise sqlite3.OperationalError("disk I/O error")

    monkeypatch.setattr(app_module, "query_db", fail_query)

    response = _student_client(course_env).get("/api/state")

    assert response.status_code == 500
    assert response.is_json
    assert response.get_json()["error"] == (
        "Something went wrong. Please try again."
    )
    assert "Retry-After" not in response.headers


def test_discussion_question_visibility_and_version(course_env):
    """Hiding a question removes it from the student list and bumps the
    version students use to refetch; showing it restores it."""
    instructor = _instructor_client(course_env)
    assert instructor.get("/api/state").status_code == 200
    _set_state(course_env, phase="discussion")
    student = _student_client(course_env, "s1")

    added = instructor.post(
        "/api/questions",
        json={
            "title": "Visibility check", "content": "discuss", "week": 1,
            "expected_phase": "discussion",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert added.status_code == 200
    target = instructor.get("/api/discussion_questions").get_json()["questions"][0]
    before_version = instructor.get(
        "/api/discussion_questions"
    ).get_json()["version"]

    # Visible to the student by default.
    assert any(
        q["key"] == target["key"]
        for q in student.get("/api/discussion_questions").get_json()["questions"]
    )

    hide = instructor.post(
        "/api/toggle_discussion_question",
        json={
            "question_key": target["key"], "visible": False,
            "expected_phase": "discussion",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert hide.status_code == 200

    after = instructor.get("/api/discussion_questions").get_json()
    assert after["version"] > before_version
    assert next(
        q for q in after["questions"] if q["key"] == target["key"]
    )["hidden"] is True
    assert not any(
        q["key"] == target["key"]
        for q in student.get("/api/discussion_questions").get_json()["questions"]
    )



def test_discussion_visibility_keys_with_shared_prefix_stay_independent(
        course_env):
    class_dir = Path(config.CLASSES_DIR) / course_env["slug"]
    (class_dir / "week-1-questions.md").write_text(
        """---
id: topic
title: Topic
---

Discuss the broad topic.

---
id: topic-detail
title: Topic Detail
---

Discuss the detailed topic.
""",
        encoding="utf-8",
    )
    _set_state(course_env, phase="discussion")
    instructor = _instructor_client(course_env)
    student = _student_client(course_env, "s1")
    topic_key = "week-1-q-topic"
    detail_key = "week-1-q-topic-detail"

    hidden = instructor.post(
        "/api/toggle_discussion_question",
        json={
            "question_key": detail_key,
            "visible": False,
            "expected_phase": "discussion",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert hidden.status_code == 200

    questions = instructor.get(
        "/api/discussion_questions"
    ).get_json()["questions"]
    by_key = {question["key"]: question for question in questions}
    assert by_key[topic_key]["hidden"] is False
    assert by_key[detail_key]["hidden"] is True

    shown = instructor.post(
        "/api/toggle_discussion_question",
        json={
            "question_key": topic_key,
            "visible": True,
            "expected_phase": "discussion",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert shown.status_code == 200
    with _connect(course_env) as db:
        stored_keys = {
            row[0] for row in db.execute(
                """SELECT question_key FROM hidden_discussion_questions
                   WHERE course_id = ? AND week_num = 1""",
                (course_env["course_id"],),
            )
        }
    assert stored_keys == {detail_key}
    student_keys = {
        question["key"] for question in student.get(
            "/api/discussion_questions"
        ).get_json()["questions"]
    }
    assert student_keys == {topic_key}

def test_hidden_bank_question_stays_hidden_after_content_edit(course_env):
    class_dir = Path(config.CLASSES_DIR) / course_env["slug"]
    question_file = class_dir / "week-1-questions.md"
    original_title = "Original title"
    original_body = "Original body."
    question_file.write_text(
        "---\nid: stable-bank-question\n"
        f"title: {original_title}\n---\n\n{original_body}\n",
        encoding="utf-8",
    )
    _set_state(course_env, phase="discussion")
    instructor = _instructor_client(course_env)
    student = _student_client(course_env, "s1")

    revision = hashlib.sha256(
        (original_title + "\0" + original_body).encode("utf-8")
    ).hexdigest()[:16]
    legacy_key = f"week-1-q-{revision}-{revision[:8]}"
    with _connect(course_env) as db:
        db.execute(
            """INSERT INTO hidden_discussion_questions
               (course_id, week_num, question_key)
               VALUES (?, 1, ?)""",
            (course_env["course_id"], legacy_key),
        )
        db.commit()

    original = instructor.get(
        "/api/discussion_questions"
    ).get_json()["questions"][0]
    assert original["key"] == "week-1-q-stable-bank-question"
    assert original["hidden"] is True
    with _connect(course_env) as db:
        stored_keys = {
            row[0] for row in db.execute(
                """SELECT question_key FROM hidden_discussion_questions
                   WHERE course_id = ? AND week_num = 1""",
                (course_env["course_id"],),
            )
        }
    assert stored_keys == {original["key"]}

    question_file.write_text(
        "---\nid: stable-bank-question\n"
        "title: Revised title\n---\n\nRevised body.\n",
        encoding="utf-8",
    )
    revised = instructor.get(
        "/api/discussion_questions"
    ).get_json()["questions"][0]
    assert revised["key"] == original["key"]
    assert revised["hidden"] is True
    assert student.get(
        "/api/discussion_questions"
    ).get_json()["questions"] == []


def test_hidden_appendix_question_stays_hidden_after_edit(course_env):
    _set_state(course_env, phase="discussion")
    instructor = _instructor_client(course_env)
    student = _student_client(course_env, "s1")
    added = instructor.post(
        "/api/questions",
        json={
            "title": "Appendix title",
            "content": "Appendix body.",
            "week": 1,
            "expected_phase": "discussion",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert added.status_code == 200

    original = next(
        question for question in instructor.get(
            "/api/discussion_questions"
        ).get_json()["questions"]
        if question.get("appendix_id") == "A1"
    )
    assert instructor.post(
        "/api/toggle_discussion_question",
        json={
            "question_key": original["key"],
            "visible": False,
            "expected_phase": "discussion",
            "expected_session_key": SESSION_KEY,
        },
    ).status_code == 200
    assert instructor.post(
        "/api/edit_appendix_question",
        json={
            "appendix_id": "A1",
            "title": "Revised appendix title",
            "content": "Revised appendix body.",
            "week": 1,
            "expected_phase": "discussion",
            "expected_session_key": SESSION_KEY,
        },
    ).status_code == 200

    revised = next(
        question for question in instructor.get(
            "/api/discussion_questions"
        ).get_json()["questions"]
        if question.get("appendix_id") == "A1"
    )
    assert revised["key"] == original["key"]
    assert revised["hidden"] is True
    assert all(
        question["key"] != original["key"]
        for question in student.get(
            "/api/discussion_questions"
        ).get_json()["questions"]
    )


def test_student_discussion_list_excludes_hidden_and_includes_appendix(course_env):
    """5 bank questions + new appendix A1 with Q2 hidden gives students
    exactly Q1, Q3, Q4, Q5, A1 — the full-week visibility model."""
    class_dir = Path(config.CLASSES_DIR) / course_env["slug"]
    blocks = [
        f"---\ntitle: Q{num}\nid: q{num}\n---\n\nBank body {num}.\n"
        for num in range(1, 6)
    ]
    (class_dir / "week-1-questions.md").write_text(
        "\n".join(blocks), encoding="utf-8"
    )
    _set_state(course_env, phase="discussion")
    instructor = _instructor_client(course_env)
    student = _student_client(course_env, "s1")

    added = instructor.post(
        "/api/questions",
        json={
            "title": "Extra", "content": "Appendix body.", "week": 1,
            "expected_phase": "discussion",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert added.status_code == 200
    assert added.get_json()["appendix_id"] == "A1"

    instructor_list = instructor.get("/api/discussion_questions").get_json()
    assert [q["title"] for q in instructor_list["questions"]] == [
        "Q1", "Q2", "Q3", "Q4", "Q5", "A1: Extra",
    ]
    assert [q["display_number"] for q in instructor_list["questions"]] == [
        1, 2, 3, 4, 5, 6,
    ]
    assert len({q["key"] for q in instructor_list["questions"]}) == 6
    q2_key = instructor_list["questions"][1]["key"]
    hide = instructor.post(
        "/api/toggle_discussion_question",
        json={
            "question_key": q2_key, "visible": False,
            "expected_phase": "discussion",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert hide.status_code == 200

    student_list = student.get("/api/discussion_questions").get_json()
    assert [q["title"] for q in student_list["questions"]] == [
        "Q1", "Q3", "Q4", "Q5", "A1: Extra",
    ]
    assert [q["display_number"] for q in student_list["questions"]] == [
        1, 3, 4, 5, 6,
    ]
    assert all(not q["hidden"] for q in student_list["questions"])


def _seed_tied_team_ratings(env):
    # Insert in reverse name order so the test also checks deterministic tie order.
    rating_specs = (
        ("Team 4", 4),
        ("Team 3", 4),
        ("Team 2", 5),
        ("Team 1", 5),
    )
    with _connect(env) as db:
        for index, (team_name, score) in enumerate(rating_specs, start=1):
            db.execute(
                f"""INSERT INTO presentation_ratings
                   (data_version, course_id, student_id, question_key, session_key,
                    presenting_team_id, presenting_team_name, question_id,
                    question_title, rater_team_id, rater_team_name,
                    q1_developed, q2_easy)
                   VALUES ('{app_module.APP_VERSION}', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    env["course_id"],
                    env["students"]["s1"],
                    f"rank-{index}",
                    SESSION_KEY,
                    env["teams"][team_name],
                    team_name,
                    env["question_id"],
                    "Question One",
                    env["teams"]["Team 1"],
                    "Team 1",
                    score,
                    score,
                ),
            )
        db.commit()


def test_tied_team_rankings_use_competition_ranks_and_include_cutoff_ties(
    course_env,
):
    _seed_tied_team_ratings(course_env)
    _set_state(course_env, phase="ended")

    with app_module.app.app_context():
        ranked = app_module._compute_top_teams(
            course_env["slug"], course_env["course_id"], [], SESSION_KEY
        )
    assert [team["name"] for team in ranked] == [
        "Team 1",
        "Team 2",
        "Team 3",
        "Team 4",
    ]
    assert [team["rank"] for team in ranked] == [1, 1, 3, 3]
    assert [team["rating_count"] for team in ranked] == [1, 1, 1, 1]

    response = _student_client(course_env).get("/api/poll")
    assert response.status_code == 200
    top_teams = response.get_json()["top_teams"]
    assert [(team["name"], team["rank"]) for team in top_teams] == [
        ("Team 1", 1),
        ("Team 2", 1),
        ("Team 3", 3),
        ("Team 4", 3),
    ]
    assert all("avg_score" not in team for team in top_teams)


def test_recorded_activity_progress_uses_aggregate_counts_only(course_env):
    old_activity = "2000-01-01 00:00:00"
    with _connect(course_env) as db:
        db.execute(
            "UPDATE students SET last_active_at = ? WHERE course_id = ?",
            (old_activity, course_env["course_id"]),
        )
        db.execute(
            f"""INSERT INTO presentation_ratings
               (data_version, course_id, student_id, question_key, session_key,
                presenting_team_id, presenting_team_name, question_id,
                question_title, rater_team_id, rater_team_name,
                q1_developed, q2_easy)
               VALUES ('{app_module.APP_VERSION}', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                course_env["course_id"],
                course_env["students"]["s4"],
                "pres-current",
                SESSION_KEY,
                course_env["teams"]["Team 1"],
                "Team 1",
                course_env["question_id"],
                "Question One",
                course_env["teams"]["Team 2"],
                "Team 2",
                4,
                5,
            ),
        )
        db.commit()

    _activate_presentation(course_env)
    presentation_state = _instructor_client(course_env).get("/api/poll").get_json()["state"]
    assert presentation_state["poll_count"] == 1
    assert presentation_state["poll_eligible_count"] == 1
    assert "poll_online_eligible_count" not in presentation_state
    assert "poll_non_raters" not in presentation_state

    _set_state(
        course_env,
        phase="discussion",
        active_team_id=None,
        active_question_id=None,
        current_discussion_key="disc-recorded",
        current_discussion_source_key="week-1-question-1",
        current_discussion_title="Question One",
        current_discussion_content="Explain the result.",
    )
    response = _student_client(course_env, "s1").post(
        "/api/grade_peer",
        json={
            "recipient_id": "s2",
            "selected": True,
            "question_key": "disc-recorded",
        },
    )
    assert response.status_code == 200
    with _connect(course_env) as db:
        db.execute(
            "UPDATE students SET last_active_at = ? WHERE student_id = ?",
            (old_activity, "s1"),
        )
        db.commit()

    discussion_state = _instructor_client(course_env).get("/api/poll").get_json()["state"]
    assert discussion_state["thumb_participant_count"] == 1
    assert discussion_state["thumb_eligible_count"] == 2
    assert "thumb_online_eligible_count" not in discussion_state


def test_student_team_api_shows_all_team_members(course_env):
    student_teams = _student_client(course_env, "s1").get("/api/teams").get_json()
    by_name = {team["name"]: team for team in student_teams}

    assert by_name["Team 1"]["members_visible"] is True
    assert {member["student_id"] for member in by_name["Team 1"]["members"]} == {
        "s1",
        "s2",
    }
    # Students can see all teams' members (not just their own)
    assert by_name["Team 2"]["members_visible"] is True
    assert by_name["Team 2"]["member_count"] == 1
    assert [member["student_id"] for member in by_name["Team 2"]["members"]] == ["s4"]

    instructor_teams = _instructor_client(course_env).get("/api/teams").get_json()
    instructor_team_2 = next(
        team for team in instructor_teams if team["name"] == "Team 2"
    )
    assert instructor_team_2["members_visible"] is True
    assert [member["student_id"] for member in instructor_team_2["members"]] == [
        "s4"
    ]


def _active_student_count(env):
    with _connect(env) as db:
        return db.execute(
            "SELECT COUNT(*) FROM students WHERE is_active = 1"
        ).fetchone()[0]


def test_roster_rejects_more_than_500_rows_without_mutation(course_env):
    rows = ["student_id,name,pin"]
    rows.extend(
        f"upload-{index},Student {index},1234"
        for index in range(app_module.MAX_ROSTER_ROWS + 1)
    )
    payload = ("\n".join(rows) + "\n").encode("utf-8")
    count_before = _active_student_count(course_env)

    response = _instructor_client(course_env).post(
        "/api/upload_roster",
        data={"file": (io.BytesIO(payload), "roster.csv")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 400
    assert "500" in response.get_json()["error"]
    assert _active_student_count(course_env) == count_before


def test_roster_rejects_file_over_one_megabyte_without_mutation(course_env):
    payload = b"student_id,name,pin\n" + b"x" * app_module.MAX_ROSTER_BYTES
    count_before = _active_student_count(course_env)

    response = _instructor_client(course_env).post(
        "/api/upload_roster",
        data={"file": (io.BytesIO(payload), "roster.csv")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 413
    assert _active_student_count(course_env) == count_before


def test_global_request_limit_rejects_oversized_upload_without_mutation(course_env):
    payload = b"x" * (app_module.app.config["MAX_CONTENT_LENGTH"] + 1)
    count_before = _active_student_count(course_env)

    response = _instructor_client(course_env).post(
        "/api/upload_roster",
        data={"file": (io.BytesIO(payload), "roster.csv")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 413
    assert _active_student_count(course_env) == count_before


def test_roster_preview_and_confirmation_are_bound_to_setup_state(course_env):
    payload = (
        "student_id,name,pin\n"
        "s1,Alice Updated,5555\n"
        "new-student,New Student,6666\n"
    ).encode("utf-8")
    client = _instructor_client(course_env)
    expected_state = {
        "expected_phase": "setup",
        "expected_session_key": str(SESSION_KEY),
        "expected_roster_version": "0",
    }

    preview = client.post(
        "/api/upload_roster",
        data={
            **expected_state,
            "file": (io.BytesIO(payload), "roster.csv"),
        },
        content_type="multipart/form-data",
    )
    assert preview.status_code == 200
    preview_data = preview.get_json()
    assert preview_data["requires_confirmation"] is True

    confirmed = client.post(
        "/api/upload_roster",
        data={
            **expected_state,
            "confirm": "true",
            "preview_token": preview_data["preview_token"],
            "file": (io.BytesIO(payload), "roster.csv"),
        },
        content_type="multipart/form-data",
    )
    assert confirmed.status_code == 200
    assert confirmed.get_json()["requires_confirmation"] is False
    assert _active_student_count(course_env) == 2
    with _connect(course_env) as db:
        roster_version = db.execute(
            "SELECT roster_version FROM course_state"
        ).fetchone()[0]
    assert roster_version == 1


def test_export_returns_a_valid_streamed_zip(course_env):
    response = _instructor_client(course_env).get(
        f"/export/{course_env['slug']}"
    )

    assert response.status_code == 200
    assert response.mimetype == "application/zip"
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        assert "course_data.xlsx" in archive.namelist()


def test_export_assets_and_filename_use_only_selected_week(course_env):
    _write_catalog_week(course_env, 1)
    _write_catalog_week(course_env, 2)
    appendix_dir = (
        course_env["data_dir"] / course_env["slug"] / "appendix"
    )
    appendix_dir.mkdir()
    (appendix_dir / "week-1-appendix.md").write_text(
        "Appendix week 1", encoding="utf-8"
    )
    legacy_appendix = (
        Path(config.CLASSES_DIR) / course_env["slug"]
        / "week-2-appendix.md"
    )
    legacy_appendix.write_text("Appendix week 2", encoding="utf-8")
    _set_state(course_env, discussion_week=2)

    response = _instructor_client(course_env).get(
        f"/export/{course_env['slug']}"
    )

    assert response.status_code == 200
    assert (
        "filename=popping_SAFE101_week_2_export.zip"
        in response.headers["Content-Disposition"]
    )
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        names = set(archive.namelist())
        appendix_text = archive.read(
            "appendix/week-2-appendix.md"
        ).decode("utf-8")
    assert {
        "course_data.xlsx",
        "questions/week-2-questions.md",
        "appendix/week-2-appendix.md",
    }.issubset(names)
    assert "questions/week2/index.md" not in names
    assert "questions/week2/q01.html" not in names
    assert appendix_text == "Appendix week 2"
    assert not any("week-1" in name or "week1/" in name for name in names)


def test_export_workbook_activity_is_scoped_to_selected_week(course_env):
    from openpyxl import load_workbook

    with _connect(course_env) as db:
        week_2_question_id = db.execute(
            """INSERT INTO questions
               (course_id, question_num, question_text, title, week_num,
                source_key)
               VALUES (?, 1, 'Week 2 question', 'Week 2 question', 2,
                       'presentation:2:1')""",
            (course_env["course_id"],),
        ).lastrowid
        thumb_rows = (
            (1, SESSION_KEY, "week-1-thumb", "week-1-q-one"),
            (2, SESSION_KEY, "week-2-thumb-a", "week-2-q-one"),
            (2, SESSION_KEY + 1, "week-2-thumb-b", "week-2-q-two"),
        )
        for index, (week, session_key, question_key, source_key) in enumerate(
            thumb_rows
        ):
            grader = course_env["students"]["s1" if index != 1 else "s2"]
            recipient = course_env["students"]["s2" if index != 1 else "s1"]
            db.execute(
                f"""INSERT INTO teammate_thumbs
                   (data_version, course_id, session_key, week_num, question_key,
                    source_question_key, question_title, grader_id,
                    recipient_id)
                   VALUES ('{app_module.APP_VERSION}', ?, ?, ?, ?, ?, 'Question', ?, ?)""",
                (
                    course_env["course_id"], session_key, week, question_key,
                    source_key, grader, recipient,
                ),
            )

        rating_rows = (
            (
                1, SESSION_KEY, "pres-week-1", course_env["question_id"],
                course_env["students"]["s1"],
                course_env["teams"]["Team 2"], "Team 2", 1, 1,
            ),
            (
                2, SESSION_KEY, "pres-week-2-a", week_2_question_id,
                course_env["students"]["s1"],
                course_env["teams"]["Team 2"], "Team 2", 4, 5,
            ),
            (
                2, SESSION_KEY + 1, "pres-week-2-b", week_2_question_id,
                course_env["students"]["s4"],
                course_env["teams"]["Team 1"], "Team 1", 3, 3,
            ),
        )
        for (
            week, session_key, question_key, question_id, student_id,
            presenting_team_id, presenting_team_name, developed, easy,
        ) in rating_rows:
            db.execute(
                f"""INSERT INTO presentation_ratings
                   (data_version, course_id, student_id, question_key, session_key, week_num,
                    presenting_team_id, presenting_team_name, question_id,
                    question_title, q1_developed, q2_easy)
                   VALUES ('{app_module.APP_VERSION}', ?, ?, ?, ?, ?, ?, ?, ?, 'Question', ?, ?)""",
                (
                    course_env["course_id"], student_id, question_key,
                    session_key, week, presenting_team_id,
                    presenting_team_name, question_id, developed, easy,
                ),
            )
        db.execute(
            "UPDATE teammate_thumbs SET data_version = ?",
            (app_module.APP_VERSION,),
        )
        db.execute(
            "UPDATE presentation_ratings SET data_version = ?",
            (app_module.APP_VERSION,),
        )
        history = [
            {
                "presentation_key": "pres-week-1",
                    "data_version": app_module.APP_VERSION,
                "week_num": 1,
                "team": "Team 2",
                "question_id": course_env["question_id"],
            },
            {
                "presentation_key": "pres-week-2-a",
                    "data_version": app_module.APP_VERSION,
                "team": "Team 2",
                "question_id": week_2_question_id,
            },
            {
                "presentation_key": "pres-week-2-b",
                    "data_version": app_module.APP_VERSION,
                "team": "Team 1",
                "question_id": None,
            },
        ]
        db.execute(
            """UPDATE course_state
               SET discussion_week = 2, presentation_history = ?""",
            (json.dumps(history),),
        )
        db.commit()

    response = _instructor_client(course_env).get(
        f"/export/{course_env['slug']}"
    )
    assert response.status_code == 200
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        workbook = load_workbook(
            io.BytesIO(archive.read("course_data.xlsx")), read_only=True
        )

    summary = {
        row[0]: row[1]
        for row in workbook["Summary"].iter_rows(values_only=True)
        if row[0]
    }
    assert summary["Lecture Week"] == 2
    assert summary["Week Peer Reviews (thumbs)"] == 2
    assert summary["Week Presentation Ratings"] == 2

    def sheet_rows(name):
        values = list(workbook[name].iter_rows(values_only=True))
        return [dict(zip(values[0], row)) for row in values[1:]]

    peer_rows = sheet_rows("Peer Reviews")
    assert {row["week"] for row in peer_rows} == {2}
    assert {row["session_key"] for row in peer_rows} == {
        SESSION_KEY, SESSION_KEY + 1,
    }
    assert {row["discussion_post_key"] for row in peer_rows} == {
        "week-2-thumb-a", "week-2-thumb-b",
    }

    presentation_rows = sheet_rows("Presentation Ratings")
    assert {row["week"] for row in presentation_rows} == {2}
    assert {row["session_key"] for row in presentation_rows} == {
        SESSION_KEY, SESSION_KEY + 1,
    }
    assert {row["presentation_key"] for row in presentation_rows} == {
        "pres-week-2-a", "pres-week-2-b",
    }

    students = {row["student_id"]: row for row in sheet_rows("Students")}
    assert students["s1"]["thumbs_given"] == 1
    assert students["s1"]["presentation_ratings_given"] == 1
    teams = {row["team_name"]: row for row in sheet_rows("Teams")}
    assert teams["Team 1"]["presentations"] == 1
    assert teams["Team 2"]["presentations"] == 1
    assert teams["Team 2"]["combined_avg"] == 4.5


def test_export_all_weeks_parameter_is_rejected(course_env):
    response = _instructor_client(course_env).get(
        f"/export/{course_env['slug']}?weeks=all"
    )

    assert response.status_code == 400
    assert b"Export one week at a time" in response.data


def test_legacy_feedback_export_contains_only_unknown_week_rows(course_env):
    with _connect(course_env) as db:
        for week_num, question_key in (
            (None, "legacy"),
            (1, "week-1-current"),
        ):
            db.execute(
                f"""INSERT INTO teammate_thumbs
                   (data_version, course_id, session_key, week_num, question_key,
                    source_question_key, grader_id, recipient_id)
                   VALUES ('{app_module.APP_VERSION}', ?, 0, ?, ?, ?, ?, ?)""",
                (
                    course_env["course_id"],
                    week_num,
                    question_key,
                    question_key,
                    course_env["students"]["s1"],
                    course_env["students"]["s2"],
                ),
            )
        for week_num, question_key in (
            (None, "legacy-rating"),
            (1, "week-1-rating"),
        ):
            db.execute(
                f"""INSERT INTO presentation_ratings
                   (data_version, course_id, student_id, question_key, session_key,
                    week_num, presenting_team_id, presenting_team_name,
                    rater_team_id, rater_team_name, q1_developed, q2_easy)
                   VALUES ('{app_module.APP_VERSION}', ?, ?, ?, 0, ?, ?, 'Team 2', ?, 'Team 1', 3, 4)""",
                (
                    course_env["course_id"],
                    course_env["students"]["s1"],
                    question_key,
                    week_num,
                    course_env["teams"]["Team 2"],
                    course_env["teams"]["Team 1"],
                ),
            )
        db.execute(
            "UPDATE teammate_thumbs SET data_version = ?",
            (app_module.APP_VERSION,),
        )
        db.execute(
            "UPDATE presentation_ratings SET data_version = ?",
            (app_module.APP_VERSION,),
        )
        db.commit()

    response = _instructor_client(course_env).get(
        f"/export/{course_env['slug']}/legacy-feedback.csv"
    )

    assert response.status_code == 200
    assert response.mimetype == "text/csv"
    assert "legacy_unknown_week_feedback.csv" in response.headers[
        "Content-Disposition"
    ]
    rows = list(csv.DictReader(io.StringIO(
        response.data.decode("utf-8-sig")
    )))
    assert len(rows) == 2
    by_type = {row["record_type"]: row for row in rows}
    thumb = by_type["teammate_thumb"]
    assert thumb["lecture_week"] == "unknown"
    assert thumb["question_key"] == "legacy"
    assert thumb["grader_id"] == "s1"
    assert thumb["recipient_id"] == "s2"
    rating = by_type["presentation_rating"]
    assert rating["lecture_week"] == "unknown"
    assert rating["question_key"] == "legacy-rating"
    assert rating["grader_id"] == "s1"
    assert rating["presenting_team"] == "Team 2"
    assert rating["q1_developed"] == "3"
    assert rating["q2_easy"] == "4"


def test_legacy_export_keeps_one_snapshot_across_concurrent_reset(
        course_env, monkeypatch):
    _set_state(course_env, phase="ended")
    with _connect(course_env) as db:
        db.execute(
            f"""INSERT INTO teammate_thumbs
               (data_version, course_id, session_key, week_num, question_key,
                source_question_key, grader_id, recipient_id)
               VALUES ('{app_module.APP_VERSION}', ?, ?, NULL, 'before-thumb', 'before-thumb', ?, ?)""",
            (
                course_env["course_id"],
                SESSION_KEY,
                course_env["students"]["s1"],
                course_env["students"]["s2"],
            ),
        )
        db.execute(
            f"""INSERT INTO presentation_ratings
               (data_version, course_id, student_id, question_key, session_key, week_num,
                presenting_team_id, presenting_team_name,
                rater_team_id, rater_team_name, q1_developed, q2_easy)
               VALUES ('{app_module.APP_VERSION}', ?, ?, 'before-rating', ?, NULL, ?, 'Team 2',
                       ?, 'Team 1', 4, 5)""",
            (
                course_env["course_id"],
                course_env["students"]["s1"],
                SESSION_KEY,
                course_env["teams"]["Team 2"],
                course_env["teams"]["Team 1"],
            ),
        )
        db.commit()

    original_get_db = database.get_db
    wrappers = {}
    reset_committed = []

    class ResetBetweenFeedbackQueries:
        def __init__(self, connection):
            self._connection = connection

        def execute(self, sql, *args, **kwargs):
            cursor = self._connection.execute(sql, *args, **kwargs)
            if (
                not reset_committed
                and "FROM teammate_thumbs p" in sql
                and "p.week_num IS NULL" in sql
            ):
                writer = sqlite3.connect(
                    course_env["db_path"], timeout=2
                )
                try:
                    writer.execute("BEGIN IMMEDIATE")
                    writer.execute(
                        "DELETE FROM teammate_thumbs WHERE course_id = ?",
                        (course_env["course_id"],),
                    )
                    writer.execute(
                        "DELETE FROM presentation_ratings WHERE course_id = ?",
                        (course_env["course_id"],),
                    )
                    writer.execute(
                        """UPDATE course_state
                           SET phase = 'setup', session_key = session_key + 1
                           WHERE course_id = ?""",
                        (course_env["course_id"],),
                    )
                    writer.commit()
                    reset_committed.append(True)
                finally:
                    writer.close()
            return cursor

        def __getattr__(self, name):
            return getattr(self._connection, name)

    def wrapped_get_db(slug):
        connection = original_get_db(slug)
        key = id(connection)
        if key not in wrappers:
            wrappers[key] = ResetBetweenFeedbackQueries(connection)
        return wrappers[key]

    monkeypatch.setattr(database, "get_db", wrapped_get_db)
    monkeypatch.setattr(app_module, "get_db", wrapped_get_db)
    response = _instructor_client(course_env).get(
        f"/export/{course_env['slug']}/legacy-feedback.csv"
    )

    assert response.status_code == 200
    assert reset_committed == [True]
    rows = list(csv.DictReader(io.StringIO(
        response.data.decode("utf-8-sig")
    )))
    assert [row["record_type"] for row in rows] == [
        "teammate_thumb",
        "presentation_rating",
    ]
    assert rows[0]["question_key"] == "before-thumb"
    assert rows[1]["question_key"] == "before-rating"
    assert _history_counts(course_env) == {"thumbs": 0, "ratings": 0}


def test_export_unknown_weeks_param_keeps_current_week_scope(course_env):
    response = _instructor_client(course_env).get(
        f"/export/{course_env['slug']}?weeks=bogus"
    )

    assert response.status_code == 200
    assert (
        "filename=popping_SAFE101_week_1_export.zip"
        in response.headers["Content-Disposition"]
    )


def test_tools_menu_marks_roster_upload_setup_only_outside_setup(course_env):
    client = _instructor_client(course_env)
    page = client.get(
        f"/instructor/{course_env['slug']}"
    ).get_data(as_text=True)
    assert "Upload Student Roster (setup only)" not in page
    assert 'onclick="uploadRoster(event)"' in page
    assert f"/export/{course_env['slug']}?weeks=all" not in page
    assert f"/export/{course_env['slug']}/legacy-feedback.csv" in page

    _set_state(course_env, phase="competition")
    page = client.get(
        f"/instructor/{course_env['slug']}"
    ).get_data(as_text=True)
    assert "Upload Student Roster (setup only)" in page
    assert 'onclick="uploadRoster(event)"' not in page


def test_export_reports_question_asset_failure(course_env, monkeypatch):
    class_dir = Path(config.CLASSES_DIR) / course_env["slug"]
    (class_dir / "week-1-questions.md").write_text(
        "question asset", encoding="utf-8"
    )

    def fail_asset_write(_archive, _path, _archive_name):
        raise OSError("asset read failed")

    monkeypatch.setattr(zipfile.ZipFile, "write", fail_asset_write)
    client = _instructor_client(course_env)
    response = client.get(f"/export/{course_env['slug']}")

    assert response.status_code == 302
    assert f"/instructor/{course_env['slug']}" in response.headers["Location"]
    with client.session_transaction() as browser_session:
        messages = [message for _level, message in browser_session["_flashes"]]
    assert any("Export failed" in message for message in messages)
    assert not any("asset read failed" in message for message in messages)


def test_export_limits_normal_team_views_but_keeps_historical_ratings(course_env):
    from openpyxl import load_workbook

    hidden_team_id = course_env["teams"]["Team 4"]
    with _connect(course_env) as db:
        db.execute("UPDATE course_state SET max_teams = 2")
        db.execute(
            f"""INSERT INTO presentation_ratings
               (data_version, course_id, student_id, question_key, session_key, week_num,
                presenting_team_id, presenting_team_name, question_id,
                question_title, rater_team_id, rater_team_name,
                q1_developed, q2_easy)
               VALUES ('{app_module.APP_VERSION}', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                course_env["course_id"],
                course_env["students"]["s1"],
                "historical-hidden-team",
                SESSION_KEY,
                1,
                hidden_team_id,
                "Team 4",
                course_env["question_id"],
                "Historical Question",
                course_env["teams"]["Team 1"],
                "Team 1",
                4,
                4,
            ),
        )
        db.execute(
            "UPDATE presentation_ratings SET data_version = ?",
            (app_module.APP_VERSION,),
        )
        db.commit()

    response = _instructor_client(course_env).get(
        f"/export/{course_env['slug']}"
    )
    assert response.status_code == 200
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        workbook = load_workbook(
            io.BytesIO(archive.read("course_data.xlsx")), read_only=True
        )

    summary = {
        row[0]: row[1]
        for row in workbook["Summary"].iter_rows(values_only=True)
        if row[0]
    }
    assert summary["Current Visible Teams"] == 2
    assert "Total Teams" not in summary
    team_names = [
        row[1]
        for row in workbook["Teams"].iter_rows(min_row=2, values_only=True)
    ]
    assert team_names == ["Team 1", "Team 2"]
    raw_presenting_teams = [
        row[6]
        for row in workbook["Presentation Ratings"].iter_rows(
            min_row=2, values_only=True
        )
    ]
    assert "Team 4" in raw_presenting_teams


def test_session_timer_is_server_authoritative_and_idempotent(course_env):
    client = _instructor_client(course_env)
    payload = {
        "expected_phase": "setup",
        "expected_session_key": SESSION_KEY,
    }

    first = client.post("/api/start_session_timer", json=payload)
    second = client.post("/api/start_session_timer", json=payload)

    assert first.status_code == 200
    assert second.status_code == 200
    started_at = first.get_json()["session_started_at"]
    assert started_at
    assert second.get_json()["session_started_at"] == started_at

    stopped = client.post("/api/stop_session_timer", json=payload)
    assert stopped.status_code == 200
    assert stopped.get_json()["session_started_at"] is None
    assert _state_row(course_env)["session_started_at"] is None


def test_poll_returns_server_derived_timer_values(course_env):
    now = datetime.utcnow()
    _activate_presentation(
        course_env,
        presentation_started_at=(now - timedelta(seconds=10)).strftime(
            "%Y-%m-%d %H:%M:%S.%f"
        ),
        presentation_time_cap=300,
        poll_active=1,
        poll_started_at=(now - timedelta(seconds=5)).strftime(
            "%Y-%m-%d %H:%M:%S.%f"
        ),
    )
    _set_state(
        course_env,
        session_started_at=(now - timedelta(seconds=20)).strftime(
            "%Y-%m-%d %H:%M:%S.%f"
        ),
    )

    state = _instructor_client(course_env).get("/api/poll").get_json()["state"]

    assert 288 <= state["presentation_remaining"] <= 291
    assert 33 <= state["poll_remaining"] <= 36
    assert 19 <= state["session_elapsed"] <= 22


def test_discussion_instructor_shows_timer_control_and_connection_status(course_env):
    _set_state(course_env, phase="discussion")

    response = _instructor_client(course_env).get(
        f"/instructor/{course_env['slug']}"
    )

    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert 'id="btn-session-timer"' in html
    assert 'id="instructor-connection-status"' in html


def test_setup_capacity_status_highlights_unassigned_students(course_env):
    client = _instructor_client(course_env)

    html = client.get(f"/instructor/{course_env['slug']}").get_data(as_text=True)
    assert 'class="capacity-summary capacity-warning"' in html
    assert 'id="setup-capacity-warning" role="status"' in html

    _assign_all_active_students(course_env)
    assigned_html = client.get(
        f"/instructor/{course_env['slug']}"
    ).get_data(as_text=True)
    assert 'class="capacity-summary capacity-success"' in assigned_html
    assert 'class="capacity-summary capacity-warning"' not in assigned_html


def test_empty_team_cannot_start_presentation(course_env):
    _set_state(course_env, phase="competition")
    response = _instructor_client(course_env).post(
        "/api/start_presentation",
        json={
            "expected_phase": "competition",
            "expected_session_key": SESSION_KEY,
            "team_id": course_env["teams"]["Team 3"],
            "question_id": course_env["question_id"],
            "time_cap": 300,
        },
    )

    assert response.status_code == 409
    state = _state_row(course_env)
    assert state["active_team_id"] is None
    assert state["active_question_id"] is None


def test_stale_question_from_another_week_cannot_start(course_env):
    _write_catalog_week(course_env, 1)
    _write_catalog_week(course_env, 2)
    client = _instructor_client(course_env)
    selected = client.post(
        "/api/set_discussion_week",
        json={
            "week": 1,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert selected.status_code == 200
    with _connect(course_env) as db:
        week_one_question = db.execute(
            """SELECT id FROM questions
               WHERE course_id = ? AND source_key = 'week-1-q-discussion-1'""",
            (course_env["course_id"],),
        ).fetchone()[0]
    _set_state(course_env, phase="competition", discussion_week=2)

    response = client.post(
        "/api/start_presentation",
        json={
            "expected_phase": "competition",
            "expected_session_key": SESSION_KEY,
            "team_id": course_env["teams"]["Team 1"],
            "question_id": week_one_question,
            "time_cap": 300,
        },
    )

    assert response.status_code == 409
    assert "different week" in response.get_json()["error"]


def test_canonical_catalog_ignores_stale_legacy_base_and_allows_appendix(
        course_env):
    _write_catalog_week(course_env, 2)
    client = _instructor_client(course_env)
    selected = client.post(
        "/api/set_discussion_week",
        json={
            "week": 2,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert selected.status_code == 200
    added = client.post(
        "/api/questions",
        json={
            "title": "Instructor follow-up",
            "content": "Discuss this follow-up.",
            "week": 2,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert added.status_code == 200
    with _connect(course_env) as db:
        stale_base_id = db.execute(
            """INSERT INTO questions
               (course_id, question_num, question_text, title, week_num,
                source_key)
               VALUES (?, 1, 'Old base', 'Old base', 2,
                       'presentation:2:1')""",
            (course_env["course_id"],),
        ).lastrowid
        appendix_id = db.execute(
            """SELECT id FROM questions
               WHERE course_id = ? AND source_key = 'appendix:2:A1'""",
            (course_env["course_id"],),
        ).fetchone()[0]
        db.commit()
    _set_state(course_env, phase="competition")

    competition_page = client.get(f"/instructor/{course_env['slug']}")
    competition_html = competition_page.get_data(as_text=True)
    assert competition_page.status_code == 200
    assert "Discussion week 2" in competition_html
    assert "Instructor follow-up" in competition_html
    assert "Old base" not in competition_html
    assert "no validated presentation question set" not in competition_html

    blocked = client.post(
        "/api/start_presentation",
        json={
            "expected_phase": "competition",
            "expected_session_key": SESSION_KEY,
            "team_id": course_env["teams"]["Team 1"],
            "question_id": stale_base_id,
            "time_cap": 300,
        },
    )
    assert blocked.status_code == 409
    assert "no longer in this week" in blocked.get_json()["error"]

    started = client.post(
        "/api/start_presentation",
        json={
            "expected_phase": "competition",
            "expected_session_key": SESSION_KEY,
            "team_id": course_env["teams"]["Team 1"],
            "question_id": appendix_id,
            "time_cap": 300,
        },
    )
    assert started.status_code == 200


def test_cancel_presentation_requires_explicit_rating_discard(course_env):
    _activate_presentation(course_env)
    with _connect(course_env) as db:
        db.execute(
            f"""INSERT INTO presentation_ratings
               (data_version, course_id, student_id, question_key, session_key,
                presenting_team_id, presenting_team_name, question_id,
                question_title, q1_developed, q2_easy)
               VALUES ('{app_module.APP_VERSION}', ?, ?, 'pres-current', ?, ?, 'Team 1', ?,
                       'Question One', 4, 4)""",
            (
                course_env["course_id"],
                course_env["students"]["s4"],
                SESSION_KEY,
                course_env["teams"]["Team 1"],
                course_env["question_id"],
            ),
        )
        db.commit()
    client = _instructor_client(course_env)

    protected = client.post(
        "/api/cancel_presentation",
        json={"presentation_key": "pres-current"},
    )
    assert protected.status_code == 409
    assert protected.get_json()["requires_discard"] is True
    assert _state_row(course_env)["active_team_id"] is not None

    cancelled = client.post(
        "/api/cancel_presentation",
        json={"presentation_key": "pres-current", "discard_ratings": True},
    )
    assert cancelled.status_code == 200
    assert _state_row(course_env)["active_team_id"] is None
    assert _state_row(course_env)["presentation_history"] == "[]"
    with _connect(course_env) as db:
        assert db.execute(
            "SELECT COUNT(*) FROM presentation_ratings"
        ).fetchone()[0] == 0


def test_login_throttle_blocks_after_three_failures_then_expires(course_env):
    client = app_module.app.test_client()
    route = f"/login/{course_env['slug']}"
    for _ in range(app_module.LOGIN_FAILURE_LIMIT):
        response = client.post(
            route, data={"student_id": "s1", "pin": "0000"}
        )
        assert response.status_code == 302

    blocked = client.post(
        route, data={"student_id": "s1", "pin": "1111"}
    )
    assert blocked.status_code == 429
    assert int(blocked.headers["Retry-After"]) > 0
    assert b"Too many failed login attempts" in blocked.data
    assert b"try again in" in blocked.data

    with _connect(course_env) as db:
        db.execute(
            """UPDATE login_attempts
               SET window_started_at = '2000-01-01 00:00:00',
                   blocked_until = '2000-01-01 00:00:00'"""
        )
        db.commit()
    success = client.post(
        route, data={"student_id": "s1", "pin": "1111"}
    )
    assert success.status_code == 302


def test_student_login_matches_id_case_insensitively(course_env):
    client = app_module.app.test_client()

    response = client.post(
        f"/login/{course_env['slug']}",
        data={"student_id": "S1", "pin": "1111"},
    )

    assert response.status_code == 302
    assert response.headers["Location"].endswith("/dashboard")
    with client.session_transaction() as flask_session:
        # The stored ID is kept verbatim, only the lookup ignores case.
        assert flask_session["student_id"] == "s1"
    assert client.get("/dashboard").status_code == 200


def test_failed_student_login_redirects_back_with_flash(course_env):
    client = app_module.app.test_client()
    route = f"/login/{course_env['slug']}"

    response = client.post(route, data={"student_id": "s1", "pin": "0000"})

    assert response.status_code == 302
    assert response.headers["Location"].endswith(route)
    page = client.get(route)
    assert b"Invalid login for this course." in page.data

    missing = client.post(route, data={"student_id": "s1", "pin": ""})
    assert missing.status_code == 302
    assert missing.headers["Location"].endswith(route)


def test_failed_instructor_login_redirects_back_with_flash(course_env):
    client = app_module.app.test_client()
    route = f"/instructor_login/{course_env['slug']}"

    response = client.post(
        route, data={"username": "instructor", "pin": "0000"}
    )

    assert response.status_code == 302
    assert response.headers["Location"].endswith(route)
    page = client.get(route)
    assert b"Invalid login for this course." in page.data

    missing = client.post(route, data={"username": "", "pin": "0000"})
    assert missing.status_code == 302
    assert missing.headers["Location"].endswith(route)


def test_question_revision_refreshes_same_question_id(course_env):
    _write_catalog_week(course_env, 1)
    instructor = _instructor_client(course_env)
    selected = instructor.post(
        "/api/set_discussion_week",
        json={
            "week": 1,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert selected.status_code == 200
    with _connect(course_env) as db:
        question_id = db.execute(
            """SELECT id FROM questions
               WHERE course_id = ?
                 AND source_key = 'week-1-q-discussion-1'""",
            (course_env["course_id"],),
        ).fetchone()[0]
    _activate_presentation(
        course_env,
        active_question_id=question_id,
        current_question="Discussion week 1",
    )
    student = _student_client(course_env, "s4")

    first = student.get("/api/poll").get_json()["state"]["active_question"]
    revision = first["revision"]
    assert first["title"] == "Discussion week 1"
    assert first["content"] == "Discuss week 1."
    assert "html_content" not in first

    compact = student.get(
        "/api/poll",
        query_string={
            "known_question_id": first["id"],
            "known_question_revision": revision,
        },
    ).get_json()["state"]["active_question"]
    assert compact["content_unchanged"] is True
    assert "content" not in compact
    assert "html_content" not in compact

    canonical_path = (
        Path(config.CLASSES_DIR) / course_env["slug"]
        / "week-1-questions.md"
    )
    canonical_path.write_text(
        "---\ntitle: Discussion week 1\nid: discussion-1\n---\n\n"
        "Version two is longer.\n",
        encoding="utf-8",
    )
    with app_module.app.app_context():
        app_module.sync_presentation_questions(
            course_env["slug"], course_env["course_id"], 1
        )
    refreshed = student.get(
        "/api/poll",
        query_string={
            "known_question_id": first["id"],
            "known_question_revision": revision,
        },
    ).get_json()["state"]["active_question"]
    assert refreshed["id"] == first["id"]
    assert refreshed["revision"] != revision
    assert refreshed["content"] == "Version two is longer."
    assert "html_content" not in refreshed


def test_question_parser_preserves_horizontal_rules_and_fenced_delimiters():
    source = """---
title: First
---

Opening paragraph.

---

Still part of the first question.

```markdown
---
title: Not frontmatter
---
```

---
title: Second
---

Second body.
"""
    entries = app_module.parse_question_blocks(source)

    assert len(entries) == 2
    assert "Still part of the first question" in entries[0][1]
    assert "title: Not frontmatter" in entries[0][1]
    assert entries[1][1] == "Second body."


def test_question_parser_rejects_unclosed_fenced_code_block():
    source = """---
title: First
---

Opening paragraph.

```markdown
---
title: Hidden by the broken fence
---
"""

    with pytest.raises(
        app_module.QuestionParseError, match="Unclosed fenced code block"
    ):
        app_module.parse_question_blocks(source)


def _write_catalog_week(env, week_num):
    class_dir = Path(config.CLASSES_DIR) / env["slug"]
    (class_dir / f"week-{week_num}-questions.md").write_text(
        f"""---
title: Discussion week {week_num}
id: discussion-{week_num}
---

Discuss week {week_num}.
""",
        encoding="utf-8",
    )


def test_runtime_question_readers_accept_utf8_bom(course_env):
    class_dir = Path(config.CLASSES_DIR) / course_env["slug"]
    (class_dir / "week-1-questions.md").write_text(
        "---\ntitle: BOM discussion\nid: bom-discussion\n---\n\nDiscuss it.\n",
        encoding="utf-8-sig",
    )
    client = _instructor_client(course_env)

    bank = client.get("/api/discussion_questions")
    assert bank.status_code == 200
    assert bank.get_json()["questions"][0]["title"] == "BOM discussion"
    selected = client.post(
        "/api/set_discussion_week",
        json={
            "week": 1,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert selected.status_code == 200
    with _connect(course_env) as db:
        row = db.execute(
            """SELECT title, content, source_key FROM questions
               WHERE course_id = ?
                 AND source_key = 'week-1-q-bom-discussion'""",
            (course_env["course_id"],),
        ).fetchone()
    assert dict(row) == {
        "title": "BOM discussion",
        "content": "Discuss it.",
        "source_key": "week-1-q-bom-discussion",
    }
    presentation = app_module.read_presentation_question_index(
        course_env["slug"], 1
    )
    assert [
        (item["title"], item["content"], item["source_key"])
        for item in presentation
    ] == [
        ("BOM discussion", "Discuss it.", "week-1-q-bom-discussion")
    ]


def test_week_selector_uses_one_file_for_both_phases(course_env):
    _write_catalog_week(course_env, 1)
    _write_catalog_week(course_env, 2)
    client = _instructor_client(course_env)

    catalog = client.get("/api/discussion_questions").get_json()
    by_week = {week["num"]: week for week in catalog["weeks"]}
    assert by_week[1]["ready"] is True
    assert by_week[2]["discussion_ready"] is True
    assert by_week[2]["presentation_ready"] is True
    assert by_week[2]["ready"] is True

    selected_week_two = client.post(
        "/api/set_discussion_week",
        json={
            "week": 2,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert selected_week_two.status_code == 200
    assert selected_week_two.get_json()["presentation_ready"] is True
    assert selected_week_two.get_json()["question_sync"] == "synced"
    assert selected_week_two.get_json()["question_count"] == 1
    assert _state_row(course_env)["discussion_week"] == 2
    with _connect(course_env) as db:
        base_count = db.execute(
            """SELECT COUNT(*) FROM questions
               WHERE course_id = ? AND COALESCE(week_num, 1) = 2
                 AND source_key LIKE 'week-2-q-%'""",
            (course_env["course_id"],),
        ).fetchone()[0]
    assert base_count == 1

    selected = client.post(
        "/api/set_discussion_week",
        json={
            "week": 1,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert selected.status_code == 200
    assert selected.get_json()["question_sync"] == "synced"


def test_week_change_rejects_current_session_teammate_thumbs_without_mutation(
        course_env):
    _write_catalog_week(course_env, 2)
    with _connect(course_env) as db:
        db.execute(
            f"""INSERT INTO teammate_thumbs
               (data_version, course_id, session_key, week_num, question_key,
                grader_id, recipient_id)
               VALUES ('{app_module.APP_VERSION}', ?, ?, 1, 'discussion', ?, ?)""",
            (
                course_env["course_id"],
                SESSION_KEY,
                course_env["students"]["s1"],
                course_env["students"]["s2"],
            ),
        )
        before_questions = [
            tuple(row) for row in db.execute(
                """SELECT id, question_num, title, content, week_num, source_key
                   FROM questions WHERE course_id = ? ORDER BY id""",
                (course_env["course_id"],),
            ).fetchall()
        ]
        before_thumb = tuple(db.execute(
            """SELECT session_key, week_num, question_key, grader_id,
                      recipient_id
               FROM teammate_thumbs WHERE course_id = ?""",
            (course_env["course_id"],),
        ).fetchone())
        db.commit()
    before_state = _state_row(course_env)

    response = _instructor_client(course_env).post(
        "/api/set_discussion_week",
        json={
            "week": 2,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )

    assert response.status_code == 409
    error = response.get_json()["error"].lower()
    assert "lecture week" in error
    assert "current session" in error
    assert _state_row(course_env) == before_state
    with _connect(course_env) as db:
        assert [
            tuple(row) for row in db.execute(
                """SELECT id, question_num, title, content, week_num, source_key
                   FROM questions WHERE course_id = ? ORDER BY id""",
                (course_env["course_id"],),
            ).fetchall()
        ] == before_questions
        assert tuple(db.execute(
            """SELECT session_key, week_num, question_key, grader_id,
                      recipient_id
               FROM teammate_thumbs WHERE course_id = ?""",
            (course_env["course_id"],),
        ).fetchone()) == before_thumb


def test_week_change_rejects_other_current_session_ratings(course_env):
    _write_catalog_week(course_env, 2)
    with _connect(course_env) as db:
        db.execute(
            f"""INSERT INTO presentation_ratings
               (data_version, course_id, student_id, question_key, session_key, week_num,
                q1_developed, q2_easy)
               VALUES ('{app_module.APP_VERSION}', ?, ?, 'pres-existing', ?, 1, 4, 5)""",
            (
                course_env["course_id"],
                course_env["students"]["s1"],
                SESSION_KEY,
            ),
        )
        db.commit()
    before_state = _state_row(course_env)

    response = _instructor_client(course_env).post(
        "/api/set_discussion_week",
        json={
            "week": 2,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )

    assert response.status_code == 409
    assert "current session" in response.get_json()["error"].lower()
    assert _state_row(course_env) == before_state
    with _connect(course_env) as db:
        rating = db.execute(
            """SELECT session_key, week_num, q1_developed, q2_easy
               FROM presentation_ratings WHERE course_id = ?""",
            (course_env["course_id"],),
        ).fetchone()
    assert tuple(rating) == (SESSION_KEY, 1, 4, 5)


def test_week_change_rejects_current_session_presentation_history_only(
        course_env):
    _write_catalog_week(course_env, 2)
    history = [{
        "presentation_key": "pres-no-ratings",
        "session_key": SESSION_KEY,
        "week_num": 1,
        "title": "Question One",
        "team_id": course_env["teams"]["Team 1"],
        "team": "Team 1",
        "responses": 0,
    }]
    _set_state(course_env, presentation_history=json.dumps(history))
    before_state = _state_row(course_env)

    response = _instructor_client(course_env).post(
        "/api/set_discussion_week",
        json={
            "week": 2,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )

    assert response.status_code == 409
    assert "current session" in response.get_json()["error"].lower()
    assert _state_row(course_env) == before_state
    assert _history_counts(course_env) == {"thumbs": 0, "ratings": 0}


def test_week_change_rejects_current_session_challenge_round(course_env):
    _write_catalog_week(course_env, 2)
    _seed_live_challenge(course_env)
    before_state = _state_row(course_env)
    with _connect(course_env) as db:
        before_round = tuple(db.execute(
            """SELECT session_key, week_num, presentation_key, challenge_key
               FROM challenge_rounds WHERE course_id = ?""",
            (course_env["course_id"],),
        ).fetchone())

    response = _instructor_client(course_env).post(
        "/api/set_discussion_week",
        json={
            "week": 2,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )

    assert response.status_code == 409
    assert "current session" in response.get_json()["error"].lower()
    assert _state_row(course_env) == before_state
    with _connect(course_env) as db:
        round_row = db.execute(
            """SELECT session_key, week_num, presentation_key, challenge_key
               FROM challenge_rounds WHERE course_id = ?""",
            (course_env["course_id"],),
        ).fetchone()
    assert tuple(round_row) == before_round


def test_old_session_presentation_history_does_not_block_week_change(
        course_env):
    _write_catalog_week(course_env, 2)
    history = [{
        "presentation_key": "pres-old-session",
        "session_key": SESSION_KEY - 1,
        "week_num": 1,
        "title": "Question One",
        "team_id": course_env["teams"]["Team 1"],
        "team": "Team 1",
        "responses": 0,
    }]
    _set_state(course_env, presentation_history=json.dumps(history))

    response = _instructor_client(course_env).post(
        "/api/set_discussion_week",
        json={
            "week": 2,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )

    assert response.status_code == 200
    state = _state_row(course_env)
    assert state["discussion_week"] == 2
    assert json.loads(state["presentation_history"]) == history
    assert _history_counts(course_env) == {"thumbs": 0, "ratings": 0}


def test_same_week_selection_is_idempotent_after_current_session_activity(
        course_env):
    _write_catalog_week(course_env, 1)
    with _connect(course_env) as db:
        db.execute(
            f"""INSERT INTO teammate_thumbs
               (data_version, course_id, session_key, week_num, question_key,
                grader_id, recipient_id)
               VALUES ('{app_module.APP_VERSION}', ?, ?, 1, 'discussion', ?, ?)""",
            (
                course_env["course_id"],
                SESSION_KEY,
                course_env["students"]["s1"],
                course_env["students"]["s2"],
            ),
        )
        db.commit()

    response = _instructor_client(course_env).post(
        "/api/set_discussion_week",
        json={
            "week": 1,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )

    assert response.status_code == 200
    assert _state_row(course_env)["discussion_week"] == 1
    with _connect(course_env) as db:
        thumb = db.execute(
            """SELECT session_key, week_num FROM teammate_thumbs
               WHERE course_id = ?""",
            (course_env["course_id"],),
        ).fetchone()
    assert tuple(thumb) == (SESSION_KEY, 1)


def test_old_session_activity_does_not_block_week_change(course_env):
    _write_catalog_week(course_env, 2)
    with _connect(course_env) as db:
        db.execute(
            f"""INSERT INTO teammate_thumbs
               (data_version, course_id, session_key, week_num, question_key,
                grader_id, recipient_id)
               VALUES ('{app_module.APP_VERSION}', ?, ?, 1, 'discussion', ?, ?)""",
            (
                course_env["course_id"],
                SESSION_KEY - 1,
                course_env["students"]["s1"],
                course_env["students"]["s2"],
            ),
        )
        db.commit()

    response = _instructor_client(course_env).post(
        "/api/set_discussion_week",
        json={
            "week": 2,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )

    assert response.status_code == 200
    assert _state_row(course_env)["discussion_week"] == 2
    with _connect(course_env) as db:
        thumb = db.execute(
            """SELECT session_key, week_num FROM teammate_thumbs
               WHERE course_id = ?""",
            (course_env["course_id"],),
        ).fetchone()
    assert tuple(thumb) == (SESSION_KEY - 1, 1)


def test_new_session_allows_week_change_after_prior_session_activity(course_env):
    _write_catalog_week(course_env, 2)
    with _connect(course_env) as db:
        db.execute(
            f"""INSERT INTO teammate_thumbs
               (data_version, course_id, session_key, week_num, question_key,
                grader_id, recipient_id)
               VALUES ('{app_module.APP_VERSION}', ?, ?, 1, 'discussion', ?, ?)""",
            (
                course_env["course_id"],
                SESSION_KEY,
                course_env["students"]["s1"],
                course_env["students"]["s2"],
            ),
        )
        db.commit()
    _set_state(course_env, phase="ended")
    instructor = _instructor_client(course_env)

    started = instructor.post(
        "/api/set_phase",
        json={
            "phase": "setup",
            "expected_phase": "ended",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert started.status_code == 200
    assert started.get_json()["session_key"] == SESSION_KEY + 1

    changed = instructor.post(
        "/api/set_discussion_week",
        json={
            "week": 2,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY + 1,
        },
    )

    assert changed.status_code == 200
    state = _state_row(course_env)
    assert state["session_key"] == SESSION_KEY + 1
    assert state["discussion_week"] == 2
    with _connect(course_env) as db:
        thumb = db.execute(
            """SELECT session_key, week_num FROM teammate_thumbs
               WHERE course_id = ?""",
            (course_env["course_id"],),
        ).fetchone()
    assert tuple(thumb) == (SESSION_KEY, 1)


def test_appendix_delete_uses_stable_id_and_unposts_current_question(course_env):
    client = _instructor_client(course_env)
    roster_version = _assign_all_active_students(course_env)
    first_add = client.post(
        "/api/questions",
        json={
            "title": "First",
            "content": "First appendix body.",
            "week": 1,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    second_add = client.post(
        "/api/questions",
        json={
            "title": "Second",
            "content": "Second appendix body.",
            "week": 1,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert first_add.get_json()["appendix_id"] == "A1"
    assert second_add.get_json()["appendix_id"] == "A2"

    phase = client.post(
        "/api/set_phase",
        json={
            "phase": "discussion",
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
            "expected_roster_version": roster_version,
        },
    )
    assert phase.status_code == 200
    bank = client.get("/api/discussion_questions").get_json()
    assert bank["current_week"] == 1
    by_id = {question.get("appendix_id"): question for question in bank["questions"]}
    assert set(by_id) == {"A1", "A2"}

    current = by_id["A2"]
    # Simulate the legacy single-question post directly in the DB; the delete
    # route still clears those columns when the deleted question was posted.
    posted_instance_key = "disc-legacy-post"
    _set_state(
        course_env,
        current_discussion_key=posted_instance_key,
        current_discussion_source_key=current["key"],
        current_discussion_title=current["title"],
        current_discussion_content=current["content"],
    )
    with _connect(course_env) as db:
        db.execute(
            f"""INSERT INTO teammate_thumbs
               (data_version, course_id, session_key, question_key, grader_id, recipient_id)
               VALUES ('{app_module.APP_VERSION}', ?, ?, ?, ?, ?)""",
            (
                course_env["course_id"],
                SESSION_KEY,
                posted_instance_key,
                course_env["students"]["s1"],
                course_env["students"]["s2"],
            ),
        )
        db.commit()

    deleted_first = client.post(
        "/api/delete_appendix_question",
        json={
            "appendix_id": "A1",
            "week": 1,
            "expected_phase": "discussion",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert deleted_first.status_code == 200
    assert deleted_first.get_json()["unposted"] is False
    state = _state_row(course_env)
    assert state["current_discussion_key"] == posted_instance_key
    assert state["current_discussion_source_key"] == current["key"]

    stale_repeat = client.post(
        "/api/delete_appendix_question",
        json={
            "appendix_id": "A1",
            "week": 1,
            "expected_phase": "discussion",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert stale_repeat.status_code == 404
    remaining = client.get("/api/discussion_questions").get_json()["questions"]
    assert [question.get("appendix_id") for question in remaining] == ["A2"]

    deleted_current = client.post(
        "/api/delete_appendix_question",
        json={
            "appendix_id": "A2",
            "week": 1,
            "expected_phase": "discussion",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert deleted_current.status_code == 200
    assert deleted_current.get_json()["unposted"] is True
    state = _state_row(course_env)
    assert state["current_discussion_key"] is None
    assert state["current_discussion_source_key"] is None
    with _connect(course_env) as db:
        assert db.execute(
            "SELECT COUNT(*) FROM teammate_thumbs"
        ).fetchone()[0] == 1


def test_appendix_edit_updates_body_and_preserves_label(course_env):
    client = _instructor_client(course_env)
    first_add = client.post(
        "/api/questions",
        json={
            "title": "First",
            "content": "First appendix body.",
            "week": 1,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    second_add = client.post(
        "/api/questions",
        json={
            "title": "Second",
            "content": "Second appendix body.",
            "week": 1,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert first_add.get_json()["appendix_id"] == "A1"
    assert second_add.get_json()["appendix_id"] == "A2"

    edited = client.post(
        "/api/edit_appendix_question",
        json={
            "appendix_id": "a1",
            "title": "First (revised)",
            "content": "Revised appendix body.",
            "week": 1,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert edited.status_code == 200
    payload = edited.get_json()
    assert payload["success"] is True
    assert payload["appendix_id"] == "A1"

    questions = client.get("/api/discussion_questions").get_json()["questions"]
    by_id = {question.get("appendix_id"): question for question in questions}
    assert set(by_id) == {"A1", "A2"}
    assert by_id["A1"]["title"] == "A1: First (revised)"
    assert by_id["A1"]["content"] == "Revised appendix body."
    assert by_id["A2"]["title"] == "A2: Second"

    # The next add continues the label sequence after the preserved labels.
    third_add = client.post(
        "/api/questions",
        json={
            "title": "Third",
            "content": "Third appendix body.",
            "week": 1,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert third_add.get_json()["appendix_id"] == "A3"

    appendix_path = (
        Path(config.DATA_DIR) / course_env["slug"] / "appendix"
        / "week-1-appendix.md"
    )
    source = appendix_path.read_text(encoding="utf-8")
    assert "A1: First (revised)" in source
    assert "Revised appendix body." in source


def test_appendix_edit_validation_and_missing_question(course_env):
    client = _instructor_client(course_env)
    added = client.post(
        "/api/questions",
        json={
            "title": "Original",
            "content": "Original body.",
            "week": 1,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert added.status_code == 200
    base = {
        "week": 1,
        "expected_phase": "setup",
        "expected_session_key": SESSION_KEY,
    }

    missing_fields = client.post(
        "/api/edit_appendix_question",
        json={**base, "appendix_id": "A1", "title": "", "content": "Body."},
    )
    assert missing_fields.status_code == 400
    assert missing_fields.get_json()["error"] == "Title and content required"

    too_long = client.post(
        "/api/edit_appendix_question",
        json={
            **base,
            "appendix_id": "A1",
            "title": "Original",
            "content": "x" * 50001,
        },
    )
    assert too_long.status_code == 400
    assert too_long.get_json()["error"] == "Title or content is too long"

    # Bank questions have no A-number label, so their keys never match.
    bank_key = client.post(
        "/api/edit_appendix_question",
        json={
            **base,
            "appendix_id": "week-1-question-1",
            "title": "Original",
            "content": "Body.",
        },
    )
    assert bank_key.status_code == 400
    assert bank_key.get_json()["error"] == "Appendix question ID required"

    unknown = client.post(
        "/api/edit_appendix_question",
        json={**base, "appendix_id": "A9", "title": "Original", "content": "Body."},
    )
    assert unknown.status_code == 404
    assert unknown.get_json()["error"] == "Appendix question not found"

    # The rejected edits left the stored question untouched.
    questions = client.get("/api/discussion_questions").get_json()["questions"]
    assert [question["title"] for question in questions] == ["A1: Original"]
    assert questions[0]["content"] == "Original body."


def test_appendix_edit_rejected_during_presentations(course_env):
    client = _instructor_client(course_env)
    added = client.post(
        "/api/questions",
        json={
            "title": "Original",
            "content": "Original body.",
            "week": 1,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert added.status_code == 200
    _set_state(course_env, phase="competition")

    edited = client.post(
        "/api/edit_appendix_question",
        json={
            "appendix_id": "A1",
            "title": "Original",
            "content": "Changed body.",
            "week": 1,
            "expected_phase": "competition",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert edited.status_code == 409
    assert "cannot be edited during presentations" in edited.get_json()["error"]


def test_start_presentation_reports_clamped_time_cap(course_env):
    _write_catalog_week(course_env, 1)
    client = _instructor_client(course_env)
    selected = client.post(
        "/api/set_discussion_week",
        json={
            "week": 1,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert selected.status_code == 200
    with _connect(course_env) as db:
        question_id = db.execute(
            """SELECT id FROM questions
               WHERE course_id = ?
                 AND source_key = 'week-1-q-discussion-1'""",
            (course_env["course_id"],),
        ).fetchone()[0]
    _set_state(course_env, phase="competition")

    clamped = client.post(
        "/api/start_presentation",
        json={
            "expected_phase": "competition",
            "expected_session_key": SESSION_KEY,
            "team_id": course_env["teams"]["Team 1"],
            "question_id": question_id,
            "time_cap": 9999,
        },
    )
    assert clamped.status_code == 200
    payload = clamped.get_json()
    assert payload["time_cap"] == 3600
    assert payload["notice"] == (
        "Time cap adjusted to 3600s (allowed range 10 to 3600 seconds)"
    )
    assert _state_row(course_env)["presentation_time_cap"] == 3600

    _set_state(
        course_env,
        active_team_id=None,
        active_question_id=None,
        presentation_started_at=None,
        presentation_created_at=None,
        poll_question_key=None,
    )
    in_range = client.post(
        "/api/start_presentation",
        json={
            "expected_phase": "competition",
            "expected_session_key": SESSION_KEY,
            "team_id": course_env["teams"]["Team 1"],
            "question_id": question_id,
            "time_cap": 120,
        },
    )
    assert in_range.status_code == 200
    payload = in_range.get_json()
    assert payload["time_cap"] == 120
    assert "notice" not in payload
    assert _state_row(course_env)["presentation_time_cap"] == 120


def test_eighty_students_sustain_compact_polling_without_locks(course_env):
    with _connect(course_env) as db:
        for index in range(5, 81):
            student_id = f"s{index}"
            db.execute(
                """INSERT INTO students
                   (course_id, student_id, name, pin, team_id)
                   VALUES (?, ?, ?, ?, ?)""",
                (
                    course_env["course_id"],
                    student_id,
                    f"Student {index}",
                    f"{index % 10000:04d}",
                    course_env["teams"][f"Team {((index - 1) % 4) + 1}"],
                ),
            )
        db.commit()
    _set_state(
        course_env,
        phase="discussion",
        current_discussion_key="traffic-q",
        current_discussion_title="Traffic question",
        current_discussion_content="x" * 32000,
    )
    clients = [
        _student_client(course_env, f"s{index}")
        for index in range(1, 81)
    ]

    def initial_poll(client):
        response = client.get("/api/poll")
        payload = response.get_json()
        return response.status_code, len(response.data), payload["state_version"]

    with ThreadPoolExecutor(max_workers=40) as pool:
        initial = list(pool.map(initial_poll, clients))
        assert all(status == 200 for status, _size, _version in initial)
        assert max(size for _status, size, _version in initial) < 2000

        versions = [version for _status, _size, version in initial]

        def compact_poll(item):
            client, version = item
            response = client.get(f"/api/poll?since={version}")
            return response.status_code, len(response.data), response.get_json()

        for _round in range(3):
            compact = list(pool.map(compact_poll, zip(clients, versions)))
            assert all(status == 200 for status, _size, _data in compact)
            assert max(size for _status, size, _data in compact) < 300
            assert all(
                data["changed"] is False
                for _status, _size, data in compact
            )


def test_instructor_templates_render_new_controls_in_each_phase(course_env):
    client = _instructor_client(course_env)

    setup_html = client.get(
        f"/instructor/{course_env['slug']}"
    ).get_data(as_text=True)
    assert f'data-session-key="{SESSION_KEY}"' in setup_html
    assert 'id="btn-session-timer"' in setup_html

    _set_state(course_env, phase="discussion")
    discussion_html = client.get(
        f"/instructor/{course_env['slug']}"
    ).get_data(as_text=True)
    assert 'id="disc-questions-list"' in discussion_html
    assert 'id="thumb-participation"' in discussion_html
    assert 'id="thumb-team-progress"' in discussion_html
    question_position = discussion_html.index('id="disc-questions-list"')
    appendix_position = discussion_html.index('id="appendix-title"')
    activity_position = discussion_html.index(
        'id="discussion-participation-panel"'
    )
    assert question_position < appendix_position < activity_position

    _activate_presentation(course_env)
    competition_html = client.get(
        f"/instructor/{course_env['slug']}"
    ).get_data(as_text=True)
    assert 'id="btn-cancel-presentation"' in competition_html
    assert 'id="competition-appendix-form"' in competition_html
    assert 'id="poll-non-raters"' not in competition_html


# ---------------------------------------------------------------------------
# state_version fast-polling signal (cheap /api/poll short-circuit)
# ---------------------------------------------------------------------------

def test_state_version_auto_bumps_on_course_state_write(course_env):
    """The schema trigger bumps state_version exactly once per UPDATE, so no
    instructor route can ever forget to signal students."""
    with _connect(course_env) as db:
        before = db.execute(
            "SELECT state_version FROM course_state WHERE course_id = ?",
            (course_env["course_id"],),
        ).fetchone()[0]
        db.execute(
            "UPDATE course_state SET teams_locked = 1 WHERE course_id = ?",
            (course_env["course_id"],),
        )
        db.commit()
        after = db.execute(
            "SELECT state_version FROM course_state WHERE course_id = ?",
            (course_env["course_id"],),
        ).fetchone()[0]
    assert after == before + 1


def test_student_poll_short_circuits_when_state_unchanged(course_env):
    """A student sending since=<current version> gets a tiny changed:false
    response with no full state, so ~1s polling stays cheap."""
    client = _student_client(course_env, "s1")

    first = client.get("/api/poll").get_json()
    assert first["changed"] is True
    assert "state" in first
    version = first["state_version"]

    same = client.get(f"/api/poll?since={version}").get_json()
    assert same["changed"] is False
    assert "state" not in same
    assert same["state_version"] == version
    assert same["poll_interval"] == 1000


def test_student_poll_returns_full_state_after_a_state_change(course_env):
    """After any course_state write bumps the version, a student's next poll
    with the stale version falls back to the full state."""
    client = _student_client(course_env, "s1")
    stale_version = client.get("/api/poll").get_json()["state_version"]

    # Any write to course_state bumps the version (here, picking a week).
    _set_state(course_env, discussion_week=2)

    after = client.get(f"/api/poll?since={stale_version}").get_json()
    assert after["changed"] is True
    assert "state" in after
    assert after["state_version"] > stale_version
    assert after["state"]["discussion_week"] == 2


def test_student_poll_uses_full_path_while_rating_window_open(course_env):
    """An open rating window expires by time without a course_state write, so
    students must keep getting the full state until it closes."""
    _activate_presentation(course_env)
    _set_state(
        course_env,
        poll_active=1,
        poll_started_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
    )
    # s4 is on Team 2; Team 1 is presenting, so s4 is an eligible rater.
    client = _student_client(course_env, "s4")
    version = client.get("/api/poll").get_json()["state_version"]

    same = client.get(f"/api/poll?since={version}").get_json()
    assert same["changed"] is True
    assert "state" in same


def test_expired_rating_poll_returns_compact_close_signal(course_env):
    _activate_presentation(course_env)
    _set_state(
        course_env,
        poll_active=1,
        poll_started_at=(
            datetime.utcnow() - timedelta(seconds=90)
        ).strftime("%Y-%m-%d %H:%M:%S"),
    )
    client = _student_client(course_env, "s4")
    version = client.get("/api/poll").get_json()["state_version"]

    same = client.get(f"/api/poll?since={version}")
    payload = same.get_json()

    assert payload["changed"] is False
    assert payload["poll_closed"] is True
    assert "state" not in payload
    assert len(same.data) < 300


def test_instructor_poll_always_uses_full_path(course_env):
    """Instructors need live participation counts (which change without a
    course_state write), so ?since must never short-circuit for them."""
    instructor = _instructor_client(course_env)
    version = instructor.get("/api/poll").get_json()["state_version"]

    again = instructor.get(f"/api/poll?since={version}").get_json()
    assert again["changed"] is True
    assert "state" in again


def test_student_activity_write_is_throttled(course_env):
    """last_active_at is written on the first poll and at most every ~30s
    after, so 1s polling can't cause a write storm but 'online' stays accurate."""
    client = _student_client(course_env, "s1")
    old = "2000-01-01 00:00:00"

    def set_last_active(value):
        with _connect(course_env) as db:
            db.execute(
                "UPDATE students SET last_active_at = ? WHERE student_id = ?",
                (value, "s1"),
            )
            db.commit()

    def get_last_active():
        with _connect(course_env) as db:
            return db.execute(
                "SELECT last_active_at FROM students WHERE student_id = ?",
                ("s1",),
            ).fetchone()[0]

    # First poll: session not yet acked -> writes, moving last_active_at up.
    set_last_active(old)
    client.get("/api/poll").get_json()
    assert get_last_active() > old

    # Immediate second poll is throttled (session acked, sync stamp fresh)
    # -> last_active_at must NOT move.
    set_last_active(old)
    client.get("/api/poll").get_json()
    assert get_last_active() == old

    # Force the sync stamp into the past; the next poll writes again.
    set_last_active(old)
    with client.session_transaction() as flask_session:
        flask_session["last_active_synced_at"] = (
            datetime.utcnow() - timedelta(seconds=90)
        ).isoformat()
    client.get("/api/poll").get_json()
    assert get_last_active() > old


def test_student_activity_lock_failure_does_not_fail_poll(course_env):
    with app_module.app.app_context():
        database.ensure_schema(course_env["slug"])
    client = _student_client(course_env, "s1")
    with client.session_transaction() as flask_session:
        flask_session["activity_session_key"] = SESSION_KEY
        flask_session["last_active_synced_at"] = (
            datetime.utcnow() - timedelta(seconds=90)
        ).isoformat()

    blocker = sqlite3.connect(course_env["db_path"], timeout=0.1)
    try:
        blocker.execute("BEGIN IMMEDIATE")
        blocker.execute(
            "UPDATE students SET last_active_at = last_active_at WHERE id = ?",
            (course_env["students"]["s1"],),
        )

        response = client.get("/api/poll")

        assert response.status_code == 200
        assert response.get_json()["changed"] is True
        with client.session_transaction() as flask_session:
            assert "last_active_sync_failed_at" in flask_session
    finally:
        blocker.rollback()
        blocker.close()


def test_failed_demo_touch_uses_short_retry_then_success_throttle(
        monkeypatch):
    slug = "demo-instance-touch-test"
    clock = {"now": 100.0}
    calls = []

    def fake_touch(_data_dir, touched_slug):
        calls.append(touched_slug)
        if len(calls) == 1:
            raise OSError("temporary marker failure")

    app_module._demo_instance_touch_last.clear()
    app_module._demo_instance_touch_failed.clear()
    app_module._demo_instance_touch_inflight.clear()
    monkeypatch.setattr(app_module.time, "monotonic", lambda: clock["now"])
    monkeypatch.setattr(app_module, "touch_demo_instance", fake_touch)

    app_module._touch_demo_instance_throttled(slug)
    clock["now"] = 102.0
    app_module._touch_demo_instance_throttled(slug)
    clock["now"] = 106.0
    app_module._touch_demo_instance_throttled(slug)
    clock["now"] = 107.0
    app_module._touch_demo_instance_throttled(slug)

    assert calls == [slug, slug]
    assert app_module._demo_instance_touch_last[slug] == 106.0
    assert slug not in app_module._demo_instance_touch_failed
    assert slug not in app_module._demo_instance_touch_inflight


def test_poll_duration_reads_from_course_yaml(course_env):
    """poll_duration is configurable via course.yaml (clamped 5-300s, default
    40) and flows through to the poll state the client sees."""
    app_module._poll_duration_cache.clear()
    slug = course_env["slug"]
    class_dir = Path(config.CLASSES_DIR) / slug

    assert app_module.get_poll_duration(slug) == 40

    (class_dir / "course.yaml").write_text(
        f"slug: {slug}\nactive: true\npoll_duration: 45\n", encoding="utf-8"
    )
    app_module._poll_duration_cache.clear()
    assert app_module.get_poll_duration(slug) == 45

    # Out-of-range falls back to the default.
    (class_dir / "course.yaml").write_text(
        f"slug: {slug}\nactive: true\npoll_duration: 3\n", encoding="utf-8"
    )
    app_module._poll_duration_cache.clear()
    assert app_module.get_poll_duration(slug) == 40

    # The configured value reaches the client through /api/poll.
    (class_dir / "course.yaml").write_text(
        f"slug: {slug}\nactive: true\npoll_duration: 45\n", encoding="utf-8"
    )
    app_module._poll_duration_cache.clear()
    _activate_presentation(course_env)
    _set_state(
        course_env,
        poll_active=1,
        poll_started_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
    )
    state = _instructor_client(course_env).get("/api/poll").get_json()["state"]
    assert state["poll_duration"] == 45


def test_initial_instructor_timer_uses_configured_poll_duration(
        course_env, monkeypatch):
    fixed_now = datetime(2026, 7, 25, 12, 0, 0)
    started_at = fixed_now - timedelta(seconds=10)
    class_dir = Path(config.CLASSES_DIR) / course_env["slug"]
    (class_dir / "course.yaml").write_text(
        (
            f"slug: {course_env['slug']}\n"
            "active: true\n"
            "poll_duration: 45\n"
        ),
        encoding="utf-8",
    )
    app_module._poll_duration_cache.clear()
    monkeypatch.setattr(app_module, "_utcnow", lambda: fixed_now)
    _activate_presentation(course_env)
    _set_state(
        course_env,
        poll_active=1,
        poll_started_at=started_at.strftime("%Y-%m-%d %H:%M:%S"),
    )

    html = _instructor_client(course_env).get(
        f"/instructor/{course_env['slug']}"
    ).get_data(as_text=True)

    assert 'data-poll-duration="45"' in html
    assert 'data-poll-remaining="35"' in html


def test_instructor_sees_per_team_discussion_thumb_progress(course_env):
    """Live rows distinguish total thumbs from participating students."""
    with _connect(course_env) as db:
        db.execute(
            """INSERT INTO students
               (course_id, student_id, name, pin, team_id)
               VALUES (?, 's5', 'Cara', '5555', ?)""",
            (course_env["course_id"], course_env["teams"]["Team 1"]),
        )
        # Historical data stays stored but must not enter this session's count.
        db.execute(
            f"""INSERT INTO teammate_thumbs
               (data_version, course_id, session_key, week_num, question_key,
                grader_id, recipient_id, grader_team_id, recipient_team_id)
               VALUES ('{app_module.APP_VERSION}', ?, ?, 1, 'discussion', ?, ?, ?, ?)""",
            (
                course_env["course_id"], SESSION_KEY - 1,
                course_env["students"]["s2"],
                course_env["students"]["s1"],
                course_env["teams"]["Team 1"],
                course_env["teams"]["Team 1"],
            ),
        )
        db.commit()
    _set_state(course_env, phase="discussion")

    alice = _student_client(course_env, "s1")
    for recipient in ("s2", "s5"):
        assert alice.post(
            "/api/grade_peer",
            json={"recipient_id": recipient, "selected": True},
        ).status_code == 200

    instructor = _instructor_client(course_env)

    def progress():
        rows = instructor.get("/api/poll").get_json()["state"][
            "thumb_team_progress"
        ]
        assert [row["team_name"] for row in rows] == [
            "Team 1", "Team 2", "Team 3", "Team 4",
        ]
        return {row["team_name"]: row for row in rows}

    rows = progress()
    assert set(rows["Team 1"]) == {
        "team_id", "team_name", "member_count", "eligible_count",
        "participant_count", "thumb_count",
    }
    assert {
        key: rows["Team 1"][key]
        for key in (
            "team_id", "team_name", "member_count", "eligible_count",
            "participant_count", "thumb_count",
        )
    } == {
        "team_id": course_env["teams"]["Team 1"],
        "team_name": "Team 1",
        "member_count": 3,
        "eligible_count": 3,
        "participant_count": 1,
        "thumb_count": 2,
    }
    assert (
        rows["Team 2"]["member_count"],
        rows["Team 2"]["eligible_count"],
        rows["Team 2"]["participant_count"],
        rows["Team 2"]["thumb_count"],
    ) == (1, 0, 0, 0)
    for team_name in ("Team 3", "Team 4"):
        row = rows[team_name]
        assert (
            row["member_count"], row["eligible_count"],
            row["participant_count"], row["thumb_count"],
        ) == (0, 0, 0, 0)

    assert alice.post(
        "/api/grade_peer",
        json={"recipient_id": "s2", "selected": False},
    ).status_code == 200
    row = progress()["Team 1"]
    assert (row["participant_count"], row["thumb_count"]) == (1, 1)

    assert alice.post(
        "/api/grade_peer",
        json={"recipient_id": "s5", "selected": False},
    ).status_code == 200
    row = progress()["Team 1"]
    assert (row["participant_count"], row["thumb_count"]) == (0, 0)
    with _connect(course_env) as db:
        remaining = db.execute(
            "SELECT session_key FROM teammate_thumbs"
        ).fetchall()
    assert [row["session_key"] for row in remaining] == [SESSION_KEY - 1]


def test_presentation_monitor_counts_one_form_per_eligible_student(course_env):
    """A resubmission updates one form and ineligible users stay excluded."""
    with _connect(course_env) as db:
        db.execute(
            """INSERT INTO students
               (course_id, student_id, name, pin, team_id, is_active)
               VALUES (?, 'inactive-rater', 'Inactive Rater', '5555', ?, 0)""",
            (course_env["course_id"], course_env["teams"]["Team 2"]),
        )
        db.commit()
    _activate_presentation(
        course_env,
        poll_active=1,
        poll_started_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S.%f"),
    )
    payload = {
        "presentation_key": "pres-current",
        "q1_developed": 3,
        "q2_easy": 4,
    }
    dana = _student_client(course_env, "s4")
    assert dana.post("/api/submit_rating", json=payload).status_code == 200
    payload.update(q1_developed=5, q2_easy=2)
    assert dana.post("/api/submit_rating", json=payload).status_code == 200

    assert _student_client(course_env, "s1").post(
        "/api/submit_rating", json=payload
    ).status_code == 403
    assert _student_client(course_env, "s3").post(
        "/api/submit_rating", json=payload
    ).status_code == 403

    state = _instructor_client(course_env).get("/api/poll").get_json()["state"]
    assert state["poll_count"] == 1
    assert state["poll_eligible_count"] == 1
    assert "poll_online_eligible_count" not in state
    assert "poll_non_raters" not in state
    with _connect(course_env) as db:
        rows = db.execute(
            """SELECT q1_developed, q2_easy FROM presentation_ratings
               WHERE course_id = ? AND question_key = 'pres-current'""",
            (course_env["course_id"],),
        ).fetchall()
    assert [(row["q1_developed"], row["q2_easy"]) for row in rows] == [(5, 2)]


def test_instructor_sees_per_challenge_submission_and_eligibility_counts(
        course_env, monkeypatch):
    """Each challenge uses its own challenger-team exclusion denominator."""
    now = datetime(2026, 8, 13, 12, 0, 0)
    recent = (now - timedelta(minutes=1)).strftime("%Y-%m-%d %H:%M:%S")
    old = (now - timedelta(minutes=10)).strftime("%Y-%m-%d %H:%M:%S")
    monkeypatch.setattr(app_module, "_utcnow", lambda: now)

    with _connect(course_env) as db:
        extra_specs = (
            ("s5", "Eli", course_env["teams"]["Team 2"], 1, old),
            ("s6", "Finn", course_env["teams"]["Team 3"], 1, recent),
            ("s7", "Gia", course_env["teams"]["Team 3"], 1, old),
            ("s8", "Hana", course_env["teams"]["Team 3"], 1, None),
            ("s9", "Iris", course_env["teams"]["Team 4"], 1, recent),
            (
                "inactive-challenge-rater", "Inactive Challenge Rater",
                course_env["teams"]["Team 4"], 0, recent,
            ),
        )
        extra_ids = {}
        for student_id, name, team_id, is_active, last_active_at in extra_specs:
            extra_ids[student_id] = db.execute(
                """INSERT INTO students
                   (course_id, student_id, name, pin, team_id, is_active,
                    last_active_at)
                   VALUES (?, ?, ?, '5555', ?, ?, ?)""",
                (
                    course_env["course_id"], student_id, name, team_id,
                    is_active, last_active_at,
                ),
            ).lastrowid
        db.execute(
            """UPDATE students SET last_active_at = CASE student_id
                   WHEN 's1' THEN ? WHEN 's2' THEN ? WHEN 's3' THEN ?
                   WHEN 's4' THEN ? ELSE last_active_at END
               WHERE course_id = ?""",
            (recent, recent, recent, recent, course_env["course_id"]),
        )
        db.commit()

    _activate_presentation(course_env)
    challenges = (
        {
            "challenge_key": "pres-current-ch1",
            "challenge_num": 1,
            "challenger_id": course_env["students"]["s4"],
            "challenger_name": "Dana",
            "challenger_team_id": course_env["teams"]["Team 2"],
            "challenger_team_name": "Team 2",
        },
        {
            "challenge_key": "pres-current-ch2",
            "challenge_num": 2,
            "challenger_id": extra_ids["s6"],
            "challenger_name": "Finn",
            "challenger_team_id": course_env["teams"]["Team 3"],
            "challenger_team_name": "Team 3",
        },
    )
    with _connect(course_env) as db:
        for challenge in challenges:
            db.execute(
                f"""INSERT INTO challenge_rounds
                   (data_version, course_id, session_key, week_num, presentation_key,
                    challenge_key, challenge_num, challenger_id,
                    challenger_name, challenger_team_id,
                    challenger_team_name, presenting_team_id,
                    presenting_team_name, question_id, question_title)
                   VALUES ('{app_module.APP_VERSION}', ?, ?, 1, 'pres-current', ?, ?, ?, ?, ?, ?, ?,
                           'Team 1', ?, 'Question One')""",
                (
                    course_env["course_id"], SESSION_KEY,
                    challenge["challenge_key"], challenge["challenge_num"],
                    challenge["challenger_id"], challenge["challenger_name"],
                    challenge["challenger_team_id"],
                    challenge["challenger_team_name"],
                    course_env["teams"]["Team 1"], course_env["question_id"],
                ),
            )
        db.execute(
            "UPDATE course_state SET active_challenges_json = ? WHERE course_id = ?",
            (json.dumps(challenges), course_env["course_id"]),
        )
        db.commit()

    def submit(student_id, challenge_key, score):
        return _student_client(course_env, student_id).post(
            "/api/submit_challenge_rating",
            json={"challenge_key": challenge_key, "score": score},
        )

    assert submit("s6", "pres-current-ch1", 4).status_code == 200
    assert submit("s9", "pres-current-ch1", 5).status_code == 200
    assert submit("s5", "pres-current-ch2", 3).status_code == 200
    assert submit("s5", "pres-current-ch2", 4).status_code == 200

    assert submit("s1", "pres-current-ch1", 3).status_code == 403
    assert submit("s4", "pres-current-ch1", 3).status_code == 403
    assert submit("s3", "pres-current-ch1", 3).status_code == 403

    state = _instructor_client(course_env).get("/api/poll").get_json()["state"]
    assert state["challenge_rating_summaries"] == {
        "pres-current-ch1": {
            "submitted_count": 2,
            "eligible_count": 4,
        },
        "pres-current-ch2": {
            "submitted_count": 1,
            "eligible_count": 3,
        },
    }
    assert "challenge_rating_counts" not in state
    assert "poll_online_eligible_count" not in state
    assert "thumb_online_eligible_count" not in state
    assert "poll_non_raters" not in state
    assert [
        (challenge["challenge_key"], challenge["challenger_name"])
        for challenge in state["active_challenges"]
    ] == [
        ("pres-current-ch1", "Dana"),
        ("pres-current-ch2", "Finn"),
    ]

    instructor_only_fields = {
        "unassigned_count", "poll_count", "poll_eligible_count",
        "thumb_participant_count", "thumb_eligible_count",
        "thumb_team_progress", "challenge_hands",
        "challenge_rating_summaries", "completed_presentation_count",
        "presentation_number",
    }
    student = _student_client(course_env, "s9")
    for route in ("/api/state", "/api/poll"):
        payload = student.get(route).get_json()
        student_state = payload["state"] if route == "/api/poll" else payload
        assert not instructor_only_fields.intersection(student_state)
        assert len(student_state["active_challenges"]) == 2

def test_presentation_monitor_uses_counts_without_rater_identities(
        course_env, monkeypatch):
    """Presentation polling sends submitted and eligible counts only."""
    # Team 1 presents; s4 (Dana, Team 2) is the only eligible rater (s3 is
    # unassigned, s1/s2 are on the presenting team).
    _activate_presentation(course_env)
    queries = []
    real_query_db = app_module.query_db

    def record_query(slug, query, args=(), one=False):
        queries.append(" ".join(query.lower().split()))
        return real_query_db(slug, query, args, one)

    monkeypatch.setattr(app_module, "query_db", record_query)
    state = _instructor_client(course_env).get("/api/poll").get_json()["state"]
    assert state["poll_count"] == 0
    assert state["poll_eligible_count"] == 1
    assert "poll_non_raters" not in state
    assert "poll_online_eligible_count" not in state
    assert "Dana" not in json.dumps(state)
    assert not any(
        "id not in" in query and "presentation_ratings" in query
        for query in queries
    )

    # Students do not receive the instructor-only progress aggregates.
    student_state = _student_client(course_env, "s4").get("/api/poll").get_json()["state"]
    assert "poll_count" not in student_state
    assert "poll_eligible_count" not in student_state
    assert "poll_non_raters" not in student_state

    # Open the window and have s4 rate; the list should clear.
    _set_state(
        course_env,
        poll_active=1,
        poll_started_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
    )
    resp = _student_client(course_env, "s4").post(
        "/api/submit_rating",
        json={"q1_developed": 4, "q2_easy": 5, "presentation_key": "pres-current"},
    )
    assert resp.status_code == 200
    state = _instructor_client(course_env).get("/api/poll").get_json()["state"]
    assert state["poll_count"] == 1
    assert state["poll_eligible_count"] == 1
    assert "poll_non_raters" not in state


def test_ended_phase_poll_interval_keeps_students_responsive(course_env):
    """After End Session students poll at ~5s (not 30s) so a newly started
    session reaches their stale results screen promptly."""
    _set_state(course_env, phase="ended")

    ended = _student_client(course_env).get("/api/poll").get_json()
    assert ended["poll_interval"] == 5000

    _set_state(course_env, phase="discussion")
    active = _student_client(course_env).get("/api/poll").get_json()
    assert active["poll_interval"] == 1000


def test_session_ended_notice_renders_on_landing(course_env):
    """Unauthenticated redirects to /?session=ended show a non-alarming
    explanation via the normal flash mechanism."""
    client = app_module.app.test_client()

    plain = client.get("/")
    assert plain.status_code == 200
    assert "session ended" not in plain.get_data(as_text=True).lower()

    noticed = client.get("/?session=ended")
    assert noticed.status_code == 200
    html = noticed.get_data(as_text=True)
    assert "your session ended" in html.lower()
    assert "flash-success" in html


def test_mathjax_loads_only_on_pages_with_question_math(course_env):
    """MathJax is heavy; login/landing pages render no math and must not load
    it, while the student dashboard (question math) still does."""
    client = app_module.app.test_client()

    login_page = client.get(f"/login/{course_env['slug']}")
    assert login_page.status_code == 200
    assert "MathJax-script" not in login_page.get_data(as_text=True)

    landing = client.get("/")
    assert landing.status_code == 200
    assert "MathJax-script" not in landing.get_data(as_text=True)

    dashboard = _student_client(course_env).get("/dashboard")
    assert dashboard.status_code == 200
    assert "MathJax-script" in dashboard.get_data(as_text=True)


def test_student_search_escapes_like_wildcards(course_env):
    """A literal % or _ in the search box must not act as a LIKE wildcard."""
    with _connect(course_env) as db:
        db.execute(
            """INSERT INTO students (course_id, student_id, name, pin)
               VALUES (?, ?, ?, ?)""",
            (course_env["course_id"], "50%_off", "Wildcard", "5555"),
        )
        db.commit()
    client = _instructor_client(course_env)

    for term in ("%", "_", "50%_off"):
        response = client.get("/api/students", query_string={"search": term})
        assert response.status_code == 200
        payload = response.get_json()
        ids = [s["student_id"] for s in payload["students"]]
        assert ids == ["50%_off"], term
        assert payload["total"] == 1


def test_appendix_add_and_edit_return_question_id_for_select_rebuild(
        course_env):
    """The competition question select is keyed by the numeric question id,
    so the add/edit responses carry it (plus the option title) for the
    client to rebuild options in place without a page reload."""
    client = _instructor_client(course_env)
    added = client.post(
        "/api/questions",
        json={
            "title": "First",
            "content": "First appendix body.",
            "week": 1,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert added.status_code == 200
    payload = added.get_json()
    with _connect(course_env) as db:
        row = db.execute(
            """SELECT id, title FROM questions
               WHERE course_id = ? AND source_key = 'appendix:1:A1'""",
            [course_env["course_id"]],
        ).fetchone()
    assert payload["question_id"] == row["id"]
    assert payload["title"] == row["title"] == "A1: First"

    edited = client.post(
        "/api/edit_appendix_question",
        json={
            "appendix_id": "A1",
            "title": "First (revised)",
            "content": "Revised appendix body.",
            "week": 1,
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert edited.status_code == 200
    edit_payload = edited.get_json()
    # Editing keeps the same questions-table row, so the id is stable.
    assert edit_payload["question_id"] == payload["question_id"]
    assert edit_payload["title"] == "A1: First (revised)"

    # The server-rendered competition select uses the same id and title.
    _set_state(course_env, phase="competition")
    html = client.get(
        f"/instructor/{course_env['slug']}"
    ).get_data(as_text=True)
    assert (
        f'<option value="{row["id"]}">'
        'Appendix — A1: First (revised)</option>'
    ) in html


def test_competition_page_renders_both_presentation_blocks(course_env):
    """Idle and active competition blocks both render server-side with the
    inactive one hidden, so the poll loop can swap them in place (and the
    no-JS initial render still shows the right block)."""
    client = _instructor_client(course_env)
    _set_state(course_env, phase="competition")
    idle_html = client.get(
        f"/instructor/{course_env['slug']}"
    ).get_data(as_text=True)
    assert '<div id="presentation-active" style="display:none">' in idle_html
    assert '<div id="presentation-idle">' in idle_html
    assert 'id="comp-team"' in idle_html

    _set_state(
        course_env,
        active_team_id=course_env["teams"]["Team 1"],
        active_question_id=course_env["question_id"],
        current_question="Question One",
    )
    active_html = client.get(
        f"/instructor/{course_env['slug']}"
    ).get_data(as_text=True)
    assert '<div id="presentation-active">' in active_html
    assert '<div id="presentation-idle" style="display:none">' in active_html
    assert 'id="timer-box"' in active_html


def test_legacy_peer_reviews_migrate_into_teammate_thumbs(course_env):
    """Databases created before teammate_thumbs existed keep their legacy
    thumbs: _ensure_schema_locked copies peer_reviews rows (score > 0) once,
    idempotently, and tolerates databases whose legacy table was dropped."""
    with _connect(course_env) as db:
        db.execute(
            "INSERT INTO peer_reviews"
            " (course_id, grader_id, recipient_id, criterion, score)"
            " VALUES (?, ?, ?, 'overall', 1)",
            (
                course_env["course_id"],
                course_env["students"]["s1"],
                course_env["students"]["s2"],
            ),
        )
        db.execute(
            "INSERT INTO peer_reviews"
            " (course_id, grader_id, recipient_id, criterion, score)"
            " VALUES (?, ?, ?, 'overall', 0)",
            (
                course_env["course_id"],
                course_env["students"]["s1"],
                course_env["students"]["s3"],
            ),
        )
        db.commit()

        database._ensure_schema_locked(db)
        db.commit()
        legacy = db.execute(
            "SELECT * FROM teammate_thumbs WHERE question_key = 'legacy'"
        ).fetchall()
        assert len(legacy) == 1
        assert legacy[0]["grader_id"] == course_env["students"]["s1"]
        assert legacy[0]["recipient_id"] == course_env["students"]["s2"]

        # Idempotent: a second schema pass does not duplicate the row.
        database._ensure_schema_locked(db)
        db.commit()
        count = db.execute(
            "SELECT COUNT(*) FROM teammate_thumbs WHERE question_key = 'legacy'"
        ).fetchone()[0]
        assert count == 1

        # Databases without the legacy table migrate cleanly (gate works).
        db.execute("DROP TABLE peer_reviews")
        db.commit()
        database._ensure_schema_locked(db)
        db.commit()

def _seed_live_challenge(env):
    challenge = {
        "challenge_key": "pres-current-ch1",
        "challenge_num": 1,
        "challenger_id": env["students"]["s4"],
        "challenger_name": "Dana",
        "challenger_team_id": env["teams"]["Team 2"],
        "challenger_team_name": "Team 2",
    }
    with _connect(env) as db:
        db.execute(
            f"""INSERT INTO challenge_rounds
               (data_version, course_id, session_key, week_num, presentation_key,
                challenge_key, challenge_num, challenger_id, challenger_name,
                challenger_team_id, challenger_team_name, presenting_team_id,
                presenting_team_name, question_id, question_title)
               VALUES ('{app_module.APP_VERSION}', ?, ?, 1, 'pres-current', ?, 1, ?, 'Dana', ?, 'Team 2',
                       ?, 'Team 1', ?, 'Question One')""",
            (
                env["course_id"], SESSION_KEY, challenge["challenge_key"],
                env["students"]["s4"], env["teams"]["Team 2"],
                env["teams"]["Team 1"], env["question_id"],
            ),
        )
        db.execute(
            """UPDATE course_state SET active_challenges_json = ?
               WHERE course_id = ?""",
            (json.dumps([challenge]), env["course_id"]),
        )
        db.commit()
    return challenge


def test_stop_poll_cutoff_is_idempotent_and_new_poll_reopens_scopes(
        course_env, monkeypatch):
    clock = {"now": datetime(2026, 8, 13, 12, 0, 0)}
    monkeypatch.setattr(app_module, "_utcnow", lambda: clock["now"])
    _activate_presentation(
        course_env,
        poll_active=1,
        poll_started_at=(clock["now"] - timedelta(seconds=10)).strftime(
            "%Y-%m-%d %H:%M:%S.%f"
        ),
    )
    _seed_live_challenge(course_env)
    instructor = _instructor_client(course_env)
    payload = {
        "presentation_key": "pres-current",
        "expected_phase": "competition",
        "expected_session_key": SESSION_KEY,
    }

    first = instructor.post("/api/stop_poll", json=payload)
    assert first.status_code == 200
    assert first.get_json()["ratings_settling_remaining"] == 3
    state = _state_row(course_env)
    cutoff = state["poll_closed_at"]
    assert state["challenge_ratings_closed_at"] == cutoff

    clock["now"] += timedelta(seconds=1)
    repeated = instructor.post("/api/stop_poll", json=payload)
    assert repeated.status_code == 200
    assert repeated.get_json()["already_stopped"] is True
    assert repeated.get_json()["ratings_settling_remaining"] == 2
    assert _state_row(course_env)["poll_closed_at"] == cutoff

    clock["now"] += timedelta(seconds=2)
    started = instructor.post("/api/start_poll", json=payload)
    assert started.status_code == 200
    state = _state_row(course_env)
    assert state["poll_closed_at"] is None
    assert state["challenge_ratings_closed_at"] is None
    live = instructor.get("/api/poll").get_json()["state"]
    assert live["challenge_ratings_open"] is True


def test_challenge_rating_uses_arrival_time_across_transition_lock(
        course_env, monkeypatch):
    cutoff = datetime(2026, 8, 13, 12, 0, 0)
    clock = {"now": cutoff}
    monkeypatch.setattr(app_module, "_utcnow", lambda: clock["now"])
    _activate_presentation(course_env)
    challenge = _seed_live_challenge(course_env)
    with _connect(course_env) as db:
        db.execute(
            """INSERT INTO students
               (course_id, student_id, name, pin, team_id)
               VALUES (?, 's5', 'Eli', '5555', ?)""",
            (course_env["course_id"], course_env["teams"]["Team 3"]),
        )
        db.commit()

    instructor = _instructor_client(course_env)
    payload = {
        "presentation_key": "pres-current",
        "expected_phase": "competition",
        "expected_session_key": SESSION_KEY,
    }
    closing = instructor.post("/api/next_presentation", json=payload)
    assert closing.status_code == 409
    assert closing.get_json()["ratings_settling_remaining"] == 3
    state = instructor.get("/api/poll").get_json()["state"]
    assert state["challenge_ratings_open"] is False

    clock["now"] = cutoff + timedelta(seconds=2)
    student = _student_client(course_env, "s5")
    with student.session_transaction() as flask_session:
        flask_session["activity_session_key"] = SESSION_KEY
        flask_session["last_active_synced_at"] = clock["now"].isoformat()
    original_get_db = app_module.get_db

    class DelayedWriteConnection:
        def __init__(self, connection):
            self._connection = connection

        def execute(self, sql, *args, **kwargs):
            if sql == "BEGIN IMMEDIATE":
                clock["now"] = cutoff + timedelta(seconds=4)
            return self._connection.execute(sql, *args, **kwargs)

        def __getattr__(self, name):
            return getattr(self._connection, name)

    monkeypatch.setattr(
        app_module,
        "get_db",
        lambda slug: DelayedWriteConnection(original_get_db(slug)),
    )
    saved = student.post(
        "/api/submit_challenge_rating",
        json={"challenge_key": challenge["challenge_key"], "score": 5},
    )
    assert saved.status_code == 200
    with _connect(course_env) as db:
        assert db.execute("SELECT COUNT(*) FROM challenge_ratings").fetchone()[0] == 1

    monkeypatch.setattr(app_module, "get_db", original_get_db)
    finished = instructor.post("/api/next_presentation", json=payload)
    assert finished.status_code == 200


def test_presentation_without_any_rating_scope_finishes_immediately(
        course_env, monkeypatch):
    monkeypatch.setattr(
        app_module, "_utcnow", lambda: datetime(2026, 8, 13, 12, 0, 0)
    )
    _activate_presentation(course_env)

    response = _instructor_client(course_env).post(
        "/api/next_presentation",
        json={"presentation_key": "pres-current"},
    )

    assert response.status_code == 200
    assert _state_row(course_env)["active_team_id"] is None



def test_roster_rejects_ids_that_differ_only_by_case(course_env):
    payload = (
        "student_id,name,pin\n"
        "new-id,First,5555\n"
        "NEW-ID,Second,6666\n"
    ).encode("utf-8")
    count_before = _active_student_count(course_env)

    response = _instructor_client(course_env).post(
        "/api/upload_roster",
        data={"file": (io.BytesIO(payload), "roster.csv")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 400
    details = response.get_json()["details"]
    assert any("duplicate student ID" in detail for detail in details)
    assert _active_student_count(course_env) == count_before


def test_roster_case_change_updates_existing_identity(course_env):
    payload = (
        "student_id,name,pin\n"
        "S1,Alice Updated,5555\n"
        "s2,Bob,2222\n"
        "s3,Unassigned,3333\n"
        "s4,Dana,4444\n"
    ).encode("utf-8")
    expected_state = {
        "expected_phase": "setup",
        "expected_session_key": str(SESSION_KEY),
        "expected_roster_version": "0",
    }
    client = _instructor_client(course_env)

    preview = client.post(
        "/api/upload_roster",
        data={
            **expected_state,
            "file": (io.BytesIO(payload), "roster.csv"),
        },
        content_type="multipart/form-data",
    )

    assert preview.status_code == 200
    preview_data = preview.get_json()
    assert {
        key: preview_data[key]
        for key in ("added", "reactivated", "updated", "removed")
    } == {
        "added": 0,
        "reactivated": 0,
        "updated": 1,
        "removed": 0,
    }

    confirmed = client.post(
        "/api/upload_roster",
        data={
            **expected_state,
            "confirm": "true",
            "preview_token": preview_data["preview_token"],
            "file": (io.BytesIO(payload), "roster.csv"),
        },
        content_type="multipart/form-data",
    )

    assert confirmed.status_code == 200
    with _connect(course_env) as db:
        rows = db.execute(
            """SELECT id, student_id, name, pin FROM students
               WHERE course_id = ? AND LOWER(student_id) = 's1'""",
            (course_env["course_id"],),
        ).fetchall()
    assert len(rows) == 1
    assert dict(rows[0]) == {
        "id": course_env["students"]["s1"],
        "student_id": "s1",
        "name": "Alice Updated",
        "pin": "5555",
    }


def _select_test_challenger(course_env, student_id="s4"):
    student = _student_client(course_env, student_id)
    raised = student.post(
        "/api/raise_hand",
        json={"presentation_key": "pres-current"},
    )
    assert raised.status_code == 200
    selected = _instructor_client(course_env).post(
        "/api/select_challenger",
        json={
            "presentation_key": "pres-current",
            "student_id": course_env["students"][student_id],
        },
    )
    assert selected.status_code == 200
    return selected.get_json()["challenge_key"]


def test_compact_poll_reads_manual_cutoff_during_active_window(
        course_env, monkeypatch):
    now = datetime(2026, 8, 13, 13, 0, 0)
    monkeypatch.setattr(app_module, "_utcnow", lambda: now)
    _activate_presentation(
        course_env,
        poll_active=1,
        poll_started_at=now.strftime("%Y-%m-%d %H:%M:%S.%f"),
    )
    client = _student_client(course_env, "s4")

    full = client.get("/api/poll")
    assert full.status_code == 200
    version = full.get_json()["state_version"]

    compact = client.get("/api/poll", query_string={"since": version})

    assert compact.status_code == 200
    payload = compact.get_json()
    assert payload["changed"] is True
    assert payload["state"]["poll_active"] is True


def test_blank_challenger_name_uses_roster_id_and_restores_saved_controls(
        course_env):
    with _connect(course_env) as db:
        db.execute("UPDATE students SET name = NULL WHERE student_id = 's4'")
        db.execute(
            """INSERT INTO students
               (course_id, student_id, name, pin, team_id)
               VALUES (?, 's5', 'Eli', '5555', ?)""",
            (course_env["course_id"], course_env["teams"]["Team 3"]),
        )
        db.commit()
    _activate_presentation(course_env)
    challenger = _student_client(course_env, "s4")
    assert challenger.post(
        "/api/raise_hand",
        json={"presentation_key": "pres-current"},
    ).status_code == 200
    raised_state = challenger.get("/api/my_responses").get_json()
    assert raised_state["challenge_hand_raised"] is True
    assert raised_state["challenge_ratings"] == {}

    # Simulate an older blank snapshot to verify the read path also falls back.
    with _connect(course_env) as db:
        db.execute("UPDATE challenge_hands SET student_name = NULL")
        db.commit()
    instructor = _instructor_client(course_env)
    hands = instructor.get("/api/poll").get_json()["state"]["challenge_hands"]
    assert hands[0]["student_name"] == "s4"

    selected = instructor.post(
        "/api/select_challenger",
        json={
            "presentation_key": "pres-current",
            "student_id": course_env["students"]["s4"],
        },
    )
    assert selected.status_code == 200
    challenge_key = selected.get_json()["challenge_key"]
    assert challenger.get(
        "/api/my_responses"
    ).get_json()["challenge_hand_raised"] is False

    rater = _student_client(course_env, "s5")
    card = rater.get("/api/poll").get_json()["state"]["active_challenges"][0]
    assert card["challenger_name"] == "s4"
    assert rater.post(
        "/api/submit_challenge_rating",
        json={"challenge_key": challenge_key, "score": 4},
    ).status_code == 200
    saved = rater.get("/api/my_responses").get_json()
    assert saved["challenge_ratings"] == {challenge_key: 4}

    with _connect(course_env) as db:
        db.execute("UPDATE challenge_ratings SET challenger_name = NULL")
        db.commit()
    _set_state(course_env, phase="ended")
    results = rater.get("/api/poll").get_json()
    assert results["top_challengers"] == [{"name": "s4", "rank": 1}]
    summary = instructor.get(
        f"/instructor/{course_env['slug']}"
    ).get_data(as_text=True)
    assert "session:</strong> 1" in summary
    assert "Top Challenger (by avg rating)" in summary
    assert "#1 s4: 4.0 average (1 submitted rating)" in summary


def test_ended_summary_keeps_all_first_place_challenger_ties(
        course_env, monkeypatch):
    monkeypatch.setattr(
        app_module,
        "_compute_top_challengers",
        lambda *_args: [
            {
                "id": 101, "name": "Leader A", "rank": 1,
                "avg_score": 4.75, "rating_count": 4,
            },
            {
                "id": 102, "name": "Leader B", "rank": 1,
                "avg_score": 4.75, "rating_count": 4,
            },
            {
                "id": 103, "name": "Runner Up", "rank": 3,
                "avg_score": 4.5, "rating_count": 4,
            },
        ],
    )
    _set_state(course_env, phase="ended")

    html = _instructor_client(course_env).get(
        f"/instructor/{course_env['slug']}"
    ).get_data(as_text=True)

    assert "Leader A" in html
    assert "Leader B" in html
    assert "Runner Up" not in html


def test_clear_challenger_settles_queued_ratings_before_discard(
        course_env, monkeypatch):
    clock = {"now": datetime(2026, 8, 13, 14, 0, 0)}
    monkeypatch.setattr(app_module, "_utcnow", lambda: clock["now"])
    with _connect(course_env) as db:
        for student_id, name, team_name in (
            ("s5", "Eli", "Team 3"),
            ("s6", "Finn", "Team 4"),
        ):
            new_student_id = db.execute(
                """INSERT INTO students
                   (course_id, student_id, name, pin, team_id)
                   VALUES (?, ?, ?, '5555', ?)""",
                (
                    course_env["course_id"], student_id, name,
                    course_env["teams"][team_name],
                ),
            ).lastrowid
            course_env["students"][student_id] = new_student_id
        db.commit()
    _activate_presentation(course_env)
    challenge_key = _select_test_challenger(course_env)
    remaining_challenge_key = _select_test_challenger(
        course_env, student_id="s5"
    )
    instructor = _instructor_client(course_env)

    settling = instructor.post(
        "/api/clear_challenger",
        json={
            "presentation_key": "pres-current",
            "challenge_key": challenge_key,
        },
    )
    assert settling.status_code == 409
    assert settling.get_json()["ratings_settling"] is True

    # This models a request that arrived before the cutoff but obtained the
    # SQLite write lock after the instructor's close request.
    clock["now"] -= timedelta(milliseconds=1)
    queued = _student_client(course_env, "s6").post(
        "/api/submit_challenge_rating",
        json={"challenge_key": challenge_key, "score": 5},
    )
    assert queued.status_code == 200

    clock["now"] += timedelta(
        seconds=app_module.POLL_SUBMISSION_GRACE_SECONDS,
        milliseconds=1,
    )
    protected = instructor.post(
        "/api/clear_challenger",
        json={
            "presentation_key": "pres-current",
            "challenge_key": challenge_key,
        },
    )
    assert protected.status_code == 409
    protected_data = protected.get_json()
    assert protected_data["requires_discard"] is True
    assert protected_data["challenge_rating_count"] == 1
    assert protected_data["rating_count"] == 1

    cleared = instructor.post(
        "/api/clear_challenger",
        json={
            "presentation_key": "pres-current",
            "challenge_key": challenge_key,
            "discard_ratings": True,
        },
    )
    assert cleared.status_code == 200
    assert cleared.get_json()["discarded_challenge_ratings"] == 1
    assert _state_row(course_env)["challenge_ratings_closed_at"] is None

    remaining_rating = _student_client(course_env, "s4").post(
        "/api/submit_challenge_rating",
        json={"challenge_key": remaining_challenge_key, "score": 3},
    )
    assert remaining_rating.status_code == 200

    duplicate = instructor.post(
        "/api/clear_challenger",
        json={
            "presentation_key": "pres-current",
            "challenge_key": challenge_key,
        },
    )
    assert duplicate.status_code == 200
    assert duplicate.get_json()["already_cleared"] is True
    assert _state_row(course_env)["challenge_ratings_closed_at"] is None
    with _connect(course_env) as db:
        rounds = db.execute(
            "SELECT challenge_key FROM challenge_rounds"
        ).fetchall()
        ratings = db.execute(
            "SELECT challenge_key, score FROM challenge_ratings"
        ).fetchall()
    assert [row["challenge_key"] for row in rounds] == [
        remaining_challenge_key
    ]
    assert [dict(row) for row in ratings] == [{
        "challenge_key": remaining_challenge_key,
        "score": 3,
    }]


def test_cancel_presentation_reports_both_rating_types_before_discard(
        course_env, monkeypatch):
    now = datetime(2026, 8, 13, 15, 0, 0)
    monkeypatch.setattr(app_module, "_utcnow", lambda: now)
    _activate_presentation(course_env)
    challenge = {
        "challenge_key": "pres-current-ch1",
        "challenge_num": 1,
        "challenger_id": course_env["students"]["s4"],
        "challenger_name": "Dana",
        "challenger_team_id": course_env["teams"]["Team 2"],
        "challenger_team_name": "Team 2",
    }
    with _connect(course_env) as db:
        db.execute(
            """UPDATE course_state
               SET active_challenges_json = ?,
                   challenge_ratings_closed_at = ?
               WHERE course_id = ?""",
            (
                json.dumps([challenge]),
                (now - timedelta(seconds=10)).strftime(
                    "%Y-%m-%d %H:%M:%S.%f"
                ),
                course_env["course_id"],
            ),
        )
        db.execute(
            f"""INSERT INTO challenge_rounds
               (data_version, course_id, session_key, week_num, presentation_key,
                challenge_key, challenge_num, challenger_id, challenger_name,
                challenger_team_id, challenger_team_name,
                presenting_team_id, presenting_team_name,
                question_id, question_title)
               VALUES ('{app_module.APP_VERSION}', ?, ?, 1, 'pres-current', 'pres-current-ch1', 1,
                       ?, 'Dana', ?, 'Team 2', ?, 'Team 1', ?,
                       'Question One')""",
            (
                course_env["course_id"], SESSION_KEY,
                course_env["students"]["s4"],
                course_env["teams"]["Team 2"],
                course_env["teams"]["Team 1"],
                course_env["question_id"],
            ),
        )
        db.execute(
            f"""INSERT INTO presentation_ratings
               (data_version, course_id, student_id, question_key, session_key, week_num,
                q1_developed, q2_easy)
               VALUES ('{app_module.APP_VERSION}', ?, ?, 'pres-current', ?, 1, 4, 5)""",
            (
                course_env["course_id"],
                course_env["students"]["s4"],
                SESSION_KEY,
            ),
        )
        db.execute(
            f"""INSERT INTO challenge_ratings
               (data_version, course_id, session_key, week_num, challenge_key,
                presentation_key, challenger_id, challenger_name,
                challenger_team_id, challenger_team_name,
                rater_id, rater_name, rater_team_id, rater_team_name, score)
               VALUES ('{app_module.APP_VERSION}', ?, ?, 1, 'pres-current-ch1', 'pres-current',
                       ?, 'Dana', ?, 'Team 2', ?, 'Alice', ?, 'Team 1', 4)""",
            (
                course_env["course_id"], SESSION_KEY,
                course_env["students"]["s4"],
                course_env["teams"]["Team 2"],
                course_env["students"]["s1"],
                course_env["teams"]["Team 1"],
            ),
        )
        db.commit()
    instructor = _instructor_client(course_env)

    protected = instructor.post(
        "/api/cancel_presentation",
        json={"presentation_key": "pres-current"},
    )

    assert protected.status_code == 409
    data = protected.get_json()
    assert data["requires_discard"] is True
    assert data["presentation_rating_count"] == 1
    assert data["challenge_rating_count"] == 1
    assert data["rating_count"] == 2

    cancelled = instructor.post(
        "/api/cancel_presentation",
        json={
            "presentation_key": "pres-current",
            "discard_ratings": True,
        },
    )
    assert cancelled.status_code == 200
    assert cancelled.get_json()["discarded_ratings"] == 2
    with _connect(course_env) as db:
        assert db.execute(
            "SELECT COUNT(*) FROM presentation_ratings"
        ).fetchone()[0] == 0
        assert db.execute(
            "SELECT COUNT(*) FROM challenge_ratings"
        ).fetchone()[0] == 0
        assert db.execute(
            "SELECT COUNT(*) FROM challenge_rounds"
        ).fetchone()[0] == 0


def test_export_includes_current_week_challenge_rounds_and_context(course_env):
    from openpyxl import load_workbook

    with _connect(course_env) as db:
        db.execute("UPDATE students SET name = NULL WHERE student_id = 's4'")
        rounds = (
            (1, "pres-current-ch1", "pres-current", 1),
            (1, "pres-current-ch2", "pres-current", 2),
            (2, "pres-old-ch1", "pres-old", 1),
        )
        for week, challenge_key, presentation_key, challenge_num in rounds:
            db.execute(
                f"""INSERT INTO challenge_rounds
                   (data_version, course_id, session_key, week_num, presentation_key,
                    challenge_key, challenge_num, challenger_id,
                    challenger_name, challenger_team_id,
                    challenger_team_name, presenting_team_id,
                    presenting_team_name, question_id, question_title)
                   VALUES ('{app_module.APP_VERSION}', ?, ?, ?, ?, ?, ?, ?, NULL, ?, 'Team 2', ?,
                           'Team 1', ?, 'Question One')""",
                (
                    course_env["course_id"], SESSION_KEY, week,
                    presentation_key, challenge_key, challenge_num,
                    course_env["students"]["s4"],
                    course_env["teams"]["Team 2"],
                    course_env["teams"]["Team 1"],
                    course_env["question_id"],
                ),
            )
        db.execute(
            f"""INSERT INTO challenge_ratings
               (data_version, course_id, session_key, week_num, challenge_key,
                presentation_key, challenger_id, challenger_name,
                challenger_team_id, challenger_team_name,
                rater_id, rater_name, rater_team_id, rater_team_name, score)
               VALUES ('{app_module.APP_VERSION}', ?, ?, 1, 'pres-current-ch1', 'pres-current', ?,
                       NULL, ?, 'Team 2', ?, 'Alice', ?, 'Team 1', 4)""",
            (
                course_env["course_id"], SESSION_KEY,
                course_env["students"]["s4"],
                course_env["teams"]["Team 2"],
                course_env["students"]["s1"],
                course_env["teams"]["Team 1"],
            ),
        )
        db.execute(
            "UPDATE challenge_rounds SET data_version = ?",
            (app_module.APP_VERSION,),
        )
        db.execute(
            "UPDATE challenge_ratings SET data_version = ?",
            (app_module.APP_VERSION,),
        )
        db.commit()

    response = _instructor_client(course_env).get(
        f"/export/{course_env['slug']}"
    )

    assert response.status_code == 200
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        workbook = load_workbook(
            io.BytesIO(archive.read("course_data.xlsx")), read_only=True
        )

    def sheet_rows(sheet_name):
        values = list(workbook[sheet_name].iter_rows(values_only=True))
        return [dict(zip(values[0], row)) for row in values[1:]]

    summary = {
        row[0]: row[1]
        for row in workbook["Summary"].iter_rows(values_only=True)
        if row[0]
    }
    assert summary["Week Challenge Rounds"] == 2
    assert summary["Week Challenge Ratings"] == 1

    round_rows = sheet_rows("Challenge Rounds")
    assert {row["challenge_key"] for row in round_rows} == {
        "pres-current-ch1", "pres-current-ch2",
    }
    rounds_by_key = {row["challenge_key"]: row for row in round_rows}
    rated_round = rounds_by_key["pres-current-ch1"]
    assert rated_round["challenger_id"] == "s4"
    assert rated_round["challenger_name"] == "s4"
    assert rated_round["presenting_team"] == "Team 1"
    assert rated_round["question_id"] == course_env["question_id"]
    assert rated_round["question_title"] == "Question One"
    assert rated_round["ratings_submitted"] == 1
    assert rated_round["average_score_1to5"] == 4
    empty_round = rounds_by_key["pres-current-ch2"]
    assert empty_round["ratings_submitted"] == 0
    assert empty_round["average_score_1to5"] is None

    rating_rows = sheet_rows("Challenge Ratings")
    assert len(rating_rows) == 1
    rating = rating_rows[0]
    assert rating["challenger_id"] == "s4"
    assert rating["challenger_name"] == "s4"
    assert rating["presenting_team"] == "Team 1"
    assert rating["question_id"] == course_env["question_id"]
    assert rating["question_title"] == "Question One"


def test_natural_poll_close_settles_presentation_and_challenge_once(
        course_env, monkeypatch):
    started_at = datetime(2026, 8, 14, 9, 0, 0)
    clock = {"now": started_at + timedelta(seconds=40)}
    monkeypatch.setattr(app_module, "_utcnow", lambda: clock["now"])
    with _connect(course_env) as db:
        for student_id in ("s5", "s6"):
            db.execute(
                """INSERT INTO students
                   (course_id, student_id, name, pin, team_id)
                   VALUES (?, ?, ?, '5555', ?)""",
                (
                    course_env["course_id"], student_id, student_id,
                    course_env["teams"]["Team 3"],
                ),
            )
        db.commit()
    _activate_presentation(
        course_env,
        poll_active=1,
        poll_started_at=started_at.strftime("%Y-%m-%d %H:%M:%S.%f"),
    )
    challenge = _seed_live_challenge(course_env)
    instructor = _instructor_client(course_env)

    state = instructor.get("/api/poll").get_json()["state"]
    assert state["poll_active"] is False
    assert state["challenge_ratings_open"] is False
    assert state["ratings_settling"] is True
    assert state["ratings_settling_remaining"] == 3
    assert _state_row(course_env)["challenge_ratings_closed_at"] is None

    clock["now"] = started_at + timedelta(seconds=42)
    accepted = _student_client(course_env, "s5").post(
        "/api/submit_challenge_rating",
        json={"challenge_key": challenge["challenge_key"], "score": 5},
    )
    assert accepted.status_code == 200

    clock["now"] = started_at + timedelta(seconds=43)
    rejected = _student_client(course_env, "s6").post(
        "/api/submit_challenge_rating",
        json={"challenge_key": challenge["challenge_key"], "score": 4},
    )
    assert rejected.status_code == 403

    finished = instructor.post(
        "/api/next_presentation",
        json={"presentation_key": "pres-current"},
    )
    assert finished.status_code == 200
    with _connect(course_env) as db:
        rows = db.execute(
            "SELECT score FROM challenge_ratings ORDER BY id"
        ).fetchall()
    assert [row["score"] for row in rows] == [5]


def test_late_challenger_waits_for_a_new_rating_poll(
        course_env, monkeypatch):
    started_at = datetime(2026, 8, 14, 10, 0, 0)
    clock = {"now": started_at + timedelta(seconds=43)}
    monkeypatch.setattr(app_module, "_utcnow", lambda: clock["now"])
    with _connect(course_env) as db:
        db.execute(
            """INSERT INTO students
               (course_id, student_id, name, pin, team_id)
               VALUES (?, 's5', 'Eli', '5555', ?)""",
            (course_env["course_id"], course_env["teams"]["Team 3"]),
        )
        db.commit()
    _activate_presentation(
        course_env,
        poll_active=1,
        poll_started_at=started_at.strftime("%Y-%m-%d %H:%M:%S.%f"),
    )
    challenger = _student_client(course_env, "s4")
    assert challenger.post(
        "/api/raise_hand",
        json={"presentation_key": "pres-current"},
    ).status_code == 200

    instructor = _instructor_client(course_env)
    selected = instructor.post(
        "/api/select_challenger",
        json={
            "presentation_key": "pres-current",
            "student_id": course_env["students"]["s4"],
        },
    )
    assert selected.status_code == 200
    selected_data = selected.get_json()
    assert selected_data["challenge_ratings_open"] is False
    challenge_key = selected_data["challenge_key"]

    rater = _student_client(course_env, "s5")
    closed = rater.post(
        "/api/submit_challenge_rating",
        json={"challenge_key": challenge_key, "score": 4},
    )
    assert closed.status_code == 403

    reopened = instructor.post(
        "/api/start_poll",
        json={"presentation_key": "pres-current"},
    )
    assert reopened.status_code == 200
    assert instructor.get(
        "/api/poll"
    ).get_json()["state"]["challenge_ratings_open"] is True
    assert rater.post(
        "/api/submit_challenge_rating",
        json={"challenge_key": challenge_key, "score": 4},
    ).status_code == 200


def test_cleared_challenge_key_is_not_reused(course_env, monkeypatch):
    clock = {"now": datetime(2026, 8, 14, 11, 0, 0)}
    monkeypatch.setattr(app_module, "_utcnow", lambda: clock["now"])
    _activate_presentation(course_env)
    first_key = _select_test_challenger(course_env)
    instructor = _instructor_client(course_env)
    payload = {
        "presentation_key": "pres-current",
        "challenge_key": first_key,
    }

    closing = instructor.post("/api/clear_challenger", json=payload)
    assert closing.status_code == 409
    assert closing.get_json()["ratings_settling"] is True
    clock["now"] += timedelta(seconds=3)
    cleared = instructor.post("/api/clear_challenger", json=payload)
    assert cleared.status_code == 200

    second_key = _select_test_challenger(course_env)
    assert second_key != first_key
    assert second_key.startswith("pres-current-ch1-")
    assert _state_row(course_env)["challenge_ratings_closed_at"] is None

    duplicate = instructor.post("/api/clear_challenger", json=payload)
    assert duplicate.status_code == 200
    assert duplicate.get_json()["already_cleared"] is True
    assert _state_row(course_env)["challenge_ratings_closed_at"] is None


def test_sixty_challenge_ratings_and_replays_are_idempotent(course_env):
    student_ids = [f"challenge-load-{index:02d}" for index in range(1, 61)]
    scores = {
        student_id: (index % 5) + 1
        for index, student_id in enumerate(student_ids)
    }
    with _connect(course_env) as db:
        for student_id in student_ids:
            db.execute(
                """INSERT INTO students
                   (course_id, student_id, name, pin, team_id)
                   VALUES (?, ?, ?, '1111', ?)""",
                (
                    course_env["course_id"], student_id, student_id,
                    course_env["teams"]["Team 3"],
                ),
            )
        db.commit()
    _activate_presentation(course_env)
    challenge = _seed_live_challenge(course_env)

    clients = []
    for student_id in student_ids:
        client = _student_client(course_env, student_id)
        with client.session_transaction() as flask_session:
            flask_session["activity_session_key"] = SESSION_KEY
            flask_session["last_active_synced_at"] = (
                datetime.utcnow().isoformat()
            )
        clients.append((client, student_id))

    def submit(item):
        client, student_id = item
        response = client.post(
            "/api/submit_challenge_rating",
            json={
                "challenge_key": challenge["challenge_key"],
                "score": scores[student_id],
            },
        )
        return response.status_code

    with ThreadPoolExecutor(max_workers=30) as pool:
        first_statuses = list(pool.map(submit, clients))
    with ThreadPoolExecutor(max_workers=30) as pool:
        replay_statuses = list(pool.map(submit, clients))

    assert first_statuses == [200] * len(student_ids)
    assert replay_statuses == [200] * len(student_ids)
    with _connect(course_env) as db:
        rows = db.execute(
            """SELECT student.student_id, rating.score,
                      rating.session_key, rating.week_num,
                      rating.rater_team_id
               FROM challenge_ratings rating
               JOIN students student ON student.id = rating.rater_id
               WHERE rating.course_id = ? AND rating.challenge_key = ?
               ORDER BY student.student_id""",
            (course_env["course_id"], challenge["challenge_key"]),
        ).fetchall()
    assert len(rows) == len(student_ids)
    assert {
        row["student_id"]: row["score"] for row in rows
    } == scores
    assert {row["session_key"] for row in rows} == {SESSION_KEY}
    assert {row["week_num"] for row in rows} == {1}
    assert {row["rater_team_id"] for row in rows} == {
        course_env["teams"]["Team 3"]
    }


def test_ended_summary_omits_top_students_by_thumbs(course_env):
    with _connect(course_env) as db:
        db.execute("UPDATE students SET name = NULL WHERE student_id = 's1'")
        db.commit()
    _set_state(course_env, phase="discussion")
    assert _student_client(course_env, "s2").post(
        "/api/grade_peer",
        json={"recipient_id": "s1", "selected": True},
    ).status_code == 200
    _set_state(course_env, phase="ended")

    html = _instructor_client(course_env).get(
        f"/instructor/{course_env['slug']}"
    ).get_data(as_text=True)

    assert "Top Students (by thumbs-up)" not in html
    assert "<li>s1: 1 thumbs-up</li>" not in html


def test_select_challenger_cannot_reopen_rating_settlement(
        course_env, monkeypatch):
    clock = {"now": datetime(2026, 8, 14, 13, 0, 0)}
    monkeypatch.setattr(app_module, "_utcnow", lambda: clock["now"])
    with _connect(course_env) as db:
        s5_db_id = db.execute(
            """INSERT INTO students
               (course_id, student_id, name, pin, team_id)
               VALUES (?, 's5', 'Eli', '5555', ?)""",
            (course_env["course_id"], course_env["teams"]["Team 3"]),
        ).lastrowid
        db.commit()
    _activate_presentation(course_env)
    _seed_live_challenge(course_env)
    assert _student_client(course_env, "s5").post(
        "/api/raise_hand",
        json={"presentation_key": "pres-current"},
    ).status_code == 200

    instructor = _instructor_client(course_env)
    closing = instructor.post(
        "/api/next_presentation",
        json={"presentation_key": "pres-current"},
    )
    assert closing.status_code == 409
    original_cutoff = _state_row(course_env)["challenge_ratings_closed_at"]

    clock["now"] += timedelta(seconds=1)
    blocked = instructor.post(
        "/api/select_challenger",
        json={
            "presentation_key": "pres-current",
            "student_id": s5_db_id,
        },
    )
    assert blocked.status_code == 409
    blocked_data = blocked.get_json()
    assert blocked_data["ratings_settling"] is True
    assert blocked_data["ratings_settling_remaining"] == 2
    assert _state_row(course_env)["challenge_ratings_closed_at"] == original_cutoff
    with _connect(course_env) as db:
        assert db.execute("SELECT COUNT(*) FROM challenge_rounds").fetchone()[0] == 1
        assert db.execute(
            "SELECT COUNT(*) FROM challenge_hands WHERE student_id = ?",
            (s5_db_id,),
        ).fetchone()[0] == 1

    clock["now"] += timedelta(seconds=2)
    selected = instructor.post(
        "/api/select_challenger",
        json={
            "presentation_key": "pres-current",
            "student_id": s5_db_id,
        },
    )
    assert selected.status_code == 200
    with _connect(course_env) as db:
        assert db.execute("SELECT COUNT(*) FROM challenge_rounds").fetchone()[0] == 2
        assert db.execute(
            "SELECT COUNT(*) FROM challenge_hands WHERE student_id = ?",
            (s5_db_id,),
        ).fetchone()[0] == 0


def test_select_challenger_remains_available_while_main_poll_is_open(
        course_env, monkeypatch):
    now = datetime(2026, 8, 14, 14, 0, 0)
    monkeypatch.setattr(app_module, "_utcnow", lambda: now)
    _activate_presentation(
        course_env,
        poll_active=1,
        poll_started_at=now.strftime("%Y-%m-%d %H:%M:%S.%f"),
    )
    assert _student_client(course_env, "s4").post(
        "/api/raise_hand",
        json={"presentation_key": "pres-current"},
    ).status_code == 200

    selected = _instructor_client(course_env).post(
        "/api/select_challenger",
        json={
            "presentation_key": "pres-current",
            "student_id": course_env["students"]["s4"],
        },
    )
    assert selected.status_code == 200
    assert selected.get_json()["challenge_ratings_open"] is True


def _seed_current_session_thumb(env):
    with _connect(env) as db:
        db.execute(
            f"""INSERT INTO teammate_thumbs
               (data_version, course_id, session_key, week_num, question_key,
                source_question_key, grader_id, recipient_id)
               VALUES ('{app_module.APP_VERSION}', ?, ?, 1, 'discussion', 'discussion', ?, ?)""",
            (
                env["course_id"],
                SESSION_KEY,
                env["students"]["s1"],
                env["students"]["s2"],
            ),
        )
        db.commit()


def test_saved_activity_freezes_every_roster_structure_route(course_env):
    _seed_current_session_thumb(course_env)
    instructor = _instructor_client(course_env)
    student = _student_client(course_env, "s1")

    no_op = instructor.post(
        "/api/assign_student",
        json=_setup_payload(
            student_id=course_env["students"]["s1"],
            team_id=course_env["teams"]["Team 1"],
        ),
    )
    assert no_op.status_code == 200
    assert no_op.get_json()["roster_version"] == 0
    student_no_op = student.post(
        "/api/join_team",
        json={"team_id": course_env["teams"]["Team 1"]},
    )
    assert student_no_op.status_code == 200
    assert student_no_op.get_json()["roster_version"] == 0

    blocked = [
        instructor.post(
            "/api/set_max_teams",
            json=_setup_payload(max_teams=3),
        ),
        instructor.post(
            "/api/set_max_members",
            json=_setup_payload(max_members=9),
        ),
        instructor.post("/api/random_assign", json=_setup_payload()),
        instructor.post("/api/unassign_all", json=_setup_payload()),
        instructor.post(
            "/api/assign_student",
            json=_setup_payload(
                student_id=course_env["students"]["s1"],
                team_id=course_env["teams"]["Team 2"],
            ),
        ),
        instructor.delete(
            f"/api/remove_student/{course_env['students']['s1']}",
            json=_setup_payload(),
        ),
        instructor.post(
            "/api/add_student",
            json=_setup_payload(
                student_id="s5", name="New Student", pin="5555"
            ),
        ),
        student.post(
            "/api/join_team",
            json={"team_id": course_env["teams"]["Team 2"]},
        ),
    ]
    assert [response.status_code for response in blocked] == [409] * len(blocked)
    assert all(
        "End Session" in response.get_json()["error"]
        for response in blocked
    )

    roster_csv = (
        "student_id,name,pin\n"
        "s1,Alice,1111\n"
        "s2,Bob,2222\n"
        "s3,Unassigned,3333\n"
        "s4,Dana,4444\n"
        "s5,New Student,5555\n"
    ).encode()
    form = {
        "expected_phase": "setup",
        "expected_session_key": str(SESSION_KEY),
        "expected_roster_version": "0",
    }
    preview = instructor.post(
        "/api/upload_roster",
        data={
            **form,
            "file": (io.BytesIO(roster_csv), "roster.csv"),
        },
        content_type="multipart/form-data",
    )
    assert preview.status_code == 200
    confirmed = instructor.post(
        "/api/upload_roster",
        data={
            **form,
            "confirm": "true",
            "preview_token": preview.get_json()["preview_token"],
            "file": (io.BytesIO(roster_csv), "roster.csv"),
        },
        content_type="multipart/form-data",
    )
    assert confirmed.status_code == 409
    assert "End Session" in confirmed.get_json()["error"]

    with _connect(course_env) as db:
        students = {
            row["student_id"]: (row["team_id"], row["is_active"])
            for row in db.execute(
                """SELECT student_id, team_id, is_active FROM students
                   WHERE course_id = ?""",
                (course_env["course_id"],),
            )
        }
        state = db.execute(
            """SELECT max_teams, max_members_per_team, roster_version
               FROM course_state"""
        ).fetchone()
    assert students == {
        "s1": (course_env["teams"]["Team 1"], 1),
        "s2": (course_env["teams"]["Team 1"], 1),
        "s3": (None, 1),
        "s4": (course_env["teams"]["Team 2"], 1),
    }
    assert dict(state) == {
        "max_teams": 4,
        "max_members_per_team": 10,
        "roster_version": 0,
    }


def test_finalized_presentation_freezes_only_its_own_session(course_env):
    history = json.dumps([
        {
            "presentation_key": "pres-zero-ratings",
            "session_key": SESSION_KEY,
            "team_id": course_env["teams"]["Team 1"],
            "team": "Team 1",
        }
    ])
    _set_state(course_env, presentation_history=history)
    instructor = _instructor_client(course_env)

    blocked = instructor.post(
        "/api/assign_student",
        json=_setup_payload(
            student_id=course_env["students"]["s1"],
            team_id=course_env["teams"]["Team 2"],
        ),
    )
    assert blocked.status_code == 409

    with _connect(course_env) as db:
        db.execute(
            "UPDATE course_state SET session_key = ?",
            (SESSION_KEY + 1,),
        )
        db.commit()
    changed = instructor.post(
        "/api/assign_student",
        json=_setup_payload(
            expected_session_key=SESSION_KEY + 1,
            student_id=course_env["students"]["s1"],
            team_id=course_env["teams"]["Team 2"],
        ),
    )
    assert changed.status_code == 200
    assert changed.get_json()["roster_version"] == 1


def test_student_management_reports_course_total_and_committed_version(
        course_env):
    instructor = _instructor_client(course_env)
    listing = instructor.get(
        "/api/students",
        query_string={"search": "s1", "team": ""},
    ).get_json()
    assert listing["total"] == 1
    assert listing["course_total"] == 4

    changed = instructor.post(
        "/api/assign_student",
        json=_setup_payload(
            student_id=course_env["students"]["s3"],
            team_id=course_env["teams"]["Team 2"],
        ),
    )
    assert changed.status_code == 200
    assert changed.get_json()["roster_version"] == 1


def test_concurrent_failed_logins_cannot_pass_the_atomic_limit(course_env):
    route = f"/login/{course_env['slug']}"

    def attempt(_):
        client = app_module.app.test_client()
        return client.post(
            route,
            data={"student_id": "s1", "pin": "0000"},
            environ_base={"REMOTE_ADDR": "203.0.113.50"},
        ).status_code

    with ThreadPoolExecutor(max_workers=8) as executor:
        statuses = list(executor.map(attempt, range(8)))

    assert statuses.count(302) == app_module.LOGIN_FAILURE_LIMIT
    assert statuses.count(429) == 8 - app_module.LOGIN_FAILURE_LIMIT
    with _connect(course_env) as db:
        attempt_row = db.execute(
            """SELECT failed_count, blocked_until FROM login_attempts
               WHERE course_id = ? AND login_type = 'student'
                 AND principal = 's1'""",
            (course_env["course_id"],),
        ).fetchone()
    assert attempt_row["failed_count"] == app_module.LOGIN_FAILURE_LIMIT
    assert attempt_row["blocked_until"] is not None


def test_ended_poll_uses_compact_unchanged_response(course_env):
    _set_state(course_env, phase="ended")
    client = _student_client(course_env)
    first = client.get("/api/poll").get_json()
    assert first["changed"] is True
    assert first["poll_interval"] == 5000

    unchanged = client.get(
        "/api/poll", query_string={"since": first["state_version"]}
    ).get_json()
    assert unchanged == {
        "changed": False,
        "state_version": first["state_version"],
        "poll_interval": 5000,
        "poll_closed": False,
    }
    saved = client.get("/api/my_responses").get_json()
    assert saved["session_key"] == SESSION_KEY


def test_exports_are_phase_gated_and_neutralize_spreadsheet_formulas(
        course_env):
    malicious_name = " =2+2"
    with _connect(course_env) as db:
        db.execute(
            "UPDATE students SET name = ? WHERE id = ?",
            (malicious_name, course_env["students"]["s1"]),
        )
        db.execute(
            f"""INSERT INTO teammate_thumbs
               (data_version, course_id, session_key, week_num, question_key,
                source_question_key, grader_id, recipient_id)
               VALUES ('{app_module.APP_VERSION}', ?, ?, NULL, 'legacy', 'legacy', ?, ?)""",
            (
                course_env["course_id"],
                SESSION_KEY,
                course_env["students"]["s1"],
                course_env["students"]["s2"],
            ),
        )
        db.commit()

    instructor = _instructor_client(course_env)
    workbook_response = instructor.get(f"/export/{course_env['slug']}")
    assert workbook_response.status_code == 200
    from openpyxl import load_workbook

    with zipfile.ZipFile(io.BytesIO(workbook_response.data)) as archive:
        workbook = load_workbook(
            io.BytesIO(archive.read("course_data.xlsx"))
        )
    student_sheet = workbook["Students"]
    exported_name = next(
        row[1] for row in student_sheet.iter_rows(values_only=True)
        if row[0] == "s1"
    )
    assert exported_name == "'" + malicious_name

    legacy = instructor.get(
        f"/export/{course_env['slug']}/legacy-feedback.csv"
    )
    assert legacy.status_code == 200
    rows = list(csv.DictReader(io.StringIO(
        legacy.data.decode("utf-8-sig")
    )))
    assert rows[0]["grader_name"] == "'" + malicious_name

    _set_state(course_env, phase="discussion")
    assert instructor.get(
        f"/export/{course_env['slug']}"
    ).status_code == 409
    assert instructor.get(
        f"/export/{course_env['slug']}/legacy-feedback.csv"
    ).status_code == 409


def test_healthz_checks_active_course_storage(course_env):
    client = app_module.app.test_client()
    healthy = client.get("/healthz")
    assert healthy.status_code == 200
    assert healthy.get_json() == {
        "status": "ok",
        "courses_checked": 1,
        "website_version": f"v{app_module.APP_VERSION}",
        "database_schema_version": f"v{app_module.SCHEMA_VERSION}",
    }

    offline_path = Path(str(course_env["db_path"]) + ".offline")
    Path(course_env["db_path"]).replace(offline_path)
    try:
        unavailable = client.get("/healthz")
        assert unavailable.status_code == 503
        assert unavailable.get_json()["status"] == "unavailable"
    finally:
        offline_path.replace(course_env["db_path"])


def test_existing_database_migration_adds_session_activity_indexes(course_env):
    expected = {
        "idx_ratings_session",
        "idx_challenge_rounds_session",
        "idx_challenge_ratings_session",
    }
    with _connect(course_env) as db:
        for index_name in expected:
            db.execute(f"DROP INDEX {index_name}")
        database._ensure_schema_locked(db)
        db.commit()
        actual = {
            row["name"] for row in db.execute(
                "SELECT name FROM sqlite_master WHERE type = 'index'"
            )
        }
    assert expected.issubset(actual)


def _participant_rows(env):
    with _connect(env) as db:
        return [
            dict(row) for row in db.execute(
                """SELECT session_key, week_num, presentation_key,
                          student_id, student_identifier, student_name,
                          team_id, team_name, data_version
                   FROM presentation_participants
                   WHERE course_id = ?
                   ORDER BY presentation_key, student_identifier""",
                (env["course_id"],),
            )
        ]


def _management_student(env, student_sid):
    response = _instructor_client(env).get(
        "/api/students",
        query_string={"search": student_sid, "per_page": 100},
    )
    assert response.status_code == 200
    matches = [
        student for student in response.get_json()["students"]
        if student["student_id"].casefold() == student_sid.casefold()
    ]
    assert len(matches) == 1
    return matches[0]


def _activate_current_schema_presentation(env, **overrides):
    preflight = _instructor_client(env).get("/api/state")
    assert preflight.status_code == 200
    _activate_presentation(env, **overrides)


def _post_after_rating_grace(client, path, payload, clock):
    settling = client.post(path, json=payload)
    assert settling.status_code == 409
    assert settling.get_json()["ratings_settling"] is True
    clock["now"] += timedelta(
        seconds=app_module.POLL_SUBMISSION_GRACE_SECONDS,
        milliseconds=1,
    )
    return client.post(path, json=payload)


def test_malformed_active_challenger_id_does_not_break_instructor_state(
        course_env):
    instructor = _instructor_client(course_env)
    assert instructor.get("/api/state").status_code == 200
    _activate_presentation(course_env)
    with _connect(course_env) as db:
        db.execute(
            "UPDATE course_state SET active_challenges_json = ?",
            (json.dumps([{
                "challenge_key": "malformed-challenger",
                "challenge_num": 1,
                "challenger_id": "not-an-integer",
                "challenger_name": "Malformed",
                "challenger_team_id": None,
                "challenger_team_name": None,
            }]),),
        )
        db.commit()

    response = instructor.get("/api/state")

    assert response.status_code == 200
    challenge = response.get_json()["active_challenges"][0]
    assert challenge["presentation_count"] == 0
    assert challenge["challenger_count"] == 0


def test_student_state_and_poll_omit_all_participation_counts(course_env):
    _activate_current_schema_presentation(course_env)
    _select_test_challenger(course_env, "s4")

    instructor_state = _instructor_client(course_env).get(
        "/api/state"
    ).get_json()
    assert len(instructor_state["active_challenges"]) == 1
    instructor_challenge = instructor_state["active_challenges"][0]
    assert instructor_challenge["presentation_count"] == 0
    assert instructor_challenge["challenger_count"] == 1

    student = _student_client(course_env, "s4")
    state_response = student.get("/api/state")
    poll_response = student.get("/api/poll")
    assert state_response.status_code == 200
    assert poll_response.status_code == 200

    student_states = (
        state_response.get_json(),
        poll_response.get_json()["state"],
    )
    for state in student_states:
        assert "presentation_count" not in state
        assert "challenger_count" not in state
        assert len(state["active_challenges"]) == 1
        challenge = state["active_challenges"][0]
        assert "presentation_count" not in challenge
        assert "challenger_count" not in challenge


def test_direct_end_session_finalizes_team_and_challenger_once(
        course_env, monkeypatch):
    clock = {"now": datetime(2026, 8, 22, 12, 0, 0)}
    monkeypatch.setattr(app_module, "_utcnow", lambda: clock["now"])
    _activate_current_schema_presentation(course_env)
    challenge_key = _select_test_challenger(course_env, "s4")
    instructor = _instructor_client(course_env)
    assert _participant_rows(course_env) == []
    assert _management_student(course_env, "s4")["challenger_count"] == 1

    end_payload = {
        "phase": "ended",
        "expected_phase": "competition",
        "expected_session_key": SESSION_KEY,
        "presentation_key": "pres-current",
        "confirm_end_session": True,
    }
    ended = _post_after_rating_grace(
        instructor, "/api/set_phase", end_payload, clock
    )

    assert ended.status_code == 200
    participants = _participant_rows(course_env)
    assert len(participants) == 2
    assert {
        (row["student_identifier"], row["presentation_key"])
        for row in participants
    } == {("s1", "pres-current"), ("s2", "pres-current")}
    with _connect(course_env) as db:
        challenge_rows = [
            dict(row) for row in db.execute(
                """SELECT challenge_key, presentation_key, challenger_id
                   FROM challenge_rounds WHERE course_id = ?""",
                (course_env["course_id"],),
            )
        ]
    assert challenge_rows == [{
        "challenge_key": challenge_key,
        "presentation_key": "pres-current",
        "challenger_id": course_env["students"]["s4"],
    }]
    assert _management_student(course_env, "s1")["presentation_count"] == 1
    assert _management_student(course_env, "s2")["presentation_count"] == 1
    assert _management_student(course_env, "s4")["challenger_count"] == 1

    repeated = instructor.post("/api/set_phase", json=end_payload)
    assert repeated.status_code == 409
    assert len(_participant_rows(course_env)) == 2
    with _connect(course_env) as db:
        assert db.execute(
            """SELECT COUNT(*) FROM challenge_rounds
               WHERE course_id = ? AND challenge_key = ?""",
            (course_env["course_id"], challenge_key),
        ).fetchone()[0] == 1
    assert _management_student(course_env, "s1")["presentation_count"] == 1
    assert _management_student(course_env, "s4")["challenger_count"] == 1


def test_finalizing_zero_rating_presentation_snapshots_members_once(
        course_env):
    _activate_current_schema_presentation(course_env)
    instructor = _instructor_client(course_env)

    finished = instructor.post(
        "/api/next_presentation",
        json={"presentation_key": "pres-current"},
    )

    assert finished.status_code == 200
    assert _participant_rows(course_env) == [
        {
            "session_key": SESSION_KEY,
            "week_num": 1,
            "presentation_key": "pres-current",
            "student_id": course_env["students"]["s1"],
            "student_identifier": "s1",
            "student_name": "Alice",
            "team_id": course_env["teams"]["Team 1"],
            "team_name": "Team 1",
            "data_version": app_module.APP_VERSION,
        },
        {
            "session_key": SESSION_KEY,
            "week_num": 1,
            "presentation_key": "pres-current",
            "student_id": course_env["students"]["s2"],
            "student_identifier": "s2",
            "student_name": "Bob",
            "team_id": course_env["teams"]["Team 1"],
            "team_name": "Team 1",
            "data_version": app_module.APP_VERSION,
        },
    ]
    assert _management_student(course_env, "s1")["presentation_count"] == 1
    assert _management_student(course_env, "s2")["presentation_count"] == 1
    assert _management_student(course_env, "s4")["presentation_count"] == 0
    sorted_presenters = instructor.get(
        "/api/students",
        query_string={
            "sort": "presentation_count", "order": "desc", "per_page": 100,
        },
    ).get_json()["students"]
    assert [student["student_id"] for student in sorted_presenters] == [
        "s1", "s2", "s3", "s4",
    ]

    repeated = instructor.post(
        "/api/next_presentation",
        json={
            "presentation_key": "pres-current",
            "expected_phase": "competition",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert repeated.status_code == 200
    assert repeated.get_json()["already_finished"] is True
    assert len(_participant_rows(course_env)) == 2

    instructor_teams = instructor.get("/api/teams").get_json()
    team_1 = next(team for team in instructor_teams if team["name"] == "Team 1")
    counts = {
        member["student_id"]: (
            member["presentation_count"], member["challenger_count"]
        )
        for member in team_1["members"]
    }
    assert counts == {"s1": (1, 0), "s2": (1, 0)}

    student_teams = _student_client(course_env, "s4").get(
        "/api/teams"
    ).get_json()
    assert all(
        "presentation_count" not in member and "challenger_count" not in member
        for team in student_teams for member in team["members"]
    )


def test_cancelling_presentation_does_not_create_participant_history(course_env):
    _activate_current_schema_presentation(course_env)

    cancelled = _instructor_client(course_env).post(
        "/api/cancel_presentation",
        json={"presentation_key": "pres-current"},
    )

    assert cancelled.status_code == 200
    assert _participant_rows(course_env) == []
    assert _management_student(course_env, "s1")["presentation_count"] == 0


def test_finish_then_cancel_reports_conflict_and_keeps_participation(course_env):
    _activate_current_schema_presentation(course_env)
    instructor = _instructor_client(course_env)
    payload = {
        "presentation_key": "pres-current",
        "expected_phase": "competition",
        "expected_session_key": SESSION_KEY,
    }

    finished = instructor.post("/api/next_presentation", json=payload)
    assert finished.status_code == 200

    missing_state_guard = instructor.post(
        "/api/cancel_presentation",
        json={"presentation_key": "pres-current"},
    )
    assert missing_state_guard.status_code == 400
    assert "expected phase" in missing_state_guard.get_json()["error"].lower()

    conflict = instructor.post("/api/cancel_presentation", json=payload)
    assert conflict.status_code == 409
    assert conflict.get_json()["outcome"] == "finished"
    assert "already finished" in conflict.get_json()["error"].lower()
    assert len(json.loads(_state_row(course_env)["presentation_history"])) == 1
    assert len(_participant_rows(course_env)) == 2


def test_cancel_then_finish_reports_conflict_without_participation(course_env):
    _activate_current_schema_presentation(course_env)
    instructor = _instructor_client(course_env)
    payload = {
        "presentation_key": "pres-current",
        "expected_phase": "competition",
        "expected_session_key": SESSION_KEY,
    }

    cancelled = instructor.post("/api/cancel_presentation", json=payload)
    assert cancelled.status_code == 200

    repeated = instructor.post("/api/cancel_presentation", json=payload)
    assert repeated.status_code == 200
    assert repeated.get_json()["already_canceled"] is True

    conflict = instructor.post("/api/next_presentation", json=payload)
    assert conflict.status_code == 409
    assert conflict.get_json()["outcome"] == "canceled"
    assert "canceled" in conflict.get_json()["error"].lower()
    assert json.loads(_state_row(course_env)["presentation_history"]) == []
    assert _participant_rows(course_env) == []


@pytest.mark.parametrize(
    "route", ("/api/next_presentation", "/api/cancel_presentation")
)
def test_inactive_presentation_key_from_prior_session_is_not_idempotent(
        course_env, route):
    _activate_current_schema_presentation(course_env)
    instructor = _instructor_client(course_env)
    assert instructor.post(
        "/api/next_presentation",
        json={"presentation_key": "pres-current"},
    ).status_code == 200
    _set_state(course_env, session_key=SESSION_KEY + 1)

    stale = instructor.post(
        route,
        json={
            "presentation_key": "pres-current",
            "expected_phase": "competition",
            "expected_session_key": SESSION_KEY + 1,
        },
    )

    assert stale.status_code == 409
    assert stale.get_json()["outcome"] == "stale_session"
    assert "earlier session" in stale.get_json()["error"].lower()
    assert len(_participant_rows(course_env)) == 2


def test_challenger_counts_follow_retained_rounds_across_clear_and_cancel(
        course_env, monkeypatch):
    clock = {"now": datetime(2026, 8, 21, 12, 0, 0)}
    monkeypatch.setattr(app_module, "_utcnow", lambda: clock["now"])
    _activate_current_schema_presentation(course_env)
    instructor = _instructor_client(course_env)

    first_key = _select_test_challenger(course_env, "s4")
    assert _management_student(course_env, "s4")["challenger_count"] == 1
    sorted_challengers = instructor.get(
        "/api/students",
        query_string={
            "sort": "challenger_count", "order": "desc", "per_page": 100,
        },
    ).get_json()["students"]
    assert [student["student_id"] for student in sorted_challengers] == [
        "s4", "s1", "s2", "s3",
    ]

    cleared = _post_after_rating_grace(
        instructor,
        "/api/clear_challenger",
        {
            "presentation_key": "pres-current",
            "challenge_key": first_key,
        },
        clock,
    )
    assert cleared.status_code == 200
    assert _management_student(course_env, "s4")["challenger_count"] == 0

    second_key = _select_test_challenger(course_env, "s4")
    raised_again = _student_client(course_env, "s4").post(
        "/api/raise_hand", json={"presentation_key": "pres-current"}
    )
    assert raised_again.status_code == 200
    state = instructor.get("/api/state").get_json()
    hand = next(
        item for item in state["challenge_hands"]
        if item["student_id"] == course_env["students"]["s4"]
    )
    assert hand["presentation_count"] == 0
    assert hand["challenger_count"] == 1

    selected_again = instructor.post(
        "/api/select_challenger",
        json={
            "presentation_key": "pres-current",
            "student_id": course_env["students"]["s4"],
        },
    )
    assert selected_again.status_code == 200
    third_key = selected_again.get_json()["challenge_key"]
    assert third_key not in {first_key, second_key}
    assert _management_student(course_env, "s4")["challenger_count"] == 2

    active = instructor.get("/api/state").get_json()["active_challenges"]
    assert {item["challenger_count"] for item in active} == {2}

    cleared_one = _post_after_rating_grace(
        instructor,
        "/api/clear_challenger",
        {
            "presentation_key": "pres-current",
            "challenge_key": second_key,
        },
        clock,
    )
    assert cleared_one.status_code == 200
    assert _management_student(course_env, "s4")["challenger_count"] == 1

    cancelled = _post_after_rating_grace(
        instructor,
        "/api/cancel_presentation",
        {"presentation_key": "pres-current"},
        clock,
    )
    assert cancelled.status_code == 200
    assert _management_student(course_env, "s4")["challenger_count"] == 0


def test_participation_counts_survive_new_session_archive_and_reactivation(
        course_env):
    _activate_current_schema_presentation(course_env)
    instructor = _instructor_client(course_env)
    assert instructor.post(
        "/api/next_presentation",
        json={"presentation_key": "pres-current"},
    ).status_code == 200

    ended = instructor.post(
        "/api/set_phase",
        json={
            "phase": "ended",
            "expected_phase": "competition",
            "expected_session_key": SESSION_KEY,
            "confirm_end_session": True,
        },
    )
    assert ended.status_code == 200
    setup = instructor.post(
        "/api/set_phase",
        json={
            "phase": "setup",
            "expected_phase": "ended",
            "expected_session_key": SESSION_KEY,
        },
    )
    assert setup.status_code == 200
    assert setup.get_json()["session_key"] == SESSION_KEY + 1

    original_id = course_env["students"]["s1"]
    archived = instructor.delete(
        f"/api/remove_student/{original_id}",
        json={
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY + 1,
            "expected_roster_version": 0,
        },
    )
    assert archived.status_code == 200
    restored = instructor.post(
        "/api/add_student",
        json={
            "expected_phase": "setup",
            "expected_session_key": SESSION_KEY + 1,
            "expected_roster_version": 1,
            "student_id": "S1",
            "name": "Alice Restored",
            "pin": "5555",
        },
    )
    assert restored.status_code == 200
    assert restored.get_json()["reactivated"] is True
    student = _management_student(course_env, "s1")
    assert student["id"] == original_id
    assert student["presentation_count"] == 1
    assert student["challenger_count"] == 0


def test_reset_erases_counts_after_preserving_them_in_recovery_backup(
        course_env, monkeypatch):
    clock = {"now": datetime(2026, 8, 21, 13, 0, 0)}
    monkeypatch.setattr(app_module, "_utcnow", lambda: clock["now"])
    _activate_current_schema_presentation(course_env)
    _select_test_challenger(course_env, "s4")
    instructor = _instructor_client(course_env)
    finalized = _post_after_rating_grace(
        instructor,
        "/api/next_presentation",
        {"presentation_key": "pres-current"},
        clock,
    )
    assert finalized.status_code == 200
    assert _management_student(course_env, "s1")["presentation_count"] == 1
    assert _management_student(course_env, "s4")["challenger_count"] == 1

    ended = instructor.post(
        "/api/set_phase",
        json={
            "phase": "ended",
            "expected_phase": "competition",
            "expected_session_key": SESSION_KEY,
            "confirm_end_session": True,
        },
    )
    assert ended.status_code == 200
    reset = instructor.post(
        "/api/reset_data", json=_reset_payload(course_env, "ended")
    )

    assert reset.status_code == 200
    backup_path = (
        Path(course_env["data_dir"])
        / course_env["slug"]
        / "reset-backups"
        / reset.get_json()["backup"]
    )
    with sqlite3.connect(backup_path) as backup:
        assert backup.execute(
            "SELECT COUNT(*) FROM presentation_participants"
        ).fetchone()[0] == 2
        assert backup.execute(
            "SELECT COUNT(*) FROM challenge_rounds"
        ).fetchone()[0] == 1

    assert _participant_rows(course_env) == []
    assert _management_student(course_env, "s1")["presentation_count"] == 0
    assert _management_student(course_env, "s4")["challenger_count"] == 0
